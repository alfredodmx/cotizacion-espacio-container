# v2.1 - titulos tarjetas abajo izquierda
import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
import io
from datetime import datetime, timedelta
import random
import re
import requests
import json
import base64
import uuid
from supabase import create_client
import urllib.parse

st.set_page_config(layout="wide", page_title="Cotizador PRO", page_icon="📊")

# =========================================================
# CONFIGURACIÓN SUPABASE
# =========================================================
SUPABASE_URL = "https://rpjktwxitceqylexcaqw.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJwamt0d3hpdGNlcXlsZXhjYXF3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzI4MzUyMzYsImV4cCI6MjA4ODQxMTIzNn0.LoZN1W7X1pjVgNLFyVRfzQ8iHFp5JN2qw2Egu5yJq0E"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJwamt0d3hpdGNlcXlsZXhjYXF3Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MjgzNTIzNiwiZXhwIjoyMDg4NDExMjM2fQ.HcXL2zeYrL6GONevt3CDDQmRZtanXymRH9PJIdOKLZk"
# ── Roles del sistema ──────────────────────────────────
# root  → acceso total, puede eliminar cualquier cuenta, ve 🛡️ SISTEMA
# admin → ve todo, puede crear ejecutivos y admins, NO ve 🛡️ SISTEMA
# ejecutivo → solo sus cotizaciones
# -----------------------------------------------------------
ROOTS = ["alfredodmx@gmail.com"]   # root fijo — agregar más si necesario

def get_rol(email, user_metadata=None):
    """Retorna el rol del usuario: 'root', 'admin' o 'ejecutivo'."""
    email_l = (email or "").lower()
    if email_l in [r.lower() for r in ROOTS]:
        return "root"
    meta_rol = (user_metadata or {}).get("rol", "ejecutivo")
    if meta_rol in ("root",):
        return "root"
    if meta_rol in ("admin", "administrador"):
        return "admin"
    return "ejecutivo"

def es_rol_superior(email, user_metadata=None):
    """True si es root o admin (acceso amplio)."""
    return get_rol(email, user_metadata) in ("root", "admin")

# API key de Anthropic para el Visor 3D (leer de secrets o env)
import os as _os_init
ANTHROPIC_API_KEY = (
    _os_init.environ.get("ANTHROPIC_API_KEY", "") or
    (st.secrets.get("ANTHROPIC_API_KEY", "") if hasattr(st, "secrets") else "")
)

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
supabase_admin = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)

# =========================================================
# SISTEMA DE AUTENTICACIÓN
# =========================================================
def login_usuario(email, password):
    """Inicia sesión con email y contraseña."""
    try:
        res = supabase.auth.sign_in_with_password({"email": email, "password": password})
        return res.user, None
    except Exception as e:
        return None, str(e)

def logout_usuario():
    """Cierra sesión — limpia TODO el session_state."""
    try:
        supabase.auth.sign_out()
    except:
        pass
    # Guardar solo la lista de claves a preservar (ninguna en este caso)
    # Limpiar ABSOLUTAMENTE todo el session_state
    _keys_to_clear = list(st.session_state.keys())
    for k in _keys_to_clear:
        try:
            del st.session_state[k]
        except:
            pass
    # Limpiar query params
    try:
        st.query_params.clear()
    except:
        pass

def crear_usuario_ejecutivo(email, password, nombre):
    """Crea un nuevo usuario ejecutivo (requiere service role)."""
    try:
        res = supabase_admin.auth.admin.create_user({
            "email": email,
            "password": password,
            "email_confirm": True,
            "user_metadata": {"nombre": nombre, "rol": "ejecutivo"}
        })
        return res.user, None
    except Exception as e:
        return None, str(e)

# =========================================================
# SISTEMA DE NOTIFICACIONES TELEGRAM
# =========================================================
TELEGRAM_BOT_TOKEN_DEFAULT = "8639597343:AAG-E3HJVmDGbbMI5oniiivLitlphTDJkCU"

def _get_notif_config(clave, default=""):
    """Lee un valor de configuración de notificaciones desde Supabase."""
    try:
        r = supabase_admin.table('notificaciones_config').select('valor').eq('clave', clave).execute()
        if r.data:
            return r.data[0]['valor'] or default
    except:
        pass
    return default

def _set_notif_config(clave, valor):
    """Guarda un valor de configuración en Supabase."""
    try:
        supabase_admin.table('notificaciones_config').upsert(
            {'clave': clave, 'valor': valor, 'updated_at': 'now()'},
            on_conflict='clave'
        ).execute()
        return True
    except:
        return False

def _enviar_telegram(chat_id, mensaje, token=None):
    """Envía un mensaje de Telegram a un chat_id."""
    if not chat_id or not str(chat_id).strip():
        return False
    try:
        _token = token or _get_notif_config('bot_token', TELEGRAM_BOT_TOKEN_DEFAULT)
        if not _token:
            return False
        _chat = str(chat_id).strip()
        _url = f"https://api.telegram.org/bot{_token}/sendMessage"
        _payload = json.dumps({
            'chat_id': _chat,
            'text': mensaje,
            'parse_mode': 'Markdown'
        }).encode('utf-8')
        try:
            # Intentar con requests primero
            import requests as _req
            _resp = _req.post(_url, data=_payload,
                              headers={'Content-Type': 'application/json'}, timeout=10)
            return _resp.status_code == 200
        except ImportError:
            # Fallback a urllib
            import urllib.request as _ur
            _req2 = _ur.Request(_url, data=_payload,
                                headers={'Content-Type': 'application/json'})
            _resp2 = _ur.urlopen(_req2, timeout=10)
            return _resp2.status == 200
    except Exception as _e:
        print(f"_enviar_telegram error: {_e}")
        return False

def _get_contactos_notif():
    """Retorna dict {user_email: chat_id} de contactos configurados."""
    try:
        raw = _get_notif_config('contactos_json', '{}')
        import json as _j
        return _j.loads(raw)
    except:
        return {}

def _get_observadores_notif():
    """Retorna lista de dicts {nombre, chat_id} de observadores."""
    try:
        raw = _get_notif_config('observadores_json', '[]')
        import json as _j
        return _j.loads(raw)
    except:
        return []

def notificar_nueva_cotizacion(ep, ejecutivo_nombre, cliente_nombre, monto, estado, ejecutivo_email):
    """Notifica a supervisores/admins/observadores cuando se guarda una cotización."""
    import traceback as _tb2
    try:
        # Obtener plantilla
        plantilla = _get_notif_config('msg_nueva_cotizacion', '🆕 *Nueva cotización para revisar*\n\n*{ep}* · {ejecutivo}\nCliente: {cliente} · Monto: *{monto}*\nEstado: {estado}')
        _fmt_monto = f"${monto:,.0f}".replace(",", ".") if monto else "$0"
        msg = plantilla.replace('{ep}', str(ep)).replace('{ejecutivo}', str(ejecutivo_nombre))                       .replace('{cliente}', str(cliente_nombre)).replace('{monto}', _fmt_monto)                       .replace('{estado}', str(estado))
        contactos = _get_contactos_notif()
        enviados = 0
        # Enviar a admins/root (no al propio ejecutivo)
        try:
            usuarios = listar_usuarios_ejecutivos()
            for u in usuarios:
                if u.get('rol') in ('admin', 'root') and u.get('email','').lower() != ejecutivo_email.lower():
                    chat_id = contactos.get(u['email'].lower(), '')
                    if chat_id and _enviar_telegram(chat_id, msg):
                        enviados += 1
        except:
            pass
        # Enviar a roots fijos
        for root_email in ROOTS:
            if root_email.lower() != ejecutivo_email.lower():
                chat_id = contactos.get(root_email.lower(), '')
                if chat_id and _enviar_telegram(chat_id, msg):
                    enviados += 1
        # Enviar a observadores
        for obs in _get_observadores_notif():
            if obs.get('chat_id') and _enviar_telegram(obs['chat_id'], msg):
                enviados += 1
        # Enviar a grupo si está configurado
        grupo_id = _get_notif_config('grupo_chat_id', '')
        grupo_filtro = _get_notif_config('grupo_filtro', 'todas')
        if grupo_id and grupo_filtro in ('todas', 'solo_nuevas'):
            _enviar_telegram(grupo_id, msg)
    except:
        pass

def notificar_cotizacion_autorizada(ep, cliente_nombre, margen, ejecutivo_email, ejecutivo_nombre, supervisor_nombre='', monto=0):
    """Notifica al ejecutivo cuando su cotización es autorizada."""
    import traceback as _tb
    try:
        _token = _get_notif_config('bot_token', TELEGRAM_BOT_TOKEN_DEFAULT)
        _margen_str = f"{float(margen):.1f}" if margen else "0"
        _sup = supervisor_nombre.upper() if supervisor_nombre else 'EL SUPERVISOR'
        _plantilla = _get_notif_config('msg_autorizada',
            "✅ *¡PRESUPUESTO AUTORIZADO!*\n\n📋 *{ep}* · {cliente}\n💰 Margen aplicado: *{margen}%*\n👤 Autorizado por: *{supervisor}*\n\nYa puedes presentárselo a tu cliente 🎉")
        _fmt_monto = f"${float(monto):,.0f}".replace(",",".") if monto else "$0"
        msg = (_plantilla
            .replace('{ep}', str(ep))
            .replace('{cliente}', str(cliente_nombre))
            .replace('{margen}', _margen_str)
            .replace('{monto}', _fmt_monto)
            .replace('{ejecutivo}', str(ejecutivo_nombre))
            .replace('{supervisor}', _sup)
        )
        contactos = _get_contactos_notif()
        enviados = 0
        # Enviar al ejecutivo si tiene chat_id
        if ejecutivo_email:
            chat_id = contactos.get(ejecutivo_email.lower(), '')
            if chat_id:
                ok = _enviar_telegram(chat_id, msg, _token)
                if ok: enviados += 1
        # Enviar a observadores
        for obs in _get_observadores_notif():
            if obs.get('chat_id'):
                if _enviar_telegram(obs['chat_id'], msg, _token):
                    enviados += 1
        # Grupo
        grupo_id = _get_notif_config('grupo_chat_id', '')
        grupo_filtro = _get_notif_config('grupo_filtro', 'todas')
        if grupo_id and grupo_filtro in ('todas', 'solo_autorizaciones'):
            _enviar_telegram(grupo_id, msg, _token)
        return enviados
    except Exception as _e:
        print(f"ERROR notificar_autorizada: {_e}\n{_tb.format_exc()}")
        return 0

def notificar_margen_removido(ep, cliente_nombre, ejecutivo_email):
    """Notifica al ejecutivo cuando se remueve el margen."""
    try:
        plantilla = _get_notif_config('msg_margen_removido', '↩️ La cotización *{ep}* volvió a estado borrador.\nEl supervisor realizó cambios. Revisa el sistema.')
        msg = plantilla.replace('{ep}', str(ep)).replace('{cliente}', str(cliente_nombre))
        contactos = _get_contactos_notif()
        chat_id = contactos.get(ejecutivo_email.lower(), '')
        if chat_id:
            _enviar_telegram(chat_id, msg)
    except:
        pass

def cambiar_rol_usuario(user_id, nuevo_rol):
    """Cambia el rol de un usuario en sus metadatos."""
    try:
        supabase_admin.auth.admin.update_user_by_id(
            user_id, {"user_metadata": {"rol": nuevo_rol}}
        )
        return True, None
    except Exception as e:
        return False, str(e)

def cambiar_password_propio(nueva_password):
    """El usuario cambia su propia contraseña (requiere sesión activa)."""
    try:
        supabase.auth.update_user({"password": nueva_password})
        return True, None
    except Exception as e:
        return False, str(e)

def resetear_password_admin(user_id, nueva_password):
    """Admin/Root resetea la contraseña de otro usuario (requiere service role)."""
    try:
        supabase_admin.auth.admin.update_user_by_id(user_id, {"password": nueva_password})
        return True, None
    except Exception as e:
        return False, str(e)

def listar_usuarios_ejecutivos():
    """Lista todos los usuarios excepto supervisores fijos."""
    try:
        res = supabase_admin.auth.admin.list_users()
        users = []
        for u in res:
            email = u.email or ""
            if email.lower() in [s.lower() for s in ROOTS]:
                continue
            meta = u.user_metadata or {}
            nombre = meta.get("nombre", email)
            rol = meta.get("rol", "ejecutivo")
            # Compatibilidad con distintas versiones de supabase-py
            try:
                _activo = not getattr(u, 'banned_until', None)
            except:
                _activo = True
            users.append({
                "id": str(u.id),
                "email": email,
                "nombre": nombre,
                "rol": rol,
                "telefono": meta.get("telefono", "") or "",
                "created_at": str(u.created_at)[:10] if u.created_at else "",
                "activo": _activo
            })
        return users
    except Exception as e:
        st.session_state['_usuarios_list_error'] = str(e)
        return []

def eliminar_usuario_ejecutivo(user_id):
    """Elimina un usuario ejecutivo."""
    try:
        supabase_admin.auth.admin.delete_user(user_id)
        return True, None
    except Exception as e:
        return False, str(e)

def verificar_conexion_supabase():
    # Solo verifica una vez por sesión, no en cada render
    if st.session_state.get('_supabase_ok'):
        return True
    try:
        supabase.table('cotizaciones').select('numero').limit(1).execute()
        st.session_state['_supabase_ok'] = True
        return True
    except Exception as e:
        st.error(f"❌ Error conectando a Supabase: {e}")
        return False

verificar_conexion_supabase()

# ── Inicializar variables de auth ANTES del check de login ──
if 'auth_user'    not in st.session_state: st.session_state.auth_user    = None
if 'auth_email'   not in st.session_state: st.session_state.auth_email   = ""
if 'auth_nombre'  not in st.session_state: st.session_state.auth_nombre  = ""
if 'es_supervisor'not in st.session_state: st.session_state.es_supervisor= False
if 'es_root'      not in st.session_state: st.session_state.es_root      = False
if 'rol_usuario'  not in st.session_state: st.session_state.rol_usuario  = "ejecutivo"
if 'modo_admin'   not in st.session_state: st.session_state.modo_admin   = False

# ── Recuperar sesión desde query param _sess (sin localStorage) ──
_sess_token = st.query_params.get("_sess")
if not st.session_state.auth_user and _sess_token:
    try:
        _sess_user = supabase.auth.get_user(_sess_token)
        if _sess_user and _sess_user.user:
            _u = _sess_user.user
            _meta = _u.user_metadata or {}
            _rol = get_rol(_u.email, _meta)
            st.session_state.auth_user    = str(_u.id)
            st.session_state.auth_email   = _u.email or ""
            st.session_state.auth_nombre  = _meta.get("nombre", _u.email or "")
            st.session_state.rol_usuario  = _rol
            st.session_state.es_supervisor= _rol in ("root", "admin")
            st.session_state.es_root      = _rol == "root"
            st.session_state.modo_admin   = _rol in ("root", "admin")
            st.query_params.clear()
            st.rerun()
    except:
        st.query_params.clear()

# =========================================================
# PANTALLA DE LOGIN — bloquea la app si no hay sesión
# =========================================================
if not st.session_state.auth_user:
    # Cargar logo3.png
    import base64 as _b64l, os as _osl
    _logo_html = ""
    for _lpath in ["logo3.png", "assets/logo3.png", "images/logo3.png"]:
        if _osl.path.exists(_lpath):
            with open(_lpath, "rb") as _lf:
                _logo_b64 = _b64l.b64encode(_lf.read()).decode()
            _logo_html = f'<img src="data:image/png;base64,{_logo_b64}" style="width:750px;max-width:100%;display:block;margin:0 auto 8px;filter:drop-shadow(0 2px 16px rgba(255,255,255,0.08));">'
            break

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@300;400;700;900&display=swap');

    [data-testid="stAppViewContainer"] {
        background: #0d0d0d !important;
    }
    [data-testid="stHeader"]      { display:none !important; }
    [data-testid="stToolbar"]     { display:none !important; }
    [data-testid="stDecoration"]  { display:none !important; }
    .stDeployButton { display:none !important; }
    #MainMenu { display:none !important; }
    footer    { display:none !important; }

    /* Línea blanca superior */
    [data-testid="stAppViewContainer"]::before {
        content:''; position:fixed; top:0; left:0; right:0; height:1px;
        background: rgba(255,255,255,0.15);
        z-index:9999;
    }

    /* Divisor */
    .login-divider {
        height:1px; margin:20px 0;
        background: rgba(255,255,255,0.08);
    }

    /* Labels */
    div[data-testid="stTextInput"] label,
    div[data-testid="stTextInput"] label p {
        color: rgba(255,255,255,0.4) !important;
        font-size: 0.68rem !important;
        letter-spacing: 0.14em !important;
        text-transform: uppercase !important;
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 400 !important;
    }
    /* Contenedor input — negro glossy */
    div[data-testid="stTextInput"] > div,
    div[data-testid="stTextInput"] > div > div,
    div[data-testid="stTextInput"] > div > div > div {
        background: #0a0a0a !important;
        border-color: rgba(255,255,255,0.15) !important;
        border-radius: 6px !important;
    }
    div[data-testid="stTextInput"] > div > div {
        background: linear-gradient(180deg, #1a1a1a 0%, #0a0a0a 100%) !important;
        border: 1px solid rgba(255,255,255,0.18) !important;
        border-radius: 6px !important;
        box-shadow:
            inset 0 2px 4px rgba(0,0,0,0.8),
            inset 0 1px 0 rgba(255,255,255,0.05),
            0 1px 0 rgba(255,255,255,0.08) !important;
    }
    div[data-testid="stTextInput"] input {
        background: transparent !important;
        color: #ffffff !important;
        font-size: 0.93rem !important;
        font-family: 'Montserrat', sans-serif !important;
        caret-color: #ffffff !important;
    }
    div[data-testid="stTextInput"] input::placeholder {
        color: rgba(255,255,255,0.18) !important;
    }
    div[data-testid="stTextInput"] > div > div:focus-within {
        border-color: rgba(255,255,255,0.45) !important;
        box-shadow:
            inset 0 2px 4px rgba(0,0,0,0.8),
            0 0 0 1px rgba(255,255,255,0.12),
            0 0 12px rgba(255,255,255,0.05) !important;
    }

    /* Botón negro glossy con texto blanco */
    div[data-testid="stButton"] > button,
    div[data-testid="stButton"] > button[kind="primary"] {
        background: linear-gradient(180deg, #2a2a2a 0%, #111111 60%, #1a1a1a 100%) !important;
        color: #ffffff !important;
        border: 1px solid rgba(255,255,255,0.2) !important;
        border-top: 1px solid rgba(255,255,255,0.3) !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        font-size: 0.75rem !important;
        letter-spacing: 0.22em !important;
        text-transform: uppercase !important;
        font-family: 'Montserrat', sans-serif !important;
        padding: 0.75rem !important;
        box-shadow:
            inset 0 1px 0 rgba(255,255,255,0.1),
            inset 0 -1px 0 rgba(0,0,0,0.5),
            0 4px 16px rgba(0,0,0,0.6) !important;
        transition: all 0.2s ease !important;
    }
    div[data-testid="stButton"] > button:hover,
    div[data-testid="stButton"] > button[kind="primary"]:hover {
        background: linear-gradient(180deg, #333333 0%, #1a1a1a 60%, #222222 100%) !important;
        border-color: rgba(255,255,255,0.35) !important;
        box-shadow:
            inset 0 1px 0 rgba(255,255,255,0.15),
            0 6px 24px rgba(0,0,0,0.7) !important;
        transform: translateY(-1px) !important;
        color: #ffffff !important;
    }
    div[data-testid="stButton"] > button:active,
    div[data-testid="stButton"] > button[kind="primary"]:active {
        transform: translateY(0) !important;
        box-shadow: inset 0 2px 6px rgba(0,0,0,0.8) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Logo
    if _logo_html:
        st.markdown(f"""
        <style>
        .login-logo-wrap img {{ width:750px !important; max-width:100% !important; display:block !important; margin:0 auto !important; }}
        </style>
        <div class="login-logo-wrap">{_logo_html}</div>
        """, unsafe_allow_html=True)
    else:
        st.markdown('<div style="text-align:center;margin-bottom:20px;color:white;font-size:3rem;">🧊</div>', unsafe_allow_html=True)

    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
    st.markdown('<div class="login-divider"></div>', unsafe_allow_html=True)
    st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)

    # Inputs centrados en columna angosta
    _lc, _mc, _rc = st.columns([1.5, 1, 1.5])
    with _mc:
        _email_in = st.text_input("Correo electrónico", key="login_email", placeholder="usuario@empresa.cl")
        _pass_in  = st.text_input("Contraseña", type="password", key="login_pass", placeholder="••••••••")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        if st.button("⚡ Ingresar al sistema", use_container_width=True, type="primary", key="btn_login"):
            if not _email_in or not _pass_in:
                st.error("Completa correo y contraseña.")
            else:
                with st.spinner("Verificando..."):
                    user, err = login_usuario(_email_in.strip(), _pass_in)
                if user:
                    st.session_state.auth_user   = str(user.id)
                    st.session_state.auth_email  = user.email or _email_in.strip()
                    meta = user.user_metadata or {}
                    st.session_state.auth_nombre = meta.get("nombre", user.email or "")
                    _meta = user.user_metadata or {}
                    _rol_login = get_rol(user.email, _meta)
                    st.session_state.rol_usuario   = _rol_login
                    st.session_state.es_supervisor = _rol_login in ("root", "admin")
                    st.session_state.es_root       = _rol_login == "root"
                    if st.session_state.es_supervisor:
                        st.session_state.modo_admin = True
                    st.session_state.pop('resultados_busqueda', None)
                    st.session_state.pop('_usuarios_cache', None)
                    # Sesión manejada via query params — sin localStorage
                    st.rerun()
                else:
                    if "Invalid login" in str(err) or "invalid_credentials" in str(err):
                        st.error("❌ Correo o contraseña incorrectos.")
                    elif "Email not confirmed" in str(err):
                        st.error("❌ Cuenta no confirmada. Contacta al administrador.")
                    else:
                        st.error(f"❌ {err}")

        st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
        st.markdown("""
        <div style="text-align:center;color:rgba(255,255,255,0.15);font-size:0.65rem;
                    letter-spacing:0.2em;text-transform:uppercase;font-family:'Montserrat',sans-serif;
                    font-weight:300;margin-top:8px;">
            Sistema de gestión · Uso interno
        </div>""", unsafe_allow_html=True)
        # (login-card ya cerrado en el bloque HTML superior)

    st.stop()

# =========================================================
# AUTO-PROVISIONAMIENTO DE EJECUTIVOS
# =========================================================
_EJECUTIVOS_INICIALES = [
    {"nombre": "BERNARD BUSTAMANTE",  "email": "balday@espaciocontainerhouse.cl",    "telefono": "+56956786366", "pass": "ECH2024!BB",  "rol": "ejecutivo"},
    {"nombre": "ANDREA OSORIO",       "email": "aosorio@espaciocontainerhouse.cl",   "telefono": "+56927619483", "pass": "ECH2024!AO",  "rol": "ejecutivo"},
    {"nombre": "REBECA CALDERON",     "email": "rcalderon@espaciocontainerhouse.cl", "telefono": "+56955286708", "pass": "ECH2024!RC",  "rol": "ejecutivo"},
    {"nombre": "MAURICIO CEVO",       "email": "mcevo@espaciocontainerhouse.cl",     "telefono": "+56971406162", "pass": "ECH2024!MC",  "rol": "ejecutivo"},
    {"nombre": "JACQUELINE PÉREZ",    "email": "jperez@espaciocontainerhouse.cl",    "telefono": "+56992286057", "pass": "ECH2024!JP",  "rol": "ejecutivo"},
    {"nombre": "JAVIER QUEZADA",      "email": "jquezada@espaciocontainerhouse.cl",  "telefono": "+56966983700", "pass": "ECH2024!JQ",  "rol": "ejecutivo"},
]

def _auto_provisionar_ejecutivos():
    """Crea las cuentas de ejecutivos si no existen. Se ejecuta solo una vez por sesión."""
    if st.session_state.get('_ejecutivos_provisionados'):
        return
    try:
        existentes = supabase_admin.auth.admin.list_users()
        emails_existentes = {u.email.lower() for u in existentes if u.email}
        for ej in _EJECUTIVOS_INICIALES:
            if ej["email"].lower() not in emails_existentes:
                try:
                    supabase_admin.auth.admin.create_user({
                        "email": ej["email"].lower(),
                        "password": ej["pass"],
                        "email_confirm": True,
                        "user_metadata": {"nombre": ej["nombre"], "telefono": ej["telefono"], "rol": ej.get("rol","ejecutivo")}
                    })
                except:
                    pass
    except:
        pass
    st.session_state['_ejecutivos_provisionados'] = True

# Solo ejecutar si hay sesión activa (supervisor)
if st.session_state.get('es_supervisor') and not st.session_state.get('_ejecutivos_provisionados'):
    _auto_provisionar_ejecutivos()

# =========================================================
# HELPERS: DESCRIPCIONES PDF CLIENTE (JSON en Storage bucket config)
# =========================================================
def cargar_descripciones_por_ep(numero, bust_cache=False):
    """Carga descripciones de un EP desde Storage bucket config."""
    try:
        import requests as _rq
        import time as _time
        _base = SUPABASE_URL.rstrip("/")
        _fname = f"pdf_desc_{numero}.json"
        url = f"{_base}/storage/v1/object/public/config/{_fname}"
        # Romper CDN cache con timestamp cuando se pide explícitamente
        if bust_cache:
            url += f"?t={int(_time.time())}"
        r = _rq.get(url, timeout=5, headers={"Cache-Control": "no-cache", "Pragma": "no-cache"})
        if r.status_code == 200:
            return r.json()
    except:
        pass
    return {}

def guardar_descripciones_por_ep(numero, descripciones: dict):
    """Guarda descripciones de un EP como JSON en Storage bucket config."""
    try:
        _fname = f"pdf_desc_{numero}.json"
        data = json.dumps(descripciones, ensure_ascii=False, indent=2).encode("utf-8")
        try:
            supabase.storage.from_("config").remove([_fname])
        except:
            pass
        supabase.storage.from_("config").upload(
            path=_fname,
            file=data,
            file_options={"content-type": "application/json", "upsert": "true"}
        )
        return True
    except Exception as e:
        st.error(f"Error al guardar descripciones: {e}")
        return False


# =========================================================
# FUNCIONES PARA MANEJO DE PDFs EN STORAGE
# =========================================================
def guardar_plano_en_storage(archivo_pdf_bytes, cotizacion_numero, nombre_original):
    try:
        file_ext = ".pdf"
        carpeta = cotizacion_numero.replace('/', '_').replace('\\', '_')
        file_name = f"cotizacion-{carpeta}/{uuid.uuid4()}{file_ext}"
        response = supabase.storage.from_('planos').upload(
            path=file_name,
            file=archivo_pdf_bytes,
            file_options={"content-type": "application/pdf"}
        )
        public_url = supabase.storage.from_('planos').get_public_url(file_name)
        return public_url, None
    except Exception as e:
        return None, str(e)

def eliminar_plano_de_storage(url_plano):
    try:
        if not url_plano:
            return True, None
        if '/planos/' in url_plano:
            path = url_plano.split('/planos/')[-1]
            supabase.storage.from_('planos').remove([path])
        return True, None
    except Exception as e:
        return False, str(e)

def descargar_plano_desde_url(url_plano):
    try:
        if not url_plano:
            return None, None
        response = requests.get(url_plano)
        if response.status_code == 200:
            return response.content, None
        else:
            return None, f"Error HTTP: {response.status_code}"
    except Exception as e:
        return None, str(e)

# =========================================================
# FUNCIÓN PARA DETECTAR NAVEGADOR
# =========================================================
def detectar_navegador():
    try:
        user_agent = st.context.headers.get('User-Agent', '')
        es_chrome = 'Chrome' in user_agent and 'Edg' not in user_agent
        es_edge = 'Edg' in user_agent
        es_safari = 'Safari' in user_agent and 'Chrome' not in user_agent
        return {
            'es_chrome': es_chrome,
            'es_edge': es_edge,
            'es_safari': es_safari,
            'es_firefox': 'Firefox' in user_agent,
            'needs_google_viewer': es_chrome or es_edge or es_safari
        }
    except:
        return {'needs_google_viewer': True}

# =========================================================
# INICIALIZAR VARIABLES DE SESIÓN
# =========================================================
if 'modo_admin' not in st.session_state:
    st.session_state.modo_admin = False
if 'mostrar_login' not in st.session_state:
    st.session_state.mostrar_login = False
if 'auth_user' not in st.session_state:
    st.session_state.auth_user = None
if 'auth_email' not in st.session_state:
    st.session_state.auth_email = ""
if 'auth_nombre' not in st.session_state:
    st.session_state.auth_nombre = ""
if 'es_supervisor' not in st.session_state:
    st.session_state.es_supervisor = False
if 'es_root' not in st.session_state:
    st.session_state.es_root = False
if 'rol_usuario' not in st.session_state:
    st.session_state.rol_usuario = "ejecutivo"

# ══════════════════════════════════════════════════════
# REGIONES Y COMUNAS DE CHILE
# ══════════════════════════════════════════════════════
REGIONES_COMUNAS = {
    "Arica y Parinacota": ["Arica","Camarones","Putre","General Lagos"],
    "Tarapacá": ["Iquique","Alto Hospicio","Pozo Almonte","Camiña","Colchane","Huara","Pica"],
    "Antofagasta": ["Antofagasta","Mejillones","Sierra Gorda","Taltal","Calama","Ollagüe","San Pedro de Atacama","Tocopilla","María Elena"],
    "Atacama": ["Copiapó","Caldera","Tierra Amarilla","Chañaral","Diego de Almagro","Vallenar","Alto del Carmen","Freirina","Huasco"],
    "Coquimbo": ["La Serena","Coquimbo","Andacollo","La Higuera","Paiguano","Vicuña","Illapel","Canela","Los Vilos","Salamanca","Ovalle","Combarbalá","Monte Patria","Punitaqui","Río Hurtado"],
    "Valparaíso": ["Valparaíso","Casablanca","Concón","Juan Fernández","Puchuncaví","Quintero","Viña del Mar","Isla de Pascua","Los Andes","Calle Larga","Rinconada","San Esteban","La Ligua","Cabildo","Papudo","Petorca","Zapallar","Quillota","Calera","Hijuelas","La Cruz","Nogales","San Antonio","Algarrobo","Cartagena","El Quisco","El Tabo","Santo Domingo","San Felipe","Catemu","Llaillay","Panquehue","Putaendo","Santa María","Quilpué","Limache","Olmué","Villa Alemana"],
    "Metropolitana": ["Santiago","Cerrillos","Cerro Navia","Conchalí","El Bosque","Estación Central","Huechuraba","Independencia","La Cisterna","La Florida","La Granja","La Pintana","La Reina","Las Condes","Lo Barnechea","Lo Espejo","Lo Prado","Macul","Maipú","Ñuñoa","Pedro Aguirre Cerda","Peñalolén","Providencia","Pudahuel","Quilicura","Quinta Normal","Recoleta","Renca","San Joaquín","San Miguel","San Ramón","Vitacura","Puente Alto","Pirque","San José de Maipo","Colina","Lampa","Tiltil","San Bernardo","Buin","Calera de Tango","Paine","Melipilla","Alhué","Curacaví","María Pinto","San Pedro","Talagante","El Monte","Isla de Maipo","Padre Hurtado","Peñaflor"],
    "O'Higgins": ["Rancagua","Codegua","Coinco","Coltauco","Doñihue","Graneros","Las Cabras","Machalí","Malloa","Mostazal","Olivar","Peumo","Pichidegua","Quinta de Tilcoco","Rengo","Requínoa","San Vicente","Pichilemu","La Estrella","Litueche","Marchihue","Navidad","Paredones","San Fernando","Chépica","Chimbarongo","Lolol","Nancagua","Palmilla","Peralillo","Placilla","Pumanque","Santa Cruz"],
    "Maule": ["Talca","Constitución","Curepto","Empedrado","Maule","Pelarco","Pencahue","Río Claro","San Clemente","San Rafael","Cauquenes","Chanco","Pelluhue","Curicó","Hualañé","Licantén","Molina","Rauco","Romeral","Sagrada Familia","Teno","Vichuquén","Linares","Colbún","Longaví","Parral","Retiro","San Javier","Villa Alegre","Yerbas Buenas"],
    "Ñuble": ["Chillán","Bulnes","Chillán Viejo","El Carmen","Pemuco","Pinto","Quillón","San Ignacio","Yungay","Cobquecura","Coelemu","Ninhue","Portezuelo","Quirihue","Ránquil","Treguaco","Coihueco","Ñiquén","San Carlos","San Fabián","San Nicolás"],
    "Biobío": ["Concepción","Coronel","Chiguayante","Florida","Hualpén","Hualqui","Lota","Penco","San Pedro de la Paz","Santa Juana","Talcahuano","Tomé","Lebu","Arauco","Cañete","Contulmo","Curanilahue","Los Álamos","Tirúa","Los Ángeles","Antuco","Cabrero","Laja","Mulchén","Nacimiento","Negrete","Quilaco","Quilleco","San Rosendo","Santa Bárbara","Tucapel","Yumbel","Alto Biobío"],
    "La Araucanía": ["Temuco","Carahue","Cunco","Curarrehue","Freire","Galvarino","Gorbea","Lautaro","Loncoche","Melipeuco","Nueva Imperial","Padre las Casas","Perquenco","Pitrufquén","Pucón","Saavedra","Teodoro Schmidt","Toltén","Vilcún","Villarrica","Cholchol","Angol","Collipulli","Curacautín","Ercilla","Lonquimay","Los Sauces","Lumaco","Purén","Renaico","Traiguén","Victoria"],
    "Los Ríos": ["Valdivia","Corral","Futrono","La Unión","Lago Ranco","Lanco","Los Lagos","Máfil","Mariquina","Paillaco","Panguipulli","Río Bueno"],
    "Los Lagos": ["Puerto Montt","Calbuco","Cochamó","Fresia","Frutillar","Los Muermos","Llanquihue","Maullín","Puerto Varas","Castro","Ancud","Chonchi","Curaco de Vélez","Dalcahue","Puqueldón","Queilén","Quellón","Quemchi","Quinchao","Osorno","Puerto Octay","Purranque","Puyehue","Río Negro","San Juan de la Costa","San Pablo","Chaitén","Futaleufú","Hualaihué","Palena"],
    "Aysén": ["Coyhaique","Lago Verde","Aysén","Cisnes","Guaitecas","Cochrane","O'Higgins","Tortel","Chile Chico","Río Ibáñez"],
    "Magallanes": ["Punta Arenas","Laguna Blanca","Río Verde","San Gregorio","Cabo de Hornos","Antártica","Porvenir","Primavera","Timaukel","Natales","Torres del Paine"],
}

# Mapa inverso: comuna → región
COMUNA_A_REGION = {c: r for r, cs in REGIONES_COMUNAS.items() for c in cs}

# Normalizar texto ingresado a nombre oficial de región/comuna
def _normalizar_nombre(texto, catalogo):
    """Busca el nombre oficial ignorando tildes, mayúsculas y 'Región de/del'."""
    import unicodedata
    def _strip(s):
        s = s.lower().strip()
        s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
        for prefix in ('region de los ', 'region de la ', 'region del ', 'region de ', 'region '):
            if s.startswith(prefix):
                s = s[len(prefix):]
        return s
    txt_norm = _strip(texto)
    for nombre in catalogo:
        if _strip(nombre) == txt_norm:
            return nombre
    # Coincidencia parcial
    for nombre in catalogo:
        if txt_norm in _strip(nombre) or _strip(nombre) in txt_norm:
            return nombre
    return texto  # devolver original si no encuentra


def selector_comuna_region(label_com, label_reg, key_com, key_reg, val_com="", val_reg="", col_layout=None):
    """
    Región filtra comunas. Elegir comuna auto-completa región.
    Default: Metropolitana / Santiago.
    """
    todas_regiones = list(REGIONES_COMUNAS.keys())

    # Determinar región inicial
    _reg_init = val_reg if val_reg in todas_regiones else (
        COMUNA_A_REGION.get(val_com, "Metropolitana")
    )
    _idx_reg = todas_regiones.index(_reg_init) if _reg_init in todas_regiones else todas_regiones.index("Metropolitana")

    region_sel = st.selectbox(
        label_reg,
        todas_regiones,
        index=_idx_reg,
        key=key_reg,
    )

    # Comunas filtradas por región seleccionada
    comunas_region = REGIONES_COMUNAS.get(region_sel, [])

    # Si la comuna guardada pertenece a esta región, preseleccionar; si no, Santiago o primera
    if val_com in comunas_region:
        _idx_com = comunas_region.index(val_com)
    elif region_sel == "Metropolitana" and "Santiago" in comunas_region:
        _idx_com = comunas_region.index("Santiago")
    else:
        _idx_com = 0

    comuna_sel = st.selectbox(
        label_com,
        comunas_region,
        index=_idx_com,
        key=key_com,
    )

    return comuna_sel, region_sel

if 'nombre_input' not in st.session_state:
    st.session_state.nombre_input = ""
if 'correo_input' not in st.session_state:
    st.session_state.correo_input = ""
if 'direccion_input' not in st.session_state:
    st.session_state.direccion_input = ""
if 'cliente_comuna' not in st.session_state:
    st.session_state.cliente_comuna = ""
if 'cliente_region' not in st.session_state:
    st.session_state.cliente_region = ""
if 'proyecto_direccion' not in st.session_state:
    st.session_state.proyecto_direccion = ""
if 'proyecto_comuna' not in st.session_state:
    st.session_state.proyecto_comuna = ""
if 'proyecto_region' not in st.session_state:
    st.session_state.proyecto_region = ""
if 'cliente_tipo' not in st.session_state:
    st.session_state.cliente_tipo = "natural"
if 'cliente_empresa' not in st.session_state:
    st.session_state.cliente_empresa = ""
if 'cliente_rut_empresa' not in st.session_state:
    st.session_state.cliente_rut_empresa = ""
    st.session_state.rut_empresa_raw     = ""
    st.session_state.rut_empresa_display = ""
    st.session_state.rut_empresa_valido  = False
if 'rut_empresa_raw' not in st.session_state:
    st.session_state.rut_empresa_raw = ""
if 'rut_empresa_display' not in st.session_state:
    st.session_state.rut_empresa_display = ""
if 'rut_empresa_valido' not in st.session_state:
    st.session_state.rut_empresa_valido = False  
if 'fecha_inicio' not in st.session_state:
    st.session_state.fecha_inicio = datetime.now().date()
if 'fecha_termino' not in st.session_state:
    st.session_state.fecha_termino = (datetime.now() + timedelta(days=15)).date()
if 'observaciones_input' not in st.session_state:
    st.session_state.observaciones_input = ""
if 'plano_adjunto' not in st.session_state:
    st.session_state.plano_adjunto = None
if 'plano_nombre' not in st.session_state:
    st.session_state.plano_nombre = ""
if 'cotizacion_seleccionada' not in st.session_state:
    st.session_state.cotizacion_seleccionada = None
if 'cotizacion_cargada' not in st.session_state:
    st.session_state.cotizacion_cargada = None
if 'carrito' not in st.session_state:
    st.session_state.carrito = []
if 'margen' not in st.session_state:
    st.session_state.margen = 0.0
if 'rut_raw' not in st.session_state:
    st.session_state.rut_raw = ""
if 'rut_display' not in st.session_state:
    st.session_state.rut_display = ""
if 'rut_valido' not in st.session_state:
    st.session_state.rut_valido = False
if 'rut_mensaje' not in st.session_state:
    st.session_state.rut_mensaje = ""
if 'telefono_raw' not in st.session_state:
    st.session_state.telefono_raw = ""
if 'asesor_seleccionado' not in st.session_state:
    st.session_state.asesor_seleccionado = "Seleccionar asesor"
if 'correo_asesor' not in st.session_state:
    st.session_state.correo_asesor = ""
if 'telefono_asesor' not in st.session_state:
    st.session_state.telefono_asesor = ""
if 'telefono_asesor_raw' not in st.session_state:
    st.session_state.telefono_asesor_raw = ""
if 'asesor_correo_temp' not in st.session_state:
    st.session_state.asesor_correo_temp = ""
# ── Leer margen desde FAB via query_params ──────────────
_mgfab = st.query_params.get("mgfab")
if _mgfab is not None:
    try:
        _mgfab_val = max(0.0, min(100.0, float(_mgfab)))
        st.session_state['margen'] = _mgfab_val
    except ValueError:
        pass
    st.query_params.clear()

# ── Leer acción guardar desde FAB via query_params ───────
if st.query_params.get("_fabg") == "1":
    st.query_params.clear()
    st.session_state['_trigger_guardar_fab'] = True
# ────────────────────────────────────────────────────────

if 'counter' not in st.session_state:
    st.session_state.counter = 0
st.session_state['_rerun_lock'] = False
if 'cargar_cotizacion_trigger' not in st.session_state:
    st.session_state.cargar_cotizacion_trigger = False
if 'cotizacion_a_cargar' not in st.session_state:
    st.session_state.cotizacion_a_cargar = None
if 'mostrar_visor' not in st.session_state:
    st.session_state.mostrar_visor = False
if 'pdf_actual' not in st.session_state:
    st.session_state.pdf_actual = None
    st.session_state.pdf_nombre = ""
if 'numero_en_visor' not in st.session_state:
    st.session_state.numero_en_visor = None
if 'pdf_url' not in st.session_state:
    st.session_state.pdf_url = None

if 'mostrar_toast_exito' not in st.session_state:
    st.session_state.mostrar_toast_exito = False

if 'toast_numero_ep' not in st.session_state:
    st.session_state.toast_numero_ep = ""

if 'recien_guardado' not in st.session_state:
    st.session_state.recien_guardado = False

if 'hash_ultimo_guardado' not in st.session_state:
    st.session_state.hash_ultimo_guardado = None

if 'recien_cargado' not in st.session_state:
    st.session_state.recien_cargado = False

if 'mostrar_advertencia_cerrar' not in st.session_state:
    st.session_state.mostrar_advertencia_cerrar = False

if 'trigger_cerrar_cotizacion' not in st.session_state:
    st.session_state.trigger_cerrar_cotizacion = False

if 'datos_pendientes_cerrar' not in st.session_state:
    st.session_state.datos_pendientes_cerrar = None

if 'numero_a_cargar_pendiente' not in st.session_state:
    st.session_state.numero_a_cargar_pendiente = None

CLAVE_ADMIN = "admin2024"

# =========================================================
# FUNCIONES DE VALIDACIÓN Y FORMATO
# =========================================================
def validar_rut(rut_completo):
    rut_limpio = re.sub(r'[^0-9kK]', '', rut_completo)
    if len(rut_limpio) < 2:
        return False, "RUT incompleto"
    cuerpo = rut_limpio[:-1]
    dv_ingresado = rut_limpio[-1].upper()
    if not cuerpo.isdigit():
        return False, "RUT inválido"
    # RUT extranjero: 9+ dígitos en cuerpo o número >= 100.000.000
    _es_extranjero = len(cuerpo) >= 9 or int(cuerpo) >= 100000000
    # Aplicar algoritmo módulo 11
    suma = 0
    multiplo = 2
    for i in range(len(cuerpo) - 1, -1, -1):
        suma += multiplo * int(cuerpo[i])
        multiplo = multiplo + 1 if multiplo < 7 else 2
    dv_esperado = 11 - (suma % 11)
    if dv_esperado == 10:
        dv_esperado = 'K'
    elif dv_esperado == 11:
        dv_esperado = '0'
    else:
        dv_esperado = str(dv_esperado)
    if dv_ingresado == dv_esperado:
        if _es_extranjero:
            return True, "RUT extranjero válido"
        return True, "RUT válido"
    else:
        # Si falla validación pero tiene formato de extranjero → advertencia en vez de error
        if _es_extranjero:
            return True, "RUT inválido o RUT extranjero"
        return False, "RUT inválido"

def formatear_rut(rut_raw):
    if not rut_raw:
        return ""
    if len(rut_raw) > 10:
        rut_raw = rut_raw[:10]
    if len(rut_raw) >= 2:
        cuerpo = rut_raw[:-1]
        dv = rut_raw[-1].upper()
        if cuerpo:
            cuerpo_formateado = ""
            for i, digito in enumerate(reversed(cuerpo)):
                if i > 0 and i % 3 == 0:
                    cuerpo_formateado = "." + cuerpo_formateado
                cuerpo_formateado = digito + cuerpo_formateado
        else:
            cuerpo_formateado = ""
        return f"{cuerpo_formateado}-{dv}"
    else:
        return rut_raw

def formatear_telefono(telefono_raw):
    if not telefono_raw:
        return ""
    if len(telefono_raw) > 9:
        telefono_raw = telefono_raw[:9]
    if len(telefono_raw) == 1:
        return f"+56 {telefono_raw}"
    elif len(telefono_raw) <= 5:
        return f"+56 {telefono_raw[:1]} {telefono_raw[1:]}"
    else:
        return f"+56 {telefono_raw[:1]} {telefono_raw[1:5]} {telefono_raw[5:]}"

def formato_clp(valor):
    return f"${valor:,.0f}".replace(",", ".")

# =========================================================
# FUNCIONES PARA PROCESAR CAMBIOS EN TIEMPO REAL
# =========================================================
def procesar_cambio_rut():
    rut_key = f"rut_input_{st.session_state.counter}"
    if rut_key in st.session_state:
        valor_actual = st.session_state[rut_key]
        raw = re.sub(r'[^0-9kK]', '', valor_actual)
        if len(raw) > 10:
            raw = raw[:10]
        st.session_state.rut_raw = raw
        if raw:
            st.session_state.rut_display = formatear_rut(raw)
        else:
            st.session_state.rut_display = ""
        if len(raw) >= 2:
            valido, mensaje = validar_rut(raw)
            st.session_state.rut_valido = valido
            st.session_state.rut_mensaje = mensaje
        else:
            st.session_state.rut_valido = False
            st.session_state.rut_mensaje = "RUT incompleto"

def procesar_cambio_rut_empresa():
    rut_emp_key = f"rut_empresa_input_{st.session_state.counter}"
    if rut_emp_key in st.session_state:
        valor_actual = st.session_state[rut_emp_key]
        raw = re.sub(r'[^0-9kK]', '', valor_actual)
        if len(raw) > 10:
            raw = raw[:10]
        st.session_state.rut_empresa_raw = raw
        if raw:
            st.session_state.rut_empresa_display = formatear_rut(raw)
            st.session_state.cliente_rut_empresa  = formatear_rut(raw)
        else:
            st.session_state.rut_empresa_display = ""
            st.session_state.cliente_rut_empresa  = ""
        if len(raw) >= 2:
            valido, mensaje = validar_rut(raw)
            st.session_state.rut_empresa_valido  = valido
            st.session_state.rut_empresa_mensaje = mensaje
        else:
            st.session_state.rut_empresa_valido  = False
            st.session_state.rut_empresa_mensaje = "RUT incompleto"

def procesar_cambio_telefono():
    telefono_key = f"telefono_input_{st.session_state.counter}"
    if telefono_key in st.session_state:
        valor_actual = st.session_state[telefono_key]
        raw = re.sub(r'[^0-9]', '', valor_actual)
        if len(raw) > 10:
            raw = raw[:10]
        st.session_state.telefono_raw = raw

def leer_datos_actuales():
    mapeo_texto = {
        'nombre_input_':    'nombre_input',
        'correo_input_':    'correo_input',
        'direccion_input_': 'direccion_input',
        'observaciones_input_': 'observaciones_input',
        'asesor_correo_input_': 'correo_asesor',
    }
    for prefijo, campo in mapeo_texto.items():
        mejor_counter = -1
        mejor_valor = None
        for key, valor in st.session_state.items():
            if isinstance(key, str) and key.startswith(prefijo):
                try:
                    c = int(key[len(prefijo):])
                    if c > mejor_counter:
                        mejor_counter = c
                        mejor_valor = valor
                except ValueError:
                    pass
        if mejor_valor is not None:
            st.session_state[campo] = mejor_valor

    mejor_counter = -1
    mejor_rut = None
    for key, valor in st.session_state.items():
        if isinstance(key, str) and key.startswith('rut_input_'):
            try:
                c = int(key[len('rut_input_'):])
                if c > mejor_counter:
                    mejor_counter = c
                    mejor_rut = valor
            except ValueError:
                pass
    if mejor_rut is not None:
        raw = re.sub(r'[^0-9kK]', '', mejor_rut)[:9]
        st.session_state.rut_raw = raw
        st.session_state.rut_display = formatear_rut(raw) if raw else ""
        if len(raw) >= 2:
            valido, mensaje = validar_rut(raw)
            st.session_state.rut_valido = valido
            st.session_state.rut_mensaje = mensaje

    mejor_counter = -1
    mejor_tel = None
    for key, valor in st.session_state.items():
        if isinstance(key, str) and key.startswith('telefono_input_'):
            try:
                c = int(key[len('telefono_input_'):])
                if c > mejor_counter:
                    mejor_counter = c
                    mejor_tel = valor
            except ValueError:
                pass
    if mejor_tel is not None:
        raw = re.sub(r'[^0-9]', '', mejor_tel)[:9]
        st.session_state.telefono_raw = raw

    mejor_counter = -1
    mejor_tel_asesor = None
    for key, valor in st.session_state.items():
        if isinstance(key, str) and key.startswith('asesor_telefono_input_'):
            try:
                c = int(key[len('asesor_telefono_input_'):])
                if c > mejor_counter:
                    mejor_counter = c
                    mejor_tel_asesor = valor
            except ValueError:
                pass
    if mejor_tel_asesor is not None:
        raw = re.sub(r'[^0-9]', '', mejor_tel_asesor)[:9]
        st.session_state.telefono_asesor = raw

    mejor_counter = -1
    mejor_fi = None
    for key, valor in st.session_state.items():
        if isinstance(key, str) and key.startswith('fecha_inicio_'):
            try:
                c = int(key[len('fecha_inicio_'):])
                if c > mejor_counter:
                    mejor_counter = c
                    mejor_fi = valor
            except ValueError:
                pass
    if mejor_fi is not None:
        st.session_state.fecha_inicio = mejor_fi

    mejor_counter = -1
    mejor_ft = None
    for key, valor in st.session_state.items():
        if isinstance(key, str) and key.startswith('fecha_termino_'):
            try:
                c = int(key[len('fecha_termino_'):])
                if c > mejor_counter:
                    mejor_counter = c
                    mejor_ft = valor
            except ValueError:
                pass
    if mejor_ft is not None:
        st.session_state.fecha_termino = mejor_ft

def calcular_hash_estado():
    """Calcula un hash del estado actual para detectar cambios no guardados."""
    import hashlib
    estado = {
        "nombre": st.session_state.get('nombre_input', ''),
        "rut": st.session_state.get('rut_display', ''),
        "correo": st.session_state.get('correo_input', ''),
        "telefono": st.session_state.get('telefono_raw', ''),
        "direccion": st.session_state.get('direccion_input', ''),
        "cliente_comuna": st.session_state.get('cliente_comuna', ''),
        "cliente_region": st.session_state.get('cliente_region', ''),
        "proyecto_direccion": st.session_state.get('proyecto_direccion', ''),
        "proyecto_comuna": st.session_state.get('proyecto_comuna', ''),
        "proyecto_region": st.session_state.get('proyecto_region', ''),
        "cliente_tipo": st.session_state.get('cliente_tipo', 'natural'),
        "cliente_empresa": st.session_state.get('cliente_empresa', ''),
        "cliente_rut_empresa": st.session_state.get('cliente_rut_empresa', ''),
        "observaciones": st.session_state.get('observaciones_input', ''),
        "asesor": st.session_state.get('asesor_seleccionado', ''),
        "correo_asesor": st.session_state.get('correo_asesor', ''),
        "telefono_asesor": st.session_state.get('telefono_asesor', ''),
        "fecha_inicio": str(st.session_state.get('fecha_inicio', '')),
        "fecha_termino": str(st.session_state.get('fecha_termino', '')),
        "carrito": json.dumps(st.session_state.get('carrito', []), sort_keys=True),
        "margen": st.session_state.get('margen', 0),
        "plano_nombre": st.session_state.get('plano_nombre', ''),
    }
    estado_str = json.dumps(estado, sort_keys=True)
    return hashlib.md5(estado_str.encode()).hexdigest()

def construir_datos_para_guardar():
    leer_datos_actuales()
    datos_cliente = {
        "Nombre": st.session_state.nombre_input or "",
        "RUT": st.session_state.rut_display or "",
        "Correo": st.session_state.correo_input or "",
        "Teléfono": st.session_state.telefono_raw or "",
        "Dirección": st.session_state.direccion_input or "",
            "ComunaCliente": st.session_state.cliente_comuna or "",
            "RegionCliente": st.session_state.cliente_region or "",
            "DireccionProyecto": st.session_state.proyecto_direccion or "",
            "ComunaProyecto": st.session_state.proyecto_comuna or "",
            "RegionProyecto": st.session_state.proyecto_region or "",
            "TipoCliente": st.session_state.cliente_tipo or "natural",
            "EmpresaCliente": st.session_state.cliente_empresa or "",
            "RutEmpresa": st.session_state.cliente_rut_empresa or "",
            "Observaciones": st.session_state.observaciones_input or ""
    }
    nombre_asesor = st.session_state.asesor_seleccionado if st.session_state.asesor_seleccionado != "Seleccionar asesor" else ""
    datos_asesor = {
        "Nombre Ejecutivo": nombre_asesor,
        "Correo Ejecutivo": st.session_state.correo_asesor or "",
        "Teléfono Ejecutivo": st.session_state.telefono_asesor or ""
    }
    proyecto = {
        'fecha_inicio': str(st.session_state.fecha_inicio),
        'fecha_termino': str(st.session_state.fecha_termino),
        'dias_validez': (st.session_state.fecha_termino - st.session_state.fecha_inicio).days,
        'observaciones': st.session_state.observaciones_input or ""
    }
    config = {
        'margen': st.session_state.margen,
        'modo_admin': st.session_state.modo_admin
    }
    if st.session_state.carrito:
        carrito_df_temp = pd.DataFrame(st.session_state.carrito)
        subtotal_base_temp = carrito_df_temp["Subtotal"].sum()
        if st.session_state.modo_admin or st.session_state.margen > 0:
            subtotal_general_temp = sum(
                item["Cantidad"] * aplicar_margen(item["Precio Unitario"], st.session_state.margen)
                for item in st.session_state.carrito
            )
            iva_temp = subtotal_general_temp * 0.19
            total_temp = subtotal_general_temp + iva_temp
            margen_valor_temp = subtotal_general_temp - subtotal_base_temp
            comision_vendedor_temp = subtotal_general_temp * 0.025
            comision_supervisor_temp = subtotal_general_temp * 0.008
            utilidad_real_temp = margen_valor_temp - comision_vendedor_temp - comision_supervisor_temp
        else:
            subtotal_general_temp = subtotal_base_temp
            iva_temp = subtotal_general_temp * 0.19
            total_temp = subtotal_general_temp + iva_temp
            margen_valor_temp = 0
            comision_vendedor_temp = 0
            comision_supervisor_temp = 0
            utilidad_real_temp = 0
    else:
        subtotal_base_temp = subtotal_general_temp = iva_temp = total_temp = 0
        margen_valor_temp = comision_vendedor_temp = comision_supervisor_temp = utilidad_real_temp = 0
    totales = {
        'subtotal_sin_margen': subtotal_base_temp,
        'subtotal_con_margen': subtotal_general_temp,
        'iva': iva_temp,
        'total': total_temp,
        'margen_valor': margen_valor_temp,
        'comision_vendedor': comision_vendedor_temp,
        'comision_supervisor': comision_supervisor_temp,
        'utilidad_real': utilidad_real_temp
    }
    plano_nombre = st.session_state.plano_nombre if st.session_state.plano_adjunto else None
    plano_datos = st.session_state.plano_adjunto if st.session_state.plano_adjunto else None
    return datos_cliente, datos_asesor, proyecto, config, totales, plano_nombre, plano_datos

# =========================================================
# FUNCIONES PARA EVALUAR ESTADO DE COTIZACIÓN
# =========================================================
def evaluar_estado_cotizacion(cotizacion):
    datos_completos = all([
        cotizacion.get('cliente_nombre', ''),
        cotizacion.get('cliente_email', '')
    ])
    asesor_completo = any([
        cotizacion.get('asesor_nombre', ''),
        cotizacion.get('asesor_email', ''),
        cotizacion.get('asesor_telefono', '')
    ])
    tiene_plano = cotizacion.get('plano_nombre') not in (None, '')
    if not datos_completos or not asesor_completo:
        return "🔴 INCOMPLETO CON PLANO" if tiene_plano else "🔴 INCOMPLETO"
    tiene_margen = cotizacion.get('config_margen', 0) > 0
    if tiene_margen:
        return "🟢 AUTORIZADO CON PLANO" if tiene_plano else "🟢 AUTORIZADO"
    else:
        return "🟠 BORRADOR CON PLANO" if tiene_plano else "🟡 BORRADOR"

def crear_badge_estado(row):
    # Soporta tanto índices numéricos como nombres de columna del DataFrame
    if hasattr(row, 'index') and 'Margen' in row.index:
        config_margen = row['Margen']
        tiene_plano   = row['Tiene_Plano']
        cliente_nombre= row['Cliente']
        cliente_email = row['Email']
        asesor_nombre = row['Asesor']
        asesor_email  = row['Asesor_Email']
        asesor_telefono = row['Asesor_Tel']
    else:
        config_margen = row[5]
        tiene_plano = row[10] if len(row) > 10 else False
        cliente_nombre = row[1]
        cliente_rut = row[6]
        cliente_email = row[7]
        asesor_nombre = row[2]
        asesor_email = row[8]
        asesor_telefono = row[9]
    datos_completos = all([cliente_nombre, cliente_email])
    asesor_completo = any([asesor_nombre, asesor_email, asesor_telefono])
    if config_margen and config_margen > 0:
        if datos_completos and asesor_completo:
            label = "🟢 AUTORIZADO CON PLANO" if tiene_plano else "🟢 AUTORIZADO"
            color = "#28a745"
            border = "#1e7e34"
        else:
            label = "🔴 INCOMPLETO CON PLANO" if tiene_plano else "🔴 INCOMPLETO"
            color = "#dc3545"
            border = "#bd2130"
    else:
        if datos_completos and asesor_completo:
            if tiene_plano:
                label = "🟠 BORRADOR CON PLANO"
                color = "#f97316"
                border = "#c2410c"
            else:
                label = "🟡 BORRADOR"
                color = "#ffc107"
                border = "#d39e00"
        else:
            label = "🔴 INCOMPLETO CON PLANO" if tiene_plano else "🔴 INCOMPLETO"
            color = "#dc3545"
            border = "#bd2130"
    text_color = "#212529" if label == "🟡 BORRADOR" else "white"
    return f'''<span style="background-color:{color};color:{text_color};padding:4px 12px;
        border-radius:20px;font-size:0.8rem;font-weight:600;display:inline-block;
        border:1px solid {border};box-shadow:0 2px 4px rgba(0,0,0,0.1);">{label}</span>'''

# =========================================================
# FUNCIONES PARA MANEJO DE MARGEN
# =========================================================
def aplicar_margen(precio_original, margen):
    return precio_original * (1 + margen / 100)

def calcular_totales_con_margen(carrito, margen):
    subtotal_con_margen = sum(
        item["Cantidad"] * aplicar_margen(item["Precio Unitario"], margen)
        for item in carrito
    )
    iva_con_margen = subtotal_con_margen * 0.19
    total_con_margen = subtotal_con_margen + iva_con_margen
    return subtotal_con_margen, iva_con_margen, total_con_margen

def calcular_comision_vendedor(subtotal_con_margen):
    return subtotal_con_margen * 0.025

def calcular_comision_supervisor(subtotal_con_margen):
    return subtotal_con_margen * 0.008

def calcular_utilidad_real(margen_valor, comision_vendedor, comision_supervisor):
    return margen_valor - comision_vendedor - comision_supervisor

@st.cache_data(ttl=3600)
def buscar_direccion(direccion):
    # Cache 1 hora — evita HTTP en cada render mientras el usuario escribe
    if not direccion or len(direccion.strip()) < 5:
        return None, None
    url = "https://nominatim.openstreetmap.org/search"
    params = {"q": direccion, "format": "json", "addressdetails": 1, "limit": 1, "countrycodes": "cl"}
    try:
        response = requests.get(url, params=params, headers={"User-Agent": "epcontainer-app"}, timeout=5)
        data = response.json()
        if data:
            address = data[0]["address"]
            comuna = address.get("city") or address.get("town") or address.get("village")
            region = address.get("state")
            return comuna, region
    except:
        pass
    return None, None


# =========================================================
# HELPER: OBTENER EXCEL ACTIVO DESDE SUPABASE
# =========================================================
import io as _io_excel

@st.cache_data(ttl=60)
@st.cache_data(ttl=60)
def _get_excel_bytes_activo():
    """Descarga el Excel activo desde Supabase Storage. Cache 60s."""
    try:
        _resp = supabase.table('excel_versiones').select('archivo_url').eq('activa', True).limit(1).execute()
        if _resp.data:
            _url = _resp.data[0]['archivo_url']
            import requests as _rq
            _r = _rq.get(_url, timeout=15)
            _r.raise_for_status()
            return _io_excel.BytesIO(_r.content)
    except:
        pass
    return "cotizador.xlsx"  # fallback local

def _excel_src():
    """Retorna la fuente del Excel (BytesIO desde Supabase o path local)."""
    if 'excel_bytes_cache' not in st.session_state:
        st.session_state.excel_bytes_cache = _get_excel_bytes_activo()
    return st.session_state.excel_bytes_cache

@st.cache_data(ttl=300)
@st.cache_data(ttl=120)
def _leer_hoja_excel(nombre_hoja):
    """Lee y cachea una hoja del Excel — evita re-parsear en cada render."""
    try:
        return pd.read_excel(_excel_src(), sheet_name=nombre_hoja)
    except:
        return pd.DataFrame()

@st.cache_data(ttl=300)
def _leer_bd_total():
    """Lee y cachea la hoja BD Total."""
    return pd.read_excel(_excel_src(), sheet_name="BD Total")[["Item", "P. Unitario real"]]

@st.cache_data(ttl=120)
def _leer_hojas_disponibles():
    """Lista de hojas disponibles con caché corto."""
    try:
        return pd.ExcelFile(_get_excel_bytes_activo()).sheet_names
    except:
        try:
            return pd.ExcelFile(_excel_src()).sheet_names
        except:
            return []

def cargar_modelo(nombre_hoja):
    df_modelo = _leer_hoja_excel(nombre_hoja)
    df_modelo = df_modelo[["Categorias", "Item", "Cantidad"]].dropna()
    df_modelo = df_modelo[df_modelo["Cantidad"] > 0]
    df_bd = _leer_bd_total()
    df_final = df_modelo.merge(df_bd, on="Item", how="left")
    carrito = []
    for _, row in df_final.iterrows():
        subtotal = row["Cantidad"] * row["P. Unitario real"]
        carrito.append({
            "Categoria": row["Categorias"], "Item": row["Item"],
            "Cantidad": row["Cantidad"], "Precio Unitario": row["P. Unitario real"], "Subtotal": subtotal
        })
    return carrito

def cargar_categoria_desde_modelo(nombre_hoja, categoria_objetivo):
    df_modelo = _leer_hoja_excel(nombre_hoja)
    df_modelo = df_modelo[["Categorias", "Item", "Cantidad"]].dropna()
    df_modelo = df_modelo[(df_modelo["Cantidad"] > 0) & (df_modelo["Categorias"] == categoria_objetivo)]
    df_bd = _leer_bd_total()
    df_final = df_modelo.merge(df_bd, on="Item", how="left")
    categoria_items = []
    for _, row in df_final.iterrows():
        subtotal = row["Cantidad"] * row["P. Unitario real"]
        categoria_items.append({
            "Categoria": row["Categorias"], "Item": row["Item"],
            "Cantidad": row["Cantidad"], "Precio Unitario": row["P. Unitario real"], "Subtotal": subtotal
        })
    return categoria_items


# =========================================================
# SISTEMA DE AUDITORÍA / LOGS DE MODIFICACIONES
# =========================================================
def registrar_log(numero, asesor, tipo_cambio, detalle_dict):
    """Inserta un registro en cotizacion_logs."""
    try:
        supabase.table('cotizacion_logs').insert({
            'numero': numero,
            'asesor': asesor,
            'tipo_cambio': tipo_cambio,
            'detalle': detalle_dict,
        }).execute()
    except Exception as e:
        pass  # El log no debe interrumpir el flujo principal

def contar_logs(numeros):
    """Devuelve dict {numero: count} para una lista de números EP."""
    if not numeros:
        return {}
    try:
        resp = supabase.table('cotizacion_logs').select('numero').in_('numero', numeros).execute()
        counts = {}
        for row in resp.data:
            n = row['numero']
            counts[n] = counts.get(n, 0) + 1
        return counts
    except:
        return {}

def obtener_logs_ep(numero):
    """Devuelve lista de logs ordenados por fecha DESC para un EP."""
    try:
        resp = supabase.table('cotizacion_logs') \
            .select('*').eq('numero', numero) \
            .order('fecha', desc=True).execute()
        return resp.data or []
    except:
        return []

def _diff_datos(anterior, nuevo):
    """Compara dos dicts y devuelve solo los campos que cambiaron."""
    LABELS = {
        'cliente_nombre': 'Nombre cliente', 'cliente_rut': 'RUT cliente',
        'cliente_email': 'Correo', 'cliente_telefono': 'Teléfono',
        'cliente_tipo': 'Tipo cliente', 'cliente_empresa': 'Empresa',
        'cliente_rut_empresa': 'RUT empresa', 'asesor_nombre': 'Asesor',
        'config_margen': 'Margen %', 'total_total': 'Total',
        'proyecto_observaciones': 'Descripción del proyecto', 'estado': 'Estado',
        'productos': 'Productos/carrito',
    }
    CAMPOS_TEXTO = {'cliente_telefono', 'cliente_rut', 'cliente_rut_empresa',
                    'cliente_email', 'cliente_nombre', 'cliente_empresa',
                    'asesor_nombre', 'proyecto_observaciones', 'estado'}

    cambios = {}

    # Comparar campos simples
    for k, label in LABELS.items():
        v_ant_raw = anterior.get(k, '') or ''
        v_new_raw = nuevo.get(k, '') or ''
        v_ant = str(v_ant_raw)
        v_new = str(v_new_raw)
        if k == 'productos':
            if v_ant != v_new:
                try:
                    import json as _j
                    _ant_list = _j.loads(v_ant_raw) if v_ant_raw else []
                    _new_list = _j.loads(v_new_raw) if v_new_raw else []
                    # Convertir a dict por Item para comparar
                    _ant_dict = {it.get('Item','?'): it.get('Cantidad',0) for it in _ant_list}
                    _new_dict = {it.get('Item','?'): it.get('Cantidad',0) for it in _new_list}
                    _detalles = []
                    # Items agregados
                    for item, cant in _new_dict.items():
                        if item not in _ant_dict:
                            _detalles.append(f"+ {item} (x{cant})")
                    # Items eliminados
                    for item, cant in _ant_dict.items():
                        if item not in _new_dict:
                            _detalles.append(f"- {item} (x{cant})")
                    # Items con cantidad modificada
                    for item in _ant_dict:
                        if item in _new_dict and _ant_dict[item] != _new_dict[item]:
                            _detalles.append(f"~ {item}: {_ant_dict[item]} → {_new_dict[item]}")
                    if _detalles:
                        cambios['Productos'] = {'antes': '—', 'despues': chr(10).join(_detalles)}
                    else:
                        cambios[label] = {'antes': '(carrito anterior)', 'despues': '(carrito actualizado)'}
                except:
                    cambios[label] = {'antes': '(carrito anterior)', 'despues': '(carrito actualizado)'}
        elif v_ant != v_new:
            cambios[label] = {'antes': v_ant or '—', 'despues': v_new or '—'}

    # Dirección cliente completa (dirección + comuna + región juntas)
    def _dir_completa(d, prefix_dir, prefix_com, prefix_reg):
        parts = [
            str(d.get(prefix_dir, '') or '').strip(),
            str(d.get(prefix_com, '') or '').strip(),
            str(d.get(prefix_reg, '') or '').strip(),
        ]
        return ', '.join(p for p in parts if p) or ''

    _dir_cli_ant = _dir_completa(anterior, 'cliente_direccion', 'cliente_comuna', 'cliente_region')
    _dir_cli_new = _dir_completa(nuevo,    'cliente_direccion', 'cliente_comuna', 'cliente_region')
    if _dir_cli_ant != _dir_cli_new:
        cambios['Dirección cliente'] = {'antes': _dir_cli_ant or '—', 'despues': _dir_cli_new or '—'}

    _dir_ins_ant = _dir_completa(anterior, 'proyecto_direccion', 'proyecto_comuna', 'proyecto_region')
    _dir_ins_new = _dir_completa(nuevo,    'proyecto_direccion', 'proyecto_comuna', 'proyecto_region')
    if _dir_ins_ant != _dir_ins_new:
        cambios['Dirección instalación'] = {'antes': _dir_ins_ant or '—', 'despues': _dir_ins_new or '—'}

    return cambios


def generar_pdf_log(numero, logs):
    """PDF de auditoría — diseño limpio tipo registro de cambios."""
    import io as _io, os as _os
    from datetime import datetime as _dt, timezone, timedelta
    from collections import OrderedDict
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, KeepTogether)

    # ── Paleta ───────────────────────────────────────────────
    C_BG       = colors.HexColor('#0d1117')
    C_AZUL     = colors.HexColor('#58a6ff')
    C_VERDE    = colors.HexColor('#3fb950')
    C_AMARILLO = colors.HexColor('#d29922')
    C_GRIS     = colors.HexColor('#8b949e')
    C_LINE     = colors.HexColor('#e2e8f0')
    C_HDR_BG   = colors.HexColor('#f6f8fa')
    C_ANTES    = colors.HexColor('#ffdcd7')
    C_DESPUES  = colors.HexColor('#ccffd8')
    C_ANT_TXT  = colors.HexColor('#cf222e')
    C_DEP_TXT  = colors.HexColor('#1a7f37')

    _tz    = timezone(timedelta(hours=-3))
    _ahora = _dt.now(_tz).strftime("%d/%m/%Y %H:%M")
    _n_mods = len([l for l in logs if l.get("tipo_cambio") == "modificacion"])
    _n_crea = len([l for l in logs if l.get("tipo_cambio") == "creacion"])
    _n_total = len(logs)
    _MESES = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",
              7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}

    # Campos que NO deben formatearse como moneda
    _CAMPOS_TEXTO = {'Teléfono','RUT cliente','RUT empresa','Correo','Nombre cliente',
                     'Asesor','Dirección cliente','Dirección instalación','Descripción del proyecto',
                     'Estado','Empresa','Tipo cliente','Comuna cliente','Región cliente'}

    def _fmt_val(v, campo=None):
        if campo and campo in _CAMPOS_TEXTO:
            return str(v) if v else "—"
        try:
            _s = str(v).strip()
            # Si ya tiene formato chileno con $ y puntos, devolverlo tal cual
            if _s.startswith("$"):
                return _s
            # Convertir directamente desde el número (float/int)
            # sin eliminar puntos que podrían ser decimales
            _n = float(_s)
            if abs(_n) > 999:
                return "$" + "{:,.0f}".format(round(_n)).replace(",",".")
            return _s
        except: return str(v) if v else "—"

    def _hora_chile(fs):
        try:
            return _dt.fromisoformat(fs.replace("Z","+00:00")).astimezone(_tz).strftime("%H:%M")
        except: return fs[11:16]

    def _fecha_chile(fs):
        try:
            return _dt.fromisoformat(fs.replace("Z","+00:00")).astimezone(_tz).strftime("%Y-%m-%d")
        except: return fs[:10]

    buf = _io.BytesIO()
    PW, PH = A4
    ML = MR = 2.2*cm
    TM = 2.8*cm   # espacio suficiente para banda + margen
    BM = 2.0*cm

    # ── Header / Footer ──────────────────────────────────────
    def _hf(cv, doc):
        cv.saveState()
        pw, ph = doc.pagesize

        # Banda oscura — altura fija 1.4cm pegada al borde superior
        banda_h = 1.4*cm
        cv.setFillColor(C_BG)
        cv.rect(0, ph - banda_h, pw, banda_h, fill=1, stroke=0)

        # Logo centrado dentro de la banda
        if _os.path.exists("logo.png"):
            from reportlab.lib.utils import ImageReader
            _img = ImageReader("logo.png")
            _iw, _ih = _img.getSize()
            _lw = 2.6*cm
            _lh = _lw * (_ih / float(_iw))
            _ly = ph - banda_h + (banda_h - _lh) / 2
            cv.drawImage(_img, x=(pw - _lw)/2, y=_ly,
                         width=_lw, height=_lh,
                         preserveAspectRatio=True, mask="auto")

        # Etiquetas laterales dentro de la banda
        cv.setFont("Helvetica-Bold", 6.5)
        cv.setFillColor(C_GRIS)
        cv.drawString(ML, ph - banda_h/2 - 2.5, "AUDIT LOG")
        cv.setFont("Helvetica", 6.5)
        cv.drawRightString(pw - MR, ph - banda_h/2 - 2.5, numero)

        # Línea azul fina separando banda del contenido
        cv.setStrokeColor(C_AZUL)
        cv.setLineWidth(1.2)
        cv.line(ML, ph - banda_h - 1, pw - MR, ph - banda_h - 1)

        # Footer
        cv.setStrokeColor(C_LINE)
        cv.setLineWidth(0.4)
        cv.line(ML, BM - 0.1*cm, pw - MR, BM - 0.1*cm)
        cv.setFont("Helvetica", 6.5)
        cv.setFillColor(C_GRIS)
        cv.drawString(ML, BM - 0.45*cm,
            "Inversiones Container House SpA  ·  RUT 78.268.851-0  ·  Documento interno confidencial")
        cv.drawRightString(pw - MR, BM - 0.45*cm, f"Pág. {doc.page}  ·  {_ahora}")
        cv.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=ML, rightMargin=MR,
                            topMargin=TM, bottomMargin=BM)

    base = getSampleStyleSheet()
    def _sty(name, **kw):
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    s_ep     = _sty("ep",  fontName="Helvetica-Bold", fontSize=20,
                            textColor=C_BG, spaceAfter=2)
    s_sub    = _sty("sub", fontName="Helvetica", fontSize=9,
                            textColor=C_GRIS, spaceAfter=14)
    s_dia    = _sty("dia", fontName="Helvetica-Bold", fontSize=7.5,
                            textColor=C_AZUL, spaceBefore=12, spaceAfter=3)
    s_normal = _sty("nor", fontName="Helvetica", fontSize=8.5, leading=13)
    s_small  = _sty("sm",  fontName="Helvetica", fontSize=7.5,
                            textColor=C_GRIS, leading=11)
    s_mono   = _sty("mn",  fontName="Courier", fontSize=7.5, leading=11)
    s_campo  = _sty("cp",  fontName="Helvetica-Bold", fontSize=7.5, leading=11)
    s_antes  = _sty("an",  fontName="Helvetica", fontSize=7.5,
                            textColor=C_ANT_TXT, leading=11)
    s_dep    = _sty("dp",  fontName="Helvetica", fontSize=7.5,
                            textColor=C_DEP_TXT, leading=11)

    story = []

    # ── Título ───────────────────────────────────────────────
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph("Historial de cambios", s_sub))
    story.append(Paragraph(numero, s_ep))
    story.append(Spacer(1, 6))

    # Línea divisoria fina bajo el título
    story.append(Table([[""]], colWidths=[doc.width],
                        style=[("LINEBELOW", (0,0), (0,0), 0.8, C_BG),
                               ("TOPPADDING", (0,0), (0,0), 0),
                               ("BOTTOMPADDING", (0,0), (0,0), 0)]))
    story.append(Spacer(1, 10))

    # ── KPIs ────────────────────────────────────────────────
    kpi_w = doc.width / 4
    kpi_data = [[
        Paragraph(f'<font size="17"><b>{_n_total}</b></font><br/>'
                  f'<font color="#8b949e" size="7">registros totales</font>', s_normal),
        Paragraph(f'<font size="17" color="#3fb950"><b>{_n_crea}</b></font><br/>'
                  f'<font color="#8b949e" size="7">creaciones</font>', s_normal),
        Paragraph(f'<font size="17" color="#58a6ff"><b>{_n_mods}</b></font><br/>'
                  f'<font color="#8b949e" size="7">modificaciones</font>', s_normal),
        Paragraph(f'<font size="8.5"><b>{_ahora}</b></font><br/>'
                  f'<font color="#8b949e" size="7">generado (hora Chile)</font>', s_normal),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=[kpi_w]*4)
    kpi_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), C_HDR_BG),
        ("LINEABOVE",     (0,0), (0,0),   2, C_BG),
        ("LINEABOVE",     (1,0), (1,0),   2, C_VERDE),
        ("LINEABOVE",     (2,0), (2,0),   2, C_AZUL),
        ("LINEABOVE",     (3,0), (3,0),   2, C_AMARILLO),
        ("LINEAFTER",     (0,0), (2,0),   0.4, C_LINE),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 18))

    if not logs:
        story.append(Paragraph("Sin registros.", s_small))
    else:
        grupos = OrderedDict()
        for lg in logs:
            _f = _fecha_chile(lg.get("fecha",""))
            grupos.setdefault(_f, []).append(lg)

        for fecha_key, items in grupos.items():
            try:
                _fd = _dt.fromisoformat(fecha_key)
                _titulo_dia = f"{_fd.day} de {_MESES[_fd.month]} de {_fd.year}".upper()
            except:
                _titulo_dia = fecha_key

            # Separador de día
            dia_tbl = Table(
                [[Paragraph(f"── {_titulo_dia}", s_dia),
                  Paragraph(f"{len(items)} evento{'s' if len(items)!=1 else ''}",
                            _sty("cnt", fontName="Helvetica", fontSize=7.5,
                                 textColor=C_GRIS, alignment=TA_RIGHT))]],
                colWidths=[doc.width*0.75, doc.width*0.25])
            dia_tbl.setStyle(TableStyle([
                ("LINEBELOW",     (0,0), (-1,0), 0.4, C_LINE),
                ("TOPPADDING",    (0,0), (-1,-1), 0),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("VALIGN",        (0,0), (-1,-1), "BOTTOM"),
                ("LEFTPADDING",   (0,0), (-1,-1), 0),
                ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ]))
            story.append(dia_tbl)

            for lg in items:
                _hora   = _hora_chile(lg.get("fecha",""))
                _asesor = lg.get("asesor","") or "Sistema"
                _tipo   = lg.get("tipo_cambio","").upper()
                _det    = lg.get("detalle", {})

                _dot_color = C_VERDE if _tipo == "CREACION" else C_AZUL
                _bdg_bg    = colors.HexColor("#d1f5db") if _tipo == "CREACION" else colors.HexColor("#dbeafe")
                _bdg_txt   = colors.HexColor("#1a7f37") if _tipo == "CREACION" else colors.HexColor("#1d4ed8")

                # Fila encabezado del evento
                hdr_tbl = Table([[
                    Paragraph(f"<b>{_hora}</b>",
                               _sty(f"hr{_hora}", fontName="Courier-Bold",
                                    fontSize=9, textColor=_dot_color)),
                    Paragraph(_tipo,
                               _sty(f"bdg{_hora}", fontName="Helvetica-Bold",
                                    fontSize=7, textColor=_bdg_txt, backColor=_bdg_bg,
                                    borderPadding=(2,5,2,5))),
                    Paragraph(f"<b>{_asesor}</b>",
                               _sty(f"as{_hora}", fontName="Helvetica",
                                    fontSize=8, textColor=C_GRIS)),
                ]], colWidths=[2.0*cm, 3.2*cm, doc.width - 5.2*cm])
                hdr_tbl.setStyle(TableStyle([
                    ("TOPPADDING",    (0,0), (-1,-1), 7),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                    ("LEFTPADDING",   (0,0), (-1,-1), 0),
                    ("RIGHTPADDING",  (0,0), (-1,-1), 4),
                    ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                    ("LINEAFTER",     (0,0), (0,-1),  1.2, _dot_color),
                    ("LEFTPADDING",   (1,0), (-1,-1), 8),
                ]))
                story.append(hdr_tbl)

                # Detalle
                if isinstance(_det, dict):
                    if "mensaje" in _det:
                        msg_tbl = Table([[
                            Paragraph("", s_small),
                            Paragraph(f"→ {_det['mensaje']}", s_small),
                        ]], colWidths=[2.0*cm, doc.width - 2.0*cm])
                        msg_tbl.setStyle(TableStyle([
                            ("TOPPADDING",    (0,0), (-1,-1), 2),
                            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
                            ("LEFTPADDING",   (0,0), (-1,-1), 0),
                            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
                            ("LINEAFTER",     (0,0), (0,-1),  1.2, _dot_color),
                            ("LEFTPADDING",   (1,0), (1,-1),  10),
                            ("VALIGN",        (0,0), (-1,-1), "TOP"),
                        ]))
                        story.append(msg_tbl)
                    else:
                        cam_rows = [[
                            Paragraph("CAMPO", s_campo),
                            Paragraph("ANTES", s_campo),
                            Paragraph("DESPUÉS", s_campo),
                        ]]
                        for _c, _v in _det.items():
                            if isinstance(_v, dict):
                                _a = _fmt_val(str(_v.get("antes","—")), campo=_c)
                                _d_raw = str(_v.get("despues","—"))
                                # Si tiene saltos de línea (detalle carrito), formatear con <br/>
                                if chr(10) in _d_raw:
                                    _d = _d_raw.replace(chr(10), '<br/>')
                                    _d_para = Paragraph(_d, s_dep)
                                else:
                                    _d = _fmt_val(_d_raw, campo=_c)
                                    _d_para = Paragraph(_d[:120], s_dep)
                                _a_para = Paragraph(_a[:70], s_antes)
                            else:
                                _a_para = Paragraph("—", s_antes)
                                _d_para = Paragraph(_fmt_val(str(_v), campo=_c)[:70], s_dep)
                            cam_rows.append([
                                Paragraph(_c, s_mono),
                                _a_para,
                                _d_para,
                            ])

                        if len(cam_rows) > 1:
                            _cw1, _cw2, _cw3 = 4.2*cm, 5.4*cm, 5.4*cm
                            cam_tbl = Table(cam_rows, colWidths=[_cw1, _cw2, _cw3])
                            cam_styles = [
                                ("BACKGROUND",    (0,0), (-1,0),  C_HDR_BG),
                                ("LINEBELOW",     (0,0), (-1,0),  0.5, C_LINE),
                                ("GRID",          (0,0), (-1,-1), 0.3, C_LINE),
                                ("TOPPADDING",    (0,0), (-1,-1), 4),
                                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                                ("LEFTPADDING",   (0,0), (-1,-1), 6),
                                ("RIGHTPADDING",  (0,0), (-1,-1), 6),
                                ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
                            ]
                            for _ri in range(1, len(cam_rows)):
                                cam_styles.append(("BACKGROUND", (1,_ri), (1,_ri), C_ANTES))
                                cam_styles.append(("BACKGROUND", (2,_ri), (2,_ri), C_DESPUES))
                            cam_tbl.setStyle(TableStyle(cam_styles))

                            wrap_tbl = Table(
                                [["", cam_tbl]],
                                colWidths=[2.0*cm, doc.width - 2.0*cm])
                            wrap_tbl.setStyle(TableStyle([
                                ("TOPPADDING",    (0,0), (-1,-1), 3),
                                ("BOTTOMPADDING", (0,0), (-1,-1), 10),
                                ("LEFTPADDING",   (0,0), (-1,-1), 0),
                                ("RIGHTPADDING",  (0,0), (-1,-1), 0),
                                ("LINEAFTER",     (0,0), (0,-1),  1.2, _dot_color),
                                ("LEFTPADDING",   (1,0), (1,-1),  10),
                                ("VALIGN",        (0,0), (-1,-1), "TOP"),
                            ]))
                            story.append(KeepTogether(wrap_tbl))

                story.append(Spacer(1, 2))

    doc.build(story, onFirstPage=_hf, onLaterPages=_hf)
    buf.seek(0)
    return buf.read()



# =========================================================
# FUNCIONES DE SUPABASE PARA COTIZACIONES
# =========================================================
def guardar_cotizacion(numero, cliente, asesor, proyecto, productos, config, totales, plano_nombre=None, plano_datos=None, usuario_logueado=None):
    try:
        fecha_actual = datetime.now().isoformat()
        productos_json = json.dumps(productos, ensure_ascii=False)
        tiene_margen = float(config.get('margen', 0) or 0) > 0
        tiene_plano = plano_datos is not None
        datos_completos = all([
            str(cliente.get('Nombre', '')).strip(),
            str(cliente.get('Correo', '')).strip()
        ])
        asesor_completo = any([
            str(asesor.get('Nombre Ejecutivo', '')).strip(),
            str(asesor.get('Correo Ejecutivo', '')).strip(),
            str(asesor.get('Teléfono Ejecutivo', '')).strip()
        ])
        if not datos_completos or not asesor_completo:
            estado = "INCOMPLETO CON PLANO" if tiene_plano else "INCOMPLETO"
        elif tiene_margen:
            estado = "AUTORIZADO CON PLANO" if tiene_plano else "AUTORIZADO"
        else:
            estado = "BORRADOR CON PLANO" if tiene_plano else "BORRADOR"

        response = supabase_admin.table('cotizaciones').select('*').eq('numero', numero).execute()
        existe = len(response.data) > 0

        plano_url = None
        if plano_datos:
            if existe and response.data[0].get('plano_url'):
                eliminar_plano_de_storage(response.data[0]['plano_url'])
            plano_url, error = guardar_plano_en_storage(plano_datos, numero, plano_nombre)
            if error:
                st.error(f"Error al subir plano: {error}")

        data = {
            'numero': numero,
            'fecha_modificacion': fecha_actual,
            'estado': estado,
            'cliente_nombre': str(cliente.get('Nombre', '') or ''),
            'cliente_rut': str(cliente.get('RUT', '') or ''),
            'cliente_email': str(cliente.get('Correo', '') or ''),
            'cliente_telefono': str(cliente.get('Teléfono', '') or '').strip(),
            'cliente_direccion': str(cliente.get('Dirección', '') or ''),
            'cliente_comuna': str(cliente.get('ComunaCliente', '') or ''),
            'cliente_region': str(cliente.get('RegionCliente', '') or ''),
            'proyecto_direccion': str(cliente.get('DireccionProyecto', '') or ''),
            'proyecto_comuna': str(cliente.get('ComunaProyecto', '') or ''),
            'proyecto_region': str(cliente.get('RegionProyecto', '') or ''),
            'cliente_tipo': str(cliente.get('TipoCliente', 'natural') or 'natural'),
            'cliente_empresa': str(cliente.get('EmpresaCliente', '') or ''),
            'cliente_rut_empresa': str(cliente.get('RutEmpresa', '') or ''),
            'asesor_nombre': str(asesor.get('Nombre Ejecutivo', '') or ''),
            'asesor_email': str(asesor.get('Correo Ejecutivo', '') or ''),
            'asesor_telefono': str(asesor.get('Teléfono Ejecutivo', '') or ''),
            'proyecto_fecha_inicio': str(proyecto.get('fecha_inicio', '') or ''),
            'proyecto_fecha_termino': str(proyecto.get('fecha_termino', '') or ''),
            'proyecto_dias_validez': int(proyecto.get('dias_validez', 0) or 0),
            'proyecto_observaciones': str(proyecto.get('observaciones', '') or ''),
            'productos': productos_json,
            'config_margen': float(config.get('margen', 0) or 0),
            'config_modo_admin': 1 if config.get('modo_admin', False) else 0,
            'total_subtotal_sin_margen': float(totales.get('subtotal_sin_margen', 0) or 0),
            'total_subtotal_con_margen': float(totales.get('subtotal_con_margen', 0) or 0),
            'total_iva': float(totales.get('iva', 0) or 0),
            'total_total': float(totales.get('total', 0) or 0),
            'total_margen_valor': float(totales.get('margen_valor', 0) or 0),
            'total_comision_vendedor': float(totales.get('comision_vendedor', 0) or 0),
            'total_comision_supervisor': float(totales.get('comision_supervisor', 0) or 0),
            'total_utilidad_real': float(totales.get('utilidad_real', 0) or 0),
            'plano_nombre': plano_nombre if plano_datos else (response.data[0].get('plano_nombre') if existe else None),
            'plano_url': plano_url if plano_datos else (response.data[0].get('plano_url') if existe else None),
            'user_id': st.session_state.get('auth_user') or None,
            'asesor_email': str(asesor.get('Correo Ejecutivo', '') or st.session_state.get('auth_email', '') or '')
        }

        if existe:
            _anterior = response.data[0]
            response = supabase_admin.table('cotizaciones').update(data).eq('numero', numero).execute()
            # Registrar log de modificación con diff
            _cambios = _diff_datos(_anterior, data)
            if _cambios:
                _actor = usuario_logueado or str(asesor.get('Nombre Ejecutivo', '') or 'Sistema')
                registrar_log(
                    numero=numero,
                    asesor=_actor,
                    tipo_cambio='modificacion',
                    detalle_dict=_cambios
                )
        else:
            data['fecha_creacion'] = fecha_actual
            response = supabase_admin.table('cotizaciones').insert(data).execute()
            # Registrar log de creación
            _actor = usuario_logueado or str(asesor.get('Nombre Ejecutivo', '') or 'Sistema')
            registrar_log(
                numero=numero,
                asesor=_actor,
                tipo_cambio='creacion',
                detalle_dict={'mensaje': f'Cotización {numero} creada'}
            )

        return True
    except Exception as e:
        st.error(f"❌ Error al guardar cotización: {e}")
        return False

def exportar_csv_completo():
    """Exporta todas las cotizaciones de Supabase a CSV."""
    try:
        response = supabase.table('cotizaciones').select(
            'numero', 'fecha_creacion', 'fecha_modificacion',
            'cliente_nombre', 'cliente_rut', 'cliente_email', 'cliente_telefono',
            'cliente_direccion', 'cliente_comuna', 'cliente_region',
            'cliente_tipo', 'cliente_empresa', 'cliente_rut_empresa',
            'proyecto_direccion', 'proyecto_comuna', 'proyecto_region',
            'asesor_nombre', 'asesor_email', 'asesor_telefono',
            'config_margen',
            'total_subtotal_sin_margen', 'total_subtotal_con_margen',
            'total_iva', 'total_total', 'total_margen_valor',
            'total_comision_vendedor', 'total_comision_supervisor', 'total_utilidad_real',
            'estado', 'plano_nombre', 'plano_url',
            'contrato_generado', 'contrato_fecha'
        ).order('fecha_creacion', desc=True).execute()
        if not response.data:
            return None
        df = pd.DataFrame(response.data)
        df.columns = [
            'N° Presupuesto', 'Fecha Creación', 'Fecha Modificación',
            'Cliente', 'RUT', 'Email Cliente', 'Teléfono Cliente',
            'Dirección Cliente', 'Comuna Cliente', 'Región Cliente',
            'Tipo Cliente', 'Empresa', 'RUT Empresa',
            'Dirección Proyecto', 'Comuna Proyecto', 'Región Proyecto',
            'Asesor', 'Email Asesor', 'Teléfono Asesor',
            'Margen %',
            'Subtotal sin Margen', 'Subtotal con Margen',
            'IVA', 'Total con IVA', 'Valor Margen',
            'Comisión Vendedor', 'Comisión Supervisor', 'Utilidad Real',
            'Estado', 'Nombre Plano', 'URL Plano',
            'Contrato Generado', 'Fecha Contrato'
        ]
        return df.to_csv(index=False).encode('utf-8-sig')
    except Exception as e:
        st.error(f"Error al exportar: {e}")
        return None

def buscar_cotizaciones(termino=None, tipo_busqueda='numero'):
    try:
        query = supabase.table('cotizaciones').select(
            'numero', 'cliente_nombre', 'asesor_nombre', 'fecha_creacion',
            'total_total', 'config_margen', 'cliente_rut', 'cliente_email',
            'asesor_email', 'asesor_telefono', 'plano_url', 'contrato_generado', 'cliente_empresa'
        )
        # Filtrar por usuario si es ejecutivo (no admin ni root)
        _rol_q = st.session_state.get('rol_usuario', 'ejecutivo')
        if _rol_q == 'ejecutivo':
            _email = st.session_state.get('auth_email', '')
            if _email:
                # ilike es case-insensitive, cubre tanto mayúsculas como minúsculas
                query = query.ilike('asesor_email', _email.strip())
        if termino and termino.strip():
            campo_map = {
                'numero': 'numero',
                'cliente': 'cliente_nombre',
                'asesor': 'asesor_nombre'
            }
            campo = campo_map.get(tipo_busqueda, 'numero')
            query = query.ilike(campo, f'%{termino}%')
        query = query.order('fecha_creacion', desc=True).limit(50)
        response = query.execute()
        resultados = []
        for row in response.data:
            resultados.append((
                row.get('numero', ''),
                row.get('cliente_nombre', '') or '',
                row.get('asesor_nombre', '') or '',
                row.get('fecha_creacion', '') or '',
                row.get('total_total', 0) or 0,
                row.get('config_margen', 0) or 0,
                row.get('cliente_rut', '') or '',
                row.get('cliente_email', '') or '',
                row.get('asesor_email', '') or '',
                row.get('asesor_telefono', '') or '',
                1 if row.get('plano_url') else 0,
                1 if row.get('contrato_generado') else 0,
                row.get('cliente_empresa', '') or ''
            ))
        # Agregar conteo de logs
        numeros_ep = [r[0] for r in resultados]
        _log_counts = contar_logs(numeros_ep)
        resultados = [r + (_log_counts.get(r[0], 0),) for r in resultados]
        return resultados
    except Exception as e:
        st.error(f"Error en búsqueda: {e}")
        return []

def cargar_cotizacion(numero):
    try:
        if not numero:
            return None
        response = supabase.table('cotizaciones').select('*').eq('numero', numero).execute()
        if response.data:
            cotizacion = response.data[0]
            # Manejar productos como string JSON o como lista directamente
            productos = cotizacion['productos']
            if isinstance(productos, str):
                cotizacion['productos'] = json.loads(productos)
            elif isinstance(productos, list):
                cotizacion['productos'] = productos
            else:
                cotizacion['productos'] = []
            if cotizacion.get('plano_url'):
                cotizacion['plano_datos'] = None
            return cotizacion
        return None
    except Exception as e:
        st.error(f"Error al cargar cotización: {e}")
        return None

def eliminar_cotizacion(numero):
    try:
        response = supabase.table('cotizaciones').select('plano_url').eq('numero', numero).execute()
        if response.data and response.data[0].get('plano_url'):
            eliminar_plano_de_storage(response.data[0]['plano_url'])
        response = supabase.table('cotizaciones').delete().eq('numero', numero).execute()
        return True
    except Exception as e:
        st.error(f"Error al eliminar: {e}")
        return False

def actualizar_estado_cotizacion(numero, estado):
    try:
        fecha_actual = datetime.now().isoformat()
        response = supabase.table('cotizaciones').update({
            'estado': estado,
            'fecha_modificacion': fecha_actual
        }).eq('numero', numero).execute()
        return True
    except Exception as e:
        st.error(f"Error al actualizar estado: {e}")
        return False

def generar_numero_unico():
    intentos = 0
    while intentos < 20:
        numero = f"EP-{random.randint(10000, 99999)}"
        try:
            response = supabase.table('cotizaciones').select('numero').eq('numero', numero).execute()
            if not response.data:
                return numero
        except:
            pass
        intentos += 1
    return f"EP-{int(datetime.now().timestamp())}"

# =========================================================
# FUNCIÓN PARA CARGAR COTIZACIÓN EN EL SISTEMA
# =========================================================
def preparar_carga_cotizacion(numero_cotizacion):
    cotizacion = cargar_cotizacion(numero_cotizacion)
    if cotizacion:
        tiene_margen = cotizacion.get('config_margen', 0) > 0
        if tiene_margen and not st.session_state.modo_admin:
            return False
        else:
            st.session_state.cotizacion_a_cargar = cotizacion
            st.session_state.cargar_cotizacion_trigger = True
            return True
    return False

def ejecutar_carga_cotizacion():
    if st.session_state.cargar_cotizacion_trigger and st.session_state.cotizacion_a_cargar:
        cotizacion = st.session_state.cotizacion_a_cargar
        st.session_state.carrito = cotizacion['productos']
        st.session_state.nombre_input = cotizacion.get('cliente_nombre', '')
        rut_valor = cotizacion.get('cliente_rut', '')
        st.session_state.rut_display = rut_valor
        st.session_state.rut_raw = re.sub(r'[^0-9kK]', '', rut_valor)
        if st.session_state.rut_raw and len(st.session_state.rut_raw) >= 2:
            valido, mensaje = validar_rut(st.session_state.rut_raw)
            st.session_state.rut_valido = valido
            st.session_state.rut_mensaje = mensaje
        else:
            st.session_state.rut_valido = False
            st.session_state.rut_mensaje = "RUT incompleto"
        st.session_state.correo_input = cotizacion.get('cliente_email', '')
        st.session_state.telefono_raw = cotizacion.get('cliente_telefono', '')
        st.session_state.direccion_input    = cotizacion.get('cliente_direccion', '')
        st.session_state.cliente_comuna      = cotizacion.get('cliente_comuna', '')
        st.session_state.cliente_region      = cotizacion.get('cliente_region', '')
        st.session_state.proyecto_direccion  = cotizacion.get('proyecto_direccion', '')
        st.session_state.proyecto_comuna     = cotizacion.get('proyecto_comuna', '')
        st.session_state.proyecto_region     = cotizacion.get('proyecto_region', '')
        st.session_state.cliente_tipo         = cotizacion.get('cliente_tipo', 'natural')
        st.session_state.cliente_empresa      = cotizacion.get('cliente_empresa', '')
        st.session_state.cliente_rut_empresa  = cotizacion.get('cliente_rut_empresa', '')
        _rut_emp_raw = re.sub(r'[^0-9kK]', '', cotizacion.get('cliente_rut_empresa', ''))
        st.session_state.rut_empresa_raw     = _rut_emp_raw
        st.session_state.rut_empresa_display = formatear_rut(_rut_emp_raw) if _rut_emp_raw else ''
        nombre_asesor = cotizacion.get('asesor_nombre', '')
        st.session_state.asesor_seleccionado = nombre_asesor if nombre_asesor else "Seleccionar asesor"
        st.session_state.correo_asesor = cotizacion.get('asesor_email', '')
        st.session_state.telefono_asesor = cotizacion.get('asesor_telefono', '')
        if cotizacion.get('proyecto_fecha_inicio'):
            try:
                st.session_state.fecha_inicio = datetime.strptime(cotizacion['proyecto_fecha_inicio'], '%Y-%m-%d').date()
            except:
                st.session_state.fecha_inicio = datetime.now().date()
        else:
            st.session_state.fecha_inicio = datetime.now().date()
        if cotizacion.get('proyecto_fecha_termino'):
            try:
                st.session_state.fecha_termino = datetime.strptime(cotizacion['proyecto_fecha_termino'], '%Y-%m-%d').date()
            except:
                st.session_state.fecha_termino = (datetime.now() + timedelta(days=15)).date()
        else:
            st.session_state.fecha_termino = (datetime.now() + timedelta(days=15)).date()
        st.session_state.observaciones_input = cotizacion.get('proyecto_observaciones', '')
        # Preservar modo_admin del usuario actual — no sobreescribir con el valor de la cotización
        # Solo activar si la cotización lo tenía guardado, pero nunca desactivar si el usuario es admin/root
        _modo_admin_cot = bool(cotizacion.get('config_modo_admin', False))
        if not st.session_state.get('es_supervisor'):
            st.session_state.modo_admin = _modo_admin_cot
        margen_valor = cotizacion.get('config_margen')
        try:
            st.session_state.margen = float(margen_valor) if margen_valor is not None else 0.0
        except (ValueError, TypeError):
            st.session_state.margen = 0.0
        plano_nombre = cotizacion.get('plano_nombre')
        plano_url = cotizacion.get('plano_url')
        if plano_nombre and plano_url:
            st.session_state.plano_nombre = plano_nombre
            st.session_state.pdf_url = plano_url
            st.session_state.plano_adjunto = None
        else:
            st.session_state.plano_nombre = ""
            st.session_state.plano_adjunto = None
            st.session_state.pdf_url = None
        st.session_state.cotizacion_cargada = cotizacion.get('numero', '')
        st.session_state.counter += 100
        st.session_state.mostrar_visor = False
        st.session_state.pdf_actual = None
        st.session_state.pdf_nombre = ""
        st.session_state.numero_en_visor = None
        st.session_state.cargar_cotizacion_trigger = False
        st.session_state.cotizacion_a_cargar = None
        # Resetear hash y marcar como recién cargado para suprimir FAB
        st.session_state.hash_ultimo_guardado = calcular_hash_estado()
        st.session_state.recien_cargado = True
        return True
    return False

# =========================================================
# EJECUTAR CARGA DE COTIZACIÓN SI HAY TRIGGER
# =========================================================
ejecutar_carga_cotizacion()

# =========================================================
# CSS PERSONALIZADO
# =========================================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,300;0,400;0,500;0,600;0,700;0,800;1,300;1,400;1,500;1,600;1,700;1,800&family=Montserrat:wght@700;800;900&family=JetBrains+Mono:wght@400;500&display=swap');



    /* ══ FAB margen iframe — sacarlo del flujo ══ */
    /* El último iframe de components antes del final se posiciona fixed */
    [data-testid="stBottom"] ~ div iframe:last-of-type {
        position: fixed !important;
        bottom: 0 !important;
        left: 0 !important;
        width: 400px !important;
        height: 300px !important;
        border: none !important;
        z-index: 999990 !important;
        pointer-events: auto !important;
        background: transparent !important;
    }

    /* ══ Sombra tabla data_editor / dataframe ══ */
    div[data-testid="stDataFrame"] > div,
    div[data-testid="stDataEditor"] > div {
        border-radius: 16px !important;
        box-shadow: 0 4px 20px rgba(91, 124, 250, 0.08), 0 1px 6px rgba(0,0,0,0.06) !important;
        border: 1px solid rgba(91,124,250,0.15) !important;
        overflow: hidden !important;
        transition: box-shadow 0.25s ease, transform 0.25s ease !important;
        background: #ffffff !important;
    }
    div[data-testid="stDataFrame"] > div:hover,
    div[data-testid="stDataEditor"] > div:hover {
        box-shadow: 0 8px 32px rgba(91, 124, 250, 0.16), 0 2px 10px rgba(0,0,0,0.08) !important;
        transform: translateY(-2px) !important;
    }

    /* ══ Sombra flotante para containers con borde ══ */
    [data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"] {
        box-shadow: 0 4px 20px rgba(91, 124, 250, 0.08), 0 1px 6px rgba(0,0,0,0.06) !important;
        border: 1px solid rgba(91,124,250,0.15) !important;
        border-radius: 16px !important;
        transition: box-shadow 0.25s ease, transform 0.25s ease !important;
        background: #ffffff !important;
    }
    [data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"]:hover {
        box-shadow: 0 8px 32px rgba(91, 124, 250, 0.16), 0 2px 10px rgba(0,0,0,0.08) !important;
        transform: translateY(-2px) !important;
    }


    #MainMenu { display: none !important; }
    footer { display: none !important; }
    [data-testid="stToolbar"] { display: none !important; }
    [data-testid="stDecoration"] { display: none !important; }
    [data-testid="stStatusWidget"] { display: none !important; }
    [data-testid="stBottomBlockContainer"] { display: none !important; }
    [class*="viewerBadge"] { display: none !important; }
    [class*="ViewerBadge"] { display: none !important; }
    [class*="_viewerBadge"] { display: none !important; }
    [class*="profileContainer"] { display: none !important; }
    [class*="_profileContainer"] { display: none !important; }
    [class*="profilePreview"] { display: none !important; }
    [class*="_profilePreview"] { display: none !important; }
    a[href*="streamlit.io"] { display: none !important; }
    a[href*="github.com"] { display: none !important; }
    button[title="View fullscreen"] { display: none !important; }

    /* ══ BASE ══ */
    html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif !important; }
    .stApp { background-color: #f0f2f8 !important; }

    /* ══ INPUTS ══ */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stNumberInput > div > div > input,
    .stDateInput > div > div > input {
        background-color: #ffffff !important; color: #1a1d2e !important;
        border: 1.5px solid #e2e6f3 !important; border-radius: 10px !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important; font-size: 0.9rem !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
        transition: border-color 0.2s, box-shadow 0.2s !important;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #5b7cfa !important;
        box-shadow: 0 0 0 3px rgba(91,124,250,0.13) !important;
    }
    [data-baseweb="select"] > div {
        background-color: #ffffff !important; border: 1.5px solid #e2e6f3 !important;
        border-radius: 10px !important; color: #1a1d2e !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
    }
    [data-baseweb="select"] span,
    .stSelectbox > div > div,
    .stSelectbox > div > div > div { color: #1a1d2e !important; }
    .stTextInput label, .stSelectbox label, .stNumberInput label,
    .stDateInput label, .stTextArea label, .stRadio label,
    .stCheckbox label, .stFileUploader label {
        color: #5a6080 !important; font-weight: 600 !important;
        font-size: 0.8rem !important; letter-spacing: 0.04em !important;
        text-transform: uppercase !important;
    }

    /* ══ BOTONES ══ */
    .stButton > button {
        background-color: #ffffff !important; color: #2a3060 !important;
        border: 1.5px solid #dde1f0 !important; border-radius: 10px !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        font-weight: 600 !important; font-size: 0.875rem !important;
        transition: all 0.2s cubic-bezier(0.4,0,0.2,1) !important;
        box-shadow: 0 1px 4px rgba(0,0,0,0.06) !important;
    }
    .stButton > button:hover {
        background-color: #eef1ff !important; border-color: #5b7cfa !important;
        color: #2a3060 !important; transform: translateY(-1px) !important;
        box-shadow: 0 4px 14px rgba(91,124,250,0.18) !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #5b7cfa 0%, #8b5cf6 100%) !important;
        color: #ffffff !important; border: none !important;
        box-shadow: 0 4px 16px rgba(91,124,250,0.4) !important;
    }
    .stButton > button[kind="primary"]:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 24px rgba(91,124,250,0.5) !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #5b7cfa 0%, #8b5cf6 100%) !important;
        color: #ffffff !important; border: none !important; border-radius: 10px !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important; font-weight: 600 !important;
        box-shadow: 0 4px 16px rgba(91,124,250,0.35) !important;
    }
    .stDownloadButton > button:hover { transform: translateY(-1px) !important; }
    .stPopover > button {
        background-color: #ffffff !important; color: #2a3060 !important;
        border: 1.5px solid #dde1f0 !important; border-radius: 10px !important;
    }

    /* ══ TABS ══ */
    /* ── Tab scroll con flechas ── */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0 !important; border-bottom: 2px solid #e2e6f3 !important;
        padding: 0 !important; margin-bottom: 0 !important;
        background: transparent !important;
        overflow-x: auto !important;
        overflow-y: hidden !important;
        scrollbar-width: none !important;
        -ms-overflow-style: none !important;
        scroll-behavior: smooth !important;
    }
    .stTabs [data-baseweb="tab-list"]::-webkit-scrollbar { display: none !important; }

    /* ── Toast personalizado ── */
    div[data-testid="stToast"] {
        background: linear-gradient(135deg, #1d4ed8 0%, #2563eb 100%) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 16px 20px !important;
        min-width: 280px !important;
        box-shadow: 0 8px 24px rgba(37,99,235,0.4) !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
        font-family: 'Plus Jakarta Sans', sans-serif !important;
    }
    div[data-testid="stToast"] p,
    div[data-testid="stToast"] span,
    div[data-testid="stToast"] div {
        color: #ffffff !important;
        font-size: 1rem !important;
        font-weight: 600 !important;
    }
    div[data-testid="stToast"] button {
        color: rgba(255,255,255,0.7) !important;
        filter: brightness(10) !important;
    }
    /* Botones flecha */
    .tab-arrow {
        position: absolute; top: 0; z-index: 999;
        background: linear-gradient(90deg, #fff 60%, transparent);
        border: none; cursor: pointer; padding: 0 10px;
        height: 100%; font-size: 1.1rem; color: #5b7cfa;
        display: flex; align-items: center;
    }
    .tab-arrow-right { right: 0; background: linear-gradient(270deg, #fff 60%, transparent); }
    .tab-arrow-left  { left: 0; }
    .tab-arrow-wrap  { position: relative; }
    .stTabs [data-baseweb="tab-panel"] {
        padding-top: 1.5rem !important;
        border-top: none !important;
    }
    .stTabs > div > div:nth-child(2) {
        border-top: none !important;
        box-shadow: none !important;
    }
    hr { display: none !important; }
    .stTabs [data-baseweb="tab"] {
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        font-size: 0.88rem !important; font-weight: 900 !important;
        color: #7c85b3 !important; padding: 0.85rem 1.6rem !important;
        background: transparent !important; border: none !important;
        border-bottom: 3px solid transparent !important;
        margin-bottom: -2px !important; letter-spacing: 0.05em !important;
        text-transform: uppercase !important;
        -webkit-font-smoothing: antialiased !important;
        transition: all 0.2s ease !important;
    }
    .stTabs [data-baseweb="tab"] p,
    .stTabs [data-baseweb="tab"] span,
    .stTabs [data-baseweb="tab"] div {
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        font-weight: 900 !important;
        font-size: 0.88rem !important;
    }
    .stTabs [data-baseweb="tab"]:hover { color: #5b7cfa !important; background: rgba(91,124,250,0.05) !important; }
    .stTabs [aria-selected="true"] {
        color: #5b7cfa !important; border-bottom: 3px solid #5b7cfa !important;
        font-weight: 900 !important; background: rgba(91,124,250,0.06) !important;
    }

    /* ══ TABLA RESULTADOS ══ */
    .resultados-table {
        width: 100%; border-collapse: collapse; border-spacing: 0;
        font-family: 'Plus Jakarta Sans', sans-serif; font-size: 0.875rem;
        background: #ffffff;
    }
    .resultados-table th {
        background: linear-gradient(135deg, #1e2447 0%, #2a3060 100%) !important;
        color: #ffffff !important; font-weight: 700 !important;
        padding: 14px 16px !important; text-align: left !important;
        font-size: 0.75rem !important; letter-spacing: 0.07em !important;
        text-transform: uppercase !important;
        position: sticky !important; top: 0 !important; z-index: 2 !important;
    }
    .resultados-table td {
        padding: 12px 16px !important; border-bottom: 1px solid #f0f2f8 !important;
        color: #3a4070 !important; background-color: #ffffff !important;
        transition: background 0.15s !important;
    }
    .resultados-table tr:hover td { background-color: #f5f7ff !important; }
    .resultados-table tr:last-child td { border-bottom: none !important; }

    /* ══ METRIC CARDS ══ */
    .metric-card {
        background: linear-gradient(150deg, #1e2447 0%, #252d5a 100%);
        border-radius: 16px; padding: 1.4rem 1.5rem;
        box-shadow: 0 8px 28px rgba(30,36,71,0.22);
        border: 1px solid rgba(255,255,255,0.07);
        transition: all 0.25s cubic-bezier(0.4,0,0.2,1);
        height: 100%; position: relative; overflow: hidden;
    }
    .metric-card::before {
        content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px;
        background: linear-gradient(90deg, #5b7cfa, #8b5cf6);
    }
    .metric-card::after {
        content: ''; position: absolute; bottom: -30px; right: -30px;
        width: 100px; height: 100px; border-radius: 50%;
        background: rgba(91,124,250,0.08);
    }
    .metric-card:hover { transform: translateY(-4px); box-shadow: 0 16px 40px rgba(30,36,71,0.3); }
    .metric-title {
        font-size: 0.7rem; font-weight: 700; text-transform: uppercase;
        letter-spacing: 0.1em; color: #7b84b0; margin-bottom: 0.7rem;
    }
    .metric-value {
        font-size: 2.5rem; font-weight: 800; line-height: 1.05;
        letter-spacing: -0.04em; color: #e8ecff;
    }
    .metric-change { font-size: 0.75rem; color: #5c6494; margin-top: 0.35rem; }

    /* ══ TARJETAS COLOREADAS ══ */
    .metric-card-special {
        border-radius: 18px; padding: 1.5rem;
        box-shadow: 0 8px 28px rgba(0,0,0,0.14);
        transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
        border: 1px solid rgba(255,255,255,0.2);
        height: 100%; display: flex; flex-direction: column;
        position: relative; overflow: hidden;
    }
    .metric-card-special::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 120px; height: 120px; border-radius: 50%;
        background: rgba(255,255,255,0.1);
    }
    .metric-card-special::after {
        content: ''; position: absolute; bottom: -20px; left: -20px;
        width: 80px; height: 80px; border-radius: 50%;
        background: rgba(255,255,255,0.06);
    }
    .metric-card-special:hover { transform: translateY(-5px); box-shadow: 0 20px 50px rgba(0,0,0,0.2); }
    .metric-card-total { background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); }
    .metric-card-comisiones { background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); }
    .metric-card-utilidad { background: linear-gradient(135deg, #10b981 0%, #059669 100%); }

    /* ══ STATS CARDS ══ */
    .stats-card {
        background: #ffffff; border-radius: 16px; padding: 1.5rem 1.6rem;
        border: 1.5px solid #e8ebf5;
        box-shadow: 0 2px 12px rgba(0,0,0,0.05);
        transition: all 0.25s cubic-bezier(0.4,0,0.2,1);
        height: 100%; position: relative; overflow: hidden;
    }
    .stats-card::after {
        content: ''; position: absolute; bottom: 0; left: 0; right: 0; height: 3px;
        background: linear-gradient(90deg, #5b7cfa, #8b5cf6);
        transform: scaleX(0); transform-origin: left;
        transition: transform 0.3s ease;
    }
    .stats-card:hover { transform: translateY(-4px); box-shadow: 0 12px 32px rgba(91,124,250,0.12); border-color: #c5ccf0; }
    .stats-card:hover::after { transform: scaleX(1); }
    .stats-title {
        font-size: 0.72rem; font-weight: 700; color: #9099be;
        text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 0.5rem;
    }
    .stats-number {
        font-size: 2.6rem; font-weight: 800; line-height: 1.1; margin: 0.5rem 0;
        letter-spacing: -0.04em; padding: 0.5rem 0; text-align: center;
        border-top: 1.5px solid #eaedf5; border-bottom: 1.5px solid #eaedf5;
    }
    .stats-number.total { color: #5b7cfa !important; }
    .stats-number.autorizadas { color: #10b981 !important; }
    .stats-number.borradores { color: #f59e0b !important; }
    .stats-number.incompletas { color: #ef4444 !important; }
    .stats-desc { font-size: 0.78rem; color: #a0a8c8; text-align: center; margin-top: 0.25rem; }

    /* ══ HEADER ══ */
    .main-title {
        font-family: 'Plus Jakarta Sans', sans-serif !important;
        font-size: 2rem !important; font-weight: 800 !important;
        background: linear-gradient(135deg, #5b7cfa 0%, #8b5cf6 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        letter-spacing: -0.04em; line-height: 1.15;
    }
    .sub-title {
        color: #9099be; font-size: 0.82rem; font-weight: 500;
        margin-top: 0.2rem; letter-spacing: 0.02em;
    }

    /* ══ STATUS BADGE ══ */
    .cotizacion-status-container {
        background: #ffffff; border-radius: 50px;
        padding: 0.5rem 1.2rem 0.5rem 1.5rem;
        margin-bottom: 1rem; border: 1.5px solid #e2e6f3;
        box-shadow: 0 2px 10px rgba(0,0,0,0.06);
        display: inline-flex; align-items: center; gap: 1rem;
    }
    .status-badge { font-size: 0.875rem; font-weight: 600; color: #2a3060; }

    /* ══ MODO ADMIN ══ */
    .modo-admin-indicator {
        background: linear-gradient(135deg, #f59e0b, #d97706);
        color: white; padding: 0.3rem 0.9rem; border-radius: 20px;
        font-size: 0.75rem; font-weight: 700; letter-spacing: 0.04em;
        box-shadow: 0 3px 10px rgba(245,158,11,0.35);
    }

    /* ══ SECCIONES ══ */
    div[data-testid="stExpander"] {
        border: 1.5px solid #e2e6f3 !important;
        border-radius: 14px !important;
        box-shadow: 0 2px 10px rgba(0,0,0,0.04) !important;
        overflow: hidden !important;
        background: #ffffff !important;
    }

    /* ══ ALERTAS ══ */
    .stAlert { border-radius: 12px !important; border: none !important; font-family: 'Plus Jakarta Sans', sans-serif !important; }
    .stSuccess { background: rgba(16,185,129,0.1) !important; color: #065f46 !important; }
    .stError { background: rgba(239,68,68,0.1) !important; color: #991b1b !important; }
    .stWarning { background: rgba(245,158,11,0.1) !important; color: #92400e !important; }
    .stInfo { background: rgba(91,124,250,0.1) !important; color: #1e3a8a !important; }

    /* ══ SEPARADORES ══ */
    hr { border: none !important; border-top: 1.5px solid #e8ebf5 !important; margin: 1.5rem 0 !important; }

    /* ══ SCROLLBAR ══ */
    ::-webkit-scrollbar { width: 5px; height: 5px; }
    ::-webkit-scrollbar-track { background: #f0f2f8; }
    ::-webkit-scrollbar-thumb { background: #c5ccf0; border-radius: 10px; }
    ::-webkit-scrollbar-thumb:hover { background: #5b7cfa; }

    /* ══ TIPOGRAFÍA GENERAL ══ */
    h1, h2, h3, h4 { font-family: 'Plus Jakarta Sans', sans-serif !important; font-weight: 700 !important; color: #1a1d2e !important; letter-spacing: -0.03em !important; }
    .stMarkdown p { color: #4a5270 !important; line-height: 1.7 !important; font-family: 'Plus Jakarta Sans', sans-serif !important; }

    /* ══ NUMBER INPUT ══ */
    .stNumberInput button { border-radius: 8px !important; }

    /* ══ FILE UPLOADER ══ */
    [data-testid="stFileUploader"] { border-radius: 12px !important; }

    /* ══ RADIO BUTTONS ══ */
    .stRadio > div { gap: 0.5rem !important; }

    /* ══ FAB GUARDAR FLOTANTE ══ */
    .fab-guardar {
        background: linear-gradient(135deg, #5b7cfa 0%, #8b5cf6 100%);
        color: white;
        border: none;
        border-radius: 50px;
        padding: 0.9rem 1.6rem;
        font-size: 0.95rem;
        font-weight: 700;
        cursor: pointer;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-family: 'Plus Jakarta Sans', sans-serif;
        letter-spacing: 0.02em;
        animation: pulse-fab 2s infinite;
        transition: all 0.3s cubic-bezier(0.4,0,0.2,1);
    }
    .fab-guardar:hover {
        transform: translateY(-3px) scale(1.05);
        box-shadow: 0 16px 40px rgba(91,124,250,0.7) !important;
        animation: none;
    }
    .fab-guardar:active {
        transform: scale(0.97);
    }
    @keyframes pulse-fab {
        0%   { box-shadow: 0 8px 24px rgba(91,124,250,0.5); }
        50%  { box-shadow: 0 8px 40px rgba(91,124,250,0.9), 0 0 0 12px rgba(91,124,250,0.15); }
        100% { box-shadow: 0 8px 24px rgba(91,124,250,0.5); }
    }
    .fab-badge {
        position: absolute;
        top: -5px;
        right: -5px;
        width: 14px;
        height: 14px;
        background: #ef4444;
        border-radius: 50%;
        border: 2px solid white;
        animation: blink-badge 1.5s infinite;
    }
    @keyframes blink-badge {
        0%, 100% { opacity: 1; }
        50%       { opacity: 0.2; }
    }
</style>
""", unsafe_allow_html=True)

st.markdown('''
<style>
.stMarkdown h3 {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 1.05rem !important;
    font-weight: 700 !important;
    color: #1e2447 !important;
    letter-spacing: -0.02em !important;
    padding-left: 0.9rem !important;
    border-left: 3.5px solid #5b7cfa !important;
    margin: 1.2rem 0 0.8rem 0 !important;
    line-height: 1.4 !important;
}
.stMarkdown h4 {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 0.95rem !important;
    font-weight: 700 !important;
    color: #2a3060 !important;
    letter-spacing: -0.01em !important;
    margin: 1rem 0 0.6rem 0 !important;
}
.block-container {
    padding-top: 0.5rem !important;
    padding-bottom: 3rem !important;
}
/* Eliminar espacio reservado del header nativo */
[data-testid="stHeader"] {
    display: none !important;
    height: 0 !important;
    min-height: 0 !important;
}
[data-testid="stAppViewContainer"] > section:first-child {
    padding-top: 65px !important;
}
/* Eliminar espacio de iframes JS-only (height=0) sin afectar FAB */
[data-testid="stCustomComponentV1"]:has(iframe[height="0"]) {
    height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
    min-height: 0 !important;
}
[data-testid="stCheckbox"] span,
[data-testid="stRadio"] span {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 500 !important;
    color: #3a4070 !important;
}
[data-testid="stMetric"] {
    background: #ffffff;
    border-radius: 14px;
    padding: 1rem 1.2rem;
    border: 1.5px solid #e8ebf5;
    box-shadow: 0 2px 10px rgba(0,0,0,0.04);
}
[data-testid="stMetricLabel"] {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-size: 0.75rem !important; font-weight: 700 !important;
    text-transform: uppercase !important; letter-spacing: 0.07em !important;
    color: #9099be !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Plus Jakarta Sans', sans-serif !important;
    font-weight: 800 !important; color: #1e2447 !important;
    letter-spacing: -0.03em !important;
}
[data-testid="stPopover"] [data-testid="stMarkdown"] h3 {
    border-left: none !important; padding-left: 0 !important;
    font-size: 1rem !important;
}
</style>
''', unsafe_allow_html=True)

_tema = st.get_option("theme.base") or "light"
if _tema == "dark":
    st.markdown('''<style>
    /* ── Headers de tabs — texto siempre blanco ── */
    .tab-header h2, .tab-header p,
    .tab-header h2 *, .tab-header p *,
    .tab-header span[style*="display:block"] { color: #ffffff !important; }
    .tab-header h2 { color: #ffffff !important; font-size:1.5rem !important; font-weight:700 !important; margin:0 !important; }
    .tab-header p  { color: rgba(255,255,255,0.75) !important; font-size:0.88rem !important; margin:6px 0 0 !important; }
    </style>''', unsafe_allow_html=True)

    st.markdown('''<style>
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stNumberInput > div > div > input,
    .stDateInput > div > div > input {
        background-color: #1e293b !important; color: #f1f5f9 !important; border: 1px solid #334155 !important;
    }
    .stSelectbox > div > div, .stSelectbox > div > div > div {
        background-color: #1e293b !important; color: #f1f5f9 !important;
    }
    [data-baseweb="select"] > div { background-color: #1e293b !important; border-color: #334155 !important; color: #f1f5f9 !important; }
    [data-baseweb="select"] span { color: #f1f5f9 !important; }
    .stTextInput label, .stSelectbox label, .stNumberInput label,
    .stDateInput label, .stTextArea label, .stRadio label, .stCheckbox label { color: #94a3b8 !important; }
    .stButton > button { background-color: #1e293b !important; color: #f1f5f9 !important; border: 1px solid #334155 !important; }
    .stButton > button:hover { background-color: #334155 !important; color: #ffffff !important; }
    .stPopover > button { background-color: #1e293b !important; color: #f1f5f9 !important; border: 1px solid #334155 !important; }
    .resultados-table { background: #1e293b !important; border-color: #334155 !important; }
    .resultados-table th { background-color: #0f172a !important; color: #f1f5f9 !important; border-bottom: 2px solid #334155 !important; }
    .resultados-table td { color: #cbd5e1 !important; background-color: #1e293b !important; border-bottom: 1px solid #334155 !important; }
    .resultados-table tr:hover td { background-color: #334155 !important; }
    .stats-card { background: #1e293b !important; border: 1px solid #334155 !important; }
    .stats-title { color: #94a3b8 !important; }
    .stats-desc { color: #64748b !important; }
    .stats-number { border-color: #334155 !important; }
    </style>''', unsafe_allow_html=True)

# =========================================================
# HEADER
# =========================================================
st.markdown('''<div id="header-marker"></div>''', unsafe_allow_html=True)
st.markdown('''
<style>
#header-marker + div [data-testid="stHorizontalBlock"] > div:last-child > div {
    display: flex !important;
    flex-direction: column !important;
    align-items: flex-end !important;
    text-align: right !important;
}
#header-marker + div [data-testid="stHorizontalBlock"] > div:last-child img {
    margin-left: auto !important;
    display: block !important;
}
#header-marker + div [data-testid="stHorizontalBlock"] > div:last-child .stPopover {
    margin-left: auto !important;
}
</style>
''', unsafe_allow_html=True)

import base64 as _b64, os as _os

_logo_html = ""
if _os.path.exists("logo.png"):
    with open("logo.png", "rb") as _f:
        _logo_b64 = _b64.b64encode(_f.read()).decode()
    _logo_html = f'<img src="data:image/png;base64,{_logo_b64}" width="350" style="display:block;margin-left:auto;">'
else:
    _logo_html = '''<svg width="350" height="48" viewBox="0 0 130 48" fill="none" xmlns="http://www.w3.org/2000/svg" style="display:block;margin-left:auto;">
        <rect width="350" height="48" rx="8" fill="url(#hg)"/>
        <path d="M26 16L32 21L26 26L20 21L26 16Z" fill="white"/>
        <circle cx="65" cy="21" r="5" fill="#FFD966"/>
        <text x="82" y="26" font-family="Inter" font-size="13" font-weight="700" fill="white">PRO</text>
        <defs><linearGradient id="hg" x1="0" y1="0" x2="130" y2="48" gradientUnits="userSpaceOnUse">
            <stop stop-color="#3B82F6"/><stop offset="1" stop-color="#8B5CF6"/>
        </linearGradient></defs>
    </svg>'''

st.markdown(f'''
<div style="
    display:flex; justify-content:space-between; align-items:center;
    padding: 1.2rem 1.8rem;
    background: #ffffff;
    border-radius: 18px;
    border: 1.5px solid #e2e6f3;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    margin-bottom: 0.8rem;
">
    <div style="display:flex; flex-direction:column; gap:0.15rem;">
        <span class="main-title">Cotizador PRO</span>
        <div class="sub-title">Sistema profesional de cotizaciones</div>
    </div>
    <div style="display:flex; flex-direction:column; align-items:flex-end; gap:0.5rem;">
        {_logo_html}
    </div>
</div>
''', unsafe_allow_html=True)

# ── Header fijo: badge + cerrar izquierda / usuario derecha ──
_nombre_display = st.session_state.get('auth_nombre', '') or st.session_state.get('auth_email', '')
_rol_disp = st.session_state.get('rol_usuario', 'ejecutivo')
if _rol_disp == 'root':
    _rol_html = '<span style="color:#f59e0b;font-weight:700;font-size:0.8rem;">🔑 ROOT</span> <span style="color:#e2e8f0;font-size:0.82rem;font-weight:600;">' + _nombre_display.upper() + '</span>'
elif _rol_disp == 'admin':
    _rol_html = '<span style="color:#a78bfa;font-weight:700;font-size:0.8rem;">👑 ADMIN</span> <span style="color:#e2e8f0;font-size:0.82rem;font-weight:600;">' + _nombre_display.upper() + '</span>'
else:
    _rol_html = '<span style="color:#94a3b8;font-weight:700;font-size:0.8rem;">👤</span> <span style="color:#e2e8f0;font-size:0.82rem;font-weight:600;">' + _nombre_display.upper() + '</span>'

# Calcular badge desde session_state directamente
_cot_cargada = st.session_state.get('cotizacion_cargada')
if _cot_cargada:
    _margen_hdr = st.session_state.get('margen', 0)
    _plano_hdr  = bool(st.session_state.get('plano_adjunto') or st.session_state.get('pdf_url') or st.session_state.get('plano_nombre'))
    _datos_hdr  = bool(st.session_state.get('nombre_input') and st.session_state.get('correo_input'))
    _ase_hdr    = st.session_state.get('asesor_seleccionado', '')
    _asesor_hdr = bool(_ase_hdr and _ase_hdr != 'Seleccionar asesor')
    if _margen_hdr > 0 and _datos_hdr and _asesor_hdr:
        _badge_hdr    = '🟢 AUTORIZADO' + (' CON PLANO' if _plano_hdr else '')
        _badge_color  = '#10b981'
        _header_color = '#064e3b'
    elif _datos_hdr and _asesor_hdr and _plano_hdr:
        _badge_hdr    = '🟠 BORRADOR CON PLANO'
        _badge_color  = '#f97316'
        _header_color = '#7c2d12'
    elif _datos_hdr and _asesor_hdr:
        _badge_hdr    = '🟡 BORRADOR'
        _badge_color  = '#eab308'
        _header_color = '#713f12'
    else:
        _badge_hdr    = '🔴 INCOMPLETO' + (' CON PLANO' if _plano_hdr else '')
        _badge_color  = '#ef4444'
        _header_color = '#7f1d1d'
    _ep_num_str = str(_cot_cargada)
    _badge_pill = ('<span id="hdr-badge-estado" data-ep="' + _ep_num_str + '" title="Click para copiar ' + _ep_num_str + '" '
                   'style="font-size:0.88rem;font-weight:700;color:#e2e8f0;cursor:pointer;'
                   'white-space:nowrap;user-select:none;-webkit-user-select:none;'
                   '-webkit-tap-highlight-color:transparent;outline:none;display:flex;align-items:center;gap:8px;">'
                   '<span>📝 ' + _ep_num_str + ' •</span>'
                   '<span style="color:' + _badge_color + ';background:rgba(0,0,0,0.3);'
                   'padding:4px 14px;border-radius:20px;border:1px solid ' + _badge_color + '55;">'
                   + _badge_hdr + '</span>'
                   '</span>')
    _ep_txt     = ''
    _cerrar_btn = ('<button id="_btn_cerrar_hdr" data-action="cerrar-cot" '
                   'style="margin-left:12px;background:rgba(239,68,68,0.15);color:#fca5a5;'
                   'border:1px solid rgba(239,68,68,0.3);border-radius:8px;padding:5px 12px;'
                   'font-size:0.85rem;font-weight:700;cursor:pointer;white-space:nowrap;">🗑️ Cerrar</button>')
    _left_html  = _badge_pill + _cerrar_btn
else:
    _header_color = '#0f172a'
    _left_html    = '<span style="font-size:0.85rem;font-weight:600;color:#ffffff;-webkit-text-fill-color:#ffffff;">Sin cotización activa</span>'

_header_bg = 'linear-gradient(90deg, ' + _header_color + ' 0%, #0f172a 65%)'

st.markdown(f"""
<style>
#_usr_header_bar {{
    position: fixed;
    top: 0; left: 0; right: 0;
    height: 65px;
    background: {_header_bg};
    border-bottom: 1px solid rgba(255,255,255,0.08);
    box-shadow: 0 4px 24px rgba(0,0,0,0.4), 0 1px 0 rgba(255,255,255,0.05);
    transition: background 0.5s ease;
    display: flex;
    align-items: center;
    padding: 0 1.5rem;
    z-index: 2147483647;
    gap: 12px;
}}
#_usr_header_bar .usr-right {{
    display: flex;
    align-items: center;
    gap: 10px;
    margin-left: auto;
    flex-shrink: 0;
}}
/* Botones del header — forzar texto blanco */
#_usr_header_bar button,
#_usr_header_bar ._hdr_btns_moved button {{
    color: #ffffff !important;
}}
/* Badge estado — sin efectos de click */
#hdr-badge-estado {{
    -webkit-tap-highlight-color: transparent !important;
    outline: none !important;
}}
#hdr-badge-estado:active {{
    opacity: 0.85 !important;
    transform: scale(0.97) !important;
}}
/* Ocultar badge y botón cerrar originales — sin espacio */
.cotizacion-status-container {{ display: none !important; }}
.st-key-btn_cerrar_cotizacion {{
    display: none !important;
    height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
}}
/* Ocultar toda la fila de columnas del badge/cerrar */
.st-key-btn_cerrar_cotizacion,
.st-key-btn_cerrar_cotizacion * {{ display: none !important; }}
/* Reducir espacio del bloque completo que contiene badge+cerrar */
[data-testid="stHorizontalBlock"]:has(.st-key-btn_cerrar_cotizacion) {{
    display: none !important;
    height: 0 !important;
    margin: 0 !important;
    padding: 0 !important;
    min-height: 0 !important;
}}
</style>
""" + '<div id="_usr_header_bar"><div style="display:flex;align-items:center;gap:4px;flex:1;min-width:0;overflow:hidden;">' + _left_html + '</div><div class="usr-right">' + _rol_html + '</div></div>', unsafe_allow_html=True)

# Dialog contraseña — se abre centrado sin interferir con popovers
if 'show_pwd_dialog' not in st.session_state:
    st.session_state.show_pwd_dialog = False

@st.dialog("🔑 Cambiar contraseña")
def _pwd_dialog():
    _nombre_usr = st.session_state.get('auth_nombre', '') or st.session_state.get('auth_email', '')
    st.markdown(f"""
    <div style='text-align:center;padding:0.8rem 0 1rem;'>
        <div style='width:52px;height:52px;border-radius:50%;
            background:linear-gradient(135deg,#5b7cfa,#8b5cf6);
            display:flex;align-items:center;justify-content:center;
            font-size:1.5rem;margin:0 auto 8px;
            box-shadow:0 4px 16px rgba(91,124,250,0.3);'>🔑</div>
        <div style='color:#64748b;font-size:0.82rem;'>
            Usuario: <strong style='color:#1e293b;'>{_nombre_usr.upper()}</strong>
        </div>
    </div>
    """, unsafe_allow_html=True)
    # Columnas para simular padding lateral
    _sp1, _mid, _sp2 = st.columns([0.08, 1, 0.08])
    with _mid:
        _pwd_actual = st.text_input("Contraseña actual", type="password", key="pwd_actual_dlg")
        _pwd_nueva  = st.text_input("Nueva contraseña", type="password", key="pwd_nueva_dlg", placeholder="Mínimo 6 caracteres")
        _pwd_repite = st.text_input("Repetir nueva contraseña", type="password", key="pwd_repite_dlg")
        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
        if st.button("🔐 Actualizar contraseña", key="btn_cambiar_pwd_dlg", use_container_width=True, type="primary"):
            if not _pwd_actual or not _pwd_nueva or not _pwd_repite:
                st.error("Completa todos los campos.")
            elif len(_pwd_nueva) < 6:
                st.error("La contraseña debe tener mínimo 6 caracteres.")
            elif _pwd_nueva != _pwd_repite:
                st.error("Las contraseñas nuevas no coinciden.")
            else:
                _u_check, _ = login_usuario(st.session_state.get('auth_email',''), _pwd_actual)
                if not _u_check:
                    st.error("❌ Contraseña actual incorrecta.")
                else:
                    _ok, _err = cambiar_password_propio(_pwd_nueva)
                    if _ok:
                        st.success("✅ ¡Contraseña actualizada correctamente!")
                        st.session_state.show_pwd_dialog = False
                    else:
                        st.error(f"❌ {_err}")
    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

if st.session_state.show_pwd_dialog:
    st.session_state.show_pwd_dialog = False  # Reset inmediato
    _pwd_dialog()

# Botones ocultos — se mueven al header via JS
# CSS para ocultar todo el bloque sin dejar espacio
st.markdown("""<style>
.st-key-btn_pwd_hdr, .st-key-btn_cerrar_sesion_header, .st-key-btn_cerrar_cotizacion {
    position:fixed!important;top:-9999px!important;left:-9999px!important;
    width:1px!important;height:1px!important;overflow:hidden!important;
}
[data-testid="stHorizontalBlock"]:has(.st-key-btn_pwd_hdr):not(:has(.st-key-btn_fab_guardar)) {
    display:none!important;height:0!important;margin:0!important;padding:0!important;min-height:0!important;
}
</style>""", unsafe_allow_html=True)

_col_esp2, _col_pwd, _col_cerrar = st.columns([12, 1, 1])
with _col_pwd:
    if st.button("🔑 Mi contraseña", key="btn_pwd_hdr", use_container_width=True):
        st.session_state.show_pwd_dialog = True
        st.rerun()
with _col_cerrar:
    if st.button("🚪 Cerrar sesión", key="btn_cerrar_sesion_header", use_container_width=True):
        logout_usuario()
        st.session_state.modo_admin = False
        st.session_state._csv_listo = None
        st.rerun()

# Calcular total+IVA para inyectar en el header via JS
_total_hdr_fmt = ''
try:
    _carrito_hdr = st.session_state.get('carrito', [])
    if _carrito_hdr:
        _margen_h = float(st.session_state.get('margen', 0) or 0)
        _sub_hdr = sum(
            float(item.get('Cantidad', 0)) * float(item.get('Precio Unitario', 0)) * (1 + _margen_h / 100)
            for item in _carrito_hdr
        )
        _total_hdr_fmt = '$' + '{:,.0f}'.format(_sub_hdr * 1.19).replace(',', '.')
except Exception:
    _total_hdr_fmt = ''

# JS consolidado — header, tabs y cerrar cotización (un solo components.html = menos espacio)
import streamlit.components.v1 as _js_global
_js_global.html("""
<script>
(function(){
(function(){
    var D = window.parent.document;
    function tagCerrarCot(){
        var btns = D.querySelectorAll('button');
        for(var i=0;i<btns.length;i++){
            var txt=(btns[i].innerText||btns[i].textContent||'').trim();
            if(txt==='🗑️ Cerrar Cotización'){
                btns[i].id='_btn_cerrar_cot_real';
                break;
            }
        }
    }
    setTimeout(tagCerrarCot,500);
    setTimeout(tagCerrarCot,1200);
})();

(function(){
    var D = window.parent.document;

    function moveButtonsToHeader() {
        var bar = D.getElementById('_usr_header_bar');
        if (!bar) return;

        var usrRight = bar.querySelector('.usr-right');
        if (!usrRight) return;

        // Limpiar wrap anterior para usar referencias frescas en cada rerun
        var oldWrap = bar.querySelector('._hdr_btns_moved');
        if (oldWrap) oldWrap.remove();
        // Restaurar visibilidad de botones originales por si quedaron ocultos
        var allHidden = D.querySelectorAll('[style*="visibility:hidden"]');
        for (var h=0; h<allHidden.length; h++) { allHidden[h].style.visibility = ''; }

        // Buscar botones por texto
        var allBtns = D.querySelectorAll('button');
        var btnPwd = null, btnCerrar = null, popoverPwd = null;

        for (var i = 0; i < allBtns.length; i++) {
            var txt = (allBtns[i].innerText || allBtns[i].textContent || '').trim();
            if (txt === '🔑 Mi contraseña') btnPwd = allBtns[i];
            if (txt === '🚪 Cerrar sesión') btnCerrar = allBtns[i];
        }

        if (!btnPwd || !btnCerrar) return;

        // Crear contenedor de botones en el header
        var wrap = D.createElement('div');
        wrap.className = '_hdr_btns_moved';
        wrap.style.cssText = 'display:flex;align-items:center;gap:6px;';

        // Estilo compartido para botones del header
        var btnStyle = [
            'background:rgba(255,255,255,0.08)!important;',
            'color:#ffffff!important;border:1px solid rgba(255,255,255,0.25)!important;',
            'border-radius:6px!important;padding:4px 12px!important;font-size:0.82rem!important;',
            'font-weight:600!important;cursor:pointer!important;white-space:nowrap!important;',
            'font-family:inherit!important;transition:background 0.2s!important;'
        ].join('');

        // Crear botones nuevos (no clonar) para evitar herencia de estilos Streamlit
        var clonePwd = D.createElement('button');
        clonePwd.textContent = '🔑 Mi contraseña';
        clonePwd.style.cssText = btnStyle;
        clonePwd.onmouseenter = function(){ this.style.background='rgba(255,255,255,0.15)'; };
        clonePwd.onmouseleave = function(){ this.style.background='rgba(255,255,255,0.08)'; };
        clonePwd.addEventListener('click', function(e){
            e.preventDefault();
            e.stopPropagation();
            btnPwd.click();
        });

        var cloneCerrar = D.createElement('button');
        cloneCerrar.textContent = '🚪 Cerrar sesión';
        cloneCerrar.style.cssText = btnStyle + 'background:rgba(239,68,68,0.25)!important;border-color:rgba(239,68,68,0.5)!important;';
        cloneCerrar.onmouseenter = function(){ this.style.background='rgba(239,68,68,0.5)!important'; };
        cloneCerrar.onmouseleave = function(){ this.style.background='rgba(239,68,68,0.25)'; };
        cloneCerrar.addEventListener('click', function(e){
            e.stopPropagation();
            btnCerrar.click();
        });

        wrap.appendChild(clonePwd);
        wrap.appendChild(cloneCerrar);
        usrRight.appendChild(wrap);

        // Ocultar visualmente — siguen en el DOM para que Streamlit registre clicks
        var colPwd = btnPwd.closest('[data-testid="stPopover"]') || btnPwd.closest('[data-testid="stButton"]');
        var colCerrar = btnCerrar.closest('[data-testid="stButton"]');
        if (colPwd) colPwd.parentElement.style.visibility = 'hidden';
        if (colCerrar) colCerrar.style.visibility = 'hidden';
    }

    // Mover badge y botón cerrar cotización al header
    function moveBadgeAndCloseToHeader() {
        var slot = D.getElementById('_badge_slot');
        if (!slot) return;

        // Badge de estado
        var badge = D.querySelector('.status-badge');
        var badgeText = badge ? (badge.innerText || badge.textContent || '').trim() : '';
        var slotBadge = slot.querySelector('._slot_badge');
        if (!slotBadge) {
            slotBadge = D.createElement('span');
            slotBadge.className = '_slot_badge';
            slotBadge.style.cssText = 'font-size:0.78rem;font-weight:700;color:#e2e8f0;white-space:nowrap;cursor:pointer;';
            slot.appendChild(slotBadge);
        }
        slotBadge.textContent = badgeText;

        // Ocultar badge original
        var badgeWrap = D.querySelector('.cotizacion-status-container');
        if (badgeWrap) { badgeWrap.style.visibility='hidden'; badgeWrap.style.height='0'; badgeWrap.style.overflow='hidden'; }

        // Botón cerrar cotización
        if (slot.querySelector('._btn_cerrar_cot')) return;
        var allBtns = D.querySelectorAll('button');
        for (var i = 0; i < allBtns.length; i++) {
            var txt = (allBtns[i].innerText || allBtns[i].textContent || '').trim();
            if (txt === '🗑️ Cerrar Cotización') {
                var orig = allBtns[i];
                var btn = D.createElement('button');
                btn.className = '_btn_cerrar_cot';
                btn.textContent = '🗑️ Cerrar';
                btn.style.cssText = 'background:rgba(239,68,68,0.12);color:#fca5a5;border:1px solid rgba(239,68,68,0.2);border-radius:6px;padding:3px 10px;font-size:0.75rem;font-weight:600;cursor:pointer;white-space:nowrap;font-family:inherit;margin-left:8px;';
                btn.onmouseenter = function(){ this.style.background='rgba(239,68,68,0.25)'; };
                btn.onmouseleave = function(){ this.style.background='rgba(239,68,68,0.12)'; };
                btn.addEventListener('click', function(){ orig.click(); });
                slot.appendChild(btn);
                // Ocultar original
                var par = orig.closest('[data-testid="stButton"]');
                if (par) par.style.cssText = 'position:fixed;top:-9999px;left:-9999px;';
                break;
            }
        }
    }

    // Listener para copiar EP al hacer click en el badge de estado
    D.addEventListener('click', function(e){
        var el = e.target && e.target.closest ? e.target.closest('#hdr-badge-estado') : null;
        if (el) {
            var ep = el.getAttribute('data-ep') || '';
            if (ep) {
                var ta = D.createElement('textarea');
                ta.value = ep; ta.style.cssText = 'position:fixed;top:-9999px;left:-9999px;';
                D.body.appendChild(ta); ta.focus(); ta.select();
                try { D.execCommand('copy'); } catch(e) {}
                ta.remove();
                var origHtml = el.innerHTML;
                var origColor = el.style.color;
                el.innerHTML = '✅ ¡Copiado!';
                el.style.setProperty('color', '#10b981', 'important');
                setTimeout(function(){
                    el.innerHTML = origHtml;
                    el.style.setProperty('color', origColor, 'important');
                }, 1200);
            }
        }
    });

    // Listener para botón cerrar cotización del header
    D.addEventListener('click', function(e){
        if (e.target && e.target.getAttribute && e.target.getAttribute('data-action') === 'cerrar-cot') {
            var btns = D.querySelectorAll('button');
            for (var i = 0; i < btns.length; i++) {
                var t = (btns[i].innerText || btns[i].textContent || '').trim();
                if (t === '🗑️ Cerrar Cotización') {
                    btns[i].click();
                    return;
                }
            }
        }
    });

    function injectTotal() {
        var bar = D.getElementById('_usr_header_bar');
        if (!bar) return;
        var existing = D.getElementById('_hdr_total_iva');
        if (existing) existing.remove();
        var totalFmt = """ + repr(_total_hdr_fmt) + """;
        if (!totalFmt) return;
        var div = D.createElement('div');
        div.id = '_hdr_total_iva';
        div.style.cssText = 'position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);display:flex;flex-direction:column;align-items:center;justify-content:center;pointer-events:none;';
        div.innerHTML = '<div style="font-size:0.58rem;font-weight:700;color:rgba(255,255,255,0.45);text-transform:uppercase;letter-spacing:0.12em;margin-bottom:2px;">Total + IVA</div>'
            + '<div style="font-size:1.25rem;font-weight:900;color:#fff;letter-spacing:-0.02em;font-family:Montserrat,sans-serif;line-height:1;">' + totalFmt + '</div>';
        bar.appendChild(div);
    }
    setTimeout(injectTotal, 700);
    setTimeout(injectTotal, 1500);
    setTimeout(moveButtonsToHeader, 800);
    setTimeout(moveButtonsToHeader, 1500);
    setTimeout(moveBadgeAndCloseToHeader, 900);
    setTimeout(moveBadgeAndCloseToHeader, 1600);
})();

(function(){
    function initTabArrows() {
        var D = window.parent.document;
        // Eliminar flechas anteriores
        D.querySelectorAll('.tab-nav-arrow').forEach(function(e){ e.remove(); });

        var tablist = D.querySelector('[data-baseweb="tab-list"]');
        if (!tablist) return;

        var wrap = tablist.parentElement;
        if (!wrap) return;
        wrap.style.position = 'relative';

        function makeArrow(dir) {
            var btn = D.createElement('button');
            btn.className = 'tab-nav-arrow';
            btn.innerHTML = dir === 'left' ? '&#8249;' : '&#8250;';
            btn.style.cssText = [
                'position:absolute;top:0;z-index:99;',
                'background:linear-gradient(' + (dir==='left'?'90':'270') + 'deg,rgba(255,255,255,0.97) 55%,rgba(255,255,255,0))',
                ';border:none;cursor:pointer;padding:0 14px;height:100%;',
                'font-size:1.4rem;font-weight:700;color:#5b7cfa;',
                dir==='left' ? 'left:0;' : 'right:0;'
            ].join('');
            btn.addEventListener('click', function(){
                tablist.scrollBy({ left: dir==='left' ? -160 : 160, behavior:'smooth' });
            });
            return btn;
        }

        var btnL = makeArrow('left');
        var btnR = makeArrow('right');
        wrap.appendChild(btnL);
        wrap.appendChild(btnR);

        function updateArrows() {
            var sl = tablist.scrollLeft;
            var maxScroll = tablist.scrollWidth - tablist.clientWidth;
            btnL.style.opacity = sl > 5 ? '1' : '0';
            btnL.style.pointerEvents = sl > 5 ? 'auto' : 'none';
            btnR.style.opacity = sl < maxScroll - 5 ? '1' : '0';
            btnR.style.pointerEvents = sl < maxScroll - 5 ? 'auto' : 'none';
        }

        tablist.addEventListener('scroll', updateArrows);
        updateArrows();
    }

    // Inicializar cuando el DOM esté listo
    setTimeout(initTabArrows, 800);
    setTimeout(initTabArrows, 1800);

    // Re-inicializar al cambiar de tab
    window.parent.addEventListener('click', function(e){
        if (e.target && e.target.getAttribute && e.target.getAttribute('data-baseweb') === 'tab') {
            setTimeout(initTabArrows, 300);
        }
    });
})();
})();
</script>
""", height=0)

# BADGE DE COTIZACIÓN CARGADA
# =========================================================
def limpiar_todo():
    st.session_state.carrito = []
    st.session_state.nombre_input = ""
    st.session_state.rut_raw = ""
    st.session_state.rut_display = ""
    st.session_state.rut_valido = False
    st.session_state.rut_mensaje = ""
    st.session_state.correo_input = ""
    st.session_state.telefono_raw = ""
    st.session_state.direccion_input = ""
    st.session_state.cliente_comuna = ""
    st.session_state.cliente_region = ""
    st.session_state.proyecto_direccion = ""
    st.session_state.proyecto_comuna = ""
    st.session_state.proyecto_region = ""
    st.session_state.cliente_tipo = "natural"
    st.session_state.cliente_empresa = ""
    st.session_state.cliente_rut_empresa = ""
    st.session_state.rut_empresa_raw     = ""
    st.session_state.rut_empresa_display = ""
    st.session_state.rut_empresa_valido  = False
    st.session_state.asesor_seleccionado = "Seleccionar asesor"
    st.session_state.correo_asesor = ""
    st.session_state.telefono_asesor = ""
    st.session_state.fecha_inicio = datetime.now().date()
    st.session_state.fecha_termino = (datetime.now() + timedelta(days=15)).date()
    st.session_state.observaciones_input = ""
    st.session_state.plano_adjunto = None
    st.session_state.plano_nombre = ""
    st.session_state.cotizacion_cargada = None
    st.session_state.cotizacion_seleccionada = None
    st.session_state.margen = 0.0
    st.session_state.mostrar_visor = False
    st.session_state.pdf_actual = None
    st.session_state.pdf_nombre = ""
    st.session_state.numero_en_visor = None
    st.session_state.pdf_url = None
    st.session_state.counter += 100


if st.session_state.cotizacion_cargada:
    datos_completos = all([
        st.session_state.nombre_input,
        st.session_state.correo_input
    ])
    asesor_completo = any([
        st.session_state.asesor_seleccionado != "Seleccionar asesor",
        st.session_state.correo_asesor,
        st.session_state.telefono_asesor
    ])

    _tiene_plano_badge = bool(st.session_state.plano_adjunto or st.session_state.get('pdf_url') or st.session_state.plano_nombre)
    if st.session_state.margen > 0:
        if datos_completos and asesor_completo:
            rol = "👑 Admin" if st.session_state.modo_admin else "🔒 Solo lectura"
            sufijo = " CON PLANO" if _tiene_plano_badge else ""
            badge_html = f"{rol} • 🟢 AUTORIZADO{sufijo} ({st.session_state.margen}%)"
        else:
            sufijo = " CON PLANO" if _tiene_plano_badge else ""
            badge_html = f"⚠️ {st.session_state.cotizacion_cargada} • 🔴 INCOMPLETO{sufijo}"
    else:
        _tiene_plano_badge = bool(st.session_state.plano_adjunto or st.session_state.get('pdf_url') or st.session_state.plano_nombre)
        if datos_completos and asesor_completo:
            if _tiene_plano_badge:
                badge_html = f"📝 {st.session_state.cotizacion_cargada} • 🟠 BORRADOR CON PLANO"
            else:
                badge_html = f"📝 {st.session_state.cotizacion_cargada} • 🟡 BORRADOR"
        else:
            sufijo = " CON PLANO" if _tiene_plano_badge else ""
            badge_html = f"⚠️ {st.session_state.cotizacion_cargada} • 🔴 INCOMPLETO{sufijo}"

    # Botón cerrar cotización — oculto visualmente, funciona desde el header
    if st.button("🗑️ Cerrar Cotización", key="btn_cerrar_cotizacion"):
        _hash_actual = calcular_hash_estado()
        _hay_cambios = (
            len(st.session_state.get('carrito', [])) > 0 and
            _hash_actual != st.session_state.get('hash_ultimo_guardado')
        )
        if _hay_cambios:
            leer_datos_actuales()
            datos_c2, datos_a2, proy2, cfg2, tots2, pnom2, pdat2 = construir_datos_para_guardar()
            st.session_state.datos_pendientes_cerrar = {
                'datos_c': datos_c2, 'datos_a': datos_a2, 'proy': proy2,
                'cfg': cfg2, 'tots': tots2, 'pnom': pnom2, 'pdat': pdat2,
                'numero': st.session_state.cotizacion_cargada
            }
            st.session_state.mostrar_advertencia_cerrar = True
            st.rerun()
        else:
            limpiar_todo()
            st.rerun()

    if st.session_state.get('mostrar_advertencia_cerrar', False):
        @st.dialog("⚠️ Cambios sin guardar")
        def dialogo_advertencia_cerrar():
            st.warning("Tienes cambios sin guardar. ¿Qué deseas hacer?")
            col_guardar, col_descartar, col_cancelar = st.columns(3)
            with col_guardar:
                if st.button("💾 Guardar y cerrar", use_container_width=True, type="primary", key="dialog_cerrar_guardar"):
                    _dp = st.session_state.datos_pendientes_cerrar
                    _usr_log2 = st.session_state.get('auth_nombre','') or st.session_state.get('auth_email','')
                    guardar_cotizacion(_dp['numero'], _dp['datos_c'], _dp['datos_a'],
                                       _dp['proy'], st.session_state.carrito,
                                       _dp['cfg'], _dp['tots'], _dp['pnom'], _dp['pdat'],
                                       usuario_logueado=_usr_log2)
                    limpiar_todo()
                    st.session_state.datos_pendientes_cerrar = None
                    st.session_state.mostrar_advertencia_cerrar = False
                    st.rerun()
            with col_descartar:
                if st.button("🗑️ Descartar y cerrar", use_container_width=True, key="dialog_cerrar_descartar"):
                    limpiar_todo()
                    st.session_state.datos_pendientes_cerrar = None
                    st.session_state.mostrar_advertencia_cerrar = False
                    st.rerun()
            with col_cancelar:
                if st.button("✖️ Cancelar", use_container_width=True, key="dialog_cerrar_cancelar"):
                    st.session_state.datos_pendientes_cerrar = None
                    st.session_state.mostrar_advertencia_cerrar = False
                    st.rerun()
        dialogo_advertencia_cerrar()

# =========================================================
# FUNCIÓN: DASHBOARD PRINCIPAL
# =========================================================
def cargar_datos_dashboard(periodo='mes'):
    """Carga todas las métricas para el dashboard."""
    try:
        from datetime import datetime as _dt, timedelta as _td
        import calendar as _cal

        _ahora = _dt.now()
        if periodo == 'mes':
            _inicio = _ahora.replace(day=1).strftime('%Y-%m-%d')
            _inicio_ant = (_ahora.replace(day=1) - _td(days=1)).replace(day=1).strftime('%Y-%m-%d')
            _fin_ant = (_ahora.replace(day=1) - _td(days=1)).strftime('%Y-%m-%d')
        elif periodo == '3meses':
            _inicio = (_ahora - _td(days=90)).strftime('%Y-%m-%d')
            _inicio_ant = (_ahora - _td(days=180)).strftime('%Y-%m-%d')
            _fin_ant = (_ahora - _td(days=91)).strftime('%Y-%m-%d')
        elif periodo == 'año':
            _inicio = _ahora.replace(month=1, day=1).strftime('%Y-%m-%d')
            _inicio_ant = (_ahora.replace(month=1, day=1).replace(year=_ahora.year-1)).strftime('%Y-%m-%d')
            _fin_ant = (_ahora.replace(month=1, day=1) - _td(days=1)).strftime('%Y-%m-%d')
        else:
            _inicio = '2000-01-01'
            _inicio_ant = None

        # Sin filtro de fecha — traer todo y filtrar en Python
        resp = supabase.table('cotizaciones').select(
            'numero, fecha_creacion, fecha_modificacion, estado, '
            'total_total, asesor_nombre, cliente_nombre, cliente_rut, '
            'cliente_comuna, cliente_region, cliente_tipo, '
            'cliente_empresa, config_margen, cliente_email, '
            'asesor_email, asesor_telefono, productos'
        ).execute()
        # Filtrar en Python según período
        if periodo != 'todo':
            _rows_all = resp.data or []
            rows = []
            for _r in _rows_all:
                _fc = (_r.get('fecha_creacion') or _r.get('fecha_modificacion') or '')[:10]
                if _fc >= _inicio:
                    rows.append(_r)
        else:
            rows = resp.data or []

        rows = resp.data or []

        # Período anterior para comparación
        rows_ant = []
        if _inicio_ant:
            resp_ant = supabase.table('cotizaciones').select(
                'total_total, estado, config_margen, cliente_nombre, cliente_email, asesor_nombre, asesor_email, asesor_telefono'
            ).gte('fecha_creacion', _inicio_ant).lte('fecha_creacion', _fin_ant).execute()
            rows_ant = resp_ant.data or []

        # ── Métricas principales ──
        total_ep       = len(rows)
        total_monto    = sum(float(r.get('total_total') or 0) for r in rows)
        promedio_monto = total_monto / total_ep if total_ep else 0

        def _clasificar_estado(r):
            e = (r.get('estado') or '').upper()
            if 'AUTORIZADO' in e:  return 'autorizado'
            if 'BORRADOR'   in e:  return 'borrador'
            return 'incompleto'

        estados     = [_clasificar_estado(r) for r in rows]
        autorizados = estados.count('autorizado')
        borradores  = estados.count('borrador')
        incompletos = estados.count('incompleto')
        pct_conv    = round((autorizados / total_ep) * 100) if total_ep else 0

        # Comparación período anterior
        total_ep_ant    = len(rows_ant)
        total_monto_ant = sum(float(r.get('total_total') or 0) for r in rows_ant)
        delta_ep        = total_ep - total_ep_ant
        delta_monto     = total_monto - total_monto_ant

        # ── Serie temporal por día/semana ──
        from collections import defaultdict as _dd
        serie = _dd(float)
        serie_n = _dd(int)
        for r in rows:
            fecha = (r.get('fecha_creacion') or r.get('fecha_modificacion') or '')[:10]
            if fecha and len(fecha) == 10:
                serie[fecha]   += float(r.get('total_total') or 0)
                serie_n[fecha] += 1
        fechas_sorted = sorted(serie.keys())
        serie_montos  = [round(serie[f]) for f in fechas_sorted]
        serie_counts  = [serie_n[f] for f in fechas_sorted]

        # ── Top categorías ──
        cat_montos = _dd(float)
        cat_counts = _dd(int)
        for r in rows:
            prods = r.get('productos') or []
            if isinstance(prods, str):
                try:
                    import json as _j; prods = _j.loads(prods)
                except: prods = []
            for p in prods:
                cat = p.get('Categoria') or 'Sin categoría'
                # Usar Subtotal directo, o calcular desde Precio Unitario * Cantidad
                subtotal = float(p.get('Subtotal') or 0)
                if subtotal == 0:
                    precio = float(p.get('Precio Unitario') or p.get('Precio Final') or 0)
                    qty    = int(p.get('Cantidad') or 1)
                    subtotal = precio * qty
                qty = int(p.get('Cantidad') or 1)
                cat_montos[cat] += subtotal
                cat_counts[cat] += qty
        top_cats = sorted(cat_montos.items(), key=lambda x: x[1], reverse=True)[:6]

        # ── Pipeline (borradores pendientes) ──
        pipeline = sum(float(r.get('total_total') or 0) for r, e in zip(rows, estados) if e == 'borrador')

        # ── Ejecutivos top ──
        ej_montos = _dd(float)
        ej_counts = _dd(int)
        for r in rows:
            ej = r.get('asesor_nombre') or 'Sin asignar'
            ej_montos[ej] += float(r.get('total_total') or 0)
            ej_counts[ej] += 1
        top_ej = sorted(ej_montos.items(), key=lambda x: x[1], reverse=True)[:5]

        # ── Top 30 productos ──
        prod_montos = _dd(float)
        prod_cantidades = _dd(int)
        prod_categoria = {}
        for r in rows:
            prods = r.get('productos') or []
            if isinstance(prods, str):
                try:
                    import json as _j2; prods = _j2.loads(prods)
                except: prods = []
            for p in prods:
                item = (p.get('Item') or '').strip()
                if not item:
                    continue
                subtotal = float(p.get('Subtotal') or 0)
                if subtotal == 0:
                    precio = float(p.get('Precio Unitario') or 0)
                    qty    = int(p.get('Cantidad') or 1)
                    subtotal = precio * qty
                qty = int(p.get('Cantidad') or 1)
                prod_montos[item]     += subtotal
                prod_cantidades[item] += qty
                if item not in prod_categoria:
                    prod_categoria[item] = (p.get('Categoria') or '').strip()
        top_productos = sorted(prod_montos.items(), key=lambda x: x[1], reverse=True)[:30]
        top_productos = [(n, v, prod_cantidades[n], prod_categoria.get(n,'')) for n, v in top_productos]

        # ── Métricas geográficas ──
        comunas  = _dd(int)
        regiones = _dd(int)
        for r in rows:
            _com = (r.get('cliente_comuna') or '').strip()
            _reg = (r.get('cliente_region') or '').strip()
            if _com: comunas[_com]   += 1
            if _reg: regiones[_reg]  += 1
        top_comunas  = sorted(comunas.items(),  key=lambda x: x[1], reverse=True)[:10]
        top_regiones = sorted(regiones.items(), key=lambda x: x[1], reverse=True)[:10]

        # ── Tipo cliente natural/jurídica ──
        n_natural   = sum(1 for r in rows if (r.get('cliente_tipo') or 'natural') == 'natural')
        n_juridica  = sum(1 for r in rows if (r.get('cliente_tipo') or '') == 'juridica')
        top_empresas = _dd(int)
        for r in rows:
            if (r.get('cliente_tipo') or '') == 'juridica':
                _emp = (r.get('cliente_empresa') or '').strip()
                if _emp: top_empresas[_emp] += 1
        top_empresas = sorted(top_empresas.items(), key=lambda x: x[1], reverse=True)[:8]

        # ── Género estimado por primer nombre ──
        _MASC = {'carlos','juan','diego','miguel','andrés','andres','pedro','luis','jorge',
                 'gabriel','rodrigo','francisco','felipe','pablo','mario','roberto','sergio',
                 'cristian','christian','nicolás','nicolas','alejandro','manuel','antonio',
                 'jose','josè','josé','daniel','matias','matías','sebastián','sebastian',
                 'gonzalo','mauricio','marcelo','ricardo','eduardo','ignacio','javier',
                 'victor','víctor','claudio','raul','raúl','alfredo','oscar','óscar',
                 'tomás','tomas','alex','alexis','ivan','iván','hugo','alberto','david'}
        _FEM  = {'maria','maría','carolina','andrea','claudia','patricia','alejandra',
                 'valentina','camila','javiera','paula','ana','rosa','carmen','lucia',
                 'lucía','fernanda','daniela','monica','mónica','paola','lorena','isabel',
                 'veronica','verónica','beatriz','sandra','laura','marcela','fabiola',
                 'natalia','jessica','pamela','viviana','pilar','francisca','constanza',
                 'nicole','yasna','ximena','soledad','teresa','angeles','ángeles',
                 'macarena','barbara','bárbara','sofia','sofía','elena','alicia','susana'}
        n_masc = n_fem = n_nd = 0
        for r in rows:
            if (r.get('cliente_tipo') or 'natural') == 'juridica': continue
            _nm = (r.get('cliente_nombre') or '').strip().lower().split()
            _primer = _nm[0] if _nm else ''
            if _primer in _MASC:   n_masc += 1
            elif _primer in _FEM:  n_fem  += 1
            else:                  n_nd   += 1

        # ── Rango etario estimado por RUT ──
        # RUT chileno < 10M ~ nacido antes 1975, 10-15M ~ 1975-1995, 15-20M ~ 1995+
        rangos = {'< 1975 (50+)': 0, '1975-1995 (30-50)': 0, '> 1995 (< 30)': 0, 'No determinado': 0}
        for r in rows:
            if (r.get('cliente_tipo') or 'natural') == 'juridica': continue
            _rut_str = re.sub(r'[^0-9]', '', str(r.get('cliente_rut') or ''))
            try:
                _rut_n = int(_rut_str[:-1]) if len(_rut_str) > 1 else 0
                if   _rut_n < 1_000_000:  rangos['No determinado'] += 1
                elif _rut_n < 10_000_000: rangos['< 1975 (50+)'] += 1
                elif _rut_n < 15_000_000: rangos['1975-1995 (30-50)'] += 1
                else:                     rangos['> 1995 (< 30)'] += 1
            except: rangos['No determinado'] += 1

        return {
            'total_ep': total_ep, 'total_monto': total_monto,
            'promedio_monto': promedio_monto, 'pipeline': pipeline,
            'autorizados': autorizados, 'borradores': borradores,
            'incompletos': incompletos, 'pct_conv': pct_conv,
            'delta_ep': delta_ep, 'delta_monto': delta_monto,
            'fechas': fechas_sorted, 'serie_montos': serie_montos,
            'serie_counts': serie_counts,
            'top_cats': top_cats, 'top_ej': top_ej,
            'top_productos': top_productos,
            'top_comunas': top_comunas, 'top_regiones': top_regiones,
            'n_natural': n_natural, 'n_juridica': n_juridica,
            'top_empresas': top_empresas,
            'n_masc': n_masc, 'n_fem': n_fem, 'n_nd': n_nd,
            'rangos_etarios': rangos,
        }
    except Exception as e:
        return None


# =========================================================
# =========================================================
# FUNCIÓN: GENERADOR PDF CONTRATO CLIENTE
# =========================================================
def num_a_palabras(n):
    """Convierte un número entero a su representación en palabras en español."""
    unidades = ['','uno','dos','tres','cuatro','cinco','seis','siete','ocho','nueve',
                'diez','once','doce','trece','catorce','quince','dieciséis','diecisiete',
                'dieciocho','diecinueve','veinte','veintiuno','veintidós','veintitrés',
                'veinticuatro','veinticinco','veintiséis','veintisiete','veintiocho','veintinueve']
    decenas = ['','diez','veinte','treinta','cuarenta','cincuenta','sesenta','setenta','ochenta','noventa']
    centenas = ['','ciento','doscientos','trescientos','cuatrocientos','quinientos',
                'seiscientos','setecientos','ochocientos','novecientos']
    if n == 0: return 'cero'
    if n < 0:  return 'menos ' + num_a_palabras(-n)
    res = ''
    if n >= 1_000_000:
        m = n // 1_000_000
        res += ('un millón' if m == 1 else num_a_palabras(m) + ' millones') + ' '
        n %= 1_000_000
    if n >= 1000:
        m = n // 1000
        res += ('mil' if m == 1 else num_a_palabras(m) + ' mil') + ' '
        n %= 1000
    if n >= 100:
        if n == 100: res += 'cien '
        else: res += centenas[n // 100] + ' '
        n %= 100
    if n >= 30:
        d, u = divmod(n, 10)
        res += decenas[d] + (' y ' + unidades[u] if u else '') + ' '
    elif n > 0:
        res += unidades[n] + ' '
    return res.strip()

def monto_a_palabras(monto):
    """Convierte monto a texto: 'X pesos'."""
    entero = int(round(monto))
    return num_a_palabras(entero) + ' pesos'

def generar_pdf_contrato(datos):
    """
    Genera PDF del contrato a partir del dict `datos` con campos:
    fecha_str, tipo_cliente (natural/juridica),
    cli_nombre, cli_rut, cli_empresa (solo jurídica), cli_rut_empresa (solo jurídica),
    cli_domicilio, cli_comuna, cli_region,
    inst_domicilio, inst_comuna, inst_region,
    ep_numero, ep_nombre,
    precio_total, plazo_dias,
    pago_50, pago_25a, pago_25b
    """
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    HRFlowable, Table, TableStyle, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_JUSTIFY
    import io
    import os as _os_c

    # ── Plantilla con header/footer en cada página ──
    buf = io.BytesIO()

    AZUL       = colors.HexColor('#0f3460')
    AZUL_LIGHT = colors.HexColor('#e8eef7')
    GRIS       = colors.HexColor('#64748b')
    NEGRO      = colors.HexColor('#0f172a')

    def _build_header_footer(canvas, doc):
        """Header con logo centrado + línea azul. Footer con número de página."""
        canvas.saveState()
        pw = doc.pagesize[0]
        # ── Header ──
        if _os_c.path.exists("logo.png"):
            from reportlab.lib.utils import ImageReader
            _img = ImageReader("logo.png")
            _iw, _ih = _img.getSize()
            _aspect = _ih / float(_iw)
            _lw = 4.5 * cm
            _lh = _lw * _aspect
            canvas.drawImage(_img,
                             x=(pw - _lw) / 2,
                             y=doc.pagesize[1] - doc.topMargin + 0.3*cm,
                             width=_lw, height=_lh,
                             preserveAspectRatio=True, mask='auto')
        # Línea azul bajo el header
        canvas.setStrokeColor(AZUL)
        canvas.setLineWidth(1.2)
        _ly = doc.pagesize[1] - doc.topMargin + 0.1*cm
        canvas.line(doc.leftMargin, _ly, pw - doc.rightMargin, _ly)
        # ── Footer ──
        canvas.setFont('Helvetica', 8)
        canvas.setFillColor(GRIS)
        _fy = doc.bottomMargin - 0.5*cm
        canvas.drawCentredString(pw/2, _fy,
            f"Inversiones Container House SpA  ·  RUT 78.268.851-0  ·  Página {doc.page}")
        canvas.setStrokeColor(GRIS)
        canvas.setLineWidth(0.4)
        canvas.line(doc.leftMargin, _fy + 0.35*cm, pw - doc.rightMargin, _fy + 0.35*cm)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=3*cm, rightMargin=3*cm,
        topMargin=3.8*cm, bottomMargin=2.2*cm,
        title=f"Contrato {datos.get('ep_numero','')}"
    )

    # ── Estilos tipográficos jurídicos ──
    base = getSampleStyleSheet()
    normal = ParagraphStyle('cNormal', parent=base['Normal'],
                            fontName='Times-Roman', fontSize=10.5,
                            leading=16, spaceAfter=6,
                            alignment=TA_JUSTIFY,
                            firstLineIndent=0)
    bold   = ParagraphStyle('cBold', parent=normal,
                            fontName='Times-Bold')
    titulo = ParagraphStyle('cTitulo', parent=base['Normal'],
                            fontName='Times-Bold', fontSize=15,
                            leading=20, spaceAfter=2,
                            spaceBefore=6,
                            alignment=TA_CENTER,
                            textColor=AZUL)
    subtit = ParagraphStyle('cSubtit', parent=base['Normal'],
                            fontName='Times-Bold', fontSize=11,
                            leading=16, spaceAfter=8,
                            alignment=TA_CENTER,
                            textColor=AZUL)
    seccion = ParagraphStyle('cSeccion', parent=base['Normal'],
                             fontName='Helvetica-Bold', fontSize=9.5,
                             leading=13, spaceBefore=14, spaceAfter=5,
                             textColor=colors.white,
                             backColor=AZUL,
                             leftIndent=-0.3*cm, rightIndent=-0.3*cm,
                             borderPadding=(4, 8, 4, 8))
    firma   = ParagraphStyle('cFirma', parent=normal,
                             fontName='Times-Roman', fontSize=10,
                             alignment=TA_CENTER)
    firma_bold = ParagraphStyle('cFirmaBold', parent=firma,
                                fontName='Times-Bold')

    def HR():
        return HRFlowable(width="100%", thickness=0.6,
                          color=AZUL_LIGHT, spaceAfter=6, spaceBefore=2)
    def SP(h=6): return Spacer(1, h)

    d = datos
    precio   = d['precio_total']
    p50      = d['pago_50']
    p25a     = d['pago_25a']
    p25b     = d['pago_25b']
    precio_p = monto_a_palabras(precio)
    p50_p    = monto_a_palabras(p50)
    p25a_p   = monto_a_palabras(p25a)
    p25b_p   = monto_a_palabras(p25b)

    fmt = lambda v: "${:,.0f}".format(v).replace(",",".")

    # ── Comparecencia cliente según tipo ──
    if d['tipo_cliente'] == 'natural':
        tratamiento = d.get('cli_tratamiento', 'Don')
        cli_bloque = (
            f"{tratamiento} <b>{d['cli_nombre']}</b>, cédula nacional de identidad "
            f"<b>N° {d['cli_rut']}</b>, con domicilio en <b>{d['cli_domicilio']}</b>, "
            f"comuna de <b>{d['cli_comuna']}</b>, Región {d['cli_region']}, "
            f"quien en adelante se denominará \"el Cliente\"."
        )
    else:
        tratamiento = d.get('cli_tratamiento', 'Don')
        cli_bloque = (
            f"{tratamiento} <b>{d['cli_nombre']}</b>, cédula nacional de identidad "
            f"<b>N° {d['cli_rut']}</b>, en representación de "
            f"<b>{d['cli_empresa']}</b>, Rol Único Tributario "
            f"<b>N° {d['cli_rut_empresa']}</b>, con domicilio en "
            f"<b>{d['cli_domicilio']}</b>, comuna de <b>{d['cli_comuna']}</b>, "
            f"Región {d['cli_region']}, quien en adelante se denominará \"el Cliente\"."
        )

    story = []

    # ── Encabezado del documento ──
    story += [
        Paragraph("CONTRATO DE FABRICACIÓN Y VENTA", titulo),
        Paragraph("VIVIENDA TIPO CONTAINER", subtit),
        SP(4),
        HRFlowable(width="100%", thickness=1.5, color=AZUL, spaceAfter=10, spaceBefore=0),
        Paragraph(f"En Santiago de Chile, a <b>{d['fecha_str']}</b>, comparecen:", normal),
        SP(4),
    ]

    # ── I. Comparecencia ──
    story += [
        Paragraph("I. COMPARECENCIA", seccion),
        Paragraph("1. EL PROVEEDOR", bold),
        Paragraph(
            "Don <b>Alan Mauricio Gatica Concha</b>, cédula nacional de identidad "
            "N° <b>13.668.157-5</b>, en representación de "
            "<b>Inversiones Container House SpA</b>, Rol Único Tributario "
            "N° <b>78.268.851-0</b>, ambos con domicilio para estos efectos en "
            "Villasana N° 2039, Departamento 51, Torre D, comuna de Quinta Normal, "
            "Región Metropolitana, quien en adelante se denominará "
            "indistintamente \"el Proveedor\".", normal),
        SP(6),
        Paragraph("2. EL CLIENTE", bold),
        Paragraph(cli_bloque, normal),
        SP(4),
        Paragraph(
            f"Se deja expresa constancia que la dirección de instalación del proyecto "
            f"será <b>{d['inst_domicilio']}</b>, comuna de <b>{d['inst_comuna']}</b>, "
            f"Región {d['inst_region']}.", normal),
        SP(4),
        Paragraph(
            "Las partes declaran ser mayores de edad, con plena capacidad legal para "
            "contratar, y acuerdan celebrar el presente <b>Contrato de Fabricación y "
            "Venta de Vivienda Tipo Container</b>, el cual se regirá por las cláusulas "
            "que se indican a continuación.", normal),
        HR(),
    ]

    # ── II. Definiciones ──
    story += [
        Paragraph("II. DEFINICIONES", seccion),
        Paragraph("Para efectos del presente contrato, se entenderá por:", normal),
        Paragraph(
            f"a) <b>Proyecto</b>: La vivienda tipo container identificada como "
            f"<b>Proyecto N° {d['ep_numero']} – \"{d['ep_nombre']}\"</b>.", normal),
        Paragraph(
            "b) <b>Anexos</b>: Los documentos técnicos y comerciales que forman parte "
            "integrante del presente contrato, en especial Anexo N°1 (Especificaciones "
            "Técnicas) y Anexo N°2 (Presupuesto Detallado).", normal),
        Paragraph(
            "c) <b>Preentrega</b>: Instancia de revisión visual del módulo previo a su "
            "despacho desde las instalaciones del Proveedor.", normal),
        HR(),
    ]

    # ── III. Objeto ──
    story += [
        Paragraph("III. OBJETO DEL CONTRATO", seccion),
        Paragraph(
            "El Cliente encarga al Proveedor la <b>fabricación y venta</b> del Proyecto "
            "individualizado precedentemente, conforme a los <b>planos entregados por el "
            "Cliente</b>, a las <b>especificaciones técnicas</b>, y al <b>presupuesto "
            "detallado contenido en el Anexo N°2</b>, documentos que el Cliente declara "
            "conocer, aceptar y que forman parte integrante e inseparable del presente "
            "contrato.", normal),
        HR(),
    ]

    # ── IV. Alcance técnico ──
    story += [
        Paragraph("IV. ALCANCE TÉCNICO Y EJECUCIÓN", seccion),
        Paragraph(
            "El Proveedor declara contar con la experiencia, conocimientos técnicos, "
            "personal calificado, herramientas e infraestructura necesarias para la "
            "correcta ejecución del Proyecto, comprometiéndose a:", normal),
        Paragraph("a) Fabricar el módulo conforme a la normativa vigente aplicable.", normal),
        Paragraph("b) Respetar las especificaciones técnicas y alcances definidos en los Anexos.", normal),
        Paragraph("c) Ejecutar los trabajos con estándares de calidad y seguridad.", normal),
        Paragraph(
            "Cualquier trabajo, modificación o prestación no contemplada expresamente en "
            "los Anexos será considerada <b>obra adicional</b>, debiendo ser cotizada y "
            "aprobada por escrito por ambas partes.", normal),
        HR(),
    ]

    # ── V. Visitas ──
    story += [
        Paragraph("V. VISITAS Y SEGUIMIENTO DEL PROYECTO", seccion),
        Paragraph(
            "El Cliente podrá realizar visitas de seguimiento a las instalaciones del "
            "Proveedor ubicadas en <b>Portezuelo, parcela 3, Colina, Región "
            "Metropolitana</b>, previa coordinación con al menos <b>48 horas hábiles de "
            "anticipación</b>, con el único objeto de verificar el avance del Proyecto, "
            "quedando expresamente prohibida cualquier interferencia en los procesos "
            "productivos o instrucciones al personal del Proveedor.", normal),
        HR(),
    ]

    # ── VI. Precio ──
    story += [
        Paragraph("VI. PRECIO", seccion),
        Paragraph(
            f"El precio total del Proyecto asciende a la suma de "
            f"<b>{fmt(precio)}</b> ({precio_p}), IVA incluido.", normal),
        HR(),
    ]

    # ── VII. Forma de pago ──
    story += [
        Paragraph("VII. FORMA Y ETAPAS DE PAGO", seccion),
        Paragraph("El precio será pagado por el Cliente al Proveedor en las siguientes etapas:", normal),
        Paragraph(
            f"a) <b>50% inicial</b>: <b>{fmt(p50)}</b> ({p50_p}), "
            f"correspondiente a la asignación del contenedor y ejecución de obra gruesa.", normal),
        Paragraph(
            f"b) <b>25% intermedio</b>: <b>{fmt(p25a)}</b> ({p25a_p}), "
            f"una vez finalizada la obra gruesa.", normal),
        Paragraph(
            f"c) <b>25% final</b>: <b>{fmt(p25b)}</b> ({p25b_p}), "
            f"luego de la preentrega del Proyecto y el mismo día del despacho del módulo.", normal),
        Paragraph(
            "El Proveedor emitirá la factura correspondiente al día hábil siguiente de "
            "recibido cada pago, bajo modalidad de <b>pago al contado</b>.", normal),
        HR(),
    ]

    # ── VIII. Inicio fabricación ──
    story += [
        Paragraph("VIII. INICIO DE FABRICACIÓN", seccion),
        Paragraph(
            "La fabricación del Proyecto se iniciará <b>única y exclusivamente</b> una "
            "vez recibido y efectivamente abonado el pago inicial del "
            "<b>50% del valor total del contrato</b>.", normal),
        HR(),
    ]

    # ── IX. Medios de pago ──
    story += [
        Paragraph("IX. MEDIOS DE PAGO", seccion),
        Paragraph(
            "Los pagos deberán efectuarse mediante <b>transferencia electrónica, "
            "cheque o vale vista</b>, a la siguiente cuenta bancaria:", normal),
    ]
    datos_banco = [
        ["Razón Social:", "Inversiones Container House SpA"],
        ["RUT:",          "78.268.851-0"],
        ["Banco:",        "Banco Itaú"],
        ["Cuenta Corriente:", "N° 230771767"],
        ["Correo de confirmación:", "jperez@espaciocontainerhouse.cl"],
    ]
    tbl = Table(datos_banco, colWidths=[4.5*cm, 11*cm])
    tbl.setStyle(TableStyle([
        ('FONTNAME',  (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',  (1,0), (1,-1), 'Helvetica'),
        ('FONTSIZE',  (0,0), (-1,-1), 10),
        ('LEADING',   (0,0), (-1,-1), 14),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.HexColor('#f8fafc'), colors.white]),
        ('TOPPADDING',    (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('LEFTPADDING',   (0,0), (-1,-1), 6),
    ]))
    story += [tbl, SP(6),
        Paragraph(
            "Cada pago deberá ser informado por el Cliente mediante correo electrónico, "
            "adjuntando el comprobante respectivo.", normal),
        HR(),
    ]

    # ── X. Plazo ──
    story += [
        Paragraph("X. PLAZO DE FABRICACIÓN Y ENTREGA", seccion),
        Paragraph(
            f"El plazo máximo de fabricación y entrega será de "
            f"<b>{d['plazo_dias']} días hábiles administrativos</b>, contados desde el "
            "día hábil siguiente a aquel en que los fondos del anticipo se encuentren "
            "efectivamente liberados.", normal),
        Paragraph(
            "El Cliente se obliga a contar con <b>radier y/o apoyos estructurales "
            "ejecutados, nivelados y aptos</b> para la instalación dentro de un plazo "
            "máximo de <b>30 días hábiles</b> desde la firma del contrato. Cualquier "
            "atraso en estas condiciones suspenderá automáticamente los plazos de entrega.",
            normal),
        HR(),
    ]

    # ── XI. Penalidad ──
    story += [
        Paragraph("XI. PENALIDAD POR ATRASO", seccion),
        Paragraph(
            "En caso de atraso imputable exclusivamente al Proveedor en los plazos "
            "establecidos para la fabricación o entrega del Proyecto, éste pagará al "
            "Cliente, a título de indemnización única y total, una suma equivalente al "
            "<b>1% del valor neto correspondiente al último 25% del Proyecto por cada "
            "7 días hábiles de atraso</b>, con un "
            "<b>tope máximo del 10% del valor neto de dicho monto</b>.", normal),
        Paragraph(
            "No se considerarán atrasos imputables al Proveedor aquellos derivados de "
            "caso fortuito, fuerza mayor, condiciones climáticas adversas, retrasos de "
            "proveedores externos, o cualquier situación no atribuible directamente al "
            "Proveedor.", normal),
        Paragraph(
            "Asimismo, en caso de que el atraso sea imputable al Cliente, ya sea por "
            "retraso en los pagos comprometidos, falta de entrega de antecedentes "
            "necesarios, impedimentos de acceso al lugar de instalación, o cualquier "
            "otra circunstancia bajo su responsabilidad, los plazos del Proyecto se "
            "extenderán automáticamente por el mismo período de tiempo que dure dicho "
            "atraso, sin que ello genere responsabilidad ni penalidad alguna para el "
            "Proveedor.", normal),
        HR(),
    ]

    # ── XII. Retiro y bodegaje ──
    story += [
        Paragraph("XII. RETIRO, DESPACHO Y BODEGAJE", seccion),
        Paragraph(
            "Una vez notificada la finalización del Proyecto, el Cliente dispondrá de "
            "un plazo máximo de <b>10 días hábiles</b> para coordinar el retiro o "
            "despacho del módulo.", normal),
        Paragraph(
            "Vencido dicho plazo, el Proveedor quedará facultado para cobrar un "
            "<b>cargo por bodegaje equivalente al 1% del valor neto del Proyecto por "
            "cada 7 días corridos</b>, hasta el retiro efectivo.", normal),
        HR(),
    ]

    # ── XIII. Garantía ──
    story += [
        Paragraph("XIII. GARANTÍA", seccion),
        Paragraph(
            "El Proveedor otorga una garantía de <b>6 meses</b> contados desde la "
            "entrega del módulo, limitada exclusivamente a <b>defectos de fabricación "
            "o construcción imputables al proceso productivo</b>.", normal),
        Paragraph("Quedan expresamente excluidos de garantía los daños derivados de:", normal),
        Paragraph("• Mal uso o uso distinto al previsto", normal),
        Paragraph("• Modificaciones no autorizadas", normal),
        Paragraph("• Transporte realizado por terceros", normal),
        Paragraph("• Vandalismo", normal),
        Paragraph("• Fenómenos naturales", normal),
        Paragraph("• Falta de mantención adecuada", normal),
        HR(),
    ]

    # ── XIV. Terminación anticipada ──
    story += [
        Paragraph("XIV. TERMINACIÓN ANTICIPADA", seccion),
        Paragraph("El presente contrato podrá terminarse anticipadamente por:", normal),
        Paragraph("a) Incumplimiento grave de cualquiera de las partes.", normal),
        Paragraph("b) Mutuo acuerdo por escrito.", normal),
        Paragraph("c) No pago oportuno de cualquiera de las etapas de pago.", normal),
        Paragraph(
            "En caso de término imputable al Cliente, los montos pagados "
            "<b>no serán reembolsables</b>, salvo acuerdo distinto por escrito.", normal),
        HR(),
    ]

    # ── XV. Domicilio y jurisdicción ──
    story += [
        Paragraph("XV. DOMICILIO Y JURISDICCIÓN", seccion),
        Paragraph(
            "Para todos los efectos legales derivados del presente contrato, las partes "
            "fijan su domicilio en la <b>ciudad de Santiago</b>, y se someten a la "
            "competencia de sus <b>Tribunales Ordinarios de Justicia</b>.", normal),
        PageBreak(),
    ]

    # ── XVI. Firma ──
    story += [
        Paragraph("XVI. FIRMA", seccion),
        Paragraph(
            "El presente contrato se firma en <b>dos ejemplares de igual tenor y "
            "fecha</b>, quedando uno en poder de cada parte.", normal),
        SP(60),
    ]

    # Bloque de firmas en tabla 2 columnas
    if d['tipo_cliente'] == 'natural':
        cli_firma_nombre = d['cli_nombre']
        cli_firma_sub    = f"RUT: {d['cli_rut']}"
    else:
        cli_firma_nombre = d['cli_nombre']
        cli_firma_sub    = f"RUT: {d['cli_rut']}\n{d['cli_empresa']}"

    firma_data = [[
        Paragraph("EL PROVEEDOR", firma_bold),
        Paragraph("EL CLIENTE", firma_bold),
    ],[
        Paragraph("_" * 34, firma),
        Paragraph("_" * 34, firma),
    ],[
        Paragraph("Alan Mauricio Gatica Concha", firma_bold),
        Paragraph(cli_firma_nombre, firma_bold),
    ],[
        Paragraph("RUT: 13.668.157-5", firma),
        Paragraph(f"RUT: {d['cli_rut']}", firma),
    ],[
        Paragraph("Inversiones Container House SpA", firma),
        Paragraph(d.get('cli_empresa', '') or '', firma),
    ]]
    firma_tbl = Table(firma_data, colWidths=[8*cm, 8*cm])
    firma_tbl.setStyle(TableStyle([
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW',  (0,0), (-1,0), 0.8, AZUL),
    ]))
    story.append(firma_tbl)

    doc.build(story,
              onFirstPage=_build_header_footer,
              onLaterPages=_build_header_footer)
    buf.seek(0)
    return buf.read()


# FUNCIÓN: RANKING DE EJECUTIVOS
# =========================================================
def cargar_ranking_ejecutivos(periodo='mes'):
    """Carga métricas de ejecutivos via RPC SECURITY DEFINER — bypasea RLS."""
    try:
        from datetime import datetime as _dt, timedelta as _td
        _inicio = None
        if periodo == 'mes':
            _inicio = _dt.now().replace(day=1).strftime('%Y-%m-%d')
        # supabase_admin usa service_role JWT que bypasea RLS completamente
        resp = supabase_admin.rpc('get_ranking_data', {'fecha_inicio': _inicio}).execute()
        resp_data = resp.data if resp.data else []
        if not resp_data:
            return []
        # Agrupar por asesor
        asesores = {}
        for row in resp_data:
            nombre = row.get('asesor_nombre') or 'Sin asignar'
            if nombre not in asesores:
                asesores[nombre] = {
                    'nombre': nombre,
                    'total_presupuestos': 0,
                    'total_generado': 0.0,
                    'autorizados': 0,
                    'borradores': 0,
                    'incompletos': 0,
                }
            a = asesores[nombre]
            a['total_presupuestos'] += 1
            a['total_generado'] += float(row.get('total_total') or 0)
            # Clasificar estado
            margen = float(row.get('config_margen') or 0)
            datos_ok = all([row.get('cliente_nombre'), row.get('cliente_email')])
            asesor_ok = any([row.get('asesor_nombre'), row.get('asesor_email'), row.get('asesor_telefono')])
            if margen > 0 and datos_ok and asesor_ok:
                a['autorizados'] += 1
            elif datos_ok and asesor_ok:
                a['borradores'] += 1
            else:
                a['incompletos'] += 1
        # Calcular métricas derivadas
        ranking = []
        for nombre, a in asesores.items():
            n = a['total_presupuestos']
            a['promedio'] = a['total_generado'] / n if n > 0 else 0
            a['pct_autorizado'] = round((a['autorizados'] / n) * 100) if n > 0 else 0
            # Score: pondera total generado (60%) + % autorizado (25%) + cantidad (15%)
            max_total = max((x['total_generado'] for x in asesores.values()), default=1) or 1
            max_n     = max((x['total_presupuestos'] for x in asesores.values()), default=1) or 1
            a['score'] = round(
                (a['total_generado'] / max_total) * 60 +
                (a['pct_autorizado'] / 100) * 25 +
                (a['total_presupuestos'] / max_n) * 15
            )
            ranking.append(a)
        # Ordenar por score desc
        ranking.sort(key=lambda x: x['score'], reverse=True)
        return ranking
    except Exception as e:
        return []


# =========================================================
# TABS
# =========================================================
_rol_actual = st.session_state.get('rol_usuario', 'ejecutivo')
if _rol_actual == 'root':
    tab_dash, tab1, tab2, tab3, tab6, tab7, tab_contrato, tab4, tab5, tab_salud, tab_usuarios, tab_notif = st.tabs(["📊 DASHBOARD", "📋 PRESUPUESTO", "👤 DATOS", "📂 COTIZACIONES", "✏️ EDICIÓN PDF", "🏆 RANKING", "📄 CONTRATO", "🧊 3D BETA", "📊 PROYECTO EXCEL", "🛡️ SISTEMA", "👥 USUARIOS", "📣 NOTIFICACIONES"])
elif _rol_actual == 'admin':
    tab_dash, tab1, tab2, tab3, tab6, tab7, tab_contrato, tab4, tab5, tab_usuarios, tab_notif = st.tabs(["📊 DASHBOARD", "📋 PRESUPUESTO", "👤 DATOS", "📂 COTIZACIONES", "✏️ EDICIÓN PDF", "🏆 RANKING", "📄 CONTRATO", "🧊 3D BETA", "📊 PROYECTO EXCEL", "👥 USUARIOS", "📣 NOTIFICACIONES"])
    tab_salud = None
else:
    tab1, tab2, tab3, tab7, tab_contrato, tab4 = st.tabs(["📋 PRESUPUESTO", "👤 DATOS", "📂 COTIZACIONES", "🏆 RANKING", "📄 CONTRATO", "🧊 3D BETA"])
    tab_dash = None
    tab_salud = None
    tab5 = None
    tab6 = None
    tab_usuarios = None
    tab_notif = None

# =========================================================
# FUNCIÓN PARA GENERAR PDF COMPLETO
# =========================================================

def _construir_texto_cliente_pdf(datos_cliente, style):
    """Construye párrafos del bloque cliente para PDFs, mostrando empresa si es jurídica."""
    d = datos_cliente
    tipo = d.get("TipoCliente", "natural")
    lines = []
    if d.get("Nombre"):       lines.append(f"<b>Nombre:</b> {d['Nombre']}")
    if d.get("RUT"):          lines.append(f"<b>RUT:</b> {d['RUT']}")
    if tipo == "juridica":
        if d.get("EmpresaCliente"): lines.append(f"<b>Empresa:</b> {d['EmpresaCliente']}")
        if d.get("RutEmpresa"):     lines.append(f"<b>RUT empresa:</b> {d['RutEmpresa']}")
    if d.get("Correo"):       lines.append(f"<b>Correo:</b> {d['Correo']}")
    if d.get("Teléfono"):     lines.append(f"<b>Teléfono:</b> {d['Teléfono']}")
    if d.get("Dirección"):
        dir_completa = d["Dirección"]
        if d.get("ComunaCliente"): dir_completa += f", {d['ComunaCliente']}"
        if d.get("RegionCliente"): dir_completa += f", {d['RegionCliente']}"
        lines.append(f"<b>Dirección:</b> {dir_completa}")
    if d.get("DireccionProyecto"):
        inst_completa = d["DireccionProyecto"]
        if d.get("ComunaProyecto"): inst_completa += f", {d['ComunaProyecto']}"
        if d.get("RegionProyecto"): inst_completa += f", {d['RegionProyecto']}"
        lines.append(f"<b>Dirección instalación:</b> {inst_completa}")
    if d.get("Observaciones"):
        lines.append(f"<b>Descripción del proyecto:</b> {d['Observaciones']}")
    return Paragraph("<br/>".join(lines), style)


@st.cache_data(ttl=300)
def cargar_visibilidad_impresion():
    """Carga hoja Impresion del Excel activo. Retorna dict {item_lower: 'Mostrar'|'Ocultar'}."""
    try:
        import openpyxl as _opx
        _src = _get_excel_bytes_activo()
        _wb  = _opx.load_workbook(_src, data_only=True, read_only=True)
        if 'Impresion' not in _wb.sheetnames:
            return {}
        _ws = _wb['Impresion']
        _vis = {}
        for _row in _ws.iter_rows(min_row=2, values_only=True):
            _item, _cat, _pdf = (_row[0], _row[1], _row[2]) if len(_row) >= 3 else (None, None, None)
            if _item and _pdf:
                _vis[str(_item).strip().lower()] = str(_pdf).strip()
        return _vis
    except:
        return {}


def _generar_qr_imagen(url, size=80):
    """Genera un QR code como imagen ReportLab desde una URL."""
    try:
        import qrcode as _qr
        import io as _io
        _qr_img = _qr.make(url)
        _buf = _io.BytesIO()
        _qr_img.save(_buf, format='PNG')
        _buf.seek(0)
        return _buf, size
    except:
        return None, size

def generar_pdf_completo(carrito_df, subtotal, iva, total, datos_cliente,
                     fecha_inicio, fecha_termino, dias_validez,
                     datos_asesor, margen=0, numero_cotizacion=None):

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           leftMargin=20, rightMargin=20,
                           topMargin=30, bottomMargin=30, allowSplitting=1)
    elements = []
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(name='SmallFont', parent=styles['Normal'], fontSize=8, leading=12, wordWrap='CJK', leftIndent=0, alignment=0))
    styles.add(ParagraphStyle(name='HeaderStyle', parent=styles['Normal'], fontSize=9, leading=11, alignment=1, textColor=colors.white, fontName='Helvetica-Bold', leftIndent=0))
    styles.add(ParagraphStyle(name='TituloPresupuesto', parent=styles['Normal'], fontSize=16, leading=20, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0))
    styles.add(ParagraphStyle(name='TextoNormal', parent=styles['Normal'], fontSize=10, leading=14, leftIndent=0, alignment=0, spaceAfter=2))
    styles.add(ParagraphStyle(name='TituloSeccion', parent=styles['Normal'], fontSize=12, leading=14, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0))
    styles.add(ParagraphStyle(name='NotasEstilo', parent=styles['Normal'], fontSize=9, leading=13, leftIndent=0, alignment=0, textColor=colors.HexColor('#0d2266'), spaceAfter=2))
    styles.add(ParagraphStyle(name='TotalLabel', parent=styles['Normal'], fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2))
    styles.add(ParagraphStyle(name='TotalValue', parent=styles['Normal'], fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2))
    styles.add(ParagraphStyle(name='TotalBold', parent=styles['Normal'], fontSize=12, leading=16, alignment=2, fontName='Helvetica-Bold', spaceAfter=2, textColor=colors.black))

    numero_presupuesto = numero_cotizacion if numero_cotizacion else f"EP-{random.randint(1000,9999)}"
    fecha_emision = datetime.now()

    # Estilos encabezado
    styles.add(ParagraphStyle(name='EmpresaNombre', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold', leading=13,
        textColor=colors.HexColor('#0d2266'), spaceAfter=2))
    styles.add(ParagraphStyle(name='QRLabel', parent=styles['Normal'],
        fontSize=7, leading=9, textColor=colors.HexColor('#6b7280'), alignment=1))
    styles.add(ParagraphStyle(name='EPTitulo', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold', leading=13,
        textColor=colors.HexColor('#0d2266'), spaceAfter=2))

    # Proporciones: col1=45%, col2=55% (iguales en fila 2 y 3)
    _col1 = doc.width * 0.45
    _col2 = doc.width * 0.55

    # ── FILA 1: vacío | logo (centrado) | QR (derecha) ──
    _qr_img, _qr_sz = _generar_qr_imagen("https://www2.sii.cl/stc/noauthz/consulta", size=70)
    _logo_cell = ""
    try:
        _logo = Image("logo.png")
        _logo_max = _col2 * 0.80
        _logo_aspect = _logo.imageHeight / float(_logo.imageWidth)
        _logo.drawWidth  = _logo_max
        _logo.drawHeight = _logo_max * _logo_aspect
        _logo_cell = _logo
    except:
        _logo_cell = Paragraph("", styles['EPTitulo'])

    if _qr_img:
        from reportlab.platypus import Image as _RLImage
        _qr_img.name = 'qr.png'
        _qr_rl = _RLImage(_qr_img, width=_qr_sz, height=_qr_sz)
        _qr_label = Paragraph("SII Verifíquenos", styles['QRLabel'])
        _qr_cell = Table([[_qr_rl], [_qr_label]], colWidths=[_qr_sz + 4])
        _qr_cell.setStyle(TableStyle([
            ('ALIGN',  (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING',  (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING',   (0,0), (-1,-1), 0),
            ('BOTTOMPADDING',(0,0), (-1,-1), 0),
        ]))
    else:
        _qr_cell = Paragraph("", styles['EPTitulo'])

    # Col2 fila1: logo centrado + QR a la derecha
    _col2_fila1 = Table([[_logo_cell, _qr_cell]],
        colWidths=[_col2 - _qr_sz - 14, _qr_sz + 10])
    _col2_fila1.setStyle(TableStyle([
        ('ALIGN',  (0,0), (0,0), 'LEFT'),
        ('ALIGN',  (1,0), (1,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))

    # Fila 1: logo izquierda | vacío centro | QR derecha
    _qr_col_w = _qr_sz + 10
    _logo_col_w = _col1
    _mid_col_w = _col2 - _qr_col_w

    _fila1 = Table([[_logo_cell, "", _qr_cell]],
        colWidths=[_logo_col_w, _mid_col_w, _qr_col_w])
    _fila1.setStyle(TableStyle([
        ('ALIGN',  (0,0), (0,0), 'LEFT'),
        ('ALIGN',  (2,0), (2,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(_fila1)

    # ── FILA 2: EP/fechas | datos empresa ──
    _txt_ep = Paragraph(
        f"<font name='Helvetica-Bold' size='14' color='#0d2266'>PRESUPUESTO Nº {numero_presupuesto}</font><br/>"
        f"<font size='9' color='#374151'><b>Fecha Emisión:</b> {fecha_emision.strftime('%d-%m-%Y')}</font><br/>"
        f"<font size='9' color='#374151'><b>Validez:</b> {fecha_inicio.strftime('%d-%m-%Y')} hasta "
        f"{fecha_termino.strftime('%d-%m-%Y')} ({dias_validez} días)</font>",
        styles['EPTitulo'])

    _txt_empresa = Paragraph(
        "<b>INVERSIONES CONTAINER HOUSE SPA</b><br/>"
        "RUT: 78.268.851-0<br/>"
        "Construcción y Acondicionamiento Containers<br/>"
        "Villasana 2039, Quinta Normal, Santiago",
        styles['EmpresaNombre'])

    _fila2 = Table([[_txt_ep, _txt_empresa]], colWidths=[_col1, _col2])
    _fila2.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (0,0), (0,-1), 'LEFT'),
        ('ALIGN',  (1,0), (1,-1), 'RIGHT'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(_fila2)
    elements.append(Spacer(1, 10))

    # Fila 3 — mismas proporciones que fila 2
    data_ca = [[Paragraph("<b>DATOS DEL CLIENTE</b>", styles['TituloSeccion']), Paragraph("<b>DATOS DEL ASESOR</b>", styles['TituloSeccion'])]]
    asesor_text = "".join(f"<b>{k}:</b> {v}<br/>" for k, v in datos_asesor.items() if v)
    data_ca.append([_construir_texto_cliente_pdf(datos_cliente, styles['TextoNormal']), Paragraph(asesor_text, styles['TextoNormal'])])
    tabla_ca = Table(data_ca, colWidths=[_col1, _col2])
    tabla_ca.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (0,0), (0,-1), 'LEFT'),
        ('ALIGN',  (1,0), (1,-1), 'LEFT'),
        ('LEFTPADDING',  (0,0), (0,-1), 0),
        ('RIGHTPADDING', (0,0), (0,-1), 10),
        ('LEFTPADDING',  (1,0), (1,-1), 0),
        ('RIGHTPADDING', (1,0), (1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
    ]))
    elements.append(tabla_ca)
    elements.append(Spacer(1, 20))

    ancho_total = doc.width
    porcentajes = [15, 50, 8, 13.5, 13.5]
    anchos = [ancho_total * p / 100 for p in porcentajes]
    data = [[
        Paragraph("<b>Categoría</b>", styles['HeaderStyle']),
        Paragraph("<b>Item</b>", styles['HeaderStyle']),
        Paragraph("<b>Cant.</b>", styles['HeaderStyle']),
        Paragraph("<b>P. Unitario</b>", styles['HeaderStyle']),
        Paragraph("<b>Subtotal</b>", styles['HeaderStyle'])
    ]]
    for _, row in carrito_df.iterrows():
        data.append([
            Paragraph(row["Categoria"], styles['SmallFont']),
            Paragraph(row["Item"], styles['SmallFont']),
            Paragraph(str(row["Cantidad"]), styles['SmallFont']),
            Paragraph(formato_clp(row["Precio Unitario"]), styles['SmallFont']),
            Paragraph(formato_clp(row["Subtotal"]), styles['SmallFont'])
        ])

    tabla_productos = Table(data, colWidths=anchos, repeatRows=1, splitByRow=1)
    tabla_productos.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.black), ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,0), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9), ('BOTTOMPADDING', (0,0), (-1,0), 8), ('TOPPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (2,1), (4,-1), 'RIGHT'), ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 2), ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,1), (-1,-1), 4), ('BOTTOMPADDING', (0,1), (-1,-1), 4),
    ]))
    for i in range(1, len(data)):
        if i % 2 == 0:
            tabla_productos.setStyle(TableStyle([('BACKGROUND', (0,i), (-1,i), colors.Color(0.95, 0.95, 0.95))]))
    elements.append(tabla_productos)
    elements.append(Spacer(1, 20))

    ancho_bloque = (doc.width - 20) / 2
    texto_transporte = "2.- Transporte y bases de apoyo <b>no incluidos</b>."
    _p50 = formato_clp(round(total * 0.50))
    _p25a = formato_clp(round(total * 0.25))
    _p25b = formato_clp(round(total * 0.25))
    notas_texto = f"""<font color='#0d2266'><b>NOTAS IMPORTANTES:</b></font><br/>1.- Valores incluyen IVA.<br/>{texto_transporte}<br/>3.- Formas de pago: transferencia - pago contado.<br/>4.- Proceso de pagos:<br/>&nbsp;&nbsp;&nbsp;<font size='10'><b>- 50% inicial</b> correspondiente a <b>{_p50}</b></font><br/>&nbsp;&nbsp;&nbsp;<font size='10'><b>- 25% obra</b> correspondiente a <b>{_p25a}</b></font><br/>&nbsp;&nbsp;&nbsp;<font size='10'><b>- 25% entrega</b> correspondiente a <b>{_p25b}</b></font>."""
    datos_transferencia = """<font color='#0d2266'><b>DATOS PARA TRANSFERENCIA:</b></font><br/>Inversiones Container House Spa<br/>RUT: 78.268.851-0<br/>Tipo de cuenta: Cuenta Corriente<br/>Banco: Itaú<br/>N° de cuenta: 230771767<br/>Correo: jperez@espaciocontainerhouse.cl"""
    # bloque_notas ya no se usa — se separa en notas_texto y datos_transferencia

    totales_data = [
        [Paragraph("Subtotal:", styles['TotalLabel']), Paragraph(formato_clp(subtotal), styles['TotalValue'])],
        [Paragraph("IVA (19%):", styles['TotalLabel']), Paragraph(formato_clp(iva), styles['TotalValue'])],
        [Paragraph("TOTAL:", styles['TotalBold']), Paragraph(formato_clp(total), styles['TotalBold'])]
    ]
    totales_tabla = Table(totales_data, colWidths=[100, 120])
    totales_tabla.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,-1), 'RIGHT'), ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2), ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2), ('LINEABOVE', (1,1), (1,1), 1, colors.grey),
        ('LINEABOVE', (1,2), (1,2), 2, colors.black),
    ]))
    # Mensaje cordial — columna central
    styles.add(ParagraphStyle(name='MensajeCordial', parent=styles['Normal'],
        fontSize=9.5, fontName='Helvetica-Oblique', leading=14,
        textColor=colors.HexColor('#4a5568'), alignment=1, spaceAfter=4))
    styles.add(ParagraphStyle(name='MensajeFirma', parent=styles['Normal'],
        fontSize=9, fontName='Helvetica-BoldOblique', leading=13,
        textColor=colors.HexColor('#0d2266'), alignment=1))
    styles.add(ParagraphStyle(name='MensajeWeb', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Oblique', leading=11,
        textColor=colors.HexColor('#6b7280'), alignment=1))

    _msg_cordial = Paragraph(
        "<para align='center'>"
        "<font name='Helvetica-Oblique' size='11'>✨</font><br/>"
        "<font name='Helvetica-Oblique' size='9' color='#718096'>¡Gracias por confiar en nosotros!<br/>"
        "Esperamos que este presupuesto<br/>"
        "supere tus expectativas.<br/><br/>"
        "Con cariño,<br/></font>"
        "<font name='Helvetica-BoldOblique' size='9' color='#0d2266'>La familia<br/>"
        "Espacio Container House<br/></font>"
        "<font name='Helvetica-Oblique' size='8' color='#9ca3af'>espaciocontainerhouse.cl</font>"
        "</para>",
        styles['NotasEstilo'])

    # Fila 1: notas (izq) | totales (der)
    # Fila 2: transferencia (izq) | mensaje cordial (der)
    _bloque_notas_p  = Paragraph(notas_texto, styles['NotasEstilo'])
    _bloque_transf_p = Paragraph(datos_transferencia, styles['NotasEstilo'])

    data_bloques = [
        [_bloque_notas_p,  totales_tabla],
        [_bloque_transf_p, _msg_cordial],
    ]
    tabla_bloques = Table(data_bloques, colWidths=[ancho_bloque, ancho_bloque])
    tabla_bloques.setStyle(TableStyle([
        ('ALIGN',  (0,0), (0,-1), 'LEFT'),
        ('ALIGN',  (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (0,0), 'TOP'),
        ('VALIGN', (0,1), (0,1), 'MIDDLE'),
        ('VALIGN', (1,0), (1,0), 'MIDDLE'),
        ('VALIGN', (1,1), (1,1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('LEFTPADDING',  (1,1), (1,1), 40),
        ('RIGHTPADDING', (1,0), (1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(tabla_bloques)
    doc.build(elements)
    buffer.seek(0)
    return buffer, numero_presupuesto

# =========================================================
# FUNCIÓN PARA GENERAR PDF CLIENTE
# =========================================================
def generar_pdf_cliente(carrito_df, subtotal, iva, total, datos_cliente,
                     fecha_inicio, fecha_termino, dias_validez,
                     datos_asesor, margen=0, numero_cotizacion=None, descripciones_ep=None):
    _descripciones_ep = descripciones_ep or {}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           leftMargin=20, rightMargin=20,
                           topMargin=30, bottomMargin=30, allowSplitting=1)
    elements = []
    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(name='SmallFont', parent=styles['Normal'], fontSize=8, leading=12, wordWrap='CJK', leftIndent=0, alignment=0))
    styles.add(ParagraphStyle(name='HeaderStyle', parent=styles['Normal'], fontSize=9, leading=11, alignment=1, textColor=colors.white, fontName='Helvetica-Bold', leftIndent=0))
    styles.add(ParagraphStyle(name='TituloPresupuesto', parent=styles['Normal'], fontSize=16, leading=20, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0))
    styles.add(ParagraphStyle(name='TextoNormal', parent=styles['Normal'], fontSize=10, leading=14, leftIndent=0, alignment=0, spaceAfter=2))
    styles.add(ParagraphStyle(name='TituloSeccion', parent=styles['Normal'], fontSize=12, leading=14, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0))
    styles.add(ParagraphStyle(name='NotasEstilo', parent=styles['Normal'], fontSize=9, leading=13, leftIndent=0, alignment=0, textColor=colors.HexColor('#0d2266'), spaceAfter=2))
    styles.add(ParagraphStyle(name='TotalLabel', parent=styles['Normal'], fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2))
    styles.add(ParagraphStyle(name='TotalValue', parent=styles['Normal'], fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2))
    styles.add(ParagraphStyle(name='TotalBold', parent=styles['Normal'], fontSize=12, leading=16, alignment=2, fontName='Helvetica-Bold', spaceAfter=2, textColor=colors.black))

    numero_presupuesto = numero_cotizacion if numero_cotizacion else f"EP-{random.randint(1000,9999)}"
    fecha_emision = datetime.now()

    # Estilos encabezado
    styles.add(ParagraphStyle(name='EmpresaNombre', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold', leading=13,
        textColor=colors.HexColor('#0d2266'), spaceAfter=2))
    styles.add(ParagraphStyle(name='QRLabel', parent=styles['Normal'],
        fontSize=7, leading=9, textColor=colors.HexColor('#6b7280'), alignment=1))
    styles.add(ParagraphStyle(name='EPTitulo', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold', leading=13,
        textColor=colors.HexColor('#0d2266'), spaceAfter=2))

    # Proporciones: col1=45%, col2=55% (iguales en fila 2 y 3)
    _col1 = doc.width * 0.45
    _col2 = doc.width * 0.55

    # ── FILA 1: vacío | logo (centrado) | QR (derecha) ──
    _qr_img, _qr_sz = _generar_qr_imagen("https://www2.sii.cl/stc/noauthz/consulta", size=70)
    _logo_cell = ""
    try:
        _logo = Image("logo.png")
        _logo_max = _col2 * 0.80
        _logo_aspect = _logo.imageHeight / float(_logo.imageWidth)
        _logo.drawWidth  = _logo_max
        _logo.drawHeight = _logo_max * _logo_aspect
        _logo_cell = _logo
    except:
        _logo_cell = Paragraph("", styles['EPTitulo'])

    if _qr_img:
        from reportlab.platypus import Image as _RLImage
        _qr_img.name = 'qr.png'
        _qr_rl = _RLImage(_qr_img, width=_qr_sz, height=_qr_sz)
        _qr_label = Paragraph("SII Verifíquenos", styles['QRLabel'])
        _qr_cell = Table([[_qr_rl], [_qr_label]], colWidths=[_qr_sz + 4])
        _qr_cell.setStyle(TableStyle([
            ('ALIGN',  (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING',  (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING',   (0,0), (-1,-1), 0),
            ('BOTTOMPADDING',(0,0), (-1,-1), 0),
        ]))
    else:
        _qr_cell = Paragraph("", styles['EPTitulo'])

    # Col2 fila1: logo centrado + QR a la derecha
    _col2_fila1 = Table([[_logo_cell, _qr_cell]],
        colWidths=[_col2 - _qr_sz - 14, _qr_sz + 10])
    _col2_fila1.setStyle(TableStyle([
        ('ALIGN',  (0,0), (0,0), 'LEFT'),
        ('ALIGN',  (1,0), (1,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))

    # Fila 1: logo izquierda | vacío centro | QR derecha
    _qr_col_w = _qr_sz + 10
    _logo_col_w = _col1
    _mid_col_w = _col2 - _qr_col_w

    _fila1 = Table([[_logo_cell, "", _qr_cell]],
        colWidths=[_logo_col_w, _mid_col_w, _qr_col_w])
    _fila1.setStyle(TableStyle([
        ('ALIGN',  (0,0), (0,0), 'LEFT'),
        ('ALIGN',  (2,0), (2,0), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(_fila1)

    # ── FILA 2: EP/fechas | datos empresa ──
    _txt_ep = Paragraph(
        f"<font name='Helvetica-Bold' size='14' color='#0d2266'>PRESUPUESTO Nº {numero_presupuesto}</font><br/>"
        f"<font size='9' color='#374151'><b>Fecha Emisión:</b> {fecha_emision.strftime('%d-%m-%Y')}</font><br/>"
        f"<font size='9' color='#374151'><b>Validez:</b> {fecha_inicio.strftime('%d-%m-%Y')} hasta "
        f"{fecha_termino.strftime('%d-%m-%Y')} ({dias_validez} días)</font>",
        styles['EPTitulo'])

    _txt_empresa = Paragraph(
        "<b>INVERSIONES CONTAINER HOUSE SPA</b><br/>"
        "RUT: 78.268.851-0<br/>"
        "Construcción y Acondicionamiento Containers<br/>"
        "Villasana 2039, Quinta Normal, Santiago",
        styles['EmpresaNombre'])

    _fila2 = Table([[_txt_ep, _txt_empresa]], colWidths=[_col1, _col2])
    _fila2.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (0,0), (0,-1), 'LEFT'),
        ('ALIGN',  (1,0), (1,-1), 'RIGHT'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0), (-1,-1), 8),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(_fila2)
    elements.append(Spacer(1, 10))

    # Fila 3 — mismas proporciones que fila 2
    data_ca = [[Paragraph("<b>DATOS DEL CLIENTE</b>", styles['TituloSeccion']), Paragraph("<b>DATOS DEL ASESOR</b>", styles['TituloSeccion'])]]
    asesor_text = "".join(f"<b>{k}:</b> {v}<br/>" for k, v in datos_asesor.items() if v)
    data_ca.append([_construir_texto_cliente_pdf(datos_cliente, styles['TextoNormal']), Paragraph(asesor_text, styles['TextoNormal'])])
    tabla_ca = Table(data_ca, colWidths=[_col1, _col2])
    tabla_ca.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN',  (0,0), (0,-1), 'LEFT'),
        ('ALIGN',  (1,0), (1,-1), 'LEFT'),
        ('LEFTPADDING',  (0,0), (0,-1), 0),
        ('RIGHTPADDING', (0,0), (0,-1), 10),
        ('LEFTPADDING',  (1,0), (1,-1), 0),
        ('RIGHTPADDING', (1,0), (1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
    ]))
    elements.append(tabla_ca)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("<b>RESUMEN POR CATEGORÍA:</b>", styles['TituloSeccion']))
    elements.append(Spacer(1, 10))

    categorias = carrito_df.groupby('Categoria')
    data_resumen = []
    for categoria, grupo in categorias:
        desc_custom = (_descripciones_ep or {}).get(categoria, '').strip()
        if desc_custom:
            descripcion_html = desc_custom.replace('\n', '<br/>')
            data_resumen.append([
                Paragraph(categoria, styles['SmallFont']),
                Paragraph(descripcion_html, styles['SmallFont'])
            ])
        else:
            _vis_map = cargar_visibilidad_impresion()
            items_lista = [
                item for item in grupo['Item'].tolist()
                if _vis_map.get(str(item).strip().lower(), 'Mostrar') == 'Mostrar'
            ]
            # Si todos los ítems dicen Ocultar → omitir categoría completa
            if not items_lista:
                continue
            descripcion_html = "<br/>".join(f"• {item}" for item in items_lista)
            data_resumen.append([
                Paragraph(categoria, styles['SmallFont']),
                Paragraph(descripcion_html, styles['SmallFont'])
            ])

    ancho_cat = doc.width * 0.25
    ancho_desc = doc.width * 0.75
    headers = [
        Paragraph("<b>Categoría</b>", styles['HeaderStyle']),
        Paragraph("<b>Descripción</b>", styles['HeaderStyle'])
    ]
    tabla_resumen = Table([headers] + data_resumen, colWidths=[ancho_cat, ancho_desc], repeatRows=1)
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.black),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
    ]))
    for i in range(1, len(data_resumen) + 1):
        if i % 2 == 0:
            tabla_resumen.setStyle(TableStyle([('BACKGROUND', (0,i), (-1,i), colors.Color(0.95, 0.95, 0.95))]))
    elements.append(tabla_resumen)
    elements.append(Spacer(1, 20))

    ancho_bloque = (doc.width - 20) / 2
    texto_transporte = "2.- Transporte y bases de apoyo <b>no incluidos</b>."
    _p50 = formato_clp(round(total * 0.50))
    _p25a = formato_clp(round(total * 0.25))
    _p25b = formato_clp(round(total * 0.25))
    notas_texto = f"""<font color='#0d2266'><b>NOTAS IMPORTANTES:</b></font><br/>1.- Valores incluyen IVA.<br/>{texto_transporte}<br/>3.- Formas de pago: transferencia - pago contado.<br/>4.- Proceso de pagos:<br/>&nbsp;&nbsp;&nbsp;<font size='10'><b>- 50% inicial</b> correspondiente a <b>{_p50}</b></font><br/>&nbsp;&nbsp;&nbsp;<font size='10'><b>- 25% obra</b> correspondiente a <b>{_p25a}</b></font><br/>&nbsp;&nbsp;&nbsp;<font size='10'><b>- 25% entrega</b> correspondiente a <b>{_p25b}</b></font>."""
    datos_transferencia = """<font color='#0d2266'><b>DATOS PARA TRANSFERENCIA:</b></font><br/>Inversiones Container House Spa<br/>RUT: 78.268.851-0<br/>Tipo de cuenta: Cuenta Corriente<br/>Banco: Itaú<br/>N° de cuenta: 230771767<br/>Correo: jperez@espaciocontainerhouse.cl"""
    # bloque_notas ya no se usa — se separa en notas_texto y datos_transferencia

    totales_data = [
        [Paragraph("Subtotal:", styles['TotalLabel']), Paragraph(formato_clp(subtotal), styles['TotalValue'])],
        [Paragraph("IVA (19%):", styles['TotalLabel']), Paragraph(formato_clp(iva), styles['TotalValue'])],
        [Paragraph("TOTAL:", styles['TotalBold']), Paragraph(formato_clp(total), styles['TotalBold'])]
    ]
    totales_tabla = Table(totales_data, colWidths=[100, 120])
    totales_tabla.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,-1), 'RIGHT'), ('ALIGN', (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'), ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2), ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2), ('LINEABOVE', (1,1), (1,1), 1, colors.grey),
        ('LINEABOVE', (1,2), (1,2), 2, colors.black),
    ]))
    # Mensaje cordial — columna central
    styles.add(ParagraphStyle(name='MensajeCordial', parent=styles['Normal'],
        fontSize=9.5, fontName='Helvetica-Oblique', leading=14,
        textColor=colors.HexColor('#4a5568'), alignment=1, spaceAfter=4))
    styles.add(ParagraphStyle(name='MensajeFirma', parent=styles['Normal'],
        fontSize=9, fontName='Helvetica-BoldOblique', leading=13,
        textColor=colors.HexColor('#0d2266'), alignment=1))
    styles.add(ParagraphStyle(name='MensajeWeb', parent=styles['Normal'],
        fontSize=8, fontName='Helvetica-Oblique', leading=11,
        textColor=colors.HexColor('#6b7280'), alignment=1))

    _msg_cordial = Paragraph(
        "<para align='center'>"
        "<font name='Helvetica-Oblique' size='11'>✨</font><br/>"
        "<font name='Helvetica-Oblique' size='9' color='#718096'>¡Gracias por confiar en nosotros!<br/>"
        "Esperamos que este presupuesto<br/>"
        "supere tus expectativas.<br/><br/>"
        "Con cariño,<br/></font>"
        "<font name='Helvetica-BoldOblique' size='9' color='#0d2266'>La familia<br/>"
        "Espacio Container House<br/></font>"
        "<font name='Helvetica-Oblique' size='8' color='#9ca3af'>espaciocontainerhouse.cl</font>"
        "</para>",
        styles['NotasEstilo'])

    # Fila 1: notas (izq) | totales (der)
    # Fila 2: transferencia (izq) | mensaje cordial (der)
    _bloque_notas_p  = Paragraph(notas_texto, styles['NotasEstilo'])
    _bloque_transf_p = Paragraph(datos_transferencia, styles['NotasEstilo'])

    data_bloques = [
        [_bloque_notas_p,  totales_tabla],
        [_bloque_transf_p, _msg_cordial],
    ]
    tabla_bloques = Table(data_bloques, colWidths=[ancho_bloque, ancho_bloque])
    tabla_bloques.setStyle(TableStyle([
        ('ALIGN',  (0,0), (0,-1), 'LEFT'),
        ('ALIGN',  (1,0), (1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (0,0), 'TOP'),
        ('VALIGN', (0,1), (0,1), 'MIDDLE'),
        ('VALIGN', (1,0), (1,0), 'MIDDLE'),
        ('VALIGN', (1,1), (1,1), 'MIDDLE'),
        ('LEFTPADDING',  (0,0), (-1,-1), 0),
        ('LEFTPADDING',  (1,1), (1,1), 40),
        ('RIGHTPADDING', (1,0), (1,-1), 0),
        ('TOPPADDING',   (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,0), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    elements.append(tabla_bloques)
    doc.build(elements)
    buffer.seek(0)
    return buffer, numero_presupuesto

# =========================================================
# FUNCIÓN LIMPIAR TODO
# =========================================================
def _ejecutar_cierre_cotizacion():
    """Limpia todo el estado al cerrar una cotización."""
    limpiar_todo()
    st.session_state.recien_guardado = True
    st.session_state.hash_ultimo_guardado = None

# Procesar triggers de cierre aquí, donde limpiar_todo ya está definida
if st.session_state.get('trigger_cerrar_cotizacion', False):
    st.session_state.trigger_cerrar_cotizacion = False
    _ejecutar_cierre_cotizacion()
    st.session_state.resultados_busqueda = buscar_cotizaciones()
    st.rerun()

# =========================================================
# TAB 2 - DATOS CLIENTE
# =========================================================
with tab2:
    st.markdown("""
    <style>
    .hdr2 {
        background: linear-gradient(135deg, #2d0d66 0%, #5b0d7a 100%);
        border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
        display: flex; align-items: center; gap: 22px;
        box-shadow: 0 8px 32px rgba(91,13,122,0.25);
        position: relative; overflow: hidden;
    }
    .hdr2::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 180px; height: 180px; border-radius: 50%;
        background: rgba(255,255,255,0.04); pointer-events: none;
    }
    .hdr2::after {
        content: ''; position: absolute; bottom: -60px; right: 80px;
        width: 240px; height: 240px; border-radius: 50%;
        background: rgba(255,255,255,0.03); pointer-events: none;
    }
    .hdr2 h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                 font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
    .hdr2 p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
    </style>
    <div class="hdr2">
      <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">👤</span>
      <div>
        <h2>Datos del Cliente</h2>
        <p>Completa la información del cliente y del proyecto antes de guardar.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    es_solo_lectura = bool(
        st.session_state.cotizacion_cargada and
        st.session_state.margen > 0 and
        not st.session_state.modo_admin
    )

    fecha_inicio = st.session_state.fecha_inicio
    fecha_termino = st.session_state.fecha_termino
    dias_validez = (fecha_termino - fecha_inicio).days

    if es_solo_lectura:
        st.warning("🔒 Modo solo lectura — cotización con márgenes aplicados.")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            with st.container(border=True):
                st.markdown("**👤 Cliente**")
                _tipo_lbl = "Persona jurídica" if st.session_state.cliente_tipo == "juridica" else "Persona natural"
                st.caption(f"🏷️ {_tipo_lbl}")
                st.text_input("Nombre", value=st.session_state.nombre_input, disabled=True, key="nombre_readonly")
                st.text_input("RUT", value=st.session_state.rut_display, disabled=True, key="rut_readonly")
                st.text_input("Correo", value=st.session_state.correo_input, disabled=True, key="correo_readonly")
                st.text_input("Teléfono", value=st.session_state.telefono_raw, disabled=True, key="telefono_readonly")
                if st.session_state.cliente_tipo == "juridica":
                    st.markdown("**🏢 Empresa**")
                    st.text_input("Razón social", value=st.session_state.cliente_empresa, disabled=True, key="empresa_readonly")
                    st.text_input("RUT empresa", value=st.session_state.cliente_rut_empresa, disabled=True, key="rut_empresa_readonly")
        with col2:
            with st.container(border=True):
                st.markdown("**📍 Cliente**")
                st.text_input("Dirección cliente", value=st.session_state.direccion_input, disabled=True, key="direccion_readonly")
                st.text_input("Comuna cliente", value=st.session_state.cliente_comuna, disabled=True, key="cliente_comuna_readonly")
                st.text_input("Región cliente", value=st.session_state.cliente_region, disabled=True, key="cliente_region_readonly")
                st.markdown("**🏗️ Proyecto**")
                st.text_input("Dirección instalación", value=st.session_state.proyecto_direccion, disabled=True, key="proyecto_dir_readonly")
                st.text_input("Comuna instalación", value=st.session_state.proyecto_comuna, disabled=True, key="proyecto_com_readonly")
                st.text_input("Región instalación", value=st.session_state.proyecto_region, disabled=True, key="proyecto_reg_readonly")
        with col3:
            with st.container(border=True):
                st.markdown("**👨‍💼 Ejecutivo**")
                st.text_input("Asesor", value=st.session_state.asesor_seleccionado, disabled=True, key="asesor_readonly")
                st.text_input("Correo Ejecutivo", value=st.session_state.correo_asesor, disabled=True, key="correo_asesor_readonly")
                st.text_input("Teléfono Ejecutivo", value=st.session_state.telefono_asesor, disabled=True, key="telefono_asesor_readonly")
        with col4:
            with st.container(border=True):
                st.markdown("**📅 Validez**")
                st.date_input("Fecha Inicio", value=fecha_inicio, disabled=True, key="fecha_inicio_readonly")
                st.date_input("Fecha Término", value=fecha_termino, disabled=True, key="fecha_termino_readonly")
                st.markdown(f"**⏱️ Duración:** {dias_validez} días")
                if dias_validez > 0:
                    st.progress(min(dias_validez/30, 1.0), text=f"{dias_validez} días")
        with st.container(border=True):
            st.markdown("**📝 Descripción del proyecto**")
            st.text_area("Descripción del proyecto", value=st.session_state.observaciones_input, disabled=True, height=80, key="observaciones_readonly")

    else:
        # ── Construir dict de asesores dinámicamente desde Supabase ──
        def _cargar_asesores_desde_supabase():
            """Carga la lista de ejecutivos/admins desde Supabase Auth."""
            try:
                _users = listar_usuarios_ejecutivos()
                _d = {"Seleccionar asesor": {"correo": "", "telefono": ""}}
                for _usr in _users:
                    _nm = (_usr.get('nombre') or _usr.get('email', '')).upper()
                    if _nm and _nm != "SELECCIONAR ASESOR":
                        _d[_nm] = {
                            "correo": _usr.get('email', '').upper(),
                            "telefono": _usr.get('telefono', '') or ''
                        }
                return _d
            except:
                # Fallback al dict estático si falla Supabase
                return {
                    "Seleccionar asesor": {"correo": "", "telefono": ""},
                    "BERNARD BUSTAMANTE": {"correo": "BALDAY@ESPACIOCONTAINERHOUSE.CL", "telefono": "+56956786366"},
                    "ANDREA OSORIO": {"correo": "AOSORIO@ESPACIOCONTAINERHOUSE.CL", "telefono": "+56927619483"},
                    "REBECA CALDERON": {"correo": "RCALDERON@ESPACIOCONTAINERHOUSE.CL", "telefono": "+56955286708"},
                    "MAURICIO CEVO": {"correo": "MCEVO@ESPACIOCONTAINERHOUSE.CL", "telefono": "+56971406162"},
                    "JACQUELINE PÉREZ": {"correo": "JPEREZ@ESPACIOCONTAINERHOUSE.CL", "telefono": "+56992286057"},
                    "JAVIER QUEZADA": {"correo": "JQUEZADA@ESPACIOCONTAINERHOUSE.CL", "telefono": "+56966983700"}
                }

        # Cachear por sesión y refrescar si se editaron usuarios
        if ('_asesores_cache' not in st.session_state or
                st.session_state.get('_asesores_cache_dirty', False)):
            st.session_state['_asesores_cache'] = _cargar_asesores_desde_supabase()
            st.session_state['_asesores_cache_dirty'] = False

        asesores = st.session_state['_asesores_cache']

        col1, col2, col3, col4 = st.columns(4)

        # ── Columna 1: Cliente ──
        with col1:
            with st.container(border=True):
                st.markdown("**👤 Cliente**")

                # Tipo de cliente
                tipo_key = f"cliente_tipo_{st.session_state.counter}"
                _tipo_options = ["natural", "juridica"]
                _tipo_labels  = ["Persona natural", "Persona jurídica"]
                _tipo_idx = _tipo_options.index(st.session_state.cliente_tipo) if st.session_state.cliente_tipo in _tipo_options else 0
                _tipo_sel = st.radio("Tipo", _tipo_labels, index=_tipo_idx,
                                     horizontal=True, key=tipo_key, label_visibility="collapsed")
                _tipo_val = _tipo_options[_tipo_labels.index(_tipo_sel)]
                if _tipo_val != st.session_state.cliente_tipo:
                    st.session_state.cliente_tipo = _tipo_val
                    st.rerun()

                nombre_key = f"nombre_input_{st.session_state.counter}"
                nombre = st.text_input("Nombre Completo*", placeholder="Ej: Juan Pérez", key=nombre_key, value=st.session_state.nombre_input)
                if nombre != st.session_state.nombre_input:
                    st.session_state.nombre_input = nombre

                correo_key = f"correo_input_{st.session_state.counter}"
                correo = st.text_input("Correo Electrónico*", placeholder="ejemplo@correo.cl", key=correo_key, value=st.session_state.correo_input)
                if correo != st.session_state.correo_input:
                    st.session_state.correo_input = correo
                if correo and "@" not in correo:
                    st.warning("⚠️ El correo debe contener @")

                rut_key = f"rut_input_{st.session_state.counter}"
                st.text_input("RUT (opcional)", value=st.session_state.rut_display, key=rut_key, placeholder="12.345.678-9", on_change=procesar_cambio_rut)
                if st.session_state.rut_raw:
                    if len(st.session_state.rut_raw) >= 2:
                        if st.session_state.rut_valido:
                            if "extranjero" in st.session_state.rut_mensaje.lower():
                                st.warning(f"⚠️ {st.session_state.rut_mensaje}")
                            else:
                                st.success("✅ RUT válido")
                        else:
                            _msg_rut = st.session_state.rut_mensaje
                            if "extranjero" in _msg_rut.lower():
                                st.warning(f"⚠️ RUT inválido o RUT extranjero")
                            else:
                                st.error(f"❌ {_msg_rut}")
                    else:
                        st.info("⏳ RUT incompleto")

                telefono_key = f"telefono_input_{st.session_state.counter}"
                st.text_input("Teléfono", value=st.session_state.telefono_raw, key=telefono_key, placeholder="961528954 (9 dígitos)", on_change=procesar_cambio_telefono)

                # Campos adicionales si es jurídica
                if st.session_state.cliente_tipo == "juridica":
                    st.markdown("---")
                    st.markdown("**🏢 Empresa**")
                    emp_key = f"cliente_empresa_{st.session_state.counter}"
                    empresa = st.text_input("Razón social*", placeholder="Ej: Constructora ABC SpA",
                                            key=emp_key, value=st.session_state.cliente_empresa)
                    if empresa != st.session_state.cliente_empresa:
                        st.session_state.cliente_empresa = empresa

                    rut_emp_key = f"rut_empresa_input_{st.session_state.counter}"
                    st.text_input("RUT empresa*", placeholder="76.123.456-7",
                                  key=rut_emp_key,
                                  value=st.session_state.rut_empresa_display,
                                  on_change=procesar_cambio_rut_empresa)
                    if st.session_state.rut_empresa_raw:
                        if len(st.session_state.rut_empresa_raw) >= 2:
                            if st.session_state.rut_empresa_valido:
                                if "extranjero" in (st.session_state.rut_empresa_mensaje or '').lower():
                                    st.warning(f"⚠️ {st.session_state.rut_empresa_mensaje}")
                                else:
                                    st.success("✅ RUT válido")
                            else:
                                st.error(f"❌ {st.session_state.rut_empresa_mensaje}")
                        else:
                            st.info("⏳ RUT incompleto")

        # ── Columna 2: Dirección ──
        with col2:
            with st.container(border=True):
                st.markdown("**📍 Cliente**")
                direccion_key = f"direccion_input_{st.session_state.counter}"
                direccion = st.text_input("Dirección cliente", placeholder="Calle, número", key=direccion_key, value=st.session_state.direccion_input)
                if direccion != st.session_state.direccion_input:
                    st.session_state.direccion_input = direccion
                _com_cli, _reg_cli = selector_comuna_region(
                    "Comuna cliente", "Región cliente",
                    f"cliente_comuna_{st.session_state.counter}",
                    f"cliente_region_{st.session_state.counter}",
                    val_com=st.session_state.cliente_comuna,
                    val_reg=st.session_state.cliente_region,
                )
                st.session_state.cliente_comuna = _com_cli
                st.session_state.cliente_region = _reg_cli

                st.markdown("**🏗️ Proyecto**")
                proy_dir_key = f"proyecto_direccion_{st.session_state.counter}"
                proy_dir = st.text_input("Dirección instalación", placeholder="Calle, número", key=proy_dir_key, value=st.session_state.proyecto_direccion)
                if proy_dir != st.session_state.proyecto_direccion:
                    st.session_state.proyecto_direccion = proy_dir
                _com_proy, _reg_proy = selector_comuna_region(
                    "Comuna instalación", "Región instalación",
                    f"proyecto_comuna_{st.session_state.counter}",
                    f"proyecto_region_{st.session_state.counter}",
                    val_com=st.session_state.proyecto_comuna,
                    val_reg=st.session_state.proyecto_region,
                )
                st.session_state.proyecto_comuna = _com_proy
                st.session_state.proyecto_region = _reg_proy

        # ── Columna 3: Ejecutivo ──
        # Si es ejecutivo, pre-cargar sus datos automáticamente y bloquear el selector
        _rol_actual_tab2 = st.session_state.get('rol_usuario', 'ejecutivo')
        _es_ejecutivo_tab2 = _rol_actual_tab2 == 'ejecutivo'

        if _es_ejecutivo_tab2:
            _email_logueado = st.session_state.get('auth_email', '').upper()
            _nombre_logueado = st.session_state.get('auth_nombre', '').upper()
            _datos_ej = None
            # Buscar por correo o nombre en el dict de asesores
            for _nm, _dat in asesores.items():
                if _nm == "Seleccionar asesor":
                    continue
                if _dat["correo"].upper() == _email_logueado or _nm.upper() == _nombre_logueado:
                    _datos_ej = (_nm, _dat)
                    break
            # Si no encontró en el dict, usar datos del perfil de Supabase
            if not _datos_ej and _nombre_logueado:
                _datos_ej = (_nombre_logueado, {
                    "correo": st.session_state.get('auth_email', ''),
                    "telefono": st.session_state.get('auth_telefono', '')
                })
            # Auto-cargar si encontró y no está ya cargado
            if _datos_ej and st.session_state.asesor_seleccionado != _datos_ej[0]:
                st.session_state.asesor_seleccionado = _datos_ej[0]
                st.session_state.correo_asesor = _datos_ej[1]["correo"]
                st.session_state.telefono_asesor = _datos_ej[1]["telefono"]

        with col3:
            with st.container(border=True):
                st.markdown("**👨‍💼 Ejecutivo**")

                if _es_ejecutivo_tab2:
                    # Ejecutivo: solo ve sus propios datos, sin dropdown
                    st.text_input("Nombre", value=st.session_state.asesor_seleccionado, disabled=True,
                                  key=f"ej_nombre_fixed_{st.session_state.counter}")
                    st.text_input("Correo Ejecutivo*", value=st.session_state.correo_asesor, disabled=True,
                                  key=f"ej_correo_fixed_{st.session_state.counter}")
                    st.text_input("Teléfono Ejecutivo", value=st.session_state.telefono_asesor, disabled=True,
                                  key=f"ej_tel_fixed_{st.session_state.counter}")
                    st.caption("🔒 Tus datos están asignados automáticamente.")
                else:
                    # Admin/root: dropdown completo
                    nombres_asesores = list(asesores.keys())
                    asesor_key = f"asesor_select_{st.session_state.counter}"
                    indice_actual = nombres_asesores.index(st.session_state.asesor_seleccionado) if st.session_state.asesor_seleccionado in nombres_asesores else 0
                    asesor_elegido = st.selectbox("Asesor", nombres_asesores, index=indice_actual, key=asesor_key, label_visibility="collapsed")
                    if asesor_elegido != st.session_state.asesor_seleccionado:
                        st.session_state.asesor_seleccionado = asesor_elegido
                        if asesor_elegido != "Seleccionar asesor":
                            st.session_state.correo_asesor = asesores[asesor_elegido]["correo"]
                            st.session_state.telefono_asesor = asesores[asesor_elegido]["telefono"]
                        else:
                            st.session_state.correo_asesor = ""
                            st.session_state.telefono_asesor = ""
                        st.session_state.counter += 1
                        st.rerun()

                    correo_asesor_key = f"asesor_correo_input_{st.session_state.counter}"
                    correo_input = st.text_input("Correo Ejecutivo*", value=st.session_state.correo_asesor, placeholder="ejecutivo@empresa.cl", key=correo_asesor_key)
                    if correo_input and "@" not in correo_input:
                        st.warning("⚠️ El correo debe contener @")
                    if correo_input != st.session_state.correo_asesor:
                        st.session_state.correo_asesor = correo_input
                        st.session_state.asesor_seleccionado = "Seleccionar asesor"
                        st.session_state.counter += 1
                        st.rerun()

                    telefono_asesor_key = f"asesor_telefono_input_{st.session_state.counter}"
                    telefono_input = st.text_input("Teléfono Ejecutivo", value=st.session_state.telefono_asesor, key=telefono_asesor_key, placeholder="912345678 (9 dígitos)")
                    if telefono_input != st.session_state.telefono_asesor:
                        raw = re.sub(r'[^0-9]', '', telefono_input)
                        if len(raw) > 9:
                            raw = raw[:9]
                        st.session_state.telefono_asesor = raw
                        st.session_state.asesor_seleccionado = "Seleccionar asesor"
                        st.session_state.counter += 1
                        st.rerun()

        # ── Columna 4: Validez ──
        with col4:
            with st.container(border=True):
                st.markdown("**📅 Validez**")
                fecha_inicio_key = f"fecha_inicio_{st.session_state.counter}"
                fecha_inicio = st.date_input("Fecha de Inicio", value=st.session_state.fecha_inicio, key=fecha_inicio_key)
                if fecha_inicio != st.session_state.fecha_inicio:
                    st.session_state.fecha_inicio = fecha_inicio

                fecha_termino_key = f"fecha_termino_{st.session_state.counter}"
                fecha_termino = st.date_input("Fecha de Término", value=st.session_state.fecha_termino, key=fecha_termino_key)
                if fecha_termino != st.session_state.fecha_termino:
                    st.session_state.fecha_termino = fecha_termino

                dias_validez = (fecha_termino - fecha_inicio).days
                if dias_validez < 0:
                    st.error("⚠️ Fecha de término anterior a inicio.")
                else:
                    st.markdown(f"**⏱️ Duración:** {dias_validez} días")
                    if dias_validez > 0:
                        st.progress(min(dias_validez/30, 1.0), text=f"{dias_validez} días de validez")

        # ── Observaciones (ancho completo) ──
        with st.container(border=True):
            st.markdown("**📝 Descripción del proyecto**")
            observaciones_key = f"observaciones_input_{st.session_state.counter}"
            observaciones = st.text_area("Descripción del proyecto", placeholder="Describe el proyecto, características especiales o información relevante...", height=80, key=observaciones_key, value=st.session_state.observaciones_input)
            if observaciones != st.session_state.observaciones_input:
                st.session_state.observaciones_input = observaciones

    nombre_asesor_final = st.session_state.asesor_seleccionado if st.session_state.asesor_seleccionado != "Seleccionar asesor" else ""
    datos_cliente = {
        "Nombre": st.session_state.nombre_input or "",
        "RUT": st.session_state.rut_display or "",
        "Correo": st.session_state.correo_input or "",
        "Teléfono": st.session_state.telefono_raw or "",
        "Dirección": st.session_state.direccion_input or "",
            "ComunaCliente": st.session_state.cliente_comuna or "",
            "RegionCliente": st.session_state.cliente_region or "",
            "DireccionProyecto": st.session_state.proyecto_direccion or "",
            "ComunaProyecto": st.session_state.proyecto_comuna or "",
            "RegionProyecto": st.session_state.proyecto_region or "",
            "TipoCliente": st.session_state.cliente_tipo or "natural",
            "EmpresaCliente": st.session_state.cliente_empresa or "",
            "RutEmpresa": st.session_state.cliente_rut_empresa or "",
            "Observaciones": st.session_state.observaciones_input or ""
    }
    nombre_asesor_final = st.session_state.asesor_seleccionado if st.session_state.asesor_seleccionado != "Seleccionar asesor" else ""
    datos_asesor = {
        "Nombre Ejecutivo": nombre_asesor_final,
        "Correo Ejecutivo": st.session_state.correo_asesor or "",
        "Teléfono Ejecutivo": st.session_state.telefono_asesor or ""
    }

# =========================================================
# TAB 1 - PREPARAR COTIZACIÓN
# =========================================================
with tab1:
    st.markdown("""
    <style>
    .hdr1 {
        background: linear-gradient(135deg, #0d2266 0%, #0d47a1 100%);
        border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
        display: flex; align-items: center; gap: 22px;
        box-shadow: 0 8px 32px rgba(37,99,235,0.25);
        position: relative; overflow: hidden;
    }
    .hdr1::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 180px; height: 180px; border-radius: 50%;
        background: rgba(255,255,255,0.04); pointer-events: none;
    }
    .hdr1::after {
        content: ''; position: absolute; bottom: -60px; right: 80px;
        width: 240px; height: 240px; border-radius: 50%;
        background: rgba(255,255,255,0.03); pointer-events: none;
    }
    .hdr1 h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                 font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
    .hdr1 p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
    </style>
    <div class="hdr1">
      <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">☑️</span>
      <div>
        <h2>Gestión de Presupuesto</h2>
        <p>Agrega productos, aplica márgenes y genera tu cotización en PDF.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    fecha_inicio = st.session_state.fecha_inicio
    fecha_termino = st.session_state.fecha_termino
    dias_validez = (fecha_termino - fecha_inicio).days

    # JS expansión popovers selectbox
    import streamlit.components.v1 as _sel_comp
    _sel_comp.html("""<script>
(function(){
    var D = window.parent.document;
    var _keys = ['modelo_select','cat_manual','item_manual','cat_eliminar','modelo_origen','cat_agregar'];

    function _expand() {
        var popovers = D.querySelectorAll('[data-baseweb="popover"]');
        if (!popovers.length) return;
        popovers.forEach(function(pop) {
            var ul = pop.querySelector('ul');
            if (!ul) return;
            var items = ul.querySelectorAll('li');
            if (!items.length) return;
            items.forEach(function(li) {
                li.style.setProperty('white-space','nowrap','important');
                li.style.setProperty('overflow','visible','important');
                li.style.setProperty('text-overflow','unset','important');
                li.querySelectorAll('*').forEach(function(ch){
                    ch.style.setProperty('white-space','nowrap','important');
                    ch.style.setProperty('overflow','visible','important');
                    ch.style.setProperty('text-overflow','unset','important');
                });
            });
            setTimeout(function(){
                var sw = ul.scrollWidth;
                if (sw < 100) return;
                var fw = Math.min(sw + 48, 1100);
                [pop, pop.firstElementChild].forEach(function(el){
                    if (el) {
                        el.style.setProperty('min-width', fw+'px','important');
                        el.style.setProperty('width', fw+'px','important');
                    }
                });
            }, 30);
        });
    }

    function _init() {
        _keys.forEach(function(k) {
            var el = D.querySelector('.st-key-' + k);
            if (!el) return;
            el.addEventListener('mousedown', function() {
                setTimeout(_expand, 100);
                setTimeout(_expand, 250);
                setTimeout(_expand, 500);
            }, true);
        });
    }
    setTimeout(_init, 900);
    setTimeout(_init, 2200);
})();
</script>""", height=0)

    es_solo_lectura = bool(
        st.session_state.cotizacion_cargada and
        st.session_state.margen > 0 and
        not st.session_state.modo_admin
    )

    if es_solo_lectura:
        st.warning("🔒 Esta cotización tiene márgenes aplicados. Modo solo lectura. Solo puedes visualizar y generar PDFs.")

    if not es_solo_lectura:
        # Calcular hojas_modelo ANTES de las columnas para que esté disponible en col_m4
        hojas_modelo = [h for h in _leer_hojas_disponibles() if h.lower().startswith("modelo")]
        col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns([1,1,1,1,0.7])

        with col_m1:
            with st.container(border=True):
                st.markdown("**📋 Modelo Predefinido**")
                try:
                    if hojas_modelo:
                        modelo_seleccionado = st.selectbox("Modelo", hojas_modelo, key="modelo_select", label_visibility="collapsed")
                        if st.button("Cargar", key="btn_modelo", use_container_width=True):
                            st.session_state.carrito = cargar_modelo(modelo_seleccionado)
                            st.session_state.modelo_base = modelo_seleccionado
                            st.session_state.margen = 0.0
                            st.session_state['_toast_msg'] = f"✅ Modelo '{modelo_seleccionado}' cargado correctamente."
                            st.rerun()
                    else:
                        st.caption("Sin modelos")
                except Exception as _e1:
                    st.caption(f"Error: {_e1}")

        with col_m2:
            with st.container(border=True):
                st.markdown("**🔍 Ítems**")
                try:
                    df = _leer_hoja_excel("BD Total")
                    categorias = df["Categorias"].dropna().unique()
                    categoria_seleccionada = st.selectbox("Categoría", categorias, key="cat_manual", label_visibility="collapsed")
                    items_filtrados = df[df["Categorias"] == categoria_seleccionada]
                    item = st.selectbox("Ítem", items_filtrados["Item"], key="item_manual", label_visibility="collapsed")
                    cantidad = st.number_input("Cantidad", min_value=1, value=1, key="cantidad_manual", label_visibility="collapsed")
                    if st.button("Agregar", key="btn_agregar_manual", use_container_width=True):
                        existe = False
                        for producto in st.session_state.carrito:
                            if producto["Item"] == item:
                                producto["Cantidad"] += cantidad
                                producto["Subtotal"] = producto["Cantidad"] * producto["Precio Unitario"]
                                existe = True
                                break
                        if not existe:
                            precio_unitario_original = items_filtrados[items_filtrados["Item"] == item]["P. Unitario real"].values[0]
                            st.session_state.carrito.append({
                                "Categoria": categoria_seleccionada, "Item": item,
                                "Cantidad": cantidad, "Precio Unitario": precio_unitario_original,
                                "Subtotal": precio_unitario_original * cantidad
                            })
                            st.session_state['_toast_msg'] = f"✅ {item} agregado exitosamente ({cantidad} un.)"
                        else:
                            st.session_state['_toast_msg'] = f"✅ {item} actualizado — {cantidad} un. más agregadas"
                        st.rerun()
                except Exception as _e2:
                    st.caption(f"Error: {_e2}")

        with col_m3:
            with st.container(border=True):
                st.markdown("**🗑️ Eliminar Categoría**")
                try:
                    if st.session_state.carrito:
                        carrito_df_temp = pd.DataFrame(st.session_state.carrito)
                        categorias_carrito = carrito_df_temp["Categoria"].unique()
                        categoria_eliminar = st.selectbox("Eliminar", ["-- Seleccionar --"] + list(categorias_carrito), key="cat_eliminar", label_visibility="collapsed")
                        if st.button("Eliminar", key="btn_eliminar_categoria", use_container_width=True):
                            if categoria_eliminar != "-- Seleccionar --":
                                st.session_state.carrito = [i for i in st.session_state.carrito if i["Categoria"] != categoria_eliminar]
                                st.session_state['_toast_msg'] = f"🗑️ Categoría '{categoria_eliminar}' eliminada del presupuesto."
                                st.rerun()
                    else:
                        st.caption("Sin categorías")
                except Exception as _e3:
                    st.caption(f"Error: {_e3}")

        with col_m4:
            with st.container(border=True):
                st.markdown("**➕ Agregar Categoría**")
                try:
                    if hojas_modelo:
                        modelo_origen = st.selectbox("Modelo", hojas_modelo, key="modelo_origen", label_visibility="collapsed")
                        df_temp = _leer_hoja_excel(modelo_origen)
                        categorias_disponibles = df_temp["Categorias"].dropna().unique()
                        categoria_agregar = st.selectbox("Categoría", categorias_disponibles, key="cat_agregar", label_visibility="collapsed")
                        if st.button("Agregar", key="btn_agregar_categoria", use_container_width=True):
                            nuevos_items = cargar_categoria_desde_modelo(modelo_origen, categoria_agregar)
                            st.session_state.carrito = [i for i in st.session_state.carrito if i["Categoria"] != categoria_agregar]
                            st.session_state.carrito.extend(nuevos_items)
                            st.session_state['_toast_msg'] = f"✅ Categoría '{categoria_agregar}' agregada al presupuesto."
                            st.rerun()
                    else:
                        st.caption("Sin modelos")
                except Exception as _e4:
                    st.caption(f"Error: {_e4}")

        with col_m5:
            with st.container(border=True):
                st.markdown("**📎 Plano PDF**")
                st.markdown('''
                <style>
                [data-testid="stFileUploader"] section {
                    border: none !important;
                    padding: 0 !important;
                    background: transparent !important;
                }
                [data-testid="stFileUploadDropzone"] {
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
                    border: none !important;
                    border-radius: 8px !important;
                    padding: 8px 16px !important;
                    min-height: 0 !important;
                }
                [data-testid="stFileUploadDropzone"]:hover {
                    opacity: 0.85 !important;
                    cursor: pointer !important;
                }
                [data-testid="stFileUploadDropzone"] span { display: none !important; }
                [data-testid="stFileUploadDropzone"] button { display: none !important; }
                [data-testid="stFileUploadDropzone"] p {
                    color: white !important;
                    font-weight: 600 !important;
                    font-size: 14px !important;
                    margin: 0 !important;
                }
                [data-testid="stFileUploadDropzone"] p::before { content: "📎 " !important; }
                div[data-testid="stFileUploader"] > label { display:none !important; }
                [data-testid="stFileUploader"] small { display:none !important; }
                </style>
                ''', unsafe_allow_html=True)
                uploaded_file = st.file_uploader("Subir Plano PDF", type=["pdf"], key=f"plano_uploader_{st.session_state.counter}", label_visibility="collapsed")
                if uploaded_file is not None:
                    if uploaded_file.name != st.session_state.plano_nombre:
                        st.session_state.plano_adjunto = uploaded_file.getvalue()
                        st.session_state.plano_nombre = uploaded_file.name
                        st.session_state['_toast_msg'] = f"📎 Plano '{uploaded_file.name}' adjuntado exitosamente."
                    st.success(f"✅ {st.session_state.plano_nombre}")
                elif st.session_state.plano_nombre:
                    st.info(f"📎 {st.session_state.plano_nombre}")
                    if st.button("❌ Quitar plano", key="btn_quitar_plano", use_container_width=True):
                        st.session_state.plano_adjunto = None
                        st.session_state.plano_nombre = ""
                        st.rerun()

    else:
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        for col, label in zip([col_m1, col_m2, col_m3, col_m4], ["MODELO PREDEFINIDO", "ITEMS", "ELIMINAR CATEGORÍA", "AGREGAR CATEGORÍA"]):
            with col:
                st.markdown(f"**{label}**")
                st.info("Modo lectura")

    st.markdown("---")

    if not st.session_state.modo_admin:
        st.markdown("#### Resumen del Presupuesto")
        if st.session_state.margen > 0:
            st.caption(f"ℹ️ Margen del {st.session_state.margen}% aplicado")

    # Input de margen en modo admin
    if st.session_state.modo_admin:
        col_res_tit, col_res_margen = st.columns([4, 1])
        with col_res_tit:
            st.markdown("#### Resumen del Presupuesto")
        with col_res_margen:
            st.caption("Margen %")
            _nuevo_margen = st.number_input(
                "Margen", min_value=0.0, max_value=100.0,
                value=float(st.session_state.margen),
                step=0.5, format="%.1f",
                key="margen_input_fijo", label_visibility="collapsed"
            )
            if _nuevo_margen != st.session_state.margen:
                st.session_state.margen = _nuevo_margen
                st.rerun()


    # Variables de métricas con valores por defecto
    utilidad_real = 0
    total_comisiones = 0
    comision_vendedor = 0
    comision_supervisor = 0
    margen_valor = 0
    subtotal_base = 0
    subtotal_general = 0
    total = 0
    iva = 0

    if st.session_state.carrito:
        # Fila buscador — solo visible con productos
        col_vacio1, col_search_c, col_fs_c, col_vacio2 = st.columns([1, 3, 0.5, 1])
        with col_search_c:
            buscar_tabla = st.text_input("🔍", placeholder="Filtrar por categoría o ítem...", key="buscar_tabla_presupuesto", label_visibility="collapsed")
        with col_fs_c:
            pantalla_completa = st.toggle("⛶", key="tabla_fullscreen", value=st.session_state.get("tabla_fullscreen_val", False), help="Expandir tabla")
            st.session_state.tabla_fullscreen_val = pantalla_completa
        carrito_df = pd.DataFrame(st.session_state.carrito)
        subtotal_base = carrito_df["Subtotal"].sum()

        if st.session_state.modo_admin or st.session_state.margen > 0:
            carrito_df_con_margen = carrito_df.copy()
            carrito_df_con_margen["Precio Unitario"] = carrito_df_con_margen["Precio Unitario"].apply(lambda x: aplicar_margen(x, st.session_state.margen))
            carrito_df_con_margen["Subtotal"] = carrito_df_con_margen["Cantidad"] * carrito_df_con_margen["Precio Unitario"]
            subtotal_general = carrito_df_con_margen["Subtotal"].sum()
        else:
            carrito_df_con_margen = carrito_df.copy()
            subtotal_general = subtotal_base

        iva = subtotal_general * 0.19
        total = subtotal_general + iva
        margen_valor = subtotal_general - subtotal_base
        tiene_margen = st.session_state.margen > 0
        comision_vendedor = subtotal_general * 0.025 if (st.session_state.modo_admin and tiene_margen) else 0
        comision_supervisor = subtotal_general * 0.008 if (st.session_state.modo_admin and tiene_margen) else 0
        total_comisiones = comision_vendedor + comision_supervisor
        utilidad_real = margen_valor - total_comisiones if (st.session_state.modo_admin and tiene_margen) else 0
        altura_tabla = 1400 if pantalla_completa else min(38 * len(carrito_df_con_margen) + 80, 420)

        if es_solo_lectura:
            carrito_df_display = carrito_df_con_margen[["Categoria", "Item", "Cantidad", "Precio Unitario", "Subtotal"]].copy()
            carrito_df_display["Precio Unitario"] = carrito_df_display["Precio Unitario"].apply(formato_clp)
            carrito_df_display["Subtotal"] = carrito_df_display["Subtotal"].apply(formato_clp)
            if buscar_tabla:
                mask = (
                    carrito_df_display["Categoria"].str.contains(buscar_tabla, case=False, na=False) |
                    carrito_df_display["Item"].str.contains(buscar_tabla, case=False, na=False)
                )
                carrito_df_display = carrito_df_display[mask]
            st.dataframe(carrito_df_display, use_container_width=True, hide_index=True, height=altura_tabla,
                column_config={"Categoria": st.column_config.TextColumn("Categoría"), "Item": st.column_config.TextColumn("Item"),
                               "Cantidad": st.column_config.NumberColumn("Cant."), "Precio Unitario": st.column_config.TextColumn("P. Unitario"),
                               "Subtotal": st.column_config.TextColumn("Subtotal")})
            st.caption("🔒 Vista de solo lectura")
        else:
            carrito_df_edit = carrito_df_con_margen.copy()
            carrito_df_edit["✏️"] = False
            carrito_df_edit["Precio Unitario"] = carrito_df_edit["Precio Unitario"].apply(formato_clp)
            carrito_df_edit["Subtotal"] = carrito_df_edit["Subtotal"].apply(formato_clp)
            if buscar_tabla:
                mask = (
                    carrito_df_edit["Categoria"].str.contains(buscar_tabla, case=False, na=False) |
                    carrito_df_edit["Item"].str.contains(buscar_tabla, case=False, na=False)
                )
                carrito_df_edit_filtrado = carrito_df_edit[mask].copy()
            else:
                carrito_df_edit_filtrado = carrito_df_edit
            edited_df = st.data_editor(carrito_df_edit_filtrado, use_container_width=True, hide_index=True, height=altura_tabla,
                key=f"data_editor_{st.session_state.counter}",
                column_config={"✏️": st.column_config.CheckboxColumn("✏️"), "Categoria": st.column_config.TextColumn("Categoría"),
                               "Item": st.column_config.TextColumn("Item"), "Cantidad": st.column_config.NumberColumn("Cant."),
                               "Precio Unitario": st.column_config.TextColumn("P. Unitario"), "Subtotal": st.column_config.TextColumn("Subtotal")})
            filas_editar = edited_df[edited_df["✏️"] == True].index.tolist()
            if st.session_state.get('_item_pendiente_eliminar') and not filas_editar:
                st.session_state.pop('_item_pendiente_eliminar', None)
                st.session_state.counter += 1
                st.rerun()
            if filas_editar:
                if not st.session_state.get('_item_pendiente_eliminar'):
                    _fila_marcada = edited_df[edited_df["✏️"] == True].iloc[0]
                    _nombre_buscar = _fila_marcada["Item"]
                    item_pendiente = next(
                        (item for item in st.session_state.carrito
                         if item["Item"] == _nombre_buscar),
                        None
                    )
                    if item_pendiente:
                        st.session_state['_item_pendiente_eliminar'] = {
                            'item': item_pendiente,
                            'nueva_cantidad': int(item_pendiente.get('Cantidad', 1))
                        }
                        st.rerun()

        # ── Panel inline editar cantidad / eliminar item ──
        if st.session_state.get('_item_pendiente_eliminar'):
            _pend          = st.session_state['_item_pendiente_eliminar']
            _item_data     = _pend['item']
            _nombre_item   = _item_data.get('Item', '')
            _cantidad_orig = int(_item_data.get('Cantidad', 1))
            _precio        = float(_item_data.get('Precio Unitario', 0))
            _categoria     = _item_data.get('Categoria', '')
            _nueva_cant    = int(_pend.get('nueva_cantidad', _cantidad_orig))
            _container_key = f"popup_container_{st.session_state.counter}"
            _css_key       = _container_key.replace('-', '_')

            st.markdown(f'''
            <style>
            .st-key-{_css_key} > div[data-testid="stVerticalBlockBorderWrapper"] {{
                background: #FCEBEB !important;
                border: 1.5px solid #E24B4A !important;
                border-radius: 14px !important;
                box-shadow: none !important;
            }}
            .st-key-{_css_key} label {{
                color: #791F1F !important; font-weight: 600 !important;
            }}
            .st-key-{_css_key} input[type="number"] {{
                background: #fff !important; border-color: #E24B4A !important;
                color: #501313 !important; font-weight: 700 !important;
            }}
            .st-key-{_css_key} button[data-testid="stNumberInputStepUp"],
            .st-key-{_css_key} button[data-testid="stNumberInputStepDown"] {{
                background: #FCEBEB !important; color: #A32D2D !important;
            }}
            .st-key-popup_cancelar_btn button {{
                background: transparent !important; border: 1px solid #F09595 !important;
                color: #791F1F !important;
            }}
            .st-key-popup_aplicar_btn button {{
                background: #fff !important; border: 1.5px solid #E24B4A !important;
                color: #A32D2D !important; font-weight: 600 !important;
            }}
            .st-key-popup_eliminar_btn button {{
                background: #E24B4A !important; border: none !important;
                color: #fff !important; font-weight: 600 !important;
            }}
            </style>
            ''', unsafe_allow_html=True)

            with st.container(border=True, key=_container_key):
                st.markdown(f'''
                <div style="margin-bottom:12px;">
                    <div style="font-size:11px;color:#A32D2D;font-weight:600;
                                text-transform:uppercase;letter-spacing:.08em;margin-bottom:3px;">{_categoria}</div>
                    <div style="font-size:17px;font-weight:700;color:#501313;margin-bottom:14px;">{_nombre_item}</div>
                    <div style="display:flex;gap:12px;margin-bottom:4px;">
                        <div style="background:#fff;border:.5px solid #F09595;border-radius:10px;padding:10px 14px;text-align:center;flex:1;">
                            <div style="font-size:11px;color:#A32D2D;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">P. unitario</div>
                            <div style="font-size:15px;font-weight:700;color:#501313;margin-top:3px;">{formato_clp(_precio)}</div>
                        </div>
                        <div style="background:#fff;border:.5px solid #F09595;border-radius:10px;padding:10px 14px;text-align:center;flex:1;">
                            <div style="font-size:11px;color:#A32D2D;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">Cant. original</div>
                            <div style="font-size:15px;font-weight:700;color:#791F1F;margin-top:3px;">{_cantidad_orig}</div>
                        </div>
                        <div style="background:#fff;border:.5px solid #E24B4A;border-radius:10px;padding:10px 14px;text-align:center;flex:1;">
                            <div style="font-size:11px;color:#A32D2D;font-weight:600;text-transform:uppercase;letter-spacing:.06em;">Subtotal nuevo</div>
                            <div style="font-size:15px;font-weight:700;color:#E24B4A;margin-top:3px;">{formato_clp(_nueva_cant * _precio)}</div>
                        </div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)

                _cant_input = st.number_input(
                    "Nueva cantidad",
                    min_value=1,
                    value=_nueva_cant,
                    step=1,
                    key=f"ni_{st.session_state.counter}"
                )
                if int(_cant_input) != _nueva_cant and not st.session_state.get('_rerun_lock'):
                    st.session_state['_item_pendiente_eliminar']['nueva_cantidad'] = int(_cant_input)
                    st.rerun()

                _ba1, _ba2, _ba3 = st.columns([1, 1.5, 1.5])
                with _ba1:
                    if st.button("✖️ Cancelar", use_container_width=True, key="popup_cancelar_btn"):
                        st.session_state.pop('_item_pendiente_eliminar', None)
                        st.session_state.pop('_rerun_lock', None)
                        st.session_state.counter += 1
                        st.rerun()
                with _ba2:
                    if st.button("✅ Aplicar cambio", use_container_width=True, key="popup_aplicar_btn"):
                        for item in st.session_state.carrito:
                            if item['Item'] == _nombre_item:
                                item['Cantidad'] = int(_cant_input)
                                item['Subtotal'] = int(_cant_input) * float(item['Precio Unitario'])
                                break
                        st.session_state.pop('_item_pendiente_eliminar', None)
                        st.session_state.pop('_rerun_lock', None)
                        st.session_state.counter += 1
                        st.rerun()
                with _ba3:
                    if st.button("🗑️ Eliminar todo", use_container_width=True, key="popup_eliminar_btn"):
                        st.session_state.carrito = [
                            i for i in st.session_state.carrito
                            if i['Item'] != _nombre_item
                        ]
                        st.session_state.pop('_item_pendiente_eliminar', None)
                        st.session_state.pop('_rerun_lock', None)
                        st.session_state.counter += 1
                        st.rerun()
        st.markdown("---")
        # Solo botón Limpiar
        col_btn_limpiar, _, _, _ = st.columns(4)
        with col_btn_limpiar:
            if not es_solo_lectura:
                if st.button("🧹 Limpiar", use_container_width=True):
                    st.session_state.pop('_item_pendiente_eliminar', None)
                    st.session_state.pop('_rerun_lock', None)
                    limpiar_todo()
                    st.rerun()
            else:
                st.button("🧹 Limpiar", use_container_width=True, disabled=True)

        correo_para_pdf = st.session_state.correo_input
        datos_cliente_pdf = {
            "Nombre": st.session_state.nombre_input,
            "RUT": st.session_state.rut_display or '',
            "Correo": st.session_state.correo_input,
            "Teléfono": st.session_state.telefono_raw or '',
            "Dirección": st.session_state.direccion_input,
            "ComunaCliente": st.session_state.cliente_comuna or "",
            "RegionCliente": st.session_state.cliente_region or "",
            "DireccionProyecto": st.session_state.proyecto_direccion or "",
            "ComunaProyecto": st.session_state.proyecto_comuna or "",
            "RegionProyecto": st.session_state.proyecto_region or "",
            "TipoCliente": st.session_state.cliente_tipo or "natural",
            "EmpresaCliente": st.session_state.cliente_empresa or "",
            "RutEmpresa": st.session_state.cliente_rut_empresa or "",
            "Observaciones": st.session_state.observaciones_input
        }
        nombre_asesor_final = st.session_state.asesor_seleccionado if st.session_state.asesor_seleccionado != "Seleccionar asesor" else ""
        datos_asesor_pdf = {
            "Nombre Ejecutivo": nombre_asesor_final,
            "Correo Ejecutivo": st.session_state.correo_asesor or "",
            "Teléfono Ejecutivo": st.session_state.telefono_asesor or ""
        }
        carrito_df_pdf = carrito_df_con_margen.copy()
        margen_actual = st.session_state.margen
        numero_para_pdf = st.session_state.cotizacion_cargada if st.session_state.cotizacion_cargada else None

        if st.session_state.modo_admin and st.session_state.margen > 0:
            st.caption(f"*Precios calculados con margen del {st.session_state.margen}%")

        # Asegurar que todas las variables de métricas estén definidas
        if 'utilidad_real' not in dir():
            utilidad_real = margen_valor - total_comisiones if st.session_state.modo_admin else 0

        st.markdown("---")
        st.markdown("#### Métricas")
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)

        total_productos = sum(item["Cantidad"] for item in st.session_state.carrito)
        categorias_unicas = len(set(item["Categoria"] for item in st.session_state.carrito))

        with col_m1:
            st.markdown(f'<div class="stats-card"><div class="stats-title">ÍTEMS</div><div class="stats-number" style="color:#3b82f6;border:none;padding:0;">{len(st.session_state.carrito)}</div><div class="stats-desc">En presupuesto</div></div>', unsafe_allow_html=True)
        with col_m2:
            st.markdown(f'<div class="stats-card"><div class="stats-title">PRODUCTOS</div><div class="stats-number" style="color:#f59e0b;border:none;padding:0;">{total_productos}</div><div class="stats-desc">Unidades</div></div>', unsafe_allow_html=True)
        with col_m3:
            st.markdown(f'<div class="stats-card"><div class="stats-title">CATEGORÍAS</div><div class="stats-number" style="color:#10b981;border:none;padding:0;">{categorias_unicas}</div><div class="stats-desc">Diferentes</div></div>', unsafe_allow_html=True)
        with col_m4:
            if st.session_state.modo_admin:
                st.markdown(f'''
                <div class="metric-card-special" style="background:linear-gradient(135deg,#ef4444,#dc2626);padding:1.5rem;display:flex;flex-direction:column;justify-content:space-between;">
                    <div style="color:rgba(255,255,255,0.85);font-size:0.9rem;">
                        <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>Costo base:</span><span>{formato_clp(subtotal_base)}</span></div>
                        <div style="display:flex;justify-content:space-between;"><span>+ Margen {st.session_state.margen}%:</span><span>{formato_clp(margen_valor)}</span></div>
                    </div>
                    <div style="border-top:2px solid rgba(255,255,255,0.5);margin-top:1rem;padding-top:0.6rem;display:flex;justify-content:space-between;align-items:center;">
                        <span style="font-size:1.4rem;font-weight:700;color:white;">📦 Total sin iva</span>
                        <span style="font-size:2.2rem;font-weight:700;color:white;">{formato_clp(subtotal_general)}</span>
                    </div>
                </div>''', unsafe_allow_html=True)
            else:
                st.markdown(f'''
                <div class="metric-card-special" style="background:linear-gradient(135deg,#ef4444,#dc2626);padding:1.5rem;display:flex;flex-direction:column;justify-content:space-between;">
                    <div style="color:rgba(255,255,255,0.85);font-size:0.9rem;">
                        <div style="display:flex;justify-content:space-between;"><span>Costo base (sin IVA):</span><span>{formato_clp(subtotal_base)}</span></div>
                    </div>
                    <div style="border-top:2px solid rgba(255,255,255,0.5);margin-top:1rem;padding-top:0.6rem;display:flex;justify-content:space-between;align-items:center;">
                        <span style="font-size:1.4rem;font-weight:700;color:white;">📦 Total sin iva</span>
                        <span style="font-size:2.2rem;font-weight:700;color:white;">{formato_clp(subtotal_base)}</span>
                    </div>
                </div>''', unsafe_allow_html=True)

        st.markdown("---")

        if st.session_state.modo_admin:
            col_total_card, col_comisiones_card, col_utilidad_card = st.columns(3)
            with col_total_card:
                    st.markdown(f'''
                    <div class="metric-card-special metric-card-total" style="padding:1.5rem;display:flex;flex-direction:column;justify-content:space-between;">
                        <div style="color:rgba(255,255,255,0.85);font-size:0.9rem;">
                            <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>Costo base:</span><span>{formato_clp(subtotal_base)}</span></div>
                            <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>+ Margen {st.session_state.margen}%:</span><span>{formato_clp(margen_valor)}</span></div>
                            <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>= Subtotal c/margen:</span><span>{formato_clp(subtotal_general)}</span></div>
                            <div style="display:flex;justify-content:space-between;"><span>+ IVA 19%:</span><span>{formato_clp(iva)}</span></div>
                        </div>
                        <div style="border-top:2px solid rgba(255,255,255,0.5);margin-top:1rem;padding-top:0.6rem;display:flex;justify-content:space-between;align-items:center;">
                            <span style="font-size:1.4rem;font-weight:700;color:white;">💰 Total con iva</span>
                            <span style="font-size:2.2rem;font-weight:700;color:white;">{formato_clp(total)}</span>
                          </div>''', unsafe_allow_html=True)
            with col_comisiones_card:
                    st.markdown(f'''
                    <div class="metric-card-special metric-card-comisiones" style="padding:1.5rem;display:flex;flex-direction:column;justify-content:space-between;">
                        <div style="color:rgba(255,255,255,0.85);font-size:0.9rem;">
                            <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>Vendedor 2.5%:</span><span>{formato_clp(comision_vendedor)}</span></div>
                            <div style="display:flex;justify-content:space-between;"><span>Supervisor 0.8%:</span><span>{formato_clp(comision_supervisor)}</span></div>
                        </div>
                        <div style="border-top:2px solid rgba(255,255,255,0.5);margin-top:1rem;padding-top:0.6rem;display:flex;justify-content:space-between;align-items:center;">
                            <span style="font-size:1.4rem;font-weight:700;color:white;">📊 Comisiones</span>
                            <span style="font-size:2.2rem;font-weight:700;color:white;">{formato_clp(total_comisiones)}</span>
                          </div>''', unsafe_allow_html=True)
            with col_utilidad_card:
                    st.markdown(f'''
                    <div class="metric-card-special metric-card-utilidad" style="padding:1.5rem;display:flex;flex-direction:column;justify-content:space-between;">
                        <div style="color:rgba(255,255,255,0.85);font-size:0.9rem;">
                            <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>Margen bruto:</span><span>{formato_clp(margen_valor)}</span></div>
                            <div style="display:flex;justify-content:space-between;"><span>- Comisiones:</span><span>{formato_clp(total_comisiones)}</span></div>
                        </div>
                        <div style="border-top:2px solid rgba(255,255,255,0.5);margin-top:1rem;padding-top:0.6rem;display:flex;justify-content:space-between;align-items:center;">
                            <span style="font-size:1.4rem;font-weight:700;color:white;">📈 Utilidad real</span>
                            <span style="font-size:2.2rem;font-weight:700;color:white;">{formato_clp(utilidad_real)}</span>
                          </div>''', unsafe_allow_html=True)

        else:
            col_t1, col_t2, col_t3 = st.columns([1, 2, 1])
            with col_t2:
                st.markdown(f'''
                <div class="metric-card-special metric-card-total" style="padding:1.5rem;display:flex;flex-direction:column;justify-content:space-between;">
                    <div style="color:rgba(255,255,255,0.85);font-size:0.9rem;">
                        <div style="display:flex;justify-content:space-between;margin-bottom:0.3rem;"><span>Costo base:</span><span>{formato_clp(subtotal_base)}</span></div>
                        <div style="display:flex;justify-content:space-between;"><span>+ IVA 19%:</span><span>{formato_clp(iva)}</span></div>
                    </div>
                    <div style="border-top:2px solid rgba(255,255,255,0.5);margin-top:1rem;padding-top:0.6rem;display:flex;justify-content:space-between;align-items:center;">
                        <span style="font-size:1.4rem;font-weight:700;color:white;">💰 Total con iva</span>
                        <span style="font-size:2.2rem;font-weight:700;color:white;">{formato_clp(total)}</span>
                    </div>
                </div>''', unsafe_allow_html=True)
            if st.session_state.margen > 0:
                st.info("🔒 Los detalles de comisiones y utilidad solo están disponibles para administradores.")
    else:
        st.info("👈 Agrega productos al presupuesto usando los controles de la izquierda")

# =========================================================
# TAB 3 - GESTIÓN DE COTIZACIONES GUARDADAS
# =========================================================
with tab3:
    st.markdown("""
    <style>
    .hdr3 {
        background: linear-gradient(135deg, #6b4e00 0%, #e65100 100%);
        border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
        display: flex; align-items: center; gap: 22px;
        box-shadow: 0 8px 32px rgba(230,81,0,0.25);
        position: relative; overflow: hidden;
    }
    .hdr3::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 180px; height: 180px; border-radius: 50%;
        background: rgba(255,255,255,0.04); pointer-events: none;
    }
    .hdr3::after {
        content: ''; position: absolute; bottom: -60px; right: 80px;
        width: 240px; height: 240px; border-radius: 50%;
        background: rgba(255,255,255,0.03); pointer-events: none;
    }
    .hdr3 h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                 font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
    .hdr3 p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
    </style>
    <div class="hdr3">
      <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">📂</span>
      <div>
        <h2>Gestión de Cotizaciones</h2>
        <p>Busca, carga y administra todas las cotizaciones del sistema.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_busqueda, col_filtros = st.columns([3, 1])
    with col_busqueda:
        with st.container(border=True):
            tipo_busqueda = st.radio("Buscar por:", ["📋 N° Presupuesto", "👤 Cliente", "👨‍💼 Asesor"], horizontal=True, key="tipo_busqueda")
            tipo_map = {"📋 N° Presupuesto": "numero", "👤 Cliente": "cliente", "👨‍💼 Asesor": "asesor"}
            termino = st.text_input("Buscar...", placeholder="Ingrese término de búsqueda...", key="buscar_cotizacion")

    with col_filtros:
        with st.container(border=True):
            st.markdown("**📅 Filtros rápidos**")
            st.button("📅 Hoy", use_container_width=True)
            st.button("📅 Esta semana", use_container_width=True)
            st.button("📅 Este mes", use_container_width=True)

    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 4])
    with col_btn1:
        buscar_btn = st.button("🔍 Buscar", type="primary", use_container_width=True)
    with col_btn2:
        limpiar_btn = st.button("🗑️ Limpiar", use_container_width=True)

    st.markdown("---")
    st.markdown("### Resultados")

    # Forzar refresco si se acaba de guardar una cotización
    if st.session_state.get('_tab3_necesita_refresh', False):
        st.session_state.resultados_busqueda = None
        st.session_state['_tab3_necesita_refresh'] = False

    if 'resultados_busqueda' not in st.session_state or st.session_state.resultados_busqueda is None:
        st.session_state.resultados_busqueda = buscar_cotizaciones()

    if buscar_btn or (termino and termino != st.session_state.get('ultimo_termino', '')):
        st.session_state.ultimo_termino = termino
        st.session_state.resultados_busqueda = buscar_cotizaciones(termino if termino else None, tipo_map[tipo_busqueda])
        st.session_state.mostrar_visor = False
        st.session_state.pdf_actual = None
        st.session_state.pdf_nombre = ""
        st.session_state.numero_en_visor = None
        st.session_state.pdf_url = None

    if limpiar_btn:
        st.session_state.resultados_busqueda = []
        st.session_state.ultimo_termino = ""
        st.session_state.mostrar_visor = False
        st.session_state.pdf_actual = None
        st.session_state.pdf_nombre = ""
        st.session_state.numero_en_visor = None
        st.session_state.pdf_url = None
        st.rerun()

    if st.session_state.resultados_busqueda:
        _cols_esperadas = ["N°", "Cliente", "Asesor", "Fecha", "Total", "Margen", "RUT", "Email", "Asesor_Email", "Asesor_Tel", "Tiene_Plano", "Tiene_Contrato", "Empresa", "NLogs"]
        _rows_norm = []
        for _r in st.session_state.resultados_busqueda:
            _r = list(_r)
            while len(_r) < len(_cols_esperadas):
                _r.append(0)
            _rows_norm.append(_r[:len(_cols_esperadas)])
        df_resultados = pd.DataFrame(_rows_norm, columns=_cols_esperadas)
        df_resultados["Total"] = df_resultados["Total"].apply(lambda x: f"${x:,.0f}".replace(",", ".") if x else "$0")
        def _fmt_fecha(x):
            """Para la tabla HTML: fecha en negrita + hora en gris — zona horaria Chile."""
            if not x: return ""
            try:
                from datetime import datetime as _dt, timezone, timedelta
                _tz_cl = timezone(timedelta(hours=-3))
                _d = _dt.fromisoformat(x.replace("Z","+00:00")).astimezone(_tz_cl)
                return f'<span style="font-weight:700;">{_d.strftime("%d/%m/%Y")}</span><br><span style="font-size:0.75em;color:#64748b;">{_d.strftime("%H:%M")}</span>'
            except: return x[:10]
        def _fmt_fecha_plana(x):
            """Para el selectbox: texto limpio sin HTML — zona horaria Chile."""
            if not x: return ""
            try:
                from datetime import datetime as _dt, timezone, timedelta
                _tz_cl = timezone(timedelta(hours=-3))
                _d = _dt.fromisoformat(x.replace("Z","+00:00")).astimezone(_tz_cl)
                return _d.strftime("%d/%m/%Y %H:%M")
            except: return x[:10]
        df_resultados["FechaPlana"] = df_resultados["Fecha"].apply(_fmt_fecha_plana)
        df_resultados["Fecha"] = df_resultados["Fecha"].apply(_fmt_fecha)
        df_resultados["Estado"] = df_resultados.apply(crear_badge_estado, axis=1)
        df_resultados["Plano"]    = df_resultados.apply(lambda row: "✅ Sí" if row["Tiene_Plano"] else "—", axis=1)
        df_resultados["MargenCol"]= df_resultados["Margen"].apply(lambda x: "✅ Sí" if x and x > 0 else "—")
        df_resultados["ContratoCol"] = df_resultados["Tiene_Contrato"].apply(lambda x: "✅ Sí" if x else "—")
        df_resultados["EmpresaCol"] = df_resultados["Empresa"].apply(lambda x: "✅ Sí" if x and x.strip() else "—")
        df_resultados["ModCol"] = df_resultados["NLogs"].apply(
            lambda x: f'<span style="font-weight:700;color:#3b82f6;">{x}</span>' if x > 0 else '<span style="color:#94a3b8;">0</span>')

        n_resultados = len(df_resultados)
        altura_tabla = min(n_resultados * 48 + 50, 530)  # ~10 filas = 530px máx

        rows_html = ""
        for _, row in df_resultados.iterrows():
            _mg_color  = 'color:#16a34a;font-weight:700;' if row['MargenCol']  == '✅ Sí' else 'color:#94a3b8;'
            _ct_color  = 'color:#16a34a;font-weight:700;' if row['ContratoCol'] == '✅ Sí' else 'color:#94a3b8;'
            _emp_color = 'color:#16a34a;font-weight:700;' if row['EmpresaCol']  == '✅ Sí' else 'color:#94a3b8;'
            _pln_color = 'color:#16a34a;font-weight:700;' if row['Plano']       == '✅ Sí' else 'color:#94a3b8;'
            rows_html += f"<tr><td data-ep=\"{row['N°']}\" style=\"cursor:pointer;font-weight:700;color:#3b82f6;\" title=\"Click para copiar {row['N°']}\">{row['N°']} 📋</td><td>{row['Cliente'] or '—'}</td><td>{row['Asesor'] or '—'}</td><td style='line-height:1.6;'>{row['Fecha']}</td><td>{row['Total']}</td><td style='text-align:center;'>{row['Estado']}</td><td style='text-align:center;{_emp_color}'>{row['EmpresaCol']}</td><td style='text-align:center;{_mg_color}'>{row['MargenCol']}</td><td style='text-align:center;{_ct_color}'>{row['ContratoCol']}</td><td style='text-align:center;{_pln_color}'>{row['Plano']}</td><td style='text-align:center;'>{row['ModCol']}</td></tr>"

        html_table = f"""
        <div style="border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.08);border:1px solid #e2e8f0;">
            <div style="overflow-y:auto;max-height:{altura_tabla}px;">
                <table class='resultados-table' style='margin:0;border-radius:0;box-shadow:none;'>
                    <thead style='position:sticky;top:0;z-index:2;'>
                        <tr><th>N° Presupuesto</th><th>Cliente</th><th>Asesor</th><th>Fecha de creación</th><th>Total</th><th>Estado</th><th>Empresa</th><th>Margen</th><th>Contrato</th><th>Plano</th><th>Modif.</th></tr>
                    </thead>
                    <tbody>{rows_html}</tbody>
                </table>
            </div>
        </div>
        <p style="font-size:0.8rem;color:#888;margin-top:6px;">Mostrando {n_resultados} resultado{'s' if n_resultados != 1 else ''}</p>
        """
        st.markdown(html_table, unsafe_allow_html=True)

        # JS para copiar EP al hacer click en la celda
        import streamlit.components.v1 as _ep_copy_comp
        _ep_copy_comp.html("""<script>
(function(){
    var D = window.parent.document;
    function initEPCopy(){
        D.addEventListener('click', function(e){
            var td = e.target && e.target.closest ? e.target.closest('td[data-ep]') : null;
            if (!td) return;
            var ep = td.getAttribute('data-ep');
            if (!ep) return;
            // Copiar al portapapeles
            var ta = D.createElement('textarea');
            ta.value = ep;
            ta.style.cssText = 'position:fixed;top:-9999px;left:-9999px;';
            D.body.appendChild(ta); ta.focus(); ta.select();
            try { D.execCommand('copy'); } catch(err) {}
            ta.remove();
            if (window.parent.navigator.clipboard) {
                window.parent.navigator.clipboard.writeText(ep).catch(function(){});
            }
            // Feedback visual
            var orig = td.innerHTML;
            var origColor = td.style.color;
            td.innerHTML = '✅ ¡Copiado!';
            td.style.color = '#10b981';
            setTimeout(function(){ td.innerHTML = orig; td.style.color = origColor; }, 1200);
        });
    }
    setTimeout(initEPCopy, 500);
})();
</script>""", height=0)

        st.markdown("### Seleccionar cotización")

        opciones = []
        for idx, row in df_resultados.iterrows():
            # Usar nombres de columna correctos (no índices numéricos)
            datos_completos = all([row['Cliente'], row['Email']])
            asesor_completo = any([row['Asesor'], row['Asesor_Email'], row['Asesor_Tel']])
            if row['Margen'] and row['Margen'] > 0:
                estado = ("🟢 AUTORIZADO CON PLANO" if row['Tiene_Plano'] else "🟢 AUTORIZADO") if (datos_completos and asesor_completo) else ("🔴 INCOMPLETO CON PLANO" if row['Tiene_Plano'] else "🔴 INCOMPLETO")
            else:
                if datos_completos and asesor_completo:
                    estado = "🟠 BORRADOR CON PLANO" if row['Tiene_Plano'] else "🟡 BORRADOR"
                else:
                    estado = "🔴 INCOMPLETO CON PLANO" if row['Tiene_Plano'] else "🔴 INCOMPLETO"
            plano_indicador = "📎" if row['Tiene_Plano'] else "❌"
            opciones.append(f"{row['N°']} - {row['Cliente'] or 'S/C'} ({row['FechaPlana']}) - {row['Total']} - {estado} {plano_indicador}")

        if opciones:
            cotizacion_seleccionada = st.selectbox("Selecciona una cotización:", options=opciones, key="selector_cotizaciones")

            if cotizacion_seleccionada:
                numero_seleccionado = cotizacion_seleccionada.split(" - ")[0]

                tiene_margen_seleccionado = False
                tiene_plano_seleccionado = False
                for row in st.session_state.resultados_busqueda:
                    if row[0] == numero_seleccionado:
                        tiene_margen_seleccionado = bool(row[5] and row[5] > 0)
                        tiene_plano_seleccionado = bool(row[10]) if len(row) > 10 else False
                        break

                if numero_seleccionado != st.session_state.numero_en_visor:
                    if tiene_plano_seleccionado and st.session_state.mostrar_visor:
                        cot_visor = cargar_cotizacion(numero_seleccionado)
                        if cot_visor and cot_visor.get('plano_url'):
                            st.session_state.pdf_url = cot_visor['plano_url']
                            st.session_state.pdf_nombre = cot_visor.get('plano_nombre', 'plano.pdf')
                            st.session_state.numero_en_visor = numero_seleccionado
                            st.rerun()
                        else:
                            st.session_state.mostrar_visor = False
                            st.session_state.pdf_actual = None
                            st.session_state.pdf_nombre = ""
                            st.session_state.numero_en_visor = None
                            st.session_state.pdf_url = None
                            st.rerun()
                    else:
                        if st.session_state.mostrar_visor:
                            st.session_state.mostrar_visor = False
                            st.session_state.pdf_actual = None
                            st.session_state.pdf_nombre = ""
                            st.session_state.numero_en_visor = None
                            st.session_state.pdf_url = None
                            st.rerun()

                if tiene_margen_seleccionado and not st.session_state.modo_admin:
                    st.warning("🔒 Cotización autorizada - Solo puedes generar PDFs")

                # ── Botón descargar log de modificaciones (PDF) ──
                _logs_ep = obtener_logs_ep(numero_seleccionado)
                if _logs_ep:
                    _n_mods = len([l for l in _logs_ep if l.get("tipo_cambio") == "modificacion"])
                    try:
                        _pdf_log_bytes = generar_pdf_log(numero_seleccionado, _logs_ep)
                        st.download_button(
                            label=f"📋 Descargar historial PDF ({len(_logs_ep)} registros · {_n_mods} modif.)",
                            data=_pdf_log_bytes,
                            file_name=f"historial_{numero_seleccionado}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key="btn_download_log"
                        )
                    except Exception as _e_log:
                        st.error(f"Error generando PDF log: {_e_log}")
                else:
                    st.caption("📋 Sin registros de modificaciones aún")

            st.markdown("---")
            st.markdown("### Acciones")
            col_acc1, col_acc2, col_acc3, col_acc4 = st.columns(4)

            with col_acc1:
                if tiene_margen_seleccionado and not st.session_state.modo_admin:
                    st.button("📂 Cargar", use_container_width=True, disabled=True)
                else:
                    if st.button("📂 Cargar", use_container_width=True):
                        # Si hay carrito sin guardar, mostrar advertencia
                        tiene_sin_guardar = (
                            len(st.session_state.carrito) > 0 and
                            st.session_state.cotizacion_cargada != numero_seleccionado
                        )
                        if tiene_sin_guardar:
                            st.session_state.mostrar_advertencia_carga = True
                            st.session_state.numero_a_cargar_pendiente = numero_seleccionado
                            st.rerun()
                        else:
                            if preparar_carga_cotizacion(numero_seleccionado):
                                st.success(f"✅ Cotización {numero_seleccionado} cargada")
                                st.rerun()

            # ── Popup advertencia productos sin guardar ──
            if st.session_state.get('mostrar_advertencia_carga', False):
                @st.dialog("⚠️ Productos sin guardar")
                def dialogo_advertencia():
                    numero_pendiente = st.session_state.get('numero_a_cargar_pendiente', '')
                    st.markdown(f"""
                    <div style="text-align:center;padding:1rem 0;">
                        <div style="font-size:3rem;margin-bottom:0.5rem;">⚠️</div>
                        <div style="font-size:1rem;font-weight:700;color:#1e2447;margin-bottom:0.5rem;">
                            Tienes productos sin guardar
                        </div>
                        <div style="font-size:0.88rem;color:#5a6080;line-height:1.6;">
                            Estás a punto de cargar la cotización <strong>{numero_pendiente}</strong>.<br/>
                            ¿Deseas guardar el presupuesto actual antes de continuar?
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    col_si, col_no, col_cancelar = st.columns(3)
                    with col_si:
                        if st.button("💾 Sí, guardar", use_container_width=True, type="primary", key="dialog_btn_si"):
                            # Guardar primero
                            datos_cliente_g, datos_asesor_g, proyecto_g, config_g, totales_g, plano_n, plano_d = construir_datos_para_guardar()
                            if st.session_state.cotizacion_cargada:
                                num_g = st.session_state.cotizacion_cargada
                            else:
                                num_g = generar_numero_unico()
                            _usr_log3 = st.session_state.get('auth_nombre','') or st.session_state.get('auth_email','')
                            guardar_cotizacion(num_g, datos_cliente_g, datos_asesor_g,
                                               proyecto_g, st.session_state.carrito,
                                               config_g, totales_g, plano_n, plano_d,
                                               usuario_logueado=_usr_log3)
                            # Luego cargar
                            st.session_state.mostrar_advertencia_carga = False
                            if preparar_carga_cotizacion(numero_pendiente):
                                st.rerun()
                    with col_no:
                        if st.button("🗑️ No, descartar", use_container_width=True, key="dialog_btn_no"):
                            # Descartar y cargar directamente
                            st.session_state.mostrar_advertencia_carga = False
                            if preparar_carga_cotizacion(numero_pendiente):
                                st.rerun()
                    with col_cancelar:
                        if st.button("✖️ Cancelar", use_container_width=True, key="dialog_btn_cancelar"):
                            st.session_state.mostrar_advertencia_carga = False
                            st.session_state.numero_a_cargar_pendiente = None
                            st.rerun()

                dialogo_advertencia()

            cotizacion_para_pdf = cargar_cotizacion(numero_seleccionado) if cotizacion_seleccionada else None

            def preparar_pdf_data(cotizacion):
                carrito_df_t = pd.DataFrame(cotizacion['productos'])
                margen_c = cotizacion.get('config_margen', 0)
                if margen_c > 0:
                    carrito_df_p = carrito_df_t.copy()
                    carrito_df_p["Precio Unitario"] = carrito_df_p["Precio Unitario"].apply(lambda x: aplicar_margen(x, margen_c))
                    carrito_df_p["Subtotal"] = carrito_df_p["Cantidad"] * carrito_df_p["Precio Unitario"]
                else:
                    carrito_df_p = carrito_df_t.copy()
                subtotal_p = carrito_df_p["Subtotal"].sum()
                iva_p = subtotal_p * 0.19
                total_p = subtotal_p + iva_p
                dc = {
                    "Nombre":            cotizacion.get('cliente_nombre',''),
                    "RUT":               cotizacion.get('cliente_rut',''),
                    "Correo":            cotizacion.get('cliente_email',''),
                    "Teléfono":          cotizacion.get('cliente_telefono',''),
                    "Dirección":         cotizacion.get('cliente_direccion',''),
                    "ComunaCliente":     cotizacion.get('cliente_comuna',''),
                    "RegionCliente":     cotizacion.get('cliente_region',''),
                    "DireccionProyecto": cotizacion.get('proyecto_direccion',''),
                    "ComunaProyecto":    cotizacion.get('proyecto_comuna',''),
                    "RegionProyecto":    cotizacion.get('proyecto_region',''),
                    "TipoCliente":       cotizacion.get('cliente_tipo','natural'),
                    "EmpresaCliente":    cotizacion.get('cliente_empresa',''),
                    "RutEmpresa":        cotizacion.get('cliente_rut_empresa',''),
                    "Observaciones":     cotizacion.get('proyecto_observaciones',''),
                }
                da = {"Nombre Ejecutivo": cotizacion.get('asesor_nombre',''),
                      "Correo Ejecutivo": cotizacion.get('asesor_email',''),
                      "Teléfono Ejecutivo": cotizacion.get('asesor_telefono','')}
                fi = datetime.strptime(cotizacion.get('proyecto_fecha_inicio', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
                ft = datetime.strptime(cotizacion.get('proyecto_fecha_termino', (datetime.now()+timedelta(days=15)).strftime('%Y-%m-%d')), '%Y-%m-%d').date()
                dv = cotizacion.get('proyecto_dias_validez', 15)
                return carrito_df_p, subtotal_p, iva_p, total_p, dc, da, fi, ft, dv, margen_c

            with col_acc2:
                if cotizacion_para_pdf:
                    carrito_df_p, subtotal_p, iva_p, total_p, dc, da, fi, ft, dv, margen_c = preparar_pdf_data(cotizacion_para_pdf)
                    pdf_buffer, _ = generar_pdf_completo(carrito_df_p, subtotal_p, iva_p, total_p, dc, fi, ft, dv, da, margen=margen_c, numero_cotizacion=numero_seleccionado)
                    st.download_button(label="📄 PDF Completo", data=pdf_buffer, file_name=f"Presupuesto_Completo_{numero_seleccionado}.pdf",
                        mime="application/pdf", use_container_width=True, key=f"pdf_completo_{numero_seleccionado}")
                else:
                    st.button("📄 PDF Completo", use_container_width=True, disabled=True)

            with col_acc3:
                if cotizacion_para_pdf:
                    carrito_df_p, subtotal_p, iva_p, total_p, dc, da, fi, ft, dv, margen_c = preparar_pdf_data(cotizacion_para_pdf)
                    _desc_ep = cargar_descripciones_por_ep(numero_seleccionado, bust_cache=True)
                    pdf_buffer, _ = generar_pdf_cliente(carrito_df_p, subtotal_p, iva_p, total_p, dc, fi, ft, dv, da, margen=margen_c, numero_cotizacion=numero_seleccionado, descripciones_ep=_desc_ep)
                    st.download_button(label="🔒 PDF Cliente", data=pdf_buffer, file_name=f"Presupuesto_Cliente_{numero_seleccionado}.pdf",
                        mime="application/pdf", use_container_width=True, key=f"pdf_cliente_{numero_seleccionado}")
                else:
                    st.button("🔒 PDF Cliente", use_container_width=True, disabled=True)

            with col_acc4:
                if cotizacion_seleccionada and tiene_plano_seleccionado:
                    label_visor = "🔄 ACTUALIZAR PLANO" if (st.session_state.mostrar_visor and st.session_state.numero_en_visor == numero_seleccionado) else "👁️ VER PLANO"
                    if st.button(label_visor, use_container_width=True, type="primary"):
                        cot_btn = cargar_cotizacion(numero_seleccionado)
                        if cot_btn and cot_btn.get('plano_url'):
                            st.session_state.pdf_url = cot_btn['plano_url']
                            st.session_state.pdf_nombre = cot_btn.get('plano_nombre', 'plano.pdf')
                            st.session_state.mostrar_visor = True
                            st.session_state.numero_en_visor = numero_seleccionado
                            st.rerun()
                else:
                    st.button("👁️ VER PLANO", use_container_width=True, disabled=True)

            # =========================================================
            # VISOR DE PDF
            # =========================================================
            if st.session_state.mostrar_visor and st.session_state.pdf_url:
                with st.expander("📄 Vista Previa del Plano", expanded=True):
                    st.markdown(f"**Archivo:** {st.session_state.pdf_nombre} — cotización `{st.session_state.numero_en_visor}`")
                    navegador = detectar_navegador()
                    pdf_url_visor = st.session_state.pdf_url
                    pdf_url_encoded = urllib.parse.quote(pdf_url_visor, safe='')
                    google_viewer_url = f"https://docs.google.com/viewer?url={pdf_url_encoded}&embedded=true"
                    usar_google = navegador['needs_google_viewer']
                    src_inicial = google_viewer_url if usar_google else pdf_url_visor

                    components.html(f"""
<style>
@keyframes spin {{from{{transform:rotate(0deg)}}to{{transform:rotate(360deg)}}}}
body,html{{margin:0;padding:0;overflow:hidden;}}
#pdf-wrap {{width:100%;height:680px;border:2px solid #e2e8f0;border-radius:12px;
            overflow:hidden;box-shadow:0 4px 12px rgba(0,0,0,0.1);background:#f0f2f5;position:relative;}}
#pdf-loading {{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;
               justify-content:center;background:#f0f2f5;z-index:2;gap:12px;
               transition:opacity 0.4s ease;}}
#pdf-spinner {{width:40px;height:40px;border:4px solid #cbd5e1;border-top-color:#5b7cfa;
               border-radius:50%;animation:spin 0.8s linear infinite;}}
#pdf-loading span {{color:#64748b;font-size:0.9rem;font-family:sans-serif;}}
#pdf-iframe {{position:absolute;inset:0;width:100%;height:100%;border:none;display:block;}}
</style>
<div id="pdf-wrap">
  <div id="pdf-loading">
    <div id="pdf-spinner"></div>
    <span id="pdf-status">Cargando PDF...</span>
  </div>
  <iframe id="pdf-iframe" src="" allow="fullscreen"></iframe>
</div>
<script>
(function() {{
  var iframe  = document.getElementById('pdf-iframe');
  var loading = document.getElementById('pdf-loading');
  var status  = document.getElementById('pdf-status');
  var googleUrl  = "{google_viewer_url}";
  var directUrl  = "{pdf_url_visor}";
  var usingGoogle = {"true" if usar_google else "false"};

  function hideLoading() {{
    loading.style.opacity = '0';
    setTimeout(function(){{ loading.style.display = 'none'; }}, 400);
  }}

  function loadDirect() {{
    usingGoogle = false;
    status.textContent = 'Cargando PDF...';
    iframe.src = directUrl;
    // Para iframe directo, ocultar spinner tras tiempo prudente según conexión
    setTimeout(hideLoading, 4000);
  }}

  // Arrancar con Google Viewer o directo según navegador
  if (usingGoogle) {{
    iframe.src = googleUrl;
    // Google Viewer: ocultar spinner a los 3s (ya renderizó o sigue cargando de fondo)
    // Si a los 12s aún no hubo contenido, caer a directo
    setTimeout(function() {{
      if (loading.style.display !== 'none') hideLoading();
    }}, 3000);
    setTimeout(function() {{
      // Intentar detectar pantalla en blanco de Google Viewer vía postMessage no es posible
      // por CORS — simplemente ofrecer fallback visual con botón
      if (usingGoogle) {{
        try {{
          var doc = iframe.contentDocument || iframe.contentWindow.document;
          if (!doc || !doc.body || doc.body.children.length === 0) loadDirect();
        }} catch(e) {{
          // CORS — no podemos leer, asumir que cargó bien
        }}
      }}
    }}, 8000);
  }} else {{
    iframe.src = directUrl;
    setTimeout(hideLoading, 4000);
  }}
}})();
</script>
""", height=710, scrolling=False)

                    try:
                        pdf_bytes = requests.get(st.session_state.pdf_url, timeout=15).content
                        st.download_button(
                            label="📥 Descargar Plano",
                            data=pdf_bytes,
                            file_name=st.session_state.pdf_nombre,
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"descargar_plano_{st.session_state.numero_en_visor}"
                        )
                    except Exception as e:
                        st.warning("⚠️ No se pudo preparar la descarga. Intenta de nuevo.")

        st.markdown("---")
        st.markdown("### 📊 Estadísticas Rápidas")

        autorizadas = autorizadas_con_plano = borradores_con_plano = borradores = 0
        incompletos_con_plano = incompletos = total_cotizado = 0

        for row in st.session_state.resultados_busqueda:
            datos_completos = all([row[1], row[6], row[7]])
            asesor_completo = any([row[2], row[8], row[9]])
            total_cotizado += row[4] if row[4] else 0
            tiene_plano = bool(row[10]) if len(row) > 10 else False

            if not datos_completos or not asesor_completo:
                if tiene_plano: incompletos_con_plano += 1
                else: incompletos += 1
            elif row[5] and row[5] > 0:
                if tiene_plano: autorizadas_con_plano += 1
                else: autorizadas += 1
            else:
                if tiene_plano: borradores_con_plano += 1
                else: borradores += 1

        autorizadas_total = autorizadas + autorizadas_con_plano

        col_e1, col_e2, col_e3, col_e4, col_e5, col_e6 = st.columns(6)
        stats = [
            (col_e1, "💰 TOTAL COTIZADO", formato_clp(total_cotizado), "total", "Total de cotizaciones"),
            (col_e2, "🟢 AUTORIZADAS", str(autorizadas_total), "autorizadas", f"{autorizadas_con_plano} con plano"),
            (col_e3, "🟠 BORRADOR C/P", str(borradores_con_plano), "color:#f97316;", "Borradores con plano"),
            (col_e4, "🟡 BORRADOR", str(borradores), "borradores", "Borradores sin plano"),
            (col_e5, "🔴 INCOMPLETO C/P", str(incompletos_con_plano), "color:#ef4444;", "Incompletos con plano"),
            (col_e6, "🔴 INCOMPLETO", str(incompletos), "incompletas", "Incompletos sin plano"),
        ]
        for col, title, number, css_class, desc in stats:
            with col:
                if len(number) > 12:
                    font_size = "1.6rem"
                elif len(number) > 8:
                    font_size = "2rem"
                else:
                    font_size = "2.8rem"
                if css_class.startswith("color:"):
                    num_html = f'<div class="stats-number" style="{css_class};font-size:{font_size};">{number}</div>'
                else:
                    num_html = f'<div class="stats-number {css_class}" style="font-size:{font_size};">{number}</div>'
                st.markdown(f'<div class="stats-card"><div class="stats-title">{title}</div>{num_html}<div class="stats-desc">{desc}</div></div>', unsafe_allow_html=True)

    else:
        st.info("💡 No hay resultados. Realice una búsqueda para ver cotizaciones guardadas.")

# =========================================================
# TOASTS GENERALES — mostrar mensaje pendiente del rerun anterior
if st.session_state.get('_toast_msg'):
    st.toast(st.session_state['_toast_msg'])
    st.session_state['_toast_msg'] = None

# TOAST ÉXITO AL GUARDAR — st.toast() nativo
# CSS solo se inyecta cuando el toast está activo, evita contenedor
# vacío pegado en pantalla entre reruns
# =========================================================
if st.session_state.get('mostrar_toast_exito', False):
    ep = st.session_state.get('toast_numero_ep', '')
    import streamlit.components.v1 as _tc
    _tc.html(f"""<script>
(function(){{
    var D=window.parent.document;
    if(D.getElementById('_toast_ep')) return;
    var t=D.createElement('div');
    t.id='_toast_ep';
    t.style.cssText='position:fixed;bottom:5rem;left:2rem;z-index:9999999;'+
        'background:linear-gradient(135deg,#10b981,#059669);color:white;'+
        'padding:14px 22px;border-radius:12px;font-size:0.95rem;font-weight:700;'+
        'font-family:Plus Jakarta Sans,sans-serif;'+
        'box-shadow:0 8px 24px rgba(16,185,129,0.4);'+
        'display:flex;align-items:center;gap:10px;'+
        'animation:slideInToast 0.3s ease;';
    t.innerHTML='<span style="font-size:1.2rem">✅</span> Cotización <b style="margin:0 4px">{ep}</b> guardada correctamente' +
        '<button onclick="this.parentElement.remove()" style="background:none;border:none;color:rgba(255,255,255,0.8);' +
        'font-size:1.1rem;cursor:pointer;margin-left:10px;padding:0;line-height:1;" title="Cerrar">✕</button>';
    var s=D.createElement('style');
    s.innerHTML='@keyframes slideInToast{{from{{transform:translateY(20px);opacity:0}}to{{transform:translateY(0);opacity:1}}}}';
    D.head.appendChild(s);
    D.body.appendChild(t);
    setTimeout(function(){{
        t.style.transition='opacity 0.4s';
        t.style.opacity='0';
        setTimeout(function(){{t.remove();}},400);
    }},3500);
}})();
</script>""", height=0)
    st.session_state.mostrar_toast_exito = False

with tab4:
    st.markdown("""
    <style>
    .hdr4 {
        background: linear-gradient(135deg, #003d52 0%, #006978 100%);
        border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
        display: flex; align-items: center; gap: 22px;
        box-shadow: 0 8px 32px rgba(0,105,120,0.25);
        position: relative; overflow: hidden;
    }
    .hdr4::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 180px; height: 180px; border-radius: 50%;
        background: rgba(255,255,255,0.04); pointer-events: none;
    }
    .hdr4::after {
        content: ''; position: absolute; bottom: -60px; right: 80px;
        width: 240px; height: 240px; border-radius: 50%;
        background: rgba(255,255,255,0.03); pointer-events: none;
    }
    .hdr4 h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                 font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
    .hdr4 p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
    </style>
    <div class="hdr4">
      <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">🧊</span>
      <div>
        <h2>Visor 3D Beta</h2>
        <p>Selecciona un presupuesto con plano adjunto para generar su prototipo 3D interactivo.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Obtener presupuestos con plano
    try:
        _resp_3d = supabase.table('cotizaciones').select(
            'numero', 'cliente_nombre', 'plano_url', 'plano_nombre'
        ).not_.is_('plano_url', 'null').order('fecha_creacion', desc=True).execute()
        _opciones_3d = _resp_3d.data or []
    except:
        _opciones_3d = []

    if not _opciones_3d:
        st.info("No hay presupuestos con plano adjunto disponibles.")
    else:
        _labels_3d = [
            f"{r['numero']} — {r['cliente_nombre'] or 'S/C'} — {r['plano_nombre'] or 'plano.pdf'}"
            for r in _opciones_3d
        ]
        _sel_3d = st.selectbox("Selecciona presupuesto con plano:", _labels_3d, key="sel_3d_presupuesto")
        _idx_3d = _labels_3d.index(_sel_3d)
        _plano_url_3d = _opciones_3d[_idx_3d]['plano_url']

        st.markdown(f"📎 **Plano:** `{_opciones_3d[_idx_3d]['plano_nombre'] or 'plano.pdf'}`")

        # Cache por URL — procesa automáticamente al seleccionar
        _cache_key = f"layout_3d_{_plano_url_3d}"
        _layout_3d = st.session_state.get(_cache_key)
        _img_b64_3d = st.session_state.get(f"img_3d_{_plano_url_3d}", "")

        # Si no hay layout en cache, procesar automáticamente
        if not _layout_3d:
            with st.spinner("⏳ Procesando plano con IA, un momento..."):
                _pdf_bytes = None
                try:
                    import requests as _req
                    _r = _req.get(_plano_url_3d, timeout=20)
                    _r.raise_for_status()
                    _pdf_bytes = _r.content
                except Exception as _e:
                    st.error(f"❌ No se pudo descargar el plano: {_e}")

            if _pdf_bytes:
                with st.spinner("🖼️ Renderizando página del plano..."):
                    try:
                        import fitz as _fitz
                        _doc = _fitz.open(stream=_pdf_bytes, filetype="pdf")
                        _page = _doc[0]
                        _pix = _page.get_pixmap(matrix=_fitz.Matrix(2.0, 2.0))
                        _img_bytes = _pix.tobytes("png")
                        _doc.close()
                        import base64 as _b64
                        _img_b64_3d = _b64.b64encode(_img_bytes).decode("utf-8")
                        st.session_state[f"img_3d_{_plano_url_3d}"] = _img_b64_3d
                    except Exception as _e:
                        st.error(f"❌ Error al renderizar PDF: {_e}. Asegúrate de instalar PyMuPDF: pip install pymupdf")
                        _img_b64_3d = ""

                if _img_b64_3d:
                    with st.spinner("🤖 Analizando plano con IA..."):
                        try:
                            import httpx as _httpx, json as _json
                            _api_key = ANTHROPIC_API_KEY
                            if not _api_key:
                                raise ValueError("ANTHROPIC_API_KEY no configurada")
                            _cv_resp = _httpx.post(
                                "https://api.anthropic.com/v1/messages",
                                headers={
                                    "x-api-key": _api_key,
                                    "anthropic-version": "2023-06-01",
                                    "content-type": "application/json"
                                },
                                json={
                                    "model": "claude-sonnet-4-20250514",
                                    "max_tokens": 2048,
                                    "messages": [
                                        {
                                            "role": "user",
                                            "content": [
                                                {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": _img_b64_3d}},
                                                {"type": "text", "text": """Analiza este plano arquitectónico de un container house.

    PASO 1 — Describe brevemente lo que ves:
    - Dimensiones totales del container (largo x ancho en metros)
    - Cuántas PUERTAS reales hay (deben tener arco de 90°) y en qué pared
    - Cuántas VENTANAS reales hay (rectángulo con doble línea o líneas internas) y en qué pared
    - IGNORA completamente: costillas estructurales, líneas de división interna, muebles, cotas

    PASO 2 — Genera el JSON final con SOLO las aberturas reales que describiste:
    {"width":<largo>,"depth":<ancho>,"wallHeight":2.8,"walls":[{"side":"front","openings":[{"type":"door","x":<x desde centro pared>,"y":1.05,"w":<ancho m>,"h":2.1},{"type":"window","x":<x>,"y":1.2,"w":<ancho m>,"h":<alto m>}]},{"side":"back","openings":[...]},{"side":"left","openings":[...]},{"side":"right","openings":[...]}]}

    LÍMITES ESTRICTOS: máximo 3 puertas y 6 ventanas en total. Una pared sin aberturas usa "openings":[]
    El JSON debe ir al final de tu respuesta, solo como bloque de código con ```json```"""}
                                            ]
                                        }
                                    ]
                                },
                                timeout=40
                            )
                            _cv_data = _cv_resp.json()
                            _cv_txt = "".join(b.get("text","") for b in _cv_data.get("content",[])).strip()

                            # Extraer JSON del bloque ```json``` si existe, sino buscar { directo
                            import re as _re
                            _json_match = _re.search(r'```json\s*(\{.*?\})\s*```', _cv_txt, _re.DOTALL)
                            if _json_match:
                                _cv_json_str = _json_match.group(1)
                            else:
                                # Fallback: encontrar el primer { hasta el último }
                                _js = _cv_txt.find('{')
                                _je = _cv_txt.rfind('}')
                                _cv_json_str = _cv_txt[_js:_je+1] if _js >= 0 else _cv_txt

                            _layout_raw = _json.loads(_cv_json_str)

                            # ── Post-procesador estricto ──────────────────────────
                            _W = float(_layout_raw.get("width", 9))
                            _D = float(_layout_raw.get("depth", 3))

                            # Límites máximos de aberturas por pared
                            _max_openings = {"front": 3, "back": 2, "left": 2, "right": 2}

                            for _wall in _layout_raw.get("walls", []):
                                _side  = _wall["side"]
                                _wlen  = _W if _side in ("front","back") else _D
                                _valid = []

                                for _op in _wall.get("openings", []):
                                    _ow  = float(_op.get("w", 0.9))
                                    _oh  = float(_op.get("h", 2.1))
                                    _ox  = float(_op.get("x", 0))
                                    _oy  = float(_op.get("y", 1.05))

                                    # Filtro 1: x dentro del rango de la pared
                                    if abs(_ox) + _ow/2 >= _wlen/2 - 0.05:
                                        continue

                                    # Filtro 2: dimensiones razonables
                                    if _op.get("type") == "door":
                                        if not (0.7 <= _ow <= 1.2 and 1.8 <= _oh <= 2.4):
                                            continue
                                    else:  # window
                                        if not (0.4 <= _ow <= 2.0 and 0.4 <= _oh <= 1.8):
                                            continue

                                    # Filtro 3: y dentro de la pared
                                    if _oy - _oh/2 < 0 or _oy + _oh/2 > 2.8:
                                        continue

                                    # Filtro 4: no solapada con otra abertura ya aceptada
                                    _overlap = False
                                    for _v in _valid:
                                        if abs(_ox - float(_v["x"])) < (_ow + float(_v["w"]))/2 + 0.05:
                                            _overlap = True; break
                                    if _overlap:
                                        continue

                                    _valid.append({**_op, "x": _ox, "y": _oy, "w": _ow, "h": _oh})

                                # Limitar cantidad máxima por pared
                                _wall["openings"] = _valid[:_max_openings.get(_side, 3)]

                            _layout_3d = _layout_raw
                            st.session_state[_cache_key] = _layout_3d
                            st.rerun()
                        except ValueError as _ve:
                            # Sin API key: análisis geométrico local
                            st.warning(f"⚠️ {_ve} — usando análisis geométrico local")
                            try:
                                import base64 as _b64c, io, json as _json
                                from PIL import Image as _PIL_Image
                                import numpy as _np

                                _img_data = _b64c.b64decode(_img_b64_3d)
                                _pil_img = _PIL_Image.open(io.BytesIO(_img_data)).convert("L")
                                _arr = _np.array(_pil_img)
                                ih, iw = _arr.shape

                                # Bounding box del contenido oscuro (paredes)
                                _dark = _arr < 100
                                _rows = _np.any(_dark, axis=1)
                                _cols = _np.any(_dark, axis=0)
                                _r0,_r1 = _np.where(_rows)[0][[0,-1]]
                                _c0,_c1 = _np.where(_cols)[0][[0,-1]]

                                # Escala: asumir que el plano representa un container HC
                                # Maitencillo HC: 6.0 x 3.0 m típico, o 9.0 x 3.0 m
                                _ratio = (_c1-_c0) / max(_r1-_r0, 1)
                                if _ratio > 2.5:
                                    _W, _D = 9.0, 3.0
                                elif _ratio > 1.8:
                                    _W, _D = 6.0, 3.0
                                else:
                                    _W, _D = 6.0, 3.0

                                # Detectar aberturas por proyección de píxeles claros en bordes
                                def _find_openings(edge_strip, wall_len, side):
                                    """Detecta zonas claras (aberturas) en un strip del borde"""
                                    bright = edge_strip > 200
                                    openings = []
                                    in_gap = False
                                    gap_start = 0
                                    min_gap = int(iw * 0.05)  # mínimo 5% del ancho
                                    for i, b in enumerate(bright):
                                        if b and not in_gap:
                                            in_gap = True; gap_start = i
                                        elif not b and in_gap:
                                            in_gap = False
                                            gap_w = i - gap_start
                                            if gap_w >= min_gap:
                                                cx = (gap_start + i/2) / len(bright) * wall_len - wall_len/2
                                                w_m = gap_w / len(bright) * wall_len
                                                if w_m < 1.5:  # puerta
                                                    openings.append({"type":"door","x":round(cx,2),"y":1.05,"w":round(min(w_m,1.0),2),"h":2.1})
                                                else:  # ventana
                                                    openings.append({"type":"window","x":round(cx,2),"y":1.2,"w":round(min(w_m,1.5),2),"h":1.0})
                                    return openings

                                _strip_h = max(5, int(ih*0.05))
                                _strip_w = max(5, int(iw*0.05))
                                _front_strip = _arr[_r1-_strip_h:_r1, _c0:_c1].mean(axis=0)
                                _back_strip  = _arr[_r0:_r0+_strip_h, _c0:_c1].mean(axis=0)
                                _left_strip  = _arr[_r0:_r1, _c0:_c0+_strip_w].mean(axis=1)
                                _right_strip = _arr[_r0:_r1, _c1-_strip_w:_c1].mean(axis=1)

                                _layout_3d = {
                                    "width": _W, "depth": _D, "wallHeight": 2.8,
                                    "walls": [
                                        {"side":"front",  "openings": _find_openings(_front_strip, _W, "front")},
                                        {"side":"back",   "openings": _find_openings(_back_strip,  _W, "back")},
                                        {"side":"left",   "openings": _find_openings(_left_strip,  _D, "left")},
                                        {"side":"right",  "openings": _find_openings(_right_strip, _D, "right")},
                                    ]
                                }
                                st.session_state[_cache_key] = _layout_3d
                                st.rerun()
                            except Exception as _e2:
                                st.error(f"❌ Error en análisis local: {_e2}")
                        except Exception as _e_api:
                            st.error(f"❌ Error API Claude: {_e_api}")
                            st.info("💡 Configura ANTHROPIC_API_KEY en los secrets de Streamlit para usar Claude Vision.")


        if _layout_3d and _img_b64_3d:
            import json as _json
            _layout_json = _json.dumps(_layout_3d)

            _visor_html = f"""<!DOCTYPE html>
    <html><head><meta charset="utf-8">
    <style>
    *{{margin:0;padding:0;box-sizing:border-box;}}
    body{{background:#0f1117;overflow:hidden;font-family:'Segoe UI',sans-serif;}}
    #wrap{{width:100%;height:600px;position:relative;}}
    #c3d{{width:100%;height:100%;display:block;}}
    #ctrl{{position:absolute;top:10px;left:10px;z-index:10;display:flex;gap:6px;flex-wrap:wrap;}}
    .btn{{background:rgba(15,17,34,0.82);color:#cdd6f4;border:1px solid rgba(255,255,255,0.15);padding:5px 12px;border-radius:18px;cursor:pointer;font-size:11px;font-weight:600;transition:all .15s;}}
    .btn:hover,.btn.on{{background:rgba(91,124,250,0.72);border-color:#5b7cfa;color:#fff;}}
    #hud{{position:absolute;bottom:10px;left:50%;transform:translateX(-50%);color:rgba(255,255,255,0.35);font-size:10px;background:rgba(0,0,0,0.5);padding:5px 14px;border-radius:18px;white-space:nowrap;}}
    </style></head>
    <body><div id="wrap">
    <canvas id="c3d"></canvas>
    <div id="ctrl">
    <button class="btn on" id="bRoof" onclick="tog('roof')">🏠 Techo</button>
    <button class="btn on" id="bPlan" onclick="tog('plan')">📐 Plano</button>
    <button class="btn on" id="bWire" onclick="tog('wire')">🔲 Wire</button>
    <button class="btn" onclick="resetCam()">🎯 Reset</button>
    <button class="btn" onclick="setV('top')">⬆ Top</button>
    <button class="btn" onclick="setV('iso')">🔷 Iso</button>
    </div>
    <div id="hud">🖱 Arrastrar: rotar │ Scroll: zoom │ Derecho: mover</div>
    </div>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/three.js/r128/three.min.js"></script>
    <script>
    const LAYOUT={_layout_json};
    const IMG_B64="{_img_b64_3d}";

    const cv=document.getElementById('c3d');
    const W0=cv.parentElement.offsetWidth,H0=600;
    const renderer=new THREE.WebGLRenderer({{canvas:cv,antialias:true}});
    renderer.setPixelRatio(Math.min(devicePixelRatio,2));
    renderer.setSize(W0,H0);
    renderer.shadowMap.enabled=true;
    const scene=new THREE.Scene();
    scene.background=new THREE.Color(0x0f1117);
    const camera=new THREE.PerspectiveCamera(42,W0/H0,0.1,300);
    let S={{th:0.6,ph:1.0,r:32}},T=new THREE.Vector3(0,1.5,0);
    function applyC(){{
      camera.position.set(T.x+S.r*Math.sin(S.ph)*Math.sin(S.th),T.y+S.r*Math.cos(S.ph),T.z+S.r*Math.sin(S.ph)*Math.cos(S.th));
      camera.lookAt(T);
    }}
    applyC();
    scene.add(new THREE.AmbientLight(0xffffff,0.5));
    const sun=new THREE.DirectionalLight(0xfff8e7,1.0);
    sun.position.set(15,25,12);sun.castShadow=true;sun.shadow.mapSize.set(2048,2048);
    scene.add(sun);
    scene.add(new THREE.HemisphereLight(0x7788cc,0x223344,0.35));
    scene.add(new THREE.GridHelper(80,80,0x1e2140,0x1a1d36));
    const gRoof=new THREE.Group(),gPlan=new THREE.Group(),gWire=new THREE.Group(),gBody=new THREE.Group();
    scene.add(gBody,gRoof,gPlan,gWire);
    const vis={{roof:true,plan:true,wire:true}};
    function tog(k){{
      vis[k]=!vis[k];
      if(k==='roof')gRoof.visible=vis[k];
      if(k==='plan')gPlan.visible=vis[k];
      if(k==='wire')gWire.visible=vis[k];
      document.getElementById('b'+k[0].toUpperCase()+k.slice(1)).classList.toggle('on',vis[k]);
    }}
    let drag=false,rDrag=false,lx=0,ly=0;
    cv.addEventListener('mousedown',e=>{{drag=true;rDrag=e.button===2;lx=e.clientX;ly=e.clientY;}});
    cv.addEventListener('contextmenu',e=>e.preventDefault());
    window.addEventListener('mouseup',()=>drag=false);
    window.addEventListener('mousemove',e=>{{
      if(!drag)return;
      const dx=e.clientX-lx,dy=e.clientY-ly;lx=e.clientX;ly=e.clientY;
      if(rDrag){{const r=new THREE.Vector3().crossVectors(new THREE.Vector3().subVectors(camera.position,T).normalize(),camera.up).normalize();T.addScaledVector(r,-dx*0.022);T.y+=dy*0.022;}}
      else{{S.th-=dx*0.007;S.ph=Math.max(0.05,Math.min(Math.PI/2.05,S.ph+dy*0.007));}}
      applyC();
    }});
    cv.addEventListener('wheel',e=>{{S.r=Math.max(3,Math.min(80,S.r+e.deltaY*0.04));applyC();}},{{passive:true}});
    function resetCam(){{S={{th:0.6,ph:1.0,r:32}};T.set(0,1.5,0);applyC();}}
    function setV(v){{if(v==='top'){{S.ph=0.04;applyC();}}else{{S={{th:0.8,ph:0.85,r:30}};applyC();}}}}
    (function loop(){{requestAnimationFrame(loop);renderer.render(scene,camera);}})();

    // Materiales
    const mWall=new THREE.MeshStandardMaterial({{color:0xd4dde3,roughness:0.45,metalness:0.2}});
    const mRoof=new THREE.MeshStandardMaterial({{color:0x546e7a,roughness:0.4,metalness:0.55}});
    const mGlass=new THREE.MeshStandardMaterial({{color:0x89cff0,transparent:true,opacity:0.4,roughness:0.05,metalness:0.1}});
    const mDoor=new THREE.MeshStandardMaterial({{color:0x37474f,roughness:0.4,metalness:0.6}});
    const mWire=new THREE.MeshBasicMaterial({{color:0x5b7cfa,wireframe:true}});
    const mRib=new THREE.MeshStandardMaterial({{color:0x8fa4ae,roughness:0.45,metalness:0.5}});

    const W=LAYOUT.width, D=LAYOUT.depth, H=LAYOUT.wallHeight||2.8, th=0.14;

    // Plano PDF como textura en el suelo
    const img=new Image();
    img.onload=()=>{{
      const tex=new THREE.Texture(img);tex.needsUpdate=true;
      const fm=new THREE.Mesh(new THREE.PlaneGeometry(W,D),new THREE.MeshStandardMaterial({{map:tex,roughness:0.85}}));
      fm.rotation.x=-Math.PI/2;fm.position.y=0.01;fm.receiveShadow=true;gPlan.add(fm);
    }};
    img.src='data:image/png;base64,'+IMG_B64;

    // Suelo sólido
    const flM=new THREE.Mesh(new THREE.BoxGeometry(W+th*2,0.1,D+th*2),
      new THREE.MeshStandardMaterial({{color:0xc8cdd4,roughness:0.95}}));
    flM.position.y=-0.05;flM.receiveShadow=true;gBody.add(flM);

    // ── makeWall: construye una pared con huecos correctos ──────────
    // wallW = largo de la pared, openings = array de {{type,x,y,w,h}}
    // x = posición desde el centro de la pared (-wallW/2 .. +wallW/2)
    // y = centro vertical desde el suelo (0..H)
    function makeWall(px, pz, rotY, wallW, openings) {{

      // 1. Clonar y sanear openings — clamp para que no salgan del borde
      const ops = openings
    .map(op => {{
      const hw = op.w / 2;
      const x  = Math.max(-wallW/2 + hw + 0.01, Math.min(wallW/2 - hw - 0.01, op.x));
      const yb = op.y - op.h/2;  // base
      const yt = op.y + op.h/2;  // tope
      const y  = Math.max(op.h/2 + 0.01, Math.min(H - op.h/2 - 0.01, op.y));
      return {{ ...op, x, y }};
    }})
    .sort((a,b) => a.x - b.x);

      const grp = new THREE.Group();
      grp.position.set(px, 0, pz);
      grp.rotation.y = rotY;

      // 2. Por cada abertura construir: panel-izq, dintel, antepecho, relleno
      let curX = -wallW / 2;

      ops.forEach(op => {{
    const opL = op.x - op.w/2;   // borde izquierdo del hueco
    const opR = op.x + op.w/2;   // borde derecho
    const opB = op.y - op.h/2;   // base del hueco
    const opT = op.y + op.h/2;   // tope del hueco

    // Panel izquierdo (lleno, de suelo a techo)
    const segW = opL - curX;
    if (segW > 0.02) {{
      const m = new THREE.Mesh(new THREE.BoxGeometry(segW, H, th), mWall);
      m.position.set(curX + segW/2, H/2, 0);
      m.castShadow = true; m.receiveShadow = true;
      grp.add(m);
    }}
    curX = opR;

    // Dintel (sobre el hueco)
    const dintelH = H - opT;
    if (dintelH > 0.02) {{
      const m = new THREE.Mesh(new THREE.BoxGeometry(op.w, dintelH, th), mWall);
      m.position.set(op.x, opT + dintelH/2, 0);
      m.castShadow = true;
      grp.add(m);
    }}

    // Antepecho (bajo el hueco — solo ventanas)
    if (op.type === 'window' && opB > 0.02) {{
      const m = new THREE.Mesh(new THREE.BoxGeometry(op.w, opB, th), mWall);
      m.position.set(op.x, opB/2, 0);
      m.castShadow = true;
      grp.add(m);
    }}

    // Relleno del hueco (vidrio o puerta)
    const mat  = op.type === 'door' ? mDoor : mGlass;
    const fill = new THREE.Mesh(new THREE.BoxGeometry(op.w, op.h, th * 0.3), mat);
    fill.position.set(op.x, op.y, 0);
    fill.castShadow = true;
    grp.add(fill);

    // Marco del hueco
    const mrkMat = mRib;
    const mkT = 0.05;
    // Jambas laterales
    [opL - mkT/2, opR + mkT/2].forEach(mx => {{
      const mk = new THREE.Mesh(new THREE.BoxGeometry(mkT, op.h, th+0.02), mrkMat);
      mk.position.set(mx, op.y, 0); grp.add(mk);
    }});
    // Dintel marco
    const mk2 = new THREE.Mesh(new THREE.BoxGeometry(op.w + mkT*2, mkT, th+0.02), mrkMat);
    mk2.position.set(op.x, opT + mkT/2, 0); grp.add(mk2);
    // Umbral (solo ventana)
    if (op.type === 'window') {{
      const mk3 = new THREE.Mesh(new THREE.BoxGeometry(op.w + mkT*2, mkT, th+0.02), mrkMat);
      mk3.position.set(op.x, opB - mkT/2, 0); grp.add(mk3);
    }}
      }});

      // Panel derecho final
      const remW = wallW/2 - curX;
      if (remW > 0.02) {{
    const m = new THREE.Mesh(new THREE.BoxGeometry(remW, H, th), mWall);
    m.position.set(curX + remW/2, H/2, 0);
    m.castShadow = true; m.receiveShadow = true;
    grp.add(m);
      }}

      gBody.add(grp);

      // Wire outline (pared completa, en grupo separado)
      const wg = new THREE.Group();
      wg.position.set(px, H/2, pz);
      wg.rotation.y = rotY;
      wg.add(new THREE.Mesh(new THREE.BoxGeometry(wallW, H, th), mWire));
      gWire.add(wg);
    }}

    // ── Construir las 4 paredes ─────────────────────────────────────
    LAYOUT.walls.forEach(w => {{
      let px, pz, rotY, ww;
      if      (w.side==='front') {{ px=0;    pz= D/2; rotY=0;          ww=W; }}
      else if (w.side==='back')  {{ px=0;    pz=-D/2; rotY=0;          ww=W; }}
      else if (w.side==='left')  {{ px=-W/2; pz=0;    rotY=Math.PI/2;  ww=D; }}
      else                       {{ px= W/2; pz=0;    rotY=Math.PI/2;  ww=D; }}
      makeWall(px, pz, rotY, ww, w.openings||[]);
    }});

    // ── Techo ───────────────────────────────────────────────────────
    // Losa principal
    const roofM = new THREE.Mesh(new THREE.BoxGeometry(W+th*2, 0.18, D+th*2), mRoof);
    roofM.position.y = H + 0.09;
    roofM.castShadow = true;
    gRoof.add(roofM);
    // Alero frontal
    const aleroM = new THREE.Mesh(new THREE.BoxGeometry(W+th*2, 0.08, 0.4), mRoof);
    aleroM.position.set(0, H+0.04, D/2+th+0.2);
    gRoof.add(aleroM);
    // Perfil metálico perimetral del techo
    [[W+th*2, 0.12, th, 0, D/2+th/2, 0],
     [W+th*2, 0.12, th, 0,-D/2-th/2, 0],
     [th, 0.12, D+th*2,-W/2-th/2, 0, 0],
     [th, 0.12, D+th*2, W/2+th/2, 0, 0]].forEach(r => {{
      const m = new THREE.Mesh(new THREE.BoxGeometry(r[0],r[1],r[2]), mRib);
      m.position.set(r[3], H, r[4]);
      gRoof.add(m);
    }});

    // ── Costillas estructurales (rasgos de container) ───────────────
    const ribHeights = [0.0, 0.55, 1.1, 1.65, 2.2, H];
    ribHeights.forEach(h => {{
      // frente y atrás
      [D/2+th*0.6, -D/2-th*0.6].forEach(pz2 => {{
    const m = new THREE.Mesh(new THREE.BoxGeometry(W+th*2, 0.05, 0.05), mRib);
    m.position.set(0, h, pz2); gBody.add(m);
      }});
      // laterales
      [-W/2-th*0.6, W/2+th*0.6].forEach(px2 => {{
    const m = new THREE.Mesh(new THREE.BoxGeometry(0.05, 0.05, D+th*2), mRib);
    m.position.set(px2, h, 0); gBody.add(m);
      }});
    }});

    // ── Columnas en las 4 esquinas ──────────────────────────────────
    [[-W/2, -D/2],[W/2, -D/2],[-W/2, D/2],[W/2, D/2]].forEach(([cx,cz]) => {{
      const col = new THREE.Mesh(new THREE.BoxGeometry(th+0.04, H+0.2, th+0.04), mRib);
      col.position.set(cx, H/2, cz);
      col.castShadow = true;
      gBody.add(col);
    }});

    // ── Ajustar cámara ──────────────────────────────────────────────
    T.set(0, H * 0.35, 0);
    S.r  = Math.max(W, D) * 2.6;
    S.th = 0.55;   // ángulo horizontal — muestra fachada frontal
    S.ph = 0.85;   // ángulo vertical — ni muy alto ni muy bajo
    applyC();
    </script></body></html>"""

            import streamlit.components.v1 as _components
            _components.html(_visor_html, height=620, scrolling=False)
            st.caption(f"⚠️ Beta — Dimensiones detectadas: {_layout_3d.get('width',0):.1f}m × {_layout_3d.get('depth',0):.1f}m × {_layout_3d.get('wallHeight',2.8):.1f}m altura")

            # Debug: mostrar JSON detectado + botón para regenerar
            import json as _json_dbg
            _col_dbg1, _col_dbg2 = st.columns([3,1])
            with _col_dbg2:
                if st.button("🔄 Regenerar", key="btn_regen_3d", help="Forzar nuevo análisis del plano"):
                    st.session_state.pop(_cache_key, None)
                    st.session_state.pop(f"img_3d_{_plano_url_3d}", None)
                    st.rerun()
            with _col_dbg1:
                with st.expander("🔍 Ver JSON detectado por Claude Vision"):
                    st.json(_layout_3d)

# =========================================================
# TAB 5 - PROYECTO EXCEL (solo admin)
# =========================================================
if st.session_state.modo_admin and tab5 is not None:
    with tab5:

        # CSS del tab5
        st.markdown("""
        <style>
        .excel-header {
            background: linear-gradient(135deg, #0f2240 0%, #1a4d33 100%);
            border-radius: 20px; padding: 32px 36px; margin-bottom: 28px; margin-top: -1rem;
            display: flex; align-items: center; gap: 22px;
            box-shadow: 0 8px 32px rgba(26,77,51,0.3);
            position: relative; overflow: hidden;
        }
        .excel-header::before {
            content: ''; position: absolute; top: -40px; right: -40px;
            width: 180px; height: 180px; border-radius: 50%;
            background: rgba(255,255,255,0.04); pointer-events: none;
        }
        .excel-header::after {
            content: ''; position: absolute; bottom: -60px; right: 80px;
            width: 240px; height: 240px; border-radius: 50%;
            background: rgba(255,255,255,0.03); pointer-events: none;
        }
        .excel-header h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                           font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
        .excel-header p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }

        /* Padding para todos los widgets del tab5 */
        [data-testid="stVerticalBlock"] [data-testid="stFileUploader"],
        [data-testid="stVerticalBlock"] [data-testid="stTextInput"] > div,
        [data-testid="stVerticalBlock"] [data-testid="stTextInput"] input {
            padding-left: 4px;
        }

        /* ── Títulos de sección industriales ── */
        .ind-titulo {
            display: flex; align-items: center; gap: 10px;
            font-size: 0.7rem; font-weight: 800; letter-spacing: 0.15em;
            text-transform: uppercase; color: #1e293b;
            border-left: 4px solid #f59e0b;
            padding: 6px 0 6px 12px;
            margin: 20px 0 14px 0;
        }
        /* ── Filas de versión ── */
        .ver-activa {
            background: linear-gradient(90deg, rgba(16,185,129,0.07), rgba(16,185,129,0.02));
            border: 1px solid #10b981;
            border-left: 5px solid #10b981;
            border-radius: 10px;
            padding: 14px 18px;
            margin-bottom: 8px;
        }
        .ver-row {
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            border-left: 5px solid #cbd5e1;
            border-radius: 10px;
            padding: 14px 18px;
            margin-bottom: 8px;
            transition: border-left-color 0.2s;
        }
        .ver-nombre-activa { font-size:1rem; font-weight:700; color:#065f46; }
        .ver-nombre       { font-size:1rem; font-weight:600; color:#1e293b; }
        .ver-badge        {
            display:inline-block; background:#10b981; color:white;
            font-size:0.6rem; font-weight:800; letter-spacing:0.1em;
            padding:2px 7px; border-radius:4px; text-transform:uppercase;
            vertical-align:middle; margin-left:8px;
        }
        .ver-meta   { font-size:0.75rem; color:#64748b; margin-top:3px; }
        .ver-archivo{ font-size:0.7rem; color:#94a3b8; font-family:monospace; margin-top:2px; }

        .version-row { margin-bottom: 4px; }
        .status-bar-green {
            background: linear-gradient(90deg,rgba(16,185,129,0.12),rgba(16,185,129,0.03));
            border: 1px solid #10b981; border-radius: 10px;
            padding: 14px 20px; margin-top: 16px;
        }
        .status-bar-warn {
            background: rgba(245,158,11,0.08); border: 1px solid #f59e0b;
            border-radius: 10px; padding: 14px 20px; margin-top: 16px;
        }
        </style>
        """, unsafe_allow_html=True)

        # Header — intacto, no se toca
        st.markdown("""
        <div class="excel-header">
          <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">📊</span>
          <div>
            <h2>Proyecto Excel — Control de Versiones</h2>
            <p>Sube nuevas versiones del cotizador.xlsx y activa la que necesites. El sistema se actualiza al instante.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="ind-titulo">⬆ Subir nueva versión</div>', unsafe_allow_html=True)

        # ── Subir nueva versión ──────────────────────────────
        # Usar key dinámica para resetear el uploader tras subir
        if "excel_upload_key" not in st.session_state:
            st.session_state.excel_upload_key = 0


        with st.container():
            st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)
            _mg_l, _titulo_col, _mg_r = st.columns([0.4, 9, 0.4])
            with _titulo_col:
                st.markdown("##### ⬆️ Subir nueva versión")
            st.markdown('<div style="height:12px"></div>', unsafe_allow_html=True)

            # Columnas con margen lateral para simular padding
            _mg, _col_up1, _col_up2, _mg2 = st.columns([0.4, 3, 2, 0.4])
            with _col_up1:
                _excel_file = st.file_uploader(
                    "Archivo cotizador.xlsx",
                    type=["xlsx"],
                    key=f"uploader_excel_{st.session_state.excel_upload_key}",
                    label_visibility="collapsed"
                )
            with _col_up2:
                st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
                _version_nombre = st.text_input(
                    "Nombre de versión",
                    placeholder="Ej: v2.1 — Abril 2025",
                    key=f"input_vnom_{st.session_state.excel_upload_key}",
                    label_visibility="collapsed"
                )
                st.caption("📝 Nombre de la versión")

            if _excel_file and _version_nombre:
                st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
                _mg3, _col_sb, _col_info, _mg4 = st.columns([0.4, 1, 3, 0.4])
                with _col_sb:
                    _btn_subir = st.button("📤 Subir versión", key="btn_subir_excel",
                                           use_container_width=True, type="primary")
                with _col_info:
                    st.info(f"📁 **{_excel_file.name}** — versión: **{_version_nombre}**")

                if _btn_subir:
                    with st.spinner("⏳ Subiendo archivo a Supabase..."):
                        try:
                            import datetime as _dt
                            import pytz as _pytz
                            _tz_cl = _pytz.timezone("America/Santiago")
                            _ts = _dt.datetime.now(_tz_cl).strftime("%Y%m%d_%H%M%S")
                            _nombre_archivo = f"cotizador_{_ts}.xlsx"
                            _excel_bytes = _excel_file.read()

                            supabase.storage.from_("config").upload(
                                path=_nombre_archivo,
                                file=_excel_bytes,
                                file_options={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
                            )
                            _url_publica = supabase.storage.from_("config").get_public_url(_nombre_archivo)
                            supabase.table("excel_versiones").insert({
                                "version_nombre": _version_nombre,
                                "archivo_url": _url_publica,
                                "archivo_nombre": _nombre_archivo,
                                "activa": False,
                                "subida_por": "admin"
                            }).execute()

                            st.session_state.excel_upload_key += 1
                            st.success(f"✅ Versión **{_version_nombre}** subida correctamente.")
                            st.rerun()
                        except Exception as _e:
                            st.error(f"❌ Error al subir: {_e}")
            elif _excel_file and not _version_nombre:
                _mg5, _col_w, _mg6 = st.columns([0.4, 5, 0.4])
                with _col_w:
                    st.warning("⚠️ Escribe un nombre para identificar esta versión.")

            st.markdown('<div style="height:16px"></div>', unsafe_allow_html=True)


        st.markdown('<div class="ind-titulo">📋 Versiones disponibles</div>', unsafe_allow_html=True)
        try:
            _versiones = supabase.table("excel_versiones").select("*").order("fecha_subida", desc=True).execute().data or []
        except:
            _versiones = []

        if not _versiones:
            st.info("📭 No hay versiones subidas aún. Sube el cotizador.xlsx para comenzar.")
        else:
            for _v in _versiones:
                _es_activa = _v.get("activa", False)
                try:
                    import pytz as _pytz_v; from datetime import datetime as _dtv
                    _tz_cl_v = _pytz_v.timezone("America/Santiago")
                    _raw_f = str(_v.get("fecha_subida",""))
                    _dtobj = _dtv.fromisoformat(_raw_f.replace("Z","+00:00"))
                    _fecha = _dtobj.astimezone(_tz_cl_v).strftime("%d/%m/%Y %H:%M")
                except:
                    _fecha = str(_v.get("fecha_subida",""))[:16].replace("T"," ")

                _cv1, _cv2, _cv3, _cv4 = st.columns([3, 2.5, 1.5, 0.6])

                with _cv1:
                    if _es_activa:
                        st.markdown(
                            f'<div class="ver-activa">'
                            f'<div class="ver-nombre-activa">🟢 {_v["version_nombre"]}<span class="ver-badge">activa</span></div>'
                            f'<div class="ver-meta">🗓 {_fecha} &nbsp;·&nbsp; 👤 {_v.get("subida_por","admin")}</div>'
                            f'<div class="ver-archivo">📁 {_v.get("archivo_nombre","")}</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                    else:
                        st.markdown(
                            f'<div class="ver-row">'
                            f'<div class="ver-nombre">⚪ {_v["version_nombre"]}</div>'
                            f'<div class="ver-meta">🗓 {_fecha} &nbsp;·&nbsp; 👤 {_v.get("subida_por","admin")}</div>'
                            f'<div class="ver-archivo">📁 {_v.get("archivo_nombre","")}</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )

                with _cv2:
                    st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)

                with _cv3:
                    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
                    if not _es_activa:
                        if st.button("⚡ Activar", key=f"btn_act_{_v['id']}",
                                     use_container_width=True, type="primary"):
                            with st.spinner("Activando..."):
                                try:
                                    supabase.table("excel_versiones").update({"activa": False}).neq("id","00000000-0000-0000-0000-000000000000").execute()
                                    supabase.table("excel_versiones").update({"activa": True}).eq("id", _v["id"]).execute()
                                    _get_excel_bytes_activo.clear()
                                    _leer_hoja_excel.clear()
                                    _leer_bd_total.clear()
                                    cargar_visibilidad_impresion.clear()
                                    st.session_state.pop("excel_bytes_cache", None)
                                    st.rerun()
                                except Exception as _e:
                                    st.error(f"❌ {_e}")
                    else:
                        st.markdown(
                            '<div style="background:#10b981;color:white;padding:8px 0;'
                            'border-radius:8px;text-align:center;font-size:12px;font-weight:700;">'
                            '🟢 ACTIVA</div>',
                            unsafe_allow_html=True
                        )

                with _cv4:
                    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
                    if not _es_activa:
                        if st.button("🗑️", key=f"btn_del_{_v['id']}", help="Eliminar esta versión"):
                            try:
                                supabase.storage.from_("config").remove([_v.get("archivo_nombre","")])
                                supabase.table("excel_versiones").delete().eq("id", _v["id"]).execute()
                                st.rerun()
                            except Exception as _e:
                                st.error(f"❌ {_e}")


        # ── Previsualizador Excel ────────────────────────────
        st.markdown('<div class="ind-titulo">👁 Vista previa del Excel activo</div>', unsafe_allow_html=True)
        with st.container(border=True):
            _prev_activa = next((_v for _v in _versiones if _v.get("activa")), None)
            if _prev_activa:
                try:
                    import pandas as pd
                    import pandas as pd
                    _prev_src = _get_excel_bytes_activo()
                    if _prev_src and hasattr(_prev_src, "read"):
                        _xls = pd.ExcelFile(_prev_src)
                        _hojas_disp = _xls.sheet_names
                        _hojas_disp = _xls.sheet_names
                        _col_sel, _col_info = st.columns([2, 3])
                        with _col_sel:
                            _hoja_sel = st.selectbox(
                                "Hoja a previsualizar",
                                options=_hojas_disp,
                                key="prev_hoja_sel",
                                label_visibility="collapsed"
                            )
                        with _col_info:
                            st.caption(f"📊 **{len(_hojas_disp)} hojas** en la versión activa · **{_prev_activa['version_nombre']}**")
                        if _hoja_sel:
                            _prev_src.seek(0)
                            _df_prev = pd.read_excel(_prev_src, sheet_name=_hoja_sel, header=None)
                            _df_prev = _df_prev.dropna(how='all').fillna('')
                            _df_str = _df_prev.astype(str).replace('nan','').replace('0.0','')
                            _altura = min(600, max(300, len(_df_str) * 35 + 50))
                            st.dataframe(
                                _df_str,
                                use_container_width=True,
                                hide_index=True,
                                height=_altura
                            )
                            st.caption(f"📋 {len(_df_prev)} filas · {len(_df_prev.columns)} columnas en esta hoja")
                    else:
                        st.info("No se pudo cargar el archivo Excel activo.", icon="⚠️")
                except Exception as _pe:
                    st.error(f"Error al previsualizar: {_pe}")
            else:
                st.info("Activa una versión para poder previsualizarla.", icon="📂")

        # ── Exportar CSV ─────────────────────────────────────
        st.markdown('<div class="ind-titulo">📦 Exportar datos</div>', unsafe_allow_html=True)
        with st.container(border=True):
            st.caption("Descarga todas las cotizaciones del sistema en formato CSV para respaldo o análisis externo.")
            _col_csv_btn, _col_csv_info = st.columns([1, 3])
            with _col_csv_btn:
                if st.button("📦 Generar CSV", key="btn_generar_csv", use_container_width=True, type="primary"):
                    st.session_state._csv_listo = exportar_csv_completo()
            with _col_csv_info:
                if st.session_state.get('_csv_listo'):
                    from datetime import datetime as _dt
                    import pytz as _pytz2
                    _tz_cl2 = _pytz2.timezone("America/Santiago")
                    _fname = f"cotizaciones_backup_{_dt.now(_tz_cl2).strftime('%Y%m%d_%H%M')}.csv"
                    st.download_button(
                        label="⬇️ Descargar CSV",
                        data=st.session_state._csv_listo,
                        file_name=_fname,
                        mime="text/csv",
                        use_container_width=True,
                        key="btn_export_csv"
                    )
                else:
                    st.info("Haz clic en **Generar CSV** para preparar el archivo.", icon="ℹ️")

        # ── Barra de estado ──────────────────────────────────
        _activa_info = next((_v for _v in _versiones if _v.get("activa")), None)
        if _activa_info:
            try:
                import pytz as _pytz_fa; from datetime import datetime as _dtfa
                _raw_fa = str(_activa_info.get("fecha_subida",""))
                _dtfa_obj = _dtfa.fromisoformat(_raw_fa.replace("Z","+00:00"))
                _fa = _dtfa_obj.astimezone(_pytz_fa.timezone("America/Santiago")).strftime("%d/%m/%Y %H:%M")
            except:
                _fa = str(_activa_info.get("fecha_subida",""))[:16].replace("T"," ")
            st.markdown(
                f'<div style="background:linear-gradient(90deg,rgba(16,185,129,0.12),rgba(16,185,129,0.03));'
                f'border:1px solid #10b981;border-radius:10px;padding:14px 20px;margin-top:12px;'
                f'display:flex;align-items:center;gap:12px;">'
                f'<span style="font-size:1.4rem">🟢</span>'
                f'<div><span style="color:#065f46;font-weight:700;">Sistema usando versión activa:</span> '
                f'<strong style="color:#059669;">{_activa_info["version_nombre"]}</strong>'
                f'<span style="color:#6b7280;font-size:0.8rem;margin-left:10px;">subida el {_fa}</span></div>'
                f'</div>',
                unsafe_allow_html=True
            )
        else:
            st.markdown(
                '<div style="background:rgba(245,158,11,0.08);border:1px solid #f59e0b;'
                'border-radius:10px;padding:14px 20px;margin-top:12px;">'
                '⚠️ <strong>Sin versión activa</strong> — el sistema usa el archivo local '
                '<code>cotizador.xlsx</code> de GitHub.</div>',
                unsafe_allow_html=True
            )



# =========================================================
# TAB SALUD - SISTEMA (solo admin)
# =========================================================
if st.session_state.get('es_root') and tab_salud is not None:
    with tab_salud:

        st.markdown("""
        <style>
        .hdr-salud {
            background: linear-gradient(135deg, #0a0a0a 0%, #1a1a1a 60%, #2d2d2d 100%);
            border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
            display: flex; align-items: center; gap: 22px;
            box-shadow: 0 8px 32px rgba(0,0,0,0.5);
            position: relative; overflow: hidden;
        }
        .hdr-salud::before {
            content: ''; position: absolute; top: -40px; right: -40px;
            width: 180px; height: 180px; border-radius: 50%;
            background: rgba(255,255,255,0.03); pointer-events: none;
        }
        .hdr-salud::after {
            content: ''; position: absolute; bottom: -60px; right: 80px;
            width: 240px; height: 240px; border-radius: 50%;
            background: rgba(255,255,255,0.02); pointer-events: none;
        }
        .hdr-salud h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                        font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
        .hdr-salud p  { color: rgba(255,255,255,0.55) !important; margin: 6px 0 0; font-size: 0.92rem; }

        .sys-card {
            background: white; border-radius: 16px; padding: 20px 24px;
            border: 1px solid rgba(226,232,240,0.8);
            box-shadow: 0 4px 20px rgba(0,0,0,0.06);
            margin-bottom: 16px;
        }
        .sys-card-title {
            font-size: 0.72rem; font-weight: 800; color: #94a3b8;
            text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 14px;
            display: flex; align-items: center; gap: 8px;
        }
        .sys-metric-val {
            font-size: 2rem; font-weight: 900; color: #0f172a;
            font-family: 'Montserrat', sans-serif; line-height: 1;
        }
        .sys-metric-sub {
            font-size: 0.78rem; color: #64748b; margin-top: 4px;
        }
        .sys-bar-wrap {
            background: #f1f5f9; border-radius: 8px; height: 12px;
            overflow: hidden; margin: 10px 0 6px;
        }
        .sys-bar-inner {
            height: 12px; border-radius: 8px; transition: width 0.5s ease;
        }
        .sys-bar-ok   { background: linear-gradient(90deg, #10b981, #34d399); }
        .sys-bar-warn { background: linear-gradient(90deg, #f59e0b, #fbbf24); }
        .sys-bar-crit { background: linear-gradient(90deg, #ef4444, #f97316); }
        .sys-pct-label {
            font-size: 0.75rem; font-weight: 700;
            display: flex; justify-content: space-between;
        }
        .sys-section-title {
            font-size: 0.78rem; font-weight: 900; color: #1e293b;
            text-transform: uppercase; letter-spacing: 0.1em;
            margin: 24px 0 14px; padding: 8px 16px;
            background: linear-gradient(90deg, rgba(15,15,15,0.07), transparent);
            border-left: 4px solid #374151; border-radius: 0 8px 8px 0;
        }
        .sys-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
        .sys-table th {
            background: #f8fafc; color: #64748b; font-weight: 700;
            font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.08em;
            padding: 10px 14px; text-align: left; border-bottom: 2px solid #e2e8f0;
        }
        .sys-table td { padding: 10px 14px; border-bottom: 1px solid #f1f5f9; color: #1e293b; }
        .sys-table tr:last-child td { border-bottom: none; }
        .sys-badge-ok   { background:#dcfce7; color:#15803d; padding:2px 8px; border-radius:4px; font-weight:700; font-size:0.7rem; }
        .sys-badge-warn { background:#fef3c7; color:#b45309; padding:2px 8px; border-radius:4px; font-weight:700; font-size:0.7rem; }
        .sys-badge-crit { background:#fee2e2; color:#dc2626; padding:2px 8px; border-radius:4px; font-weight:700; font-size:0.7rem; }
        </style>
        <div class="hdr-salud">
          <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.5));">🛡️</span>
          <div>
            <h2>Salud del Sistema</h2>
            <p>Monitoreo de capacidad y estado de Supabase — Plan Core (actualizado al cargar)</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        import datetime as _dt_sys

        with st.spinner("Consultando métricas del sistema..."):

            # ── 1. Tamaño BD via SQL directo a pg_database_size ──
            _db_size_mb = 0
            _db_rows = {}
            try:
                # Usar execute_sql via rpc si existe, sino estimamos por filas
                import httpx as _hx, json as _js_sys
                _headers_sys = {
                    "apikey": SUPABASE_KEY,
                    "Authorization": f"Bearer {SUPABASE_KEY}",
                    "Content-Type": "application/json",
                }
                _sql_resp = _hx.post(
                    f"{SUPABASE_URL}/rest/v1/rpc/get_db_stats",
                    headers=_headers_sys,
                    json={},
                    timeout=10
                )
                if _sql_resp.status_code == 200:
                    _db_size_mb = round(_sql_resp.json() / (1024*1024), 2)
            except:
                pass

            # Contar filas por tabla (siempre funciona)
            for _tbl in ['cotizaciones', 'cotizacion_logs', 'excel_versiones']:
                try:
                    _r = supabase.table(_tbl).select('id', count='exact').execute()
                    _db_rows[_tbl] = _r.count or 0
                except:
                    _db_rows[_tbl] = 0

            # Estimación de tamaño BD basada en filas si no pudimos leer pg_database_size
            if _db_size_mb == 0:
                _total_filas = sum(_db_rows.values())
                # Estimación conservadora: ~2KB por fila promedio
                _db_size_mb = round((_total_filas * 2048) / (1024*1024), 2)
                _db_size_estimado = True
            else:
                _db_size_estimado = False

            # ── 2. Storage — contar archivos desde tablas + estimar tamaño ──
            _storage_info = {}
            try:
                # Bucket "planos" — contar cotizaciones con plano adjunto
                _r_planos = supabase.table('cotizaciones')                    .select('numero', count='exact')                    .not_.is_('plano_url', 'null')                    .execute()
                _n_planos = _r_planos.count or 0
                # PDFs de planos: estimado ~500KB por archivo
                _mb_planos = round((_n_planos * 500 * 1024) / (1024*1024), 2)
                _storage_info['planos'] = {'archivos': _n_planos, 'mb': _mb_planos, 'estimado': True}
            except:
                _storage_info['planos'] = {'archivos': 0, 'mb': 0, 'estimado': True}

            try:
                # Bucket "config" — versiones Excel + JSONs de descripción
                _r_excel = supabase.table('excel_versiones').select('id', count='exact').execute()
                _n_excel = _r_excel.count or 0
                # PDFs desc por cotización autorizada
                _r_auth = supabase.table('cotizaciones')                    .select('numero', count='exact')                    .eq('estado', 'Autorizado')                    .execute()
                _n_jsons = _r_auth.count or 0
                # Excel ~2MB c/u, JSONs ~5KB c/u
                _mb_config = round((_n_excel * 2 * 1024 + _n_jsons * 5) / 1024, 2)
                _storage_info['config'] = {
                    'archivos': _n_excel + _n_jsons,
                    'mb': _mb_config,
                    'estimado': True
                }
            except:
                _storage_info['config'] = {'archivos': 0, 'mb': 0, 'estimado': True}

            _storage_total_mb = sum(v['mb'] for v in _storage_info.values())

        # ── LÍMITES FREE ──
        _DB_LIMIT_MB  = 500
        _STG_LIMIT_MB = 1024  # 1 GB

        _db_pct  = min(round((_db_size_mb / _DB_LIMIT_MB) * 100, 1), 100) if _db_size_mb > 0 else 0
        _stg_pct = min(round((_storage_total_mb / _STG_LIMIT_MB) * 100, 1), 100)

        def _bar_class(pct):
            if pct >= 80: return "sys-bar-crit"
            if pct >= 50: return "sys-bar-warn"
            return "sys-bar-ok"

        def _badge(pct):
            if pct >= 80: return "sys-badge-crit", "⚠️ Crítico"
            if pct >= 50: return "sys-badge-warn", "🟡 Atención"
            return "sys-badge-ok", "🟢 Normal"

        # ── FILA 1: BD y Storage ──
        st.markdown('<div class="sys-section-title">💾 Capacidad de almacenamiento</div>', unsafe_allow_html=True)
        _col_db, _col_stg = st.columns(2)

        with _col_db:
            _bc, _bl = _badge(_db_pct)
            st.markdown(f"""
            <div class="sys-card">
              <div class="sys-card-title">🗄️ Base de datos PostgreSQL</div>
              <div class="sys-metric-val">{_db_size_mb} MB</div>
              <div class="sys-metric-sub">Límite Core: {_DB_LIMIT_MB} MB &nbsp;·&nbsp; <span class="{_bc}">{_bl}</span>{" &nbsp;· <i style='color:#94a3b8;font-size:0.7rem'>estimado</i>" if _db_size_estimado else ""}</div>
              <div class="sys-bar-wrap">
                <div class="sys-bar-inner {_bar_class(_db_pct)}" style="width:{_db_pct}%"></div>
              </div>
              <div class="sys-pct-label"><span>{_db_pct}% usado</span><span>{round(_DB_LIMIT_MB - _db_size_mb, 1)} MB libres</span></div>
            </div>
            """, unsafe_allow_html=True)

        with _col_stg:
            _bc2, _bl2 = _badge(_stg_pct)
            st.markdown(f"""
            <div class="sys-card">
              <div class="sys-card-title">📦 Storage (todos los buckets)</div>
              <div class="sys-metric-val">{round(_storage_total_mb, 1)} MB</div>
              <div class="sys-metric-sub">Límite Core: {_STG_LIMIT_MB} MB (1 GB) &nbsp;·&nbsp; <span class="{_bc2}">{_bl2}</span></div>
              <div class="sys-bar-wrap">
                <div class="sys-bar-inner {_bar_class(_stg_pct)}" style="width:{_stg_pct}%"></div>
              </div>
              <div class="sys-pct-label"><span>{_stg_pct}% usado</span><span>{round(_STG_LIMIT_MB - _storage_total_mb, 1)} MB libres</span></div>
            </div>
            """, unsafe_allow_html=True)

        # ── FILA 2: Tablas y Buckets ──
        _col_tbl, _col_bkt = st.columns(2)

        with _col_tbl:
            st.markdown('<div class="sys-section-title">📋 Filas por tabla</div>', unsafe_allow_html=True)
            _tbl_html = '<div class="sys-card"><table class="sys-table"><thead><tr><th>Tabla</th><th>Filas</th></tr></thead><tbody>'
            _tbl_labels = {'cotizaciones': '📄 Cotizaciones', 'cotizacion_logs': '📝 Logs auditoría', 'excel_versiones': '📊 Versiones Excel'}
            for _t, _cnt in _db_rows.items():
                _lbl = _tbl_labels.get(_t, _t)
                _tbl_html += f'<tr><td>{_lbl}</td><td><b>{_cnt:,}</b></td></tr>'
            _tbl_html += '</tbody></table></div>'
            st.markdown(_tbl_html, unsafe_allow_html=True)

        with _col_bkt:
            st.markdown('<div class="sys-section-title">🗂️ Archivos por bucket</div>', unsafe_allow_html=True)
            if _storage_info:
                _bkt_html = '<div class="sys-card"><table class="sys-table"><thead><tr><th>Bucket</th><th>Archivos</th><th>Tamaño</th></tr></thead><tbody>'
                _bkt_icons = {'planos': '📐 planos', 'config': '⚙️ config'}
                for _bn, _bv in _storage_info.items():
                    _blbl = _bkt_icons.get(_bn, f'📁 {_bn}')
                    _est = ' <i style="color:#94a3b8;font-size:0.7rem">(est.)</i>' if _bv.get('estimado') else ''
                    _bkt_html += f'<tr><td>{_blbl}</td><td>{_bv["archivos"]}</td><td>{_bv["mb"]} MB{_est}</td></tr>'
                _bkt_html += '</tbody></table></div>'
                st.markdown(_bkt_html, unsafe_allow_html=True)
            else:
                st.markdown('<div class="sys-card"><p style="color:#94a3b8;font-size:0.85rem;">No se pudieron leer los buckets.</p></div>', unsafe_allow_html=True)

        # ── FILA 3: Info del plan ──
        st.markdown('<div class="sys-section-title">ℹ️ Límites del plan Core</div>', unsafe_allow_html=True)
        _plan_html = """
        <div class="sys-card">
          <table class="sys-table">
            <thead><tr><th>Recurso</th><th>Límite Core</th><th>Estado</th></tr></thead>
            <tbody>
              <tr><td>🗄️ Base de datos</td><td>500 MB</td><td><span class="{_bc_db}">{_bl_db}</span></td></tr>
              <tr><td>📦 File Storage</td><td>1 GB</td><td><span class="{_bc_s}">{_bl_s}</span></td></tr>
              <tr><td>👥 Usuarios auth</td><td>50,000</td><td><span class="sys-badge-ok">🟢 Normal</span></td></tr>
              <tr><td>📡 Requests API</td><td>500K / mes</td><td><span class="sys-badge-ok">🟢 Normal</span></td></tr>
              <tr><td>⚡ Ancho de banda</td><td>5 GB / mes</td><td><span class="sys-badge-ok">🟢 Normal</span></td></tr>
              <tr><td>😴 Inactividad</td><td>Pausa tras 7 días sin uso</td><td><span class="sys-badge-warn">🟡 Atención</span></td></tr>
            </tbody>
          </table>
        </div>
        """.replace('{_bc_db}', _badge(_db_pct)[0]).replace('{_bl_db}', _badge(_db_pct)[1])            .replace('{_bc_s}', _badge(_stg_pct)[0]).replace('{_bl_s}', _badge(_stg_pct)[1])
        st.markdown(_plan_html, unsafe_allow_html=True)

        # ── Nota actualización ──
        _now_cl = _dt_sys.datetime.now(_dt_sys.timezone(_dt_sys.timedelta(hours=-3)))
        st.caption(f"🕐 Última actualización: {_now_cl.strftime('%d/%m/%Y %H:%M')} hora Chile · Las métricas de tamaño de BD en Supabase Core se actualizan diariamente.")

        if st.button("🔄 Actualizar métricas", key="btn_refresh_salud"):
            st.rerun()


# =========================================================
# FAB - BOTÓN GUARDAR FLOTANTE
# =========================================================
_es_solo_lectura = (
    st.session_state.cotizacion_cargada and
    st.session_state.margen > 0 and
    not st.session_state.modo_admin
)

_hash_actual = calcular_hash_estado()
_hay_cambios = _hash_actual != st.session_state.get('hash_ultimo_guardado')

_mostrar_fab = (
    len(st.session_state.get('carrito', [])) > 0 and
    not _es_solo_lectura and
    not st.session_state.get('recien_guardado', False) and
    not st.session_state.get('recien_cargado', False) and
    _hay_cambios
)


if st.session_state.get('recien_guardado', False):
    st.session_state.recien_guardado = False
if st.session_state.get('recien_cargado', False):
    st.session_state.recien_cargado = False

if _mostrar_fab:
    # CSS puro — posiciona el botón Streamlit como FAB flotante
    st.markdown("""
<style>
@keyframes pfab{
    0%{box-shadow:0 8px 24px rgba(91,124,250,0.5);}
    50%{box-shadow:0 8px 40px rgba(91,124,250,0.9),0 0 0 12px rgba(91,124,250,0.15);}
    100%{box-shadow:0 8px 24px rgba(91,124,250,0.5);}
}
.st-key-btn_fab_guardar {
    position: fixed !important;
    bottom: 1.5rem !important;
    left: 2rem !important;
    z-index: 999999 !important;
}
.st-key-btn_fab_guardar button {
    background: linear-gradient(135deg,#5b7cfa,#8b5cf6) !important;
    color: #fff !important;
    border: none !important;
    border-radius: 50px !important;
    padding: 0.85rem 1.6rem !important;
    font-size: 0.95rem !important;
    font-weight: 700 !important;
    animation: pfab 2s infinite !important;
    white-space: nowrap !important;
    min-width: 140px !important;
}
.st-key-btn_fab_guardar button:hover {
    transform: translateY(-3px) !important;
    animation: none !important;
}
</style>
""", unsafe_allow_html=True)
    if st.button("💾 Guardar", key="btn_fab_guardar"):
        leer_datos_actuales()
        datos_c, datos_a, proy, cfg, tots, pl_n, pl_d = construir_datos_para_guardar()
        num_g = st.session_state.cotizacion_cargada or generar_numero_unico()
        _usr_log = st.session_state.get('auth_nombre','') or st.session_state.get('auth_email','')
        guardar_cotizacion(num_g, datos_c, datos_a, proy,
                           st.session_state.carrito, cfg, tots, pl_n, pl_d,
                           usuario_logueado=_usr_log)
        st.session_state.cotizacion_cargada = num_g
        st.session_state.hash_ultimo_guardado = calcular_hash_estado()
        st.session_state.recien_guardado = True
        st.session_state.mostrar_toast_exito = True
        st.session_state.toast_numero_ep = num_g
        st.session_state.resultados_busqueda = None
        try:
            _margen_notif = st.session_state.get('margen', 0)
            _cli_nombre   = st.session_state.get('nombre_input', '')
            _tiene_plano  = bool(st.session_state.get('plano_adjunto') or st.session_state.get('pdf_url') or st.session_state.get('plano_nombre'))
            _monto        = tots.get('total', 0) if tots else 0
            _ej_email  = ''
            _ej_nombre = ''
            try:
                _rd = supabase_admin.table('cotizaciones').select('asesor_email','asesor_nombre','cliente_nombre').eq('numero', num_g).execute()
                if _rd.data:
                    _ej_email  = _rd.data[0].get('asesor_email', '') or ''
                    _ej_nombre = _rd.data[0].get('asesor_nombre', '') or ''
                    _cli_nombre = _rd.data[0].get('cliente_nombre', '') or _cli_nombre
            except: pass
            import threading as _thr
            if _tiene_plano:
                if _margen_notif > 0:
                    _sup_nombre = st.session_state.get('auth_nombre','') or st.session_state.get('auth_email','')
                    _thr.Thread(
                        target=notificar_cotizacion_autorizada,
                        args=(num_g, _cli_nombre, _margen_notif, _ej_email, _ej_nombre, _sup_nombre, _monto),
                        daemon=True
                    ).start()
                else:
                    _autor_email = st.session_state.get('auth_email','')
                    _autor_nombre = st.session_state.get('auth_nombre', _ej_nombre)
                    _thr.Thread(
                        target=notificar_nueva_cotizacion,
                        args=(num_g, _autor_nombre, _cli_nombre, _monto, '🟠 Borrador con plano', _autor_email),
                        daemon=True
                    ).start()
        except:
            pass
        st.rerun()

# =========================================================

# =========================================================
# INDICADOR DE PROGRESO FLOTANTE
# =========================================================
_mostrar_progreso = bool(
    st.session_state.get('cotizacion_cargada') or
    len(st.session_state.get('carrito', [])) > 0
)

if _mostrar_progreso:
    _ss = st.session_state
    _es_juridica = _ss.get('cliente_tipo', 'natural') == 'juridica'
    _asesor_ok = bool(_ss.get('asesor_seleccionado','') and _ss.get('asesor_seleccionado') != 'Seleccionar asesor')
    _campos = [
        ('Carrito',        25, bool(len(_ss.get('carrito',[])) > 0)),
        ('Plano PDF',      10, bool(_ss.get('plano_adjunto') or _ss.get('pdf_url') or _ss.get('plano_nombre'))),
        ('Asesor',         10, _asesor_ok),
        ('Nombre cliente', 10, bool(str(_ss.get('nombre_input','')).strip())),
        ('Correo',          8, bool(str(_ss.get('correo_input','')).strip())),
        ('RUT',             8, bool(str(_ss.get('rut_display','')).strip())),
        ('Teléfono',        5, bool(str(_ss.get('telefono_raw','')).strip())),
        ('Descripción proyecto', 5, bool(str(_ss.get('observaciones_input','')).strip())),
        ('Dir. cliente',    5, bool(str(_ss.get('direccion_input','')).strip())),
        ('Dir. proyecto',   5, bool(str(_ss.get('proyecto_direccion','')).strip())),
    ]
    if _es_juridica:
        _campos.append(('Empresa',      5, bool(str(_ss.get('cliente_empresa','')).strip())))
        _campos.append(('RUT empresa',  4, bool(str(_ss.get('cliente_rut_empresa','')).strip())))
    _total_peso = sum(p for _,p,_ in _campos)
    _peso_ok    = sum(p for _,p,v in _campos if v)
    _pct        = int(round(_peso_ok / _total_peso * 100)) if _total_peso > 0 else 0
    if _pct == 100:
        _pc, _pb = '#10b981', 'rgba(16,185,129,0.15)'
    elif _pct >= 70:
        _pc, _pb = '#f97316', 'rgba(249,115,22,0.15)'
    elif _pct >= 40:
        _pc, _pb = '#eab308', 'rgba(234,179,8,0.15)'
    else:
        _pc, _pb = '#ef4444', 'rgba(239,68,68,0.15)'
    _items_parts = []
    for _lbl, _peso, _ok in _campos:
        _ic  = '✅' if _ok else '⬜'
        _col = '#374151' if _ok else '#9ca3af'
        _fw  = '600' if _ok else '400'
        _items_parts.append(
            '<div style="display:flex;align-items:center;gap:6px;padding:2px 0;">'
            '<span style="font-size:0.7rem;">' + _ic + '</span>'
            '<span style="font-size:0.7rem;color:' + _col + ';font-weight:' + _fw + ';">' + _lbl + '</span>'
            '</div>'
        )
    _items_html = ''.join(_items_parts)
    # Panel principal + botón toggle al fondo
    _barra = (
        '<div id="_prog_panel" style="position:fixed;right:0;top:50%;transform:translateY(-50%);'
        'z-index:99997;background:#ffffff;border-radius:14px 0 0 14px;padding:12px 10px;width:148px;'
        'box-shadow:0 4px 24px rgba(0,0,0,0.12),0 1px 4px rgba(0,0,0,0.06);'
        'border:1px solid #e2e8f0;">'
        '<div style="text-align:center;margin-bottom:8px;">'
        '<div style="font-size:1.4rem;font-weight:900;color:' + _pc + ';line-height:1;">' + str(_pct) + '%</div>'
        '<div style="font-size:0.62rem;color:#9ca3af;margin-top:1px;text-transform:uppercase;letter-spacing:0.05em;">Completado</div>'
        '</div>'
        '<div style="background:#f1f5f9;border-radius:99px;height:6px;margin-bottom:10px;overflow:hidden;">'
        '<div style="width:' + str(_pct) + '%;height:100%;border-radius:99px;'
        'background:linear-gradient(90deg,' + _pc + ',' + _pc + 'cc);transition:width 0.4s ease;"></div>'
        '</div>'
        '<div id="_prog_items" style="display:flex;flex-direction:column;gap:1px;">' + _items_html + '</div>'
        # Botón ocultar al fondo del panel
        '<div id="_prog_toggle" data-action="prog-toggle" style="margin-top:8px;text-align:center;'
        'cursor:pointer;font-size:0.65rem;color:#9ca3af;padding:3px 0;'
        'border-top:1px solid #f1f5f9;user-select:none;" title="Ocultar">'
        '› Ocultar</div>'
        '</div>'
        # Botón mini cuando está oculto
        '<div id="_prog_mini" style="display:none;position:fixed;right:0;top:50%;'
        'transform:translateY(-50%);z-index:99997;background:' + _pc + ';'
        'border-radius:10px 0 0 10px;padding:14px 8px;cursor:pointer;'
        'box-shadow:0 4px 20px rgba(0,0,0,0.2);text-align:center;width:54px;"'
        ' data-action="prog-show">'
        '<div style="font-size:1.15rem;font-weight:900;color:#fff;line-height:1;">' + str(_pct) + '%</div>'
        '<div style="font-size:0.7rem;color:rgba(255,255,255,0.85);margin-top:5px;">📊</div>'
        '<div style="font-size:0.58rem;font-weight:700;color:rgba(255,255,255,0.75);margin-top:3px;letter-spacing:0.06em;">VER</div>'
        '</div>'
    )
    st.markdown(_barra, unsafe_allow_html=True)
    # JS listener para toggle — en el mismo DOM
    import streamlit.components.v1 as _prog_comp
    _prog_comp.html("""<script>
(function(){
    var D = window.parent.document;
    function initToggle(){
        D.addEventListener('click', function(e){
            var el = e.target && e.target.closest ? e.target.closest('[data-action]') : null;
            if (!el) return;
            var action = el.getAttribute('data-action');
            var panel = D.getElementById('_prog_panel');
            var mini  = D.getElementById('_prog_mini');
            if (!panel || !mini) return;
            if (action === 'prog-toggle') {
                panel.style.display = 'none';
                mini.style.display  = 'block';
            } else if (action === 'prog-show') {
                panel.style.display = 'block';
                mini.style.display  = 'none';
            }
        });
    }
    setTimeout(initToggle, 300);
})();
</script>""", height=1)

# FAB - MARGEN FLOTANTE (st.popover nativo — 100% confiable)
# =========================================================
_margen_actual = st.session_state.margen
_mstr = f"{_margen_actual:.1f}"

if st.session_state.modo_admin:
    _color_fab = '#10b981' if _margen_actual > 0 else '#6b7280'
    st.markdown(f"""
<style>
/* Identificar el popover de margen por ser el ÚNICO popover fuera del header */
section[data-testid="stMain"] div[data-testid="stPopover"] {{
    position: fixed !important;
    bottom: 1.5rem !important;
    left: 12rem !important;
    z-index: 99998 !important;
}}
section[data-testid="stMain"] div[data-testid="stPopover"] > div > button {{
    background: linear-gradient(135deg, {_color_fab}, {_color_fab}dd) !important;
    color: white !important;
    border: none !important;
    border-radius: 50px !important;
    padding: 0.8rem 1.4rem !important;
    font-size: 0.9rem !important;
    font-weight: 700 !important;
    white-space: nowrap !important;
    box-shadow: 0 6px 20px rgba(16,185,129,.4) !important;
    min-height: unset !important;
    height: auto !important;
}}
</style>
""", unsafe_allow_html=True)

    with st.popover(f"📊 Margen: {_mstr}%"):
        st.markdown("**Aplicar margen**")
        _mg_pop = st.number_input(
            "Margen %", min_value=0.0, max_value=100.0,
            value=float(_margen_actual),
            step=0.5, format="%.1f",
            key="margen_popover"
        )
        if st.button("✅ Aplicar", key="btn_aplicar_margen", use_container_width=True):
            _margen_anterior = st.session_state.margen
            st.session_state.margen = _mg_pop
            # Notificar via Telegram
            try:
                _ep_notif = st.session_state.get('cotizacion_cargada', '')
                _cli_notif = st.session_state.get('nombre_input', '') or 'Cliente'
                # Obtener email ejecutivo — desde session_state o desde la BD
                _ej_email_notif = st.session_state.get('correo_asesor', '') or st.session_state.get('asesor_correo_temp', '')
                _ej_nombre_notif = st.session_state.get('asesor_seleccionado', '')
                # Si no hay email en session, buscarlo en la cotización cargada
                if not _ej_email_notif and _ep_notif:
                    try:
                        _cot_data = supabase_admin.table('cotizaciones').select('asesor_email','asesor_nombre').eq('numero', _ep_notif).execute()
                        if _cot_data.data:
                            _ej_email_notif = _cot_data.data[0].get('asesor_email', '')
                            _ej_nombre_notif = _ej_nombre_notif or _cot_data.data[0].get('asesor_nombre', '')
                    except:
                        pass
            except Exception as _ne:
                pass
            st.rerun()

else:
    components.html("""<script>
(function(){
  var D=window.parent.document;
  ['_fm_s','_fm_b','_fm_p'].forEach(function(id){
    var e=D.getElementById(id); if(e) e.remove();
  });
})();
</script>""", height=0)

# =========================================================
# TAB DASHBOARD — visible para todos
# =========================================================
if tab_dash is not None:
 with tab_dash:
    st.markdown("""
    <style>
    .dash-hdr {
        background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 50%, #2563eb 100%);
        border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
        display: flex; align-items: center; gap: 22px;
        box-shadow: 0 8px 32px rgba(37,99,235,0.25);
        position: relative; overflow: hidden;
    }
    .dash-hdr::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 180px; height: 180px; border-radius: 50%;
        background: rgba(255,255,255,0.04); pointer-events: none;
    }
    .dash-hdr::after {
        content: ''; position: absolute; bottom: -60px; right: 80px;
        width: 240px; height: 240px; border-radius: 50%;
        background: rgba(255,255,255,0.03); pointer-events: none;
    }
    .dash-hdr h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                 font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
    .dash-hdr p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
    </style>
    <div class="dash-hdr">
      <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">📊</span>
      <div>
        <h2>Dashboard</h2>
        <p>Resumen ejecutivo del rendimiento comercial en tiempo real.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)


    # ── CSS métricas dashboard ──
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@700;800;900&display=swap');
    .kpi-card {
        background: white; border-radius: 18px; padding: 22px 24px;
        border: 1px solid rgba(226,232,240,0.8);
        box-shadow: 0 4px 24px rgba(0,0,0,0.07), 0 1px 4px rgba(0,0,0,0.04);
        height: 100%; transition: transform 0.2s;
        position: relative; overflow: hidden;
    }
    .kpi-card::after {
        content: ''; position: absolute; top: 0; left: 0; right: 0;
        height: 3px; border-radius: 18px 18px 0 0;
        background: linear-gradient(90deg, #2563eb, #06b6d4);
    }
    .kpi-label { font-size: 0.72rem; font-weight: 800; color: #94a3b8;
                 text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 8px; }
    .kpi-value { font-size: 2.1rem; font-weight: 900; color: #0f172a;
                 font-family: 'Montserrat', sans-serif; line-height: 1; }
    .kpi-delta-pos { font-size: 0.75rem; font-weight: 700; color: #16a34a; margin-top: 8px;
                     display: flex; align-items: center; gap: 4px; }
    .kpi-delta-neg { font-size: 0.75rem; font-weight: 700; color: #dc2626; margin-top: 8px;
                     display: flex; align-items: center; gap: 4px; }
    .kpi-delta-neu { font-size: 0.75rem; font-weight: 600; color: #94a3b8; margin-top: 8px;
                     display: flex; align-items: center; gap: 4px; }
    .section-title {
        font-size: 0.78rem; font-weight: 900; color: #1e293b;
        text-transform: uppercase; letter-spacing: 0.1em;
        margin: 24px 0 14px; padding: 8px 16px;
        background: linear-gradient(90deg, rgba(37,99,235,0.07), transparent);
        border-left: 4px solid #2563eb; border-radius: 0 8px 8px 0;
    }
    .dash-panel {
        background: white; border-radius: 16px; padding: 20px 22px;
        border: 1px solid rgba(226,232,240,0.8);
        box-shadow: 0 4px 20px rgba(0,0,0,0.06);
    }
    .funnel-bar-wrap { background: #f1f5f9; border-radius: 10px; overflow: hidden; height: 10px; }
    .funnel-bar-inner { height: 10px; border-radius: 10px; }
    .cat-row { display: flex; align-items: center; gap: 12px; margin-bottom: 12px; }
    .cat-name { font-size: 0.82rem; font-weight: 700; color: #334155; min-width: 130px; }
    .cat-bar-wrap { flex: 1; background: #f1f5f9; border-radius: 6px; height: 8px; overflow: hidden; }
    .cat-bar-inner { height: 8px; border-radius: 6px; background: linear-gradient(90deg,#2563eb,#06b6d4); }
    .cat-monto { font-size: 0.8rem; font-weight: 800; color: #3b82f6; min-width: 70px; text-align: right; }
    .ej-row { display: flex; align-items: center; gap: 12px; margin-bottom: 12px; }
    .ej-pos { font-size: 1.2rem; min-width: 28px; }
    .ej-name { font-size: 0.85rem; font-weight: 700; color: #1e293b; flex: 1; }
    .ej-monto { font-size: 0.83rem; font-weight: 900; color: #2563eb; min-width: 80px; text-align: right; }
    .kpi-card .prod-divider { border: none; border-top: 1px solid #f1f5f9; margin: 10px 0; }
    </style>
    """, unsafe_allow_html=True)

    # ── Filtro período ──
    _periodo_opciones = {"Este mes": "mes", "Últimos 3 meses": "3meses", "Este año": "año", "Todos los tiempos": "todo"}
    _periodo_label = st.radio("Período", list(_periodo_opciones.keys()),
                               horizontal=True, index=0, key="dash_periodo",
                               label_visibility="collapsed")
    _periodo = _periodo_opciones[_periodo_label]

    with st.spinner("Cargando datos..."):
        _d = cargar_datos_dashboard(_periodo)

    if not _d:
        st.error("No se pudieron cargar los datos del dashboard.")
    else:
        def _fmt_monto(v):
            if v >= 1_000_000: return f"${v/1_000_000:.1f}M"
            if v >= 1_000:     return f"${v/1_000:.0f}K"
            return f"${v:,.0f}"

        def _delta_html(val, prefix=""):
            if val > 0:   return f'<div class="kpi-delta-pos">▲ {prefix}{abs(val):,} vs período ant.</div>'
            if val < 0:   return f'<div class="kpi-delta-neg">▼ {prefix}{abs(val):,} vs período ant.</div>'
            return f'<div class="kpi-delta-neu">— Sin cambio</div>'

        # ── KPIs principales ──
        st.markdown('<div class="section-title">Métricas clave</div>', unsafe_allow_html=True)
        k1, k2, k3, k4 = st.columns(4)
        kpis = [
            (k1, "💼 Presupuestos", str(_d['total_ep']),
             _delta_html(_d['delta_ep'])),
            (k2, "💰 Monto total", _fmt_monto(_d['total_monto']),
             _delta_html(int(_d['delta_monto']), prefix="$")),
            (k3, "📈 Ticket promedio", _fmt_monto(_d['promedio_monto']),
             '<div class="kpi-delta-neu">por cotización</div>'),
            (k4, "🔄 Pipeline", _fmt_monto(_d['pipeline']),
             '<div class="kpi-delta-neu">borradores activos</div>'),
        ]
        for col, label, val, delta in kpis:
            with col:
                st.markdown(f"""
                <div class="kpi-card">
                  <div class="kpi-label">{label}</div>
                  <div class="kpi-value">{val}</div>
                  {delta}
                </div>""", unsafe_allow_html=True)

        st.markdown("")

        # ── Embudo de conversión ──
        st.markdown('<div class="section-title">Embudo de conversión</div>', unsafe_allow_html=True)
        _total_ep = _d['total_ep'] or 1
        _funnel_data = [
            ("🟢 Autorizados", _d['autorizados'], "#16a34a"),
            ("🟠 Borradores",  _d['borradores'],  "#f59e0b"),
            ("🔴 Incompletos", _d['incompletos'],  "#ef4444"),
        ]
        col_funnel, col_donut = st.columns([3, 2])
        with col_funnel:
            _funnel_html = """
            <div class="dash-panel">
              <div style="display:flex;justify-content:space-between;margin-bottom:20px;">
                <span style="font-size:0.78rem;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:0.08em;">Estado</span>
                <span style="font-size:0.78rem;font-weight:700;color:#94a3b8;text-transform:uppercase;letter-spacing:0.08em;">% del total</span>
              </div>"""
            for label_f, count_f, color_f in _funnel_data:
                pct_f = round((count_f / _total_ep) * 100) if _total_ep else 0
                _funnel_html += f"""
                <div style="margin-bottom:16px;">
                  <div style="display:flex;justify-content:space-between;margin-bottom:5px;">
                    <span style="font-size:0.85rem;font-weight:700;color:#1e293b;">{label_f}</span>
                    <span style="font-size:0.85rem;font-weight:800;color:{color_f};">{count_f} &nbsp;({pct_f}%)</span>
                  </div>
                  <div class="funnel-bar-wrap">
                    <div class="funnel-bar-inner" style="width:{pct_f}%;background:{color_f};opacity:0.85;"></div>
                  </div>
                </div>"""
            _funnel_html += "</div>"
            st.markdown(_funnel_html, unsafe_allow_html=True)

        with col_donut:
            with st.container(border=True):
                if _d['autorizados'] + _d['borradores'] + _d['incompletos'] > 0:
                    import plotly.graph_objects as go
                    _fig_donut = go.Figure(go.Pie(
                        labels=["Autorizados", "Borradores", "Incompletos"],
                        values=[_d['autorizados'], _d['borradores'], _d['incompletos']],
                        hole=0.62,
                        marker=dict(colors=["#16a34a", "#f59e0b", "#ef4444"],
                                    line=dict(color='white', width=3)),
                        textinfo='percent',
                        textfont=dict(size=11, family='Montserrat', color='white'),
                        hovertemplate='<b>%{label}</b><br>%{value} cotizaciones<br>%{percent}<extra></extra>',
                    ))
                    _fig_donut.add_annotation(
                        text=f"<b>{_d['pct_conv']}%</b><br><span style='font-size:10px'>conv.</span>",
                        x=0.5, y=0.5, showarrow=False, font=dict(size=18, family='Montserrat'),
                        xref="paper", yref="paper", align="center"
                    )
                    _fig_donut.update_layout(
                        showlegend=True, margin=dict(t=16, b=16, l=16, r=16),
                        height=240, paper_bgcolor='rgba(0,0,0,0)',
                        plot_bgcolor='rgba(0,0,0,0)',
                        legend=dict(font=dict(size=11, color='#475569'),
                                    orientation='h', yanchor='bottom',
                                    y=-0.18, xanchor='center', x=0.5,
                                    bgcolor='rgba(0,0,0,0)'),
                    )
                    st.plotly_chart(_fig_donut, use_container_width=True, config={'displayModeBar': False})

        # ── Evolución temporal ──
        if _d['fechas']:
            st.markdown('<div class="section-title">Evolución de cotizaciones</div>', unsafe_allow_html=True)
            with st.container(border=True):
                import plotly.graph_objects as go
                _fig_line = go.Figure()
                _fig_line.add_trace(go.Bar(
                    x=_d['fechas'], y=_d['serie_counts'],
                    name='Nº EP',
                    marker=dict(color='rgba(99,102,241,0.25)', line=dict(width=0)),
                    yaxis='y2',
                    hovertemplate='<b>%{x}</b><br>%{y} EP<extra></extra>',
                ))
                _fig_line.add_trace(go.Scatter(
                    x=_d['fechas'], y=_d['serie_montos'],
                    mode='lines+markers',
                    name='Monto ($)',
                    line=dict(color='#2563eb', width=3.5, shape='spline'),
                    marker=dict(size=8, color='#2563eb',
                                line=dict(color='white', width=2.5)),
                    fill='tozeroy',
                    fillcolor='rgba(37,99,235,0.07)',
                    hovertemplate='<b>%{x}</b><br>$%{y:,.0f}<extra></extra>',
                ))
                _fig_line.update_layout(
                    height=320, margin=dict(t=20, b=40, l=70, r=60),
                    paper_bgcolor='rgba(0,0,0,0)',
                    plot_bgcolor='rgba(248,250,252,0.6)',
                    xaxis=dict(showgrid=False, tickfont=dict(size=10, color='#64748b'),
                               linecolor='#e2e8f0'),
                    yaxis=dict(showgrid=True, gridcolor='rgba(226,232,240,0.6)',
                               tickformat='$,.0f', tickfont=dict(size=10, color='#64748b'),
                               zeroline=False),
                    yaxis2=dict(overlaying='y', side='right', showgrid=False,
                                tickfont=dict(size=10, color='#94a3b8'), title=''),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                xanchor='right', x=1, font=dict(size=11),
                                bgcolor='rgba(0,0,0,0)'),
                    hovermode='x unified',
                    hoverlabel=dict(bgcolor='white', font_size=12,
                                    bordercolor='#e2e8f0'),
                )
                st.plotly_chart(_fig_line, use_container_width=True, config={'displayModeBar': False})

        # ── Top categorías + Top ejecutivos ──
        col_cats, col_ejs = st.columns(2)

        with col_cats:
            st.markdown('<div class="section-title">Top categorías</div>', unsafe_allow_html=True)
            if _d['top_cats']:
                _max_cat = _d['top_cats'][0][1] or 1
                html_cats = '<div class="dash-panel">'
                for cat_name, cat_val in _d['top_cats']:
                    pct_c = round((cat_val / _max_cat) * 100)
                    html_cats += f"""
                    <div class="cat-row">
                      <div class="cat-name">{cat_name[:18]}</div>
                      <div class="cat-bar-wrap"><div class="cat-bar-inner" style="width:{pct_c}%;"></div></div>
                      <div class="cat-monto">{_fmt_monto(cat_val)}</div>
                    </div>"""
                html_cats += '</div>'
                st.markdown(html_cats, unsafe_allow_html=True)
            else:
                st.info("Sin datos de categorías.")

        with col_ejs:
            st.markdown('<div class="section-title">Top ejecutivos</div>', unsafe_allow_html=True)
            if _d['top_ej']:
                _medallas_d = {0:"🥇",1:"🥈",2:"🥉",3:"4️⃣",4:"5️⃣"}
                html_ejs = '<div class="dash-panel">'
                for idx_e, (ej_name, ej_val) in enumerate(_d['top_ej']):
                    ej_count = next((v for k,v in zip(
                        [e[0] for e in _d['top_ej']],
                        [sum(1 for r in [] if True)]
                    ) if k == ej_name), '—')
                    html_ejs += f"""
                    <div class="ej-row">
                      <div class="ej-pos">{_medallas_d.get(idx_e, str(idx_e+1))}</div>
                      <div class="ej-name">{ej_name[:22]}</div>
                      <div class="ej-monto">{_fmt_monto(ej_val)}</div>
                    </div>"""
                html_ejs += '</div>'
                st.markdown(html_ejs, unsafe_allow_html=True)
            else:
                st.info("Sin datos de ejecutivos.")


        st.markdown('<div class="section-title">👥 Perfil de Clientes</div>', unsafe_allow_html=True)
        _tc = _d.get('top_comunas', [])
        _tr = _d.get('top_regiones', [])
        _nn = _d.get('n_natural', 0)
        _nj = _d.get('n_juridica', 0)
        _nm = _d.get('n_masc', 0)
        _nf = _d.get('n_fem', 0)
        _nd_g = _d.get('n_nd', 0)
        _re = _d.get('rangos_etarios', {})
        _te = _d.get('top_empresas', [])

        import plotly.graph_objects as go
        import plotly.express as px

        _TMPL = dict(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(family='Inter, sans-serif', color='#1e293b'),
            margin=dict(l=10, r=10, t=35, b=10),
        )

        # ── Fila 1: Comunas + Regiones ──────────────────────
        col_com, col_reg = st.columns(2)

        with col_com:
            st.markdown('<div class="section-title">📍 Top Comunas</div>', unsafe_allow_html=True)
            with st.container(border=True):
                if _tc:
                    _coms   = [x[0] for x in _tc]
                    _vals_c = [x[1] for x in _tc]
                    _fig_com = go.Figure(go.Bar(
                        x=_vals_c[::-1], y=_coms[::-1], orientation='h',
                        marker=dict(color=_vals_c[::-1],
                                    colorscale=[[0,'#bfdbfe'],[1,'#1d4ed8']],
                                    showscale=False),
                        text=[str(v) for v in _vals_c[::-1]], textposition='outside',
                        hovertemplate='<b>%{y}</b><br>%{x} cotizaciones<extra></extra>',
                    ))
                    _fig_com.update_layout(**_TMPL,
                        xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
                        yaxis=dict(tickfont=dict(size=11)), height=320)
                    st.plotly_chart(_fig_com, use_container_width=True, config={'displayModeBar': False})
                else:
                    st.info("Sin datos de comunas aún")

        with col_reg:
            st.markdown('<div class="section-title">🗺️ Top Regiones</div>', unsafe_allow_html=True)
            with st.container(border=True):
                if _tr:
                    _regs   = [x[0] for x in _tr]
                    _vals_r = [x[1] for x in _tr]
                    _fig_reg = go.Figure(go.Bar(
                        x=_vals_r[::-1], y=_regs[::-1], orientation='h',
                        marker=dict(color=_vals_r[::-1],
                                    colorscale=[[0,'#bbf7d0'],[1,'#15803d']],
                                    showscale=False),
                        text=[str(v) for v in _vals_r[::-1]], textposition='outside',
                        hovertemplate='<b>%{y}</b><br>%{x} cotizaciones<extra></extra>',
                    ))
                    _fig_reg.update_layout(**_TMPL,
                        xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
                        yaxis=dict(tickfont=dict(size=11)), height=320)
                    st.plotly_chart(_fig_reg, use_container_width=True, config={'displayModeBar': False})
                else:
                    st.info("Sin datos de regiones aún")

        # ── Fila 2: Tipo cliente + Género + Rango etario ────
        col_tipo, col_gen, col_edad = st.columns(3)

        with col_tipo:
            st.markdown('<div class="section-title">🏢 Tipo Cliente</div>', unsafe_allow_html=True)
            with st.container(border=True):
                _total_tipo = _nn + _nj
                if _total_tipo > 0:
                    _fig_tipo = go.Figure(go.Pie(
                        labels=['Persona Natural', 'Persona Jurídica'],
                        values=[_nn, _nj], hole=0.55,
                        marker=dict(colors=['#3b82f6','#f59e0b'],
                                    line=dict(color='white', width=2)),
                        textinfo='percent',
                        hovertemplate='<b>%{label}</b><br>%{value} (%{percent})<extra></extra>',
                    ))
                    _fig_tipo.add_annotation(text=f"<b>{_total_tipo}</b><br>clientes",
                        x=0.5, y=0.5, showarrow=False,
                        font=dict(size=13, color='#0f172a'), align='center')
                    _fig_tipo.update_layout(**_TMPL,
                        showlegend=True,
                        legend=dict(orientation='h', y=-0.15, font=dict(size=9)),
                        height=280)
                    st.plotly_chart(_fig_tipo, use_container_width=True, config={'displayModeBar': False})
                else:
                    st.info("Sin datos")

        with col_gen:
            st.markdown('<div class="section-title">⚤ Género Estimado</div>', unsafe_allow_html=True)
            with st.container(border=True):
                _total_gen = _nm + _nf + _nd_g
                if _total_gen > 0:
                    _lbl_g, _val_g, _col_g = [], [], []
                    if _nm: _lbl_g.append('Masculino'); _val_g.append(_nm); _col_g.append('#3b82f6')
                    if _nf: _lbl_g.append('Femenino');  _val_g.append(_nf); _col_g.append('#ec4899')
                    if _nd_g: _lbl_g.append('No determinado'); _val_g.append(_nd_g); _col_g.append('#94a3b8')
                    _fig_gen = go.Figure(go.Pie(
                        labels=_lbl_g, values=_val_g, hole=0.55,
                        marker=dict(colors=_col_g, line=dict(color='white', width=2)),
                        textinfo='percent',
                        hovertemplate='<b>%{label}</b><br>%{value} (%{percent})<extra></extra>',
                    ))
                    _fig_gen.add_annotation(text=f"<b>{_nm+_nf}</b><br>detect.",
                        x=0.5, y=0.5, showarrow=False,
                        font=dict(size=13, color='#0f172a'), align='center')
                    _fig_gen.update_layout(**_TMPL,
                        showlegend=True,
                        legend=dict(orientation='h', y=-0.15, font=dict(size=9)),
                        height=280)
                    st.plotly_chart(_fig_gen, use_container_width=True, config={'displayModeBar': False})
                    st.caption("⚠️ Estimado por primer nombre")
                else:
                    st.info("Sin datos")

        with col_edad:
            st.markdown('<div class="section-title">📅 Rango Etario Est.</div>', unsafe_allow_html=True)
            with st.container(border=True):
                _re_f = {k: v for k, v in _re.items() if v > 0}
                if _re_f:
                    _orden_e = ['< 1975 (50+)', '1975-1995 (30-50)', '> 1995 (< 30)', 'No determinado']
                    _lbl_e = [k for k in _orden_e if k in _re_f]
                    _val_e = [_re_f[k] for k in _lbl_e]
                    _col_e = ['#7c3aed','#2563eb','#0891b2','#94a3b8'][:len(_lbl_e)]
                    _fig_edad = go.Figure(go.Bar(
                        x=_val_e, y=_lbl_e,
                        orientation='h',
                        marker=dict(color=_col_e, line=dict(color='white', width=1)),
                        text=_val_e, textposition='outside',
                        hovertemplate='<b>%{y}</b><br>%{x} clientes<extra></extra>',
                    ))
                    _fig_edad.update_layout(**_TMPL,
                        xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
                        yaxis=dict(tickfont=dict(size=10), automargin=True),
                        height=280)
                    st.plotly_chart(_fig_edad, use_container_width=True, config={'displayModeBar': False})
                    st.caption("⚠️ Estimado por correlación RUT")
                else:
                    st.info("Sin datos")

        # ── Fila 3: Top empresas ─────────────────────────────
        if _te:
            st.markdown('<div class="section-title">🏢 Top Empresas Cotizantes</div>', unsafe_allow_html=True)
            with st.container(border=True):
                _emp_n = [x[0] for x in _te]
                _emp_v = [x[1] for x in _te]
                _fig_emp = go.Figure(go.Bar(
                    x=_emp_v[::-1], y=_emp_n[::-1], orientation='h',
                    marker=dict(color=_emp_v[::-1],
                                colorscale=[[0,'#fde68a'],[1,'#d97706']],
                                showscale=False),
                    text=[str(v) for v in _emp_v[::-1]], textposition='outside',
                    hovertemplate='<b>%{y}</b><br>%{x} cotizaciones<extra></extra>',
                ))
                _fig_emp.update_layout(**_TMPL,
                    xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
                    yaxis=dict(tickfont=dict(size=11)),
                    height=max(200, len(_te) * 38))
                st.plotly_chart(_fig_emp, use_container_width=True, config={'displayModeBar': False})





        # ── Top 30 productos ──
        st.markdown('<div class="section-title">🏅 Top 30 productos más cotizados</div>', unsafe_allow_html=True)
        if _d.get('top_productos'):
            _max_prod = _d['top_productos'][0][1] or 1
            _html_prods = '<div class="dash-panel">'
            _html_prods += """
            <div style="display:flex;gap:8px;margin-bottom:14px;padding-bottom:10px;
                        border-bottom:1px solid #f1f5f9;">
              <span style="font-size:0.72rem;font-weight:800;color:#94a3b8;
                           text-transform:uppercase;letter-spacing:0.08em;min-width:24px;">#</span>
              <span style="font-size:0.72rem;font-weight:800;color:#94a3b8;
                           text-transform:uppercase;letter-spacing:0.08em;flex:1;">Producto</span>
              <span style="font-size:0.72rem;font-weight:800;color:#94a3b8;
                           text-transform:uppercase;letter-spacing:0.08em;min-width:55px;text-align:center;">Cant.</span>
              <span style="font-size:0.72rem;font-weight:800;color:#94a3b8;
                           text-transform:uppercase;letter-spacing:0.08em;min-width:100px;text-align:right;">Monto</span>
            </div>"""
            for idx_p, (prod_name, prod_val, prod_qty, prod_cat) in enumerate(_d['top_productos'], 1):
                pct_p    = round((prod_val / _max_prod) * 100)
                _color_p = "#3b82f6" if idx_p <= 3 else "#6366f1" if idx_p <= 10 else "#94a3b8"
                _bold_p  = "800" if idx_p <= 3 else "600"
                _html_prods += f"""
                <div style="display:flex;align-items:center;gap:8px;margin-bottom:9px;">
                  <span style="font-size:0.78rem;font-weight:700;color:{_color_p};
                               min-width:24px;text-align:center;">{idx_p}</span>
                  <div style="flex:1;min-width:0;">
                    <div style="font-size:0.82rem;font-weight:{_bold_p};color:#1e293b;
                                white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">{prod_name}</div>
                    <div style="font-size:0.72rem;color:#94a3b8;margin-bottom:3px;">{prod_cat}</div>
                    <div style="background:#f1f5f9;border-radius:4px;height:5px;overflow:hidden;">
                      <div style="width:{pct_p}%;height:5px;border-radius:4px;
                                  background:{_color_p};opacity:0.7;"></div>
                    </div>
                  </div>
                  <span style="font-size:0.8rem;font-weight:700;color:#475569;
                               min-width:55px;text-align:center;">{prod_qty:,}</span>
                  <span style="font-size:0.82rem;font-weight:800;color:{_color_p};
                               min-width:100px;text-align:right;">{_fmt_monto(prod_val)}</span>
                </div>"""
            _html_prods += '</div>'
            st.markdown(_html_prods, unsafe_allow_html=True)
        else:
            st.info("Sin datos de productos para el período seleccionado.")



            st.caption("Score = 60% total generado + 25% % conversión + 15% cantidad de presupuestos")

        st.caption(f"Datos actualizados al abrir la pestaña · Período: {_periodo_label}")


# =========================================================
# TAB 6 - EDICIÓN PDF (visible para todos)
# =========================================================
if tab6 is not None:
    with tab6:
        st.markdown("""
        <style>
        .hdr6 {
            background: linear-gradient(135deg, #b91c1c 0%, #dc2626 100%);
            border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
            display: flex; align-items: center; gap: 22px;
            box-shadow: 0 8px 32px rgba(220,38,38,0.25);
            position: relative; overflow: hidden;
        }
        .hdr6::before {
            content: ''; position: absolute; top: -40px; right: -40px;
            width: 180px; height: 180px; border-radius: 50%;
            background: rgba(255,255,255,0.04); pointer-events: none;
        }
        .hdr6::after {
            content: ''; position: absolute; bottom: -60px; right: 80px;
            width: 240px; height: 240px; border-radius: 50%;
            background: rgba(255,255,255,0.03); pointer-events: none;
        }
        .hdr6 h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                     font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
        .hdr6 p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
        </style>
        <div class="hdr6">
          <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">✏️</span>
          <div>
            <h2>Edición PDF Cliente</h2>
            <p>Busca tu cotización por número EP y personaliza la descripción de cada categoría para el cliente.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # Buscar cotización por EP
        with st.container(border=True):
            st.markdown("#### 🔍 Buscar cotización")
            col_ep, col_btn = st.columns([3, 1])
            with col_ep:
                _ep_buscar = st.text_input("Número EP", placeholder="Ej: EP-22286",
                                           key="pdf_edit_ep_input",
                                           label_visibility="collapsed")
            with col_btn:
                _btn_buscar_ep = st.button("🔍 Buscar", use_container_width=True,
                                           key="pdf_edit_btn_buscar", type="primary")

        # Estado de cotización cargada para edición
        if 'pdf_edit_cotizacion' not in st.session_state:
            st.session_state.pdf_edit_cotizacion = None
        if 'pdf_edit_numero' not in st.session_state:
            st.session_state.pdf_edit_numero = None

        if _btn_buscar_ep and _ep_buscar.strip():
            _cot_found = cargar_cotizacion(_ep_buscar.strip().upper())
            if _cot_found:
                st.session_state.pdf_edit_cotizacion = _cot_found
                st.session_state.pdf_edit_numero = _ep_buscar.strip().upper()
                st.success(f"✅ Cotización {st.session_state.pdf_edit_numero} encontrada — {_cot_found.get('cliente_nombre','S/C')}")
            else:
                st.error("❌ No se encontró la cotización. Verifica el número EP.")
                st.session_state.pdf_edit_cotizacion = None
                st.session_state.pdf_edit_numero = None

        # Si hay cotización cargada, mostrar editor
        if st.session_state.pdf_edit_cotizacion and st.session_state.pdf_edit_numero:
            _cot_edit = st.session_state.pdf_edit_cotizacion
            _num_edit = st.session_state.pdf_edit_numero

            # Info de la cotización
            st.markdown(f"""
            <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:10px;
                        padding:12px 16px;margin:12px 0;">
                <b>📋 {_num_edit}</b> — {_cot_edit.get('cliente_nombre','S/C')} &nbsp;|&nbsp;
                Asesor: {_cot_edit.get('asesor_nombre','—')} &nbsp;|&nbsp;
                Fecha: {(_cot_edit.get('fecha_creacion','')[:10])}
            </div>
            """, unsafe_allow_html=True)

            # Cargar descripciones desde Storage solo la primera vez que se abre este EP
            _init_key = f"_desc_init_{_num_edit}"
            if _init_key not in st.session_state:
                _desc_storage = cargar_descripciones_por_ep(_num_edit, bust_cache=True)
                # Inicializar cada text_area con el valor de Storage
                for _k, _v in _desc_storage.items():
                    _ta_key = f"pdf_edit_desc_{_num_edit}_{_k}"
                    if _ta_key not in st.session_state:
                        st.session_state[_ta_key] = _v
                st.session_state[_init_key] = True

            # Obtener categorías de los productos de esta cotización
            _productos = _cot_edit.get('productos', [])
            if _productos:
                _cats_ep = sorted(list({p.get('Categoria','') for p in _productos if p.get('Categoria','')}))
            else:
                _cats_ep = []

            if not _cats_ep:
                st.warning("Esta cotización no tiene productos con categorías definidas.")
            else:
                st.markdown(f"#### 📝 Editar descripciones ({len(_cats_ep)} categorías)")
                st.caption("Escribe la descripción que verá el cliente en el PDF. Si la dejas vacía, se mostrarán los ítems del carrito.")

                _desc_editadas = {}
                for _cat in _cats_ep:
                    _key_ta = f"pdf_edit_desc_{_num_edit}_{_cat}"
                    _tiene_desc = bool(st.session_state.get(_key_ta, '').strip())
                    with st.container(border=True):
                        col_cat, col_estado, col_limpiar_uno = st.columns([3, 1, 1])
                        with col_cat:
                            st.markdown(f"**{_cat}**")
                        with col_estado:
                            if _tiene_desc:
                                st.markdown("🟣 Personalizada")
                            else:
                                st.markdown("⬜ Por defecto")
                        with col_limpiar_uno:
                            if _tiene_desc:
                                if st.button("🗑️ Limpiar", key=f"pdf_limpiar_{_num_edit}_{_cat}",
                                             use_container_width=True):
                                    # Borrar widget y guardar en Storage sin esta categoría
                                    st.session_state[_key_ta] = ''
                                    _dict_sin = {
                                        _c: st.session_state.get(f"pdf_edit_desc_{_num_edit}_{_c}", '')
                                        for _c in _cats_ep
                                        if _c != _cat and st.session_state.get(f"pdf_edit_desc_{_num_edit}_{_c}", '').strip()
                                    }
                                    guardar_descripciones_por_ep(_num_edit, _dict_sin)
                                    st.rerun()

                        _nueva = st.text_area(
                            f"Descripción para {_cat}",
                            height=80,
                            placeholder=f"Ej: Incluye todos los elementos de {_cat.lower()}...",
                            key=_key_ta,
                            label_visibility="collapsed"
                        )
                        _desc_editadas[_cat] = _nueva

                st.markdown("")
                col_guardar, col_limpiar = st.columns([2, 1])
                with col_guardar:
                    if st.button("💾 Guardar todas las descripciones", type="primary",
                                 use_container_width=True, key="pdf_edit_guardar_todo"):
                        _dict_final = {k: v.strip() for k, v in _desc_editadas.items() if v.strip()}
                        if guardar_descripciones_por_ep(_num_edit, _dict_final):
                            st.success("✅ Descripciones guardadas. Se usarán al generar el PDF cliente.")
                            # Limpiar init para que recargue desde Storage al reabrir
                            if _init_key in st.session_state:
                                del st.session_state[_init_key]
                            st.session_state.pdf_edit_cotizacion = None
                            st.session_state.pdf_edit_numero = None
                            st.rerun()

                with col_limpiar:
                    if st.button("🗑️ Limpiar todas", use_container_width=True,
                                 key="pdf_edit_limpiar_todo"):
                        guardar_descripciones_por_ep(_num_edit, {})
                        # Limpiar todos los widgets
                        for _c in _cats_ep:
                            _kw = f"pdf_edit_desc_{_num_edit}_{_c}"
                            if _kw in st.session_state:
                                del st.session_state[_kw]
                        if _init_key in st.session_state:
                            del st.session_state[_init_key]
                        st.session_state.pdf_edit_cotizacion = None
                        st.session_state.pdf_edit_numero = None
                        st.rerun()
        else:
            st.info("🔍 Ingresa el número EP y presiona Buscar para editar las descripciones de una cotización.")

# =========================================================
# TAB 7 - RANKING EJECUTIVOS (visible para todos)
# =========================================================
if tab7 is not None:
    with tab7:
        st.markdown("""
        <style>
        .hdr7 {
            background: linear-gradient(135deg, #78350f 0%, #d97706 100%);
            border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
            display: flex; align-items: center; gap: 22px;
            box-shadow: 0 8px 32px rgba(217,119,6,0.25);
            position: relative; overflow: hidden;
        }
        .hdr7 h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                     font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
        .hdr7 p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
        .rank-chart-box {
            background: #ffffff; border-radius: 18px; padding: 22px 24px;
            border: 1px solid #e8edf5; margin-bottom: 20px;
            box-shadow: 0 6px 32px rgba(0,0,0,0.08), 0 2px 8px rgba(0,0,0,0.04);
        }
        /* Sombra solo en containers del tab ranking */
        .hdr7 ~ div [data-testid="stVerticalBlockBorderWrapper"],
        .hdr7 + div [data-testid="stVerticalBlockBorderWrapper"] {
            box-shadow: 0 6px 32px rgba(0,0,0,0.08), 0 2px 8px rgba(0,0,0,0.04) !important;
            border-radius: 18px !important;
            border: 1px solid #e8edf5 !important;
            transition: box-shadow 0.2s, transform 0.2s !important;
        }
        .hdr7 ~ div [data-testid="stVerticalBlockBorderWrapper"]:hover {
            box-shadow: 0 10px 40px rgba(0,0,0,0.13) !important;
            transform: translateY(-1px) !important;
        }
        /* Centrar verticalmente columnas dentro de los containers del ranking */
        .hdr7 ~ div [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stHorizontalBlock"] {
            align-items: center !important;
        }
        .rank-chart-title {
            font-size: 0.78rem; font-weight: 800; color: #64748b;
            text-transform: uppercase; letter-spacing: 0.08em;
            margin-bottom: 14px; padding-bottom: 12px;
            border-bottom: 2px solid #f1f5f9;
        }

        .rank-kpi {
            background: #ffffff; border-radius: 16px; padding: 22px 20px;
            border: 1px solid #e8edf5; text-align: center;
            box-shadow: 0 6px 32px rgba(0,0,0,0.08), 0 2px 8px rgba(0,0,0,0.04);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .rank-kpi:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 40px rgba(0,0,0,0.12);
        }
        /* Centrar verticalmente columnas card ejecutivo */
        .hdr7 ~ div [data-testid="stVerticalBlockBorderWrapper"] [data-testid="column"] {
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
        }
        .hdr7 ~ div [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stMarkdownContainer"] {
            display: flex !important;
            align-items: center !important;
            height: 100% !important;
        }
        .rank-kpi-label { font-size: 0.72rem; font-weight: 700; color: #94a3b8;
                          text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 8px; }
        .rank-kpi-value { font-size: 2rem; font-weight: 900; color: #0f172a;
                          font-family: 'Montserrat', sans-serif; line-height: 1; }
        .rank-section {
            font-size: 0.75rem; font-weight: 900; color: #1e293b;
            text-transform: uppercase; letter-spacing: 0.1em;
            margin: 24px 0 14px; padding: 8px 16px;
            background: linear-gradient(90deg, rgba(217,119,6,0.1), transparent);
            border-left: 4px solid #d97706; border-radius: 0 8px 8px 0;
        }
        .rank-card {
            background: #ffffff; border-radius: 18px; padding: 20px 24px;
            border: 1px solid #e2e8f0; margin-bottom: 0px;
            box-shadow: 0 6px 32px rgba(0,0,0,0.08), 0 2px 8px rgba(0,0,0,0.04);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .rank-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 10px 40px rgba(0,0,0,0.12);
        }
        .rank-1 { border-left: 5px solid #f59e0b; background: linear-gradient(135deg, #fffbeb 0%, #ffffff 40%); }
        .rank-2 { border-left: 5px solid #94a3b8; background: linear-gradient(135deg, #f8fafc 0%, #ffffff 40%); }
        .rank-3 { border-left: 5px solid #b45309; background: linear-gradient(135deg, #fef3c7 0%, #ffffff 40%); }
        .rank-other { border-left: 5px solid #e2e8f0; }
        </style>
        <div class="hdr7">
          <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">🏆</span>
          <div>
            <h2>Ranking de Ejecutivos</h2>
            <p>Desempeño del equipo de ventas — este mes.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        with st.spinner("Cargando ranking..."):
            _ranking = cargar_ranking_ejecutivos(periodo='mes')

        if not _ranking:
            st.info("No hay cotizaciones registradas este mes.")
        else:
            import plotly.graph_objects as go
            from datetime import datetime as _dt

            def _fmt_r(v):
                if v >= 1_000_000: return f"${v/1_000_000:.1f}M"
                if v >= 1_000:     return f"${v/1_000:.0f}K"
                return f"${v:,.0f}"

            _mes_actual   = _dt.now().strftime("%B %Y").capitalize()
            _total_mes    = sum(a['total_generado'] for a in _ranking)
            _total_presup = sum(a['total_presupuestos'] for a in _ranking)
            _total_autori = sum(a['autorizados'] for a in _ranking)
            _pct_g        = round((_total_autori / _total_presup) * 100) if _total_presup else 0
            _n_ejs        = len(_ranking)
            _h_bar        = max(280, 52 * _n_ejs)
            _nombres      = [a['nombre'].split()[0] for a in _ranking]
            _colores      = ["#f59e0b" if i==0 else "#94a3b8" if i==1 else "#cd7c3a" if i==2 else "#cbd5e1"
                             for i in range(_n_ejs)]

            # ── KPIs globales ──
            st.markdown(f'<div class="rank-section">📅 {_mes_actual} · {_n_ejs} ejecutivo{"s" if _n_ejs!=1 else ""}</div>', unsafe_allow_html=True)
            rk1, rk2, rk3, rk4 = st.columns(4)
            for _col, _lbl, _val in [
                (rk1, "💰 Total generado", _fmt_r(_total_mes)),
                (rk2, "📋 Presupuestos",   str(_total_presup)),
                (rk3, "🟢 Autorizados",    str(_total_autori)),
                (rk4, "📈 % Conversión",   f"{_pct_g}%"),
            ]:
                with _col:
                    with st.container(border=True):
                        st.markdown(f'''<div style="text-align:center;padding:8px 4px;">
                          <div class="rank-kpi-label">{_lbl}</div>
                          <div class="rank-kpi-value">{_val}</div>
                        </div>''', unsafe_allow_html=True)

            st.markdown("")

            # ── Gráfico barras HORIZONTALES: monto generado ──
            st.markdown('<div class="rank-section">💰 Monto generado por ejecutivo</div>', unsafe_allow_html=True)
            _fig_bar = go.Figure(go.Bar(
                y=_nombres[::-1],
                x=[a['total_generado'] for a in _ranking[::-1]],
                orientation='h',
                marker_color=_colores[::-1],
                marker_line_width=0,
                text=[_fmt_r(a['total_generado']) for a in _ranking[::-1]],
                textposition='outside',
                textfont=dict(size=11, family='Montserrat', color='#1e293b'),
                hovertemplate='<b>%{y}</b><br>%{text}<extra></extra>',
            ))
            _fig_bar.update_layout(
                height=_h_bar, margin=dict(t=10, b=10, l=10, r=90),
                paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(showgrid=True, gridcolor='#f1f5f9', visible=False),
                yaxis=dict(showgrid=False, tickfont=dict(size=12, family='Montserrat')),
                showlegend=False,
            )
            with st.container(border=True):
                st.markdown('<div class="rank-chart-title">💰 Monto total generado</div>', unsafe_allow_html=True)
                st.plotly_chart(_fig_bar, use_container_width=True, config={'displayModeBar': False})

            # ── Gráficos HORIZONTALES: presupuestos + conversión ──
            st.markdown('<div class="rank-section">📊 Presupuestos y conversión por ejecutivo</div>', unsafe_allow_html=True)
            col_bars, col_conv = st.columns([3, 2])

            with col_bars:
                _fig_combo = go.Figure()
                _fig_combo.add_trace(go.Bar(
                    name='Total EP', y=_nombres[::-1],
                    x=[a['total_presupuestos'] for a in _ranking[::-1]],
                    orientation='h',
                    marker_color='rgba(99,102,241,0.75)',
                    text=[a['total_presupuestos'] for a in _ranking[::-1]],
                    textposition='auto',
                    hovertemplate='<b>%{y}</b><br>%{x} EP<extra></extra>',
                ))
                _fig_combo.add_trace(go.Bar(
                    name='Autorizados', y=_nombres[::-1],
                    x=[a['autorizados'] for a in _ranking[::-1]],
                    orientation='h',
                    marker_color='rgba(22,163,74,0.8)',
                    text=[a['autorizados'] for a in _ranking[::-1]],
                    textposition='auto',
                    hovertemplate='<b>%{y}</b><br>%{x} autorizados<extra></extra>',
                ))
                _fig_combo.update_layout(
                    barmode='group', height=_h_bar,
                    margin=dict(t=10, b=10, l=10, r=20),
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=True, gridcolor='#f1f5f9', tickfont=dict(size=10)),
                    yaxis=dict(showgrid=False, tickfont=dict(size=11)),
                    legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                xanchor='right', x=1, font=dict(size=11)),
                )
                with st.container(border=True):
                    st.markdown('<div class="rank-chart-title">📋 EP totales vs autorizados</div>', unsafe_allow_html=True)
                    st.plotly_chart(_fig_combo, use_container_width=True, config={'displayModeBar': False})

            with col_conv:
                _fig_conv = go.Figure(go.Bar(
                    y=_nombres[::-1],
                    x=[a['pct_autorizado'] for a in _ranking[::-1]],
                    orientation='h',
                    marker=dict(
                        color=[a['pct_autorizado'] for a in _ranking[::-1]],
                        colorscale=[[0,'#ef4444'],[0.5,'#f59e0b'],[1,'#16a34a']],
                        showscale=False,
                    ),
                    text=[f"{a['pct_autorizado']}%" for a in _ranking[::-1]],
                    textposition='outside',
                    hovertemplate='<b>%{y}</b><br>%{text}<extra></extra>',
                ))
                _fig_conv.update_layout(
                    height=_h_bar, margin=dict(t=10, b=10, l=10, r=50),
                    paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)',
                    xaxis=dict(showgrid=True, gridcolor='#f1f5f9', range=[0,115],
                               ticksuffix='%', tickfont=dict(size=10)),
                    yaxis=dict(showgrid=False, tickfont=dict(size=11)),
                )
                with st.container(border=True):
                    st.markdown('<div class="rank-chart-title">📈 % Conversión</div>', unsafe_allow_html=True)
                    st.plotly_chart(_fig_conv, use_container_width=True, config={'displayModeBar': False})


            # ── Cards con gráfico circular por ejecutivo ──
            st.markdown('<div class="rank-section">🎖️ Detalle por ejecutivo</div>', unsafe_allow_html=True)
            _medallas = {1:"🥇", 2:"🥈", 3:"🥉"}
            _clases   = {1:"rank-1", 2:"rank-2", 3:"rank-3"}

            for i, ej in enumerate(_ranking, 1):
                cls       = _clases.get(i, "rank-other")
                total_fmt = _fmt_r(ej['total_generado'])
                prom_fmt  = _fmt_r(ej['promedio'])
                score     = ej['score']
                _pct_aut  = ej['pct_autorizado']
                _pc_color = "#16a34a" if _pct_aut >= 50 else "#f59e0b" if _pct_aut >= 25 else "#ef4444"

                # Card completa en un solo markdown con SVG donut — centrado vertical perfecto
                _r, _r2 = 45, 28  # radio exterior e interior del donut
                _cx, _cy = 55, 55
                _circum = 2 * 3.14159 * _r
                _dash = _circum * _pct_aut / 100
                _gap  = _circum - _dash
                _svg_donut = (
                    f'<svg width="110" height="110" viewBox="0 0 110 110" xmlns="http://www.w3.org/2000/svg">' +
                    f'<circle cx="{_cx}" cy="{_cy}" r="{_r}" fill="none" stroke="#f1f5f9" stroke-width="14"/>' +
                    f'<circle cx="{_cx}" cy="{_cy}" r="{_r}" fill="none" stroke="{_pc_color}" stroke-width="14" ' +
                    f'stroke-dasharray="{_dash:.1f} {_gap:.1f}" stroke-dashoffset="{_circum/4:.1f}" stroke-linecap="round"/>' +
                    f'<text x="{_cx}" y="{_cy-4}" text-anchor="middle" font-size="14" font-weight="bold" fill="{_pc_color}">{_pct_aut}%</text>' +
                    f'<text x="{_cx}" y="{_cy+12}" text-anchor="middle" font-size="9" fill="#94a3b8">Conv.</text>' +
                    f'</svg>'
                )
                _border_color = '#f59e0b' if i==1 else '#94a3b8' if i==2 else '#b45309' if i==3 else '#e2e8f0'
                st.markdown(f'''
                <div style="display:flex;align-items:center;gap:16px;padding:12px 16px;
                    background:#fff;border-radius:16px;border:1px solid #e2e8f0;
                    border-left:5px solid {_border_color};
                    box-shadow:0 4px 20px rgba(0,0,0,0.07);">
                  <span style="font-size:1.8rem;flex-shrink:0;">{_medallas.get(i, f"#{i}")}</span>
                  <div style="flex:1;min-width:160px;">
                    <div style="font-size:1rem;font-weight:800;color:#1e293b;font-family:'Montserrat',sans-serif;">{ej['nombre']}</div>
                    <div style="background:#f1f5f9;border-radius:8px;height:10px;margin:6px 0 2px;overflow:hidden;">
                      <div style="width:{score}%;height:10px;border-radius:8px;background:linear-gradient(90deg,#f59e0b,#d97706);"></div>
                    </div>
                    <div style="font-size:0.72rem;color:#94a3b8;">Score {score}/100</div>
                  </div>
                  <div style="display:flex;gap:18px;align-items:center;flex-wrap:wrap;">
                    <div style="text-align:center;"><div style="font-size:1.2rem;font-weight:800;color:#0f172a;">{ej['total_presupuestos']}</div><div style="font-size:0.7rem;color:#64748b;">EP</div></div>
                    <div style="text-align:center;"><div style="font-size:1.2rem;font-weight:800;color:#0f172a;">{total_fmt}</div><div style="font-size:0.7rem;color:#64748b;">Total</div></div>
                    <div style="text-align:center;"><div style="font-size:1.2rem;font-weight:800;color:#0f172a;">{prom_fmt}</div><div style="font-size:0.7rem;color:#64748b;">Promedio</div></div>
                    <div style="text-align:center;"><div style="font-size:1.2rem;font-weight:800;color:#16a34a;">{ej['autorizados']}</div><div style="font-size:0.7rem;color:#64748b;">🟢 Auth.</div></div>
                    <div style="text-align:center;"><div style="font-size:1.2rem;font-weight:800;color:#f59e0b;">{ej['borradores']}</div><div style="font-size:0.7rem;color:#64748b;">🟡 Borr.</div></div>
                  </div>
                  <div style="flex-shrink:0;">{_svg_donut}</div>
                </div>
                ''', unsafe_allow_html=True)
                st.markdown("<div style='margin-bottom:6px;'></div>", unsafe_allow_html=True)
                st.markdown("<div style='margin-bottom:8px;'></div>", unsafe_allow_html=True)

# TAB CONTRATO CLIENTE
# =========================================================
with tab_contrato:
    st.markdown("""
    <style>
    .hdr-contrato {
        background: linear-gradient(135deg, #0f3460 0%, #16213e 60%, #1a1a2e 100%);
        border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
        display: flex; align-items: center; gap: 22px;
        box-shadow: 0 8px 32px rgba(15,52,96,0.3);
        position: relative; overflow: hidden;
    }
    .hdr-contrato::before {
        content: ''; position: absolute; top: -40px; right: -40px;
        width: 180px; height: 180px; border-radius: 50%;
        background: rgba(255,255,255,0.04); pointer-events: none;
    }
    .hdr-contrato::after {
        content: ''; position: absolute; bottom: -60px; right: 80px;
        width: 240px; height: 240px; border-radius: 50%;
        background: rgba(255,255,255,0.03); pointer-events: none;
    }
    .hdr-contrato h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                 font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
    .hdr-contrato p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
    </style>
    <div class="hdr-contrato">
      <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">📄</span>
      <div>
        <h2>Contrato Cliente</h2>
        <p>Genera el contrato de fabricación y venta listo para imprimir y firmar.</p>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Paso 1: Buscar EP ──
    st.markdown('<div class="cont-section">🔍 Paso 1 — Buscar cotización</div>', unsafe_allow_html=True)
    with st.container():
        _col_ep, _col_btn = st.columns([3, 1])
        with _col_ep:
            _ep_input = st.text_input("Número EP", placeholder="EP-22286",
                                      key="cont_ep_input",
                                      label_visibility="collapsed")
        with _col_btn:
            _buscar = st.button("🔍 Buscar EP", use_container_width=True, key="cont_buscar")

    if _buscar and _ep_input.strip():
        _num = _ep_input.strip().upper()
        if not _num.startswith("EP-"):
            _num = "EP-" + _num.lstrip("EP- ")
        try:
            _res = supabase.table("cotizaciones").select("*") \
                           .eq("numero", _num).limit(1).execute()
            if _res.data:
                _cot = _res.data[0]
                st.session_state["cont_cot"] = _cot
                st.session_state["cont_ep"]  = _num
                # Mapeo con campos reales de la tabla cotizaciones
                _dir = _cot.get("cliente_direccion", "")
                # Intentar extraer comuna de la dirección si está disponible
                _total = float(_cot.get("total_total", 0) or _cot.get("total", 0) or 0)
                st.session_state["cont_cli_nombre"]     = _cot.get("cliente_nombre", "")
                st.session_state["cont_cli_rut"]        = _cot.get("cliente_rut", "")
                st.session_state["cont_cli_domicilio"]  = _dir
                st.session_state["cont_cli_comuna"]     = _cot.get("cliente_comuna", "")
                st.session_state["cont_cli_region"]     = _cot.get("cliente_region", "")
                st.session_state["cont_inst_domicilio"] = _cot.get("proyecto_direccion", "") or _dir
                st.session_state["cont_inst_comuna"]    = _cot.get("proyecto_comuna", "")
                st.session_state["cont_inst_region"]    = _cot.get("proyecto_region", "")
                st.session_state["cont_tipo_cli_val"]   = _cot.get("cliente_tipo", "natural")
                st.session_state["cont_cli_empresa"]    = _cot.get("cliente_empresa", "")
                st.session_state["cont_cli_rut_empresa"]= _cot.get("cliente_rut_empresa", "")
                # Si ya tiene contrato guardado, cargar sus datos
                if _cot.get("contrato_datos"):
                    import json as _json
                    try:
                        _cd = _json.loads(_cot["contrato_datos"])
                        st.session_state["cont_cli_domicilio"]  = _cd.get("cli_domicilio", st.session_state["cont_cli_domicilio"])
                        st.session_state["cont_cli_comuna"]     = _cd.get("cli_comuna", st.session_state.get("cont_cli_comuna",""))
                        st.session_state["cont_cli_region"]     = _cd.get("cli_region", st.session_state.get("cont_cli_region",""))
                        st.session_state["cont_inst_domicilio"] = _cd.get("inst_domicilio", st.session_state.get("cont_inst_domicilio",""))
                        st.session_state["cont_inst_comuna"]    = _cd.get("inst_comuna", st.session_state.get("cont_inst_comuna",""))
                        st.session_state["cont_inst_region"]    = _cd.get("inst_region", st.session_state.get("cont_inst_region",""))
                        st.session_state["cont_ep_nombre"]      = _cd.get("ep_nombre", st.session_state.get("cont_ep_nombre",""))
                        st.session_state["cont_tiene_contrato"] = True
                    except:
                        st.session_state["cont_tiene_contrato"] = False
                else:
                    st.session_state["cont_tiene_contrato"] = False
                st.session_state["cont_ep_nombre"]      = ""  # campo libre, no viene de observaciones
                st.session_state["cont_precio"]         = _total
                st.rerun()
            else:
                st.error(f"No se encontró la cotización {_num}")
        except Exception as _e:
            st.error(f"Error al buscar: {_e}")

    # ── Formulario editable (solo si hay EP cargado) ──
    if st.session_state.get("cont_cot"):
        _cot    = st.session_state["cont_cot"]
        _ep_num = st.session_state.get("cont_ep", "")

        _tiene_cont = st.session_state.get("cont_tiene_contrato", False)
        if _tiene_cont:
            st.success(f"✅ Cotización **{_ep_num}** cargada — {_cot.get('cliente_nombre','')} · 📄 Ya tiene contrato generado")
        else:
            st.success(f"✅ Cotización **{_ep_num}** cargada — {_cot.get('cliente_nombre','')} · Sin contrato previo")

        # ── Datos del contrato ──
        st.markdown('<div class="cont-section">📝 Paso 2 — Completar datos del contrato</div>', unsafe_allow_html=True)

        # CSS compacto para el formulario
        st.markdown("""
        <style>
        div[data-testid="stForm"] { padding: 0 !important; }
        .cont-form-panel {
            background: white; border-radius: 14px; padding: 16px 20px 10px;
            border: 1px solid #e8edf5;
            box-shadow: 0 2px 12px rgba(0,0,0,0.05);
            margin-bottom: 10px;
        }
        .cont-form-title {
            font-size: 0.68rem; font-weight: 900; color: #64748b;
            text-transform: uppercase; letter-spacing: 0.1em;
            margin-bottom: 10px; display: flex; align-items: center; gap: 6px;
        }
        /* reducir espacio entre widgets dentro del formulario */
        .cont-form-panel div[data-testid="stVerticalBlock"] > div { margin-bottom: -8px !important; }
        </style>
        """, unsafe_allow_html=True)

        from datetime import date as _date_t
        _meses_es = {1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",
                     7:"julio",8:"agosto",9:"septiembre",10:"octubre",
                     11:"noviembre",12:"diciembre"}

        # ── Layout: col izquierda (formulario) | col derecha (pagos) ──
        _fcol, _pcol = st.columns([3, 2])

        with _fcol:
            # Panel 1 — Contrato
            _hoy = _date_t.today()
            _fecha_obj   = st.date_input("Fecha del contrato", value=_hoy, key="cont_fecha")
            _fecha_str   = f"{_fecha_obj.day} de {_meses_es[_fecha_obj.month]} de {_fecha_obj.year}"
            _ep_num_input = _ep_num  # solo lectura — viene del EP buscado

            _fa, _fb = st.columns([3, 1])
            with _fa:
                _ep_nombre = st.text_input("📝 Nombre / descripción del proyecto",
                    value=st.session_state.get("cont_ep_nombre",""),
                    placeholder="Ej: Casa container 2 módulos, 45m²",
                    key="cont_ep_nombre_input")
            with _fb:
                _plazo = st.number_input("Plazo (días)", min_value=1, max_value=180, value=45, key="cont_plazo")

            # N° EP — solo lectura, mostrado como card
            _html_ep = (
                "<div style='background:linear-gradient(135deg,#0f3460,#16213e);border-radius:14px;"
                "padding:14px 18px;margin-bottom:8px;display:flex;justify-content:space-between;align-items:center;'>"
                "<div>"
                "<div style='font-size:0.6rem;font-weight:900;color:rgba(255,255,255,0.5);"
                "text-transform:uppercase;letter-spacing:0.1em;margin-bottom:4px;'>📋 Presupuesto</div>"
                f"<div style='font-size:1.2rem;font-weight:900;color:#fff;'>{_ep_num_input or '—'}</div>"
                "</div>"
                f"<div style='font-size:0.8rem;color:rgba(255,255,255,0.5);'>{_fecha_str}</div>"
                "</div>"
            )
            st.markdown(_html_ep, unsafe_allow_html=True)

            # Panel 2 — Cliente (solo Don/Doña editable, resto lectura)
            _tipo_pre      = st.session_state.get("cont_tipo_cli_val", st.session_state.get("cliente_tipo", "natural"))
            _es_juridica   = (_tipo_pre == "juridica")
            _cli_nombre    = st.session_state.get("cont_cli_nombre",    st.session_state.get("nombre_input", ""))
            _cli_rut       = st.session_state.get("cont_cli_rut",       st.session_state.get("rut_display", ""))
            _cli_empresa   = st.session_state.get("cont_cli_empresa",   st.session_state.get("cliente_empresa", ""))
            _cli_rut_empresa = st.session_state.get("cont_cli_rut_empresa", st.session_state.get("cliente_rut_empresa", ""))
            _tipo_lbl      = "Persona jurídica" if _es_juridica else "Persona natural"

            _g1, _g2 = st.columns([1, 4])
            with _g1:
                _tratamiento = st.selectbox("Trato", ["Don", "Doña"], key="cont_tratamiento", label_visibility="collapsed")
            with _g2:
                _tag_tipo = "🏢 Persona jurídica" if _es_juridica else "👤 Persona natural"
                _emp_html = (
                    f"<div style='display:flex;gap:8px;margin-top:6px;'>"
                    f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>{_tag_tipo}</span>"
                    f"</div>"
                ) if not _es_juridica else (
                    f"<div style='display:flex;gap:8px;flex-wrap:wrap;margin-top:6px;'>"
                    f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>🏢 {_cli_empresa or '—'}</span>"
                    f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>RUT empresa: {_cli_rut_empresa or '—'}</span>"
                    f"</div>"
                )
                _sin_nombre_cli = "<span style='color:rgba(255,255,255,0.3);font-style:italic;font-weight:400;'>Sin nombre</span>"
                _html_cli = (
                    "<div style='background:linear-gradient(135deg,#0f3460,#16213e);border-radius:14px;padding:16px 18px;margin-bottom:4px;'>"
                    "<div style='font-size:0.65rem;font-weight:900;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:12px;'>👤 Datos del cliente</div>"
                    "<div style='display:flex;flex-direction:column;gap:10px;'>"
                    "<div style='background:rgba(255,255,255,0.07);border-radius:10px;padding:12px 14px;'>"
                    f"<div style='font-size:1rem;font-weight:800;color:#fff;'>{_tratamiento} {_cli_nombre if _cli_nombre else _sin_nombre_cli}</div>"
                    f"<div style='font-size:0.75rem;color:rgba(255,255,255,0.6);margin-top:4px;'>RUT: {_cli_rut or '—'}</div>"
                    + _emp_html +
                    "</div>"
                    "</div>"
                    "</div>"
                )
                st.markdown(_html_cli, unsafe_allow_html=True)

            # Panel 3 — Domicilios (solo lectura — vienen de pestaña DATOS)
            _cli_dom  = st.session_state.get("cont_cli_domicilio",  st.session_state.get("direccion_input", ""))
            _cli_com  = st.session_state.get("cont_cli_comuna",     st.session_state.get("cliente_comuna", ""))
            _cli_reg  = st.session_state.get("cont_cli_region",     st.session_state.get("cliente_region", ""))
            _inst_dom = st.session_state.get("cont_inst_domicilio", st.session_state.get("proyecto_direccion", ""))
            _inst_com = st.session_state.get("cont_inst_comuna",    st.session_state.get("proyecto_comuna", ""))
            _inst_reg = st.session_state.get("cont_inst_region",    st.session_state.get("proyecto_region", ""))
            _todas_regiones = list(REGIONES_COMUNAS.keys())
            _todas_comunas  = [c for cs in REGIONES_COMUNAS.values() for c in cs]
            if _cli_com:  _cli_com  = _normalizar_nombre(_cli_com,  _todas_comunas)
            if _cli_reg:  _cli_reg  = _normalizar_nombre(_cli_reg,  _todas_regiones)
            if _inst_com: _inst_com = _normalizar_nombre(_inst_com, _todas_comunas)
            if _inst_reg: _inst_reg = _normalizar_nombre(_inst_reg, _todas_regiones)

            _sin_dir = "<span style='color:rgba(255,255,255,0.25);font-style:italic;font-weight:400;'>Sin dirección</span>"
            _html_dom = (
                "<div style='background:linear-gradient(135deg,#0f3460,#16213e);border-radius:14px;padding:18px 20px;margin-bottom:12px;'>"
                "<div style='font-size:0.65rem;font-weight:900;color:rgba(255,255,255,0.5);text-transform:uppercase;letter-spacing:0.1em;margin-bottom:14px;'>📍 Domicilios</div>"
                "<div style='display:flex;flex-direction:column;gap:10px;'>"
                # Tarjeta cliente
                "<div style='background:rgba(255,255,255,0.07);border-radius:10px;padding:12px 14px;'>"
                "<div style='font-size:0.6rem;font-weight:800;color:rgba(255,255,255,0.45);text-transform:uppercase;letter-spacing:0.08em;margin-bottom:4px;'>Cliente</div>"
                f"<div style='font-size:0.95rem;font-weight:700;color:#fff;margin-bottom:2px;'>{_cli_dom or _sin_dir}</div>"
                "<div style='display:flex;gap:8px;margin-top:4px;'>"
                f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>🏙️ {_cli_com or '—'}</span>"
                f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>🗺️ {_cli_reg or '—'}</span>"
                "</div></div>"
                # Tarjeta instalación
                "<div style='background:rgba(255,255,255,0.07);border-radius:10px;padding:12px 14px;'>"
                "<div style='font-size:0.6rem;font-weight:800;color:rgba(255,255,255,0.45);text-transform:uppercase;letter-spacing:0.08em;margin-bottom:4px;'>Instalación</div>"
                f"<div style='font-size:0.95rem;font-weight:700;color:#fff;margin-bottom:2px;'>{_inst_dom or _sin_dir}</div>"
                "<div style='display:flex;gap:8px;margin-top:4px;'>"
                f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>🏙️ {_inst_com or '—'}</span>"
                f"<span style='font-size:0.7rem;background:rgba(255,255,255,0.1);border-radius:6px;padding:2px 8px;color:rgba(255,255,255,0.7);'>🗺️ {_inst_reg or '—'}</span>"
                "</div></div>"
                "</div></div>"
            )
            st.markdown(_html_dom, unsafe_allow_html=True)

        with _pcol:
            # El precio viene del EP cargado, sin input visible
            _precio = int(st.session_state.get("cont_precio", 0))

            _pago50  = round(_precio * 0.50)
            _pago25a = round(_precio * 0.25)
            _pago25b = _precio - _pago50 - _pago25a

            def _fp(v): return "${:,.0f}".format(int(v)).replace(",",".")

            st.markdown(f"""
            <div style="background:linear-gradient(135deg,#0f3460,#16213e);border-radius:14px;
                        padding:18px 20px;margin-top:4px;">
              <div style="font-size:0.65rem;font-weight:900;color:rgba(255,255,255,0.5);
                          text-transform:uppercase;letter-spacing:0.1em;margin-bottom:14px;">
                Etapas de pago
              </div>
              <div style="display:flex;flex-direction:column;gap:12px;">
                <div style="background:rgba(255,255,255,0.07);border-radius:10px;padding:12px 14px;">
                  <div style="font-size:0.65rem;font-weight:800;color:rgba(255,255,255,0.5);
                              text-transform:uppercase;letter-spacing:0.08em;">50% Inicial</div>
                  <div style="font-size:1.3rem;font-weight:900;color:#fff;margin:3px 0;">{_fp(_pago50)}</div>
                  <div style="font-size:0.65rem;color:rgba(255,255,255,0.4);">Asignación + obra gruesa</div>
                </div>
                <div style="background:rgba(255,255,255,0.07);border-radius:10px;padding:12px 14px;">
                  <div style="font-size:0.65rem;font-weight:800;color:rgba(255,255,255,0.5);
                              text-transform:uppercase;letter-spacing:0.08em;">25% Intermedio</div>
                  <div style="font-size:1.3rem;font-weight:900;color:#fff;margin:3px 0;">{_fp(_pago25a)}</div>
                  <div style="font-size:0.65rem;color:rgba(255,255,255,0.4);">Finalizada obra gruesa</div>
                </div>
                <div style="background:rgba(255,255,255,0.07);border-radius:10px;padding:12px 14px;">
                  <div style="font-size:0.65rem;font-weight:800;color:rgba(255,255,255,0.5);
                              text-transform:uppercase;letter-spacing:0.08em;">25% Final</div>
                  <div style="font-size:1.3rem;font-weight:900;color:#fff;margin:3px 0;">{_fp(_pago25b)}</div>
                  <div style="font-size:0.65rem;color:rgba(255,255,255,0.4);">Día del despacho</div>
                </div>
              </div>
              <div style="border-top:1px solid rgba(255,255,255,0.15);margin-top:14px;padding-top:12px;
                          display:flex;justify-content:space-between;align-items:center;">
                <div>
                  <div style="font-size:0.65rem;color:rgba(255,255,255,0.45);font-weight:800;
                               text-transform:uppercase;letter-spacing:0.08em;">Total IVA incluido</div>
                </div>
                <span style="font-size:1.15rem;font-weight:900;color:#fff;">{_fp(_precio)}</span>
              </div>
            </div>
            """, unsafe_allow_html=True)

        # ── Paso 3: Generar ──
        st.markdown('<div class="cont-section">📤 Paso 3 — Generar contrato</div>', unsafe_allow_html=True)

        # Preview resumen
        _fmt_p = lambda v: "${:,.0f}".format(int(v)).replace(",",".")
        st.markdown(f"""
        <div class="cont-preview">
          <b>Contrato {_ep_num_input}</b> — {_fecha_str}<br>
          <b>Cliente:</b> {_tratamiento} {_cli_nombre or '—'} · RUT {_cli_rut or '—'}<br>
          {"<b>Empresa:</b> " + _cli_empresa + " · RUT " + _cli_rut_empresa + "<br>" if _es_juridica else ""}
          <b>Domicilio cliente:</b> {_cli_dom or '—'}, {_cli_com or '—'}, Región {_cli_reg}<br>
          <b>Instalación:</b> {_inst_dom or '—'}, {_inst_com or '—'}, Región {_inst_reg}<br>
          <b>Proyecto:</b> {_ep_nombre or '—'}<br>
          <b>Precio total:</b> {_fmt_p(_precio)} · Plazo: {_plazo} días hábiles<br>
          <b>Pagos:</b> 50% {_fmt_p(_pago50)} / 25% {_fmt_p(_pago25a)} / 25% {_fmt_p(_pago25b)}
        </div>
        """, unsafe_allow_html=True)

        st.markdown("")
        _gen_col, _ = st.columns([2, 3])
        with _gen_col:
            _generar = st.button("📄 Generar contrato PDF", type="primary",
                                 use_container_width=True, key="cont_generar")

        if _generar:
            _campos_req = [_cli_nombre, _cli_rut, _cli_dom, _cli_com, _ep_nombre]
            if _es_juridica:
                _campos_req += [_cli_empresa, _cli_rut_empresa]
            if not all(_campos_req):
                st.error("⚠️ Completa todos los campos obligatorios antes de generar.")
            else:
                with st.spinner("Generando contrato..."):
                    _datos_contrato = {
                        "fecha_str":       _fecha_str,
                        "tipo_cliente":    "juridica" if _es_juridica else "natural",
                        "cli_tratamiento": _tratamiento,
                        "cli_nombre":      _cli_nombre,
                        "cli_rut":         _cli_rut,
                        "cli_empresa":     _cli_empresa,
                        "cli_rut_empresa": _cli_rut_empresa,
                        "cli_domicilio":   _cli_dom,
                        "cli_comuna":      _cli_com,
                        "cli_region":      _cli_reg,
                        "inst_domicilio":  _inst_dom,
                        "inst_comuna":     _inst_com,
                        "inst_region":     _inst_reg,
                        "ep_numero":       _ep_num_input,
                        "ep_nombre":       _ep_nombre,
                        "precio_total":    _precio,
                        "plazo_dias":      _plazo,
                        "pago_50":         _pago50,
                        "pago_25a":        _pago25a,
                        "pago_25b":        _pago25b,
                    }
                    try:
                        _pdf_bytes = generar_pdf_contrato(_datos_contrato)
                        st.session_state["cont_pdf_bytes"] = _pdf_bytes
                        st.session_state["cont_pdf_nombre"] = f"Contrato_{_ep_num_input.replace('-','_')}.pdf"
                        # Guardar contrato en Supabase
                        import json as _json
                        try:
                            _cont_payload = {
                                "contrato_generado": True,
                                "contrato_datos": _json.dumps(_datos_contrato, ensure_ascii=False),
                                "contrato_fecha": _datos_contrato["fecha_str"],
                            }
                            supabase.table("cotizaciones").update(_cont_payload)                                     .eq("numero", _ep_num_input).execute()
                            st.success("✅ Contrato generado y guardado exitosamente.")
                        except Exception as _es:
                            st.success("✅ Contrato generado.")
                            st.warning(f"No se pudo guardar en BD: {_es}")
                    except Exception as _e:
                        st.error(f"Error al generar PDF: {_e}")

        # ── Descarga e impresión ──
        if st.session_state.get("cont_pdf_bytes"):
            import base64 as _b64
            _pdf_b64 = _b64.b64encode(st.session_state["cont_pdf_bytes"]).decode()
            _pdf_nom = st.session_state.get("cont_pdf_nombre","contrato.pdf")

            _dl_col, _pr_col, _sp = st.columns([1.5, 1.5, 3])
            with _dl_col:
                st.download_button(
                    label="⬇️ Descargar PDF",
                    data=st.session_state["cont_pdf_bytes"],
                    file_name=_pdf_nom,
                    mime="application/pdf",
                    use_container_width=True,
                    key="cont_download"
                )
            with _pr_col:
                st.markdown(f"""
                <a href="data:application/pdf;base64,{_pdf_b64}" target="_blank"
                   style="display:block;background:linear-gradient(135deg,#0f3460,#16213e);
                          color:white;text-align:center;padding:0.55rem 1rem;border-radius:8px;
                          font-weight:700;font-size:0.88rem;text-decoration:none;
                          box-shadow:0 2px 8px rgba(15,52,96,0.3);">
                  🖨️ Abrir e imprimir
                </a>""", unsafe_allow_html=True)

    else:
        st.info("👆 Ingresa un número EP y presiona **Buscar EP** para cargar los datos del cliente.")


# =========================================================
# TAB USUARIOS — solo admin/supervisor
# =========================================================
if st.session_state.modo_admin and tab_usuarios is not None:
    with tab_usuarios:
        st.markdown("""
        <style>
        .hdr-usuarios {
            background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
            border-radius: 20px; padding: 32px 36px; margin-bottom: 28px;
            display: flex; align-items: center; gap: 22px;
            box-shadow: 0 8px 32px rgba(15,52,96,0.4);
            position: relative; overflow: hidden;
        }
        .hdr-usuarios::before {
            content: ''; position: absolute; top: -40px; right: -40px;
            width: 180px; height: 180px; border-radius: 50%;
            background: rgba(255,255,255,0.04); pointer-events: none;
        }
        .hdr-usuarios::after {
            content: ''; position: absolute; bottom: -60px; right: 80px;
            width: 240px; height: 240px; border-radius: 50%;
            background: rgba(255,255,255,0.03); pointer-events: none;
        }
        .hdr-usuarios h2 { color: #fff !important; margin: 0; font-size: 1.8rem; font-weight: 900;
                           font-family: 'Montserrat', sans-serif; letter-spacing: -0.02em; }
        .hdr-usuarios p  { color: rgba(255,255,255,0.65) !important; margin: 6px 0 0; font-size: 0.92rem; }
        .usr-card {
            background: var(--background-color, #fff);
            border: 1px solid rgba(0,0,0,0.08);
            border-radius: 14px; padding: 18px 20px;
            margin-bottom: 10px;
            display: flex; align-items: center; gap: 16px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }
        .usr-avatar {
            width: 44px; height: 44px; border-radius: 50%;
            background: linear-gradient(135deg, #5b7cfa, #8b5cf6);
            display: flex; align-items: center; justify-content: center;
            font-size: 1.2rem; font-weight: 700; color: white;
            flex-shrink: 0;
        }
        .usr-info { flex: 1; }
        .usr-nombre { font-weight: 700; font-size: 0.97rem; color: #1e2447; }
        .usr-email  { font-size: 0.82rem; color: #64748b; margin-top: 2px; }
        .usr-fecha  { font-size: 0.78rem; color: #94a3b8; }
        .usr-badge-ok   { background:#dcfce7; color:#166534; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:700; }
        .usr-badge-off  { background:#fee2e2; color:#991b1b; padding:3px 10px; border-radius:20px; font-size:0.75rem; font-weight:700; }
        </style>
        <div class="hdr-usuarios">
          <span style="font-size:2.8rem;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.3));">👥</span>
          <div>
            <h2>Gestión de Usuarios</h2>
            <p>Crea y administra las cuentas de acceso de los ejecutivos.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Crear nuevo usuario ──
        with st.expander("➕ Crear nuevo usuario", expanded=False):
            # Root puede crear ejecutivos y admins; admin solo ejecutivos
            _tabs_crear = ["👤 Nuevo ejecutivo", "👑 Nuevo administrador"]  # admin y root pueden crear ambos
            _tabs_obj = st.tabs(_tabs_crear)
            _tab_ej = _tabs_obj[0]
            _tab_adm = _tabs_obj[1] if len(_tabs_obj) > 1 else None

            with _tab_ej:
                col_n, col_e, col_t, col_p = st.columns(4)
                with col_n:
                    _u_nombre = st.text_input("Nombre completo", key="new_usr_nombre", placeholder="Juan Pérez")
                with col_e:
                    _u_email = st.text_input("Correo electrónico", key="new_usr_email", placeholder="juan@empresa.cl")
                with col_t:
                    _u_tel = st.text_input("Teléfono", key="new_usr_tel", placeholder="+56912345678")
                with col_p:
                    _u_pass = st.text_input("Contraseña", type="password", key="new_usr_pass", placeholder="Mínimo 6 caracteres")

                if st.button("✅ Crear ejecutivo", type="primary", key="btn_crear_usuario"):
                    if not _u_nombre.strip():
                        st.error("Ingresa el nombre del ejecutivo.")
                    elif not _u_email.strip() or "@" not in _u_email:
                        st.error("Ingresa un correo válido.")
                    elif len(_u_pass) < 6:
                        st.error("La contraseña debe tener al menos 6 caracteres.")
                    elif es_rol_superior(_u_email):
                        st.error("Ese correo ya está registrado como supervisor/admin.")
                    else:
                        with st.spinner("Creando cuenta..."):
                            try:
                                _res_ej = supabase_admin.auth.admin.create_user({
                                    "email": _u_email.strip().lower(),
                                    "password": _u_pass,
                                    "email_confirm": True,
                                    "user_metadata": {
                                        "nombre": _u_nombre.strip().upper(),
                                        "telefono": _u_tel.strip(),
                                        "rol": "ejecutivo"
                                    }
                                })
                                user = _res_ej.user
                                err = None
                            except Exception as _ex:
                                user = None
                                err = str(_ex)
                        if user:
                            st.success(f"✅ Cuenta creada para **{_u_nombre}** ({_u_email})")
                            st.session_state.pop('_usuarios_cache', None)
                            st.session_state['_asesores_cache_dirty'] = True
                            st.rerun()
                        else:
                            if "already registered" in str(err) or "already been registered" in str(err):
                                st.error("❌ Ya existe una cuenta con ese correo.")
                            else:
                                st.error(f"❌ Error: {err}")

            if _tab_adm is not None:
                with _tab_adm:
                    st.info("Los administradores ven todas las cotizaciones, pueden crear ejecutivos y otros admins, pero no pueden crear ni eliminar cuentas Root. El tab 🛡️ Sistema es exclusivo de Root.")
                    col_an, col_ae, col_at, col_ap = st.columns(4)
                    with col_an:
                        _a_nombre = st.text_input("Nombre completo", key="new_adm_nombre", placeholder="Ana Rodríguez")
                    with col_ae:
                        _a_email = st.text_input("Correo electrónico", key="new_adm_email", placeholder="ana@empresa.cl")
                    with col_at:
                        _a_tel = st.text_input("Teléfono", key="new_adm_tel", placeholder="+56912345678")
                    with col_ap:
                        _a_pass = st.text_input("Contraseña", type="password", key="new_adm_pass", placeholder="Mínimo 6 caracteres")

                    if st.button("👑 Crear administrador", type="primary", key="btn_crear_admin"):
                        if not _a_nombre.strip():
                            st.error("Ingresa el nombre del administrador.")
                        elif not _a_email.strip() or "@" not in _a_email:
                            st.error("Ingresa un correo válido.")
                        elif len(_a_pass) < 6:
                            st.error("La contraseña debe tener al menos 6 caracteres.")
                        else:
                            with st.spinner("Creando cuenta..."):
                                try:
                                    _res_adm = supabase_admin.auth.admin.create_user({
                                        "email": _a_email.strip().lower(),
                                        "password": _a_pass,
                                        "email_confirm": True,
                                        "user_metadata": {
                                            "nombre": _a_nombre.strip().upper(),
                                            "telefono": _a_tel.strip(),
                                            "rol": "admin"
                                        }
                                    })
                                    _adm_user = _res_adm.user
                                    _adm_err = None
                                except Exception as _ex:
                                    _adm_user = None
                                    _adm_err = str(_ex)
                            if _adm_user:
                                st.success(f"✅ Administrador creado: **{_a_nombre}** ({_a_email})")
                                st.info("El acceso de administrador es permanente — guardado en Supabase.")
                                st.session_state.pop('_usuarios_cache', None)
                                st.session_state['_asesores_cache_dirty'] = True
                                st.rerun()
                            else:
                                if "already registered" in str(_adm_err) or "already been registered" in str(_adm_err):
                                    st.error("❌ Ya existe una cuenta con ese correo.")
                                else:
                                    st.error(f"❌ Error: {_adm_err}")

        st.markdown("---")

        # ── CSS adicional para la lista ──
        st.markdown("""
        <style>
        .usr-card2 {
            background: var(--background-color, #fff);
            border-radius: 14px;
            border: 1.5px solid #e8eaf0;
            padding: 16px 20px;
            display: flex; align-items: center; gap: 16px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            transition: box-shadow 0.2s;
        }
        .usr-card2:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.10); }
        .usr-card2.es-admin { border-left: 4px solid #8b5cf6; }
        .usr-card2.es-ejecutivo { border-left: 4px solid #3b82f6; }
        .usr-card2.es-root { border-left: 4px solid #f59e0b; }
        .usr-av2 {
            width: 46px; height: 46px; border-radius: 50%;
            display: flex; align-items: center; justify-content: center;
            font-size: 1.25rem; font-weight: 800; color: white; flex-shrink: 0;
        }
        .usr-av2.av-admin    { background: linear-gradient(135deg,#7c3aed,#a78bfa); }
        .usr-av2.av-ejecutivo{ background: linear-gradient(135deg,#2563eb,#60a5fa); }
        .usr-av2.av-root     { background: linear-gradient(135deg,#b45309,#f59e0b); }
        .usr-nm2  { font-weight: 700; font-size: 0.97rem; color: #1e2447; }
        .usr-em2  { font-size: 0.81rem; color: #64748b; margin-top: 2px; }
        .usr-meta { font-size: 0.75rem; color: #94a3b8; margin-top: 2px; }
        .rol-pill {
            display: inline-block; padding: 2px 10px; border-radius: 20px;
            font-size: 0.72rem; font-weight: 700; letter-spacing: 0.04em;
        }
        .rol-admin    { background:#ede9fe; color:#6d28d9; }
        .rol-ejecutivo{ background:#dbeafe; color:#1d4ed8; }
        .rol-root     { background:#fef3c7; color:#b45309; }
        .accion-panel {
            background: #f8f9fc;
            border: 1px solid #e8eaf0;
            border-radius: 12px;
            padding: 16px 18px;
            margin-top: 6px;
            margin-bottom: 8px;
        }
        .accion-title { font-size:0.78rem; font-weight:700; color:#64748b;
                        text-transform:uppercase; letter-spacing:0.06em; margin-bottom:10px; }
        </style>
        """, unsafe_allow_html=True)

        # ── Header con contador y refresh ──
        _hcol1, _hcol2 = st.columns([4, 1])
        with _hcol1:
            if '_usuarios_cache' not in st.session_state:
                with st.spinner("Cargando usuarios..."):
                    st.session_state['_usuarios_cache'] = listar_usuarios_ejecutivos()
            _usuarios = st.session_state.get('_usuarios_cache', [])
            _n_adm = sum(1 for u in _usuarios if u.get('rol') in ('admin','administrador'))
            _n_ej  = sum(1 for u in _usuarios if u.get('rol','ejecutivo') == 'ejecutivo')
            st.markdown(f"""
            <div style="display:flex;gap:12px;align-items:center;margin-bottom:4px;">
                <span style="font-size:1.1rem;font-weight:800;color:#1e2447;">👥 Usuarios registrados</span>
                <span style="background:#ede9fe;color:#6d28d9;padding:3px 10px;border-radius:20px;font-size:0.75rem;font-weight:700;">👑 {_n_adm} Admin</span>
                <span style="background:#dbeafe;color:#1d4ed8;padding:3px 10px;border-radius:20px;font-size:0.75rem;font-weight:700;">👤 {_n_ej} Ejecutivo</span>
            </div>
            """, unsafe_allow_html=True)
        with _hcol2:
            if st.button("🔄 Actualizar", key="btn_refresh_usuarios", use_container_width=True):
                st.session_state.pop('_usuarios_cache', None)
                st.rerun()

        if st.session_state.get('_usuarios_list_error'):
            st.error(f"❌ Error al cargar usuarios: {st.session_state['_usuarios_list_error']}")
            st.session_state.pop('_usuarios_list_error', None)

        if not _usuarios:
            st.info("No hay usuarios registrados aún.")
        else:
            for _u in _usuarios:
                _inicial   = (_u['nombre'] or _u['email'] or "?")[0].upper()
                _rol_obj   = _u.get('rol', 'ejecutivo')
                _es_root_u = _rol_obj == 'root'
                _es_admin_u= _rol_obj in ('admin','administrador')

                # Clases CSS según rol
                _card_cls = 'es-root' if _es_root_u else ('es-admin' if _es_admin_u else 'es-ejecutivo')
                _av_cls   = 'av-root'  if _es_root_u else ('av-admin'  if _es_admin_u else 'av-ejecutivo')
                _rol_pill = ('rol-root'  if _es_root_u else ('rol-admin' if _es_admin_u else 'rol-ejecutivo'))
                _rol_txt  = ('🔑 Root'   if _es_root_u else ('👑 Admin'  if _es_admin_u else '👤 Ejecutivo'))
                _u_tel_disp = _u.get('telefono', '') or ''  

                # Permisos
                _puede = (
                    st.session_state.get('es_root') or
                    (st.session_state.get('rol_usuario') == 'admin' and _rol_obj in ('ejecutivo','admin'))
                ) and not _es_root_u

                # ── Tarjeta principal ──
                st.markdown(f"""
                <div class="usr-card2 {_card_cls}">
                    <div class="usr-av2 {_av_cls}">{_inicial}</div>
                    <div style="flex:1;min-width:0;">
                        <div class="usr-nm2">{_u['nombre']}</div>
                        <div class="usr-em2">✉️ {_u['email']}{(' &nbsp;·&nbsp; 📞 ' + _u_tel_disp) if _u_tel_disp else ''}</div>
                        <div class="usr-meta">📅 {_u['created_at']} &nbsp;·&nbsp;
                            <span class="rol-pill {_rol_pill}">{_rol_txt}</span>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                if _puede:
                    # Botones de acción en línea
                    _ba1, _ba2, _ba3, _ba4, _ba5 = st.columns([1, 1, 1, 1, 2])
                    with _ba1:
                        _nuevo_rol  = 'admin' if _rol_obj == 'ejecutivo' else 'ejecutivo'
                        _lbl_rol    = '⬆️ Admin' if _nuevo_rol == 'admin' else '⬇️ Ejecutivo'
                        if st.button(_lbl_rol, key=f"rol_{_u['id']}", use_container_width=True):
                            st.session_state[f'_accion_{_u["id"]}'] = 'rol'
                            st.rerun()
                    with _ba2:
                        if st.button("✏️ Editar", key=f"edit_btn_{_u['id']}", use_container_width=True):
                            st.session_state[f'_accion_{_u["id"]}'] = 'edit'
                            st.rerun()
                    with _ba3:
                        if st.button("🔑 Contraseña", key=f"pwd_btn_{_u['id']}", use_container_width=True):
                            st.session_state[f'_accion_{_u["id"]}'] = 'pwd'
                            st.rerun()
                    with _ba4:
                        if st.button("🗑️ Eliminar", key=f"del_btn_{_u['id']}", use_container_width=True):
                            st.session_state[f'_accion_{_u["id"]}'] = 'del'
                            st.rerun()

                    # Panel de acción expandido según selección
                    _accion = st.session_state.get(f'_accion_{_u["id"]}')

                    if _accion == 'edit':
                        st.markdown(f'<div class="accion-panel"><div class="accion-title">✏️ Editar datos de {_u["nombre"]}</div>', unsafe_allow_html=True)
                        _ec1, _ec2, _ec3 = st.columns(3)
                        with _ec1:
                            _edit_nombre = st.text_input("Nombre completo", value=_u['nombre'],
                                                          key=f"edit_nm_{_u['id']}")
                        with _ec2:
                            _edit_email = st.text_input("Correo electrónico", value=_u['email'],
                                                         key=f"edit_em_{_u['id']}")
                        with _ec3:
                            _edit_tel = st.text_input("Teléfono", value=_u.get('telefono',''),
                                                       key=f"edit_tel_{_u['id']}", placeholder="+56912345678")
                        _ep1, _ep2 = st.columns(2)
                        with _ep1:
                            if st.button("✅ Guardar cambios", key=f"edit_ok_{_u['id']}", use_container_width=True, type="primary"):
                                if not _edit_nombre.strip():
                                    st.error("El nombre no puede estar vacío.")
                                elif not _edit_email.strip() or "@" not in _edit_email:
                                    st.error("Ingresa un correo válido.")
                                else:
                                    try:
                                        supabase_admin.auth.admin.update_user_by_id(_u['id'], {
                                            "email": _edit_email.strip().lower(),
                                            "user_metadata": {
                                                "nombre": _edit_nombre.strip().upper(),
                                                "telefono": _edit_tel.strip(),
                                                "rol": _rol_obj
                                            }
                                        })
                                        st.success(f"✅ Datos de {_edit_nombre.upper()} actualizados.")
                                        st.session_state.pop('_usuarios_cache', None)
                                        st.session_state['_asesores_cache_dirty'] = True
                                        st.session_state.pop(f'_accion_{_u["id"]}', None)
                                        st.rerun()
                                    except Exception as _edit_err:
                                        st.error(f"❌ {_edit_err}")
                        with _ep2:
                            if st.button("✖️ Cancelar", key=f"edit_no_{_u['id']}", use_container_width=True):
                                st.session_state.pop(f'_accion_{_u["id"]}', None)
                                st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)

                    elif _accion == 'rol':
                        st.markdown(f'<div class="accion-panel"><div class="accion-title">Cambiar rol de {_u["nombre"]}</div>', unsafe_allow_html=True)
                        st.info(f"Rol actual: **{_rol_txt}** → Nuevo rol: **{'Admin' if _nuevo_rol == 'admin' else 'Ejecutivo'}**")
                        _cr1, _cr2 = st.columns(2)
                        with _cr1:
                            if st.button(f"✅ Confirmar cambio", key=f"rol_ok_{_u['id']}", use_container_width=True, type="primary"):
                                _ok_r, _err_r = cambiar_rol_usuario(_u['id'], _nuevo_rol)
                                if _ok_r:
                                    st.success(f"✅ {_u['nombre']} ahora es {_nuevo_rol}.")
                                    st.session_state.pop('_usuarios_cache', None)
                                    st.session_state['_asesores_cache_dirty'] = True
                                    st.session_state.pop(f'_accion_{_u["id"]}', None)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {_err_r}")
                        with _cr2:
                            if st.button("✖️ Cancelar", key=f"rol_no_{_u['id']}", use_container_width=True):
                                st.session_state.pop(f'_accion_{_u["id"]}', None)
                                st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)

                    elif _accion == 'pwd':
                        st.markdown(f'<div class="accion-panel"><div class="accion-title">Resetear contraseña de {_u["nombre"]}</div>', unsafe_allow_html=True)
                        _nueva_pwd = st.text_input("Nueva contraseña", type="password",
                                                   key=f"pwd_inp_{_u['id']}", placeholder="Mínimo 6 caracteres")
                        _pp1, _pp2 = st.columns(2)
                        with _pp1:
                            if st.button("✅ Guardar contraseña", key=f"pwd_ok_{_u['id']}", use_container_width=True, type="primary"):
                                if len(_nueva_pwd or '') < 6:
                                    st.error("Mínimo 6 caracteres.")
                                else:
                                    _ok_p, _err_p = resetear_password_admin(_u['id'], _nueva_pwd)
                                    if _ok_p:
                                        st.success(f"✅ Contraseña actualizada.")
                                        st.session_state.pop(f'_accion_{_u["id"]}', None)
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {_err_p}")
                        with _pp2:
                            if st.button("✖️ Cancelar", key=f"pwd_no_{_u['id']}", use_container_width=True):
                                st.session_state.pop(f'_accion_{_u["id"]}', None)
                                st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)

                    elif _accion == 'del':
                        st.markdown(f'<div class="accion-panel"><div class="accion-title">Eliminar usuario</div>', unsafe_allow_html=True)
                        st.warning(f"⚠️ ¿Estás seguro que deseas eliminar a **{_u['nombre']}** ({_u['email']})? Esta acción no se puede deshacer.")
                        _dd1, _dd2 = st.columns(2)
                        with _dd1:
                            if st.button("🗑️ Sí, eliminar", key=f"del_ok_{_u['id']}", use_container_width=True, type="primary"):
                                _ok_d, _err_d = eliminar_usuario_ejecutivo(_u['id'])
                                if _ok_d:
                                    st.success("✅ Usuario eliminado.")
                                    st.session_state.pop('_usuarios_cache', None)
                                    st.session_state['_asesores_cache_dirty'] = True
                                    st.session_state.pop(f'_accion_{_u["id"]}', None)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {_err_d}")
                        with _dd2:
                            if st.button("✖️ Cancelar", key=f"del_no_{_u['id']}", use_container_width=True):
                                st.session_state.pop(f'_accion_{_u["id"]}', None)
                                st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)

                elif _es_root_u:
                    st.markdown("<div style='padding:4px 0 8px;'><span style='color:#f59e0b;font-size:0.8rem;font-weight:600;'>🔑 Cuenta Root — protegida</span></div>", unsafe_allow_html=True)

                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)


# =========================================================
# TAB NOTIFICACIONES — solo admin y root
# =========================================================
if tab_notif is not None and st.session_state.get('es_supervisor'):
    with tab_notif:
        import json as _json_notif

        # Header
        st.markdown("""
        <style>
        .hdr-notif {
            background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 60%, #0f172a 100%);
            border-radius: 20px; padding: 28px 32px; margin-bottom: 20px;
            display: flex; align-items: center; gap: 18px;
        }
        .hdr-notif h2 { color:#fff !important; margin:0; font-size:1.6rem; font-weight:900; }
        .hdr-notif p  { color:rgba(255,255,255,0.5) !important; margin:4px 0 0; font-size:0.88rem; }
        </style>
        <div class="hdr-notif">
          <span style="font-size:2.4rem">📣</span>
          <div>
            <h2>Notificaciones</h2>
            <p>Configura Telegram, contactos, observadores y mensajes automáticos.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── 1. Configuración del Bot ──
        with st.container(border=True):
            st.markdown("**⚙️ 1 · Configuración del Bot**")
            _token_actual = _get_notif_config('bot_token', TELEGRAM_BOT_TOKEN_DEFAULT)
            _bot_nombre   = _get_notif_config('bot_nombre', 'Cotizador ECH Bot')
            _c1, _c2, _c3 = st.columns([2, 1.5, 1])
            with _c1:
                _token_inp = st.text_input("Token del Bot", value=_token_actual, type="password", key="notif_token")
            with _c2:
                _nombre_inp = st.text_input("Nombre del Bot", value=_bot_nombre, key="notif_nombre")
            with _c3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                _cb1, _cb2 = st.columns(2)
                with _cb1:
                    if st.button("🔌 Probar", key="btn_probar_bot", use_container_width=True):
                        with st.spinner("Probando..."):
                            try:
                                import requests as _rtest
                                _r = _rtest.get(f"https://api.telegram.org/bot{_token_inp}/getMe", timeout=10)
                                _rdata = _r.json()
                                _bot_ok = _rdata.get('ok', False)
                            except:
                                _bot_ok = False; _rdata = {}
                        if _bot_ok:
                            st.success(f"✅ @{_rdata.get('result',{}).get('username','')}")
                        else:
                            st.error("❌ Token inválido")
                with _cb2:
                    if st.button("💾 Guardar", key="btn_guardar_bot", use_container_width=True, type="primary"):
                        _set_notif_config('bot_token', _token_inp)
                        _set_notif_config('bot_nombre', _nombre_inp)
                        st.success("✅ Guardado")

        # ── 2. Contactos del sistema ──
        with st.container(border=True):
            st.markdown("**👥 2 · Contactos del sistema**")
            st.caption("Cada usuario debe escribirle /start al bot una vez para obtener su Chat ID")

            # ── Panel getUpdates — detectar quién escribió /start ──
            with st.expander("📡 Detectar usuarios desde Telegram (getUpdates)", expanded=False):
                st.caption("Muestra las personas que le han escrito al bot. Haz click en ➕ para asignar su Chat ID.")
                if st.button("🔄 Obtener usuarios del bot", key="btn_get_updates"):
                    st.session_state['_tg_updates'] = None
                    try:
                        import requests as _ru
                        _tok = _get_notif_config('bot_token', TELEGRAM_BOT_TOKEN_DEFAULT)
                        _resp = _ru.get(f"https://api.telegram.org/bot{_tok}/getUpdates?limit=100", timeout=10)
                        _data = _resp.json()
                        if _data.get('ok'):
                            _vistos = {}
                            for _upd in _data.get('result', []):
                                _msg = _upd.get('message', {})
                                _ch  = _msg.get('chat', {})
                                if _ch.get('id'):
                                    _uid = str(_ch['id'])
                                    _vistos[_uid] = {
                                        'chat_id': _uid,
                                        'nombre':  _ch.get('first_name','') + (' ' + _ch.get('last_name','') if _ch.get('last_name') else ''),
                                        'username': _ch.get('username','')
                                    }
                            st.session_state['_tg_updates'] = list(_vistos.values())
                        else:
                            st.error("❌ Error al consultar el bot")
                    except Exception as _ue:
                        st.error(f"❌ {_ue}")

                if st.session_state.get('_tg_updates') is not None:
                    _updates = st.session_state['_tg_updates']
                    if not _updates:
                        st.info("No hay usuarios registrados aún. Pídeles que escriban /start al bot.")
                    else:
                        st.markdown(f"**{len(_updates)} persona(s) encontradas:**")
                        for _up in _updates:
                            _uc1, _uc2, _uc3 = st.columns([2, 1.5, 1])
                            with _uc1:
                                _uname = f"@{_up['username']}" if _up['username'] else "sin username"
                                st.markdown(f"**{_up['nombre']}** · `{_uname}`")
                            with _uc2:
                                st.code(_up['chat_id'], language=None)
                            with _uc3:
                                if st.button(f"➕ Asignar", key=f"asignar_{_up['chat_id']}"):
                                    st.session_state[f"_asignar_chat_{_up['chat_id']}"] = _up['chat_id']
                                    st.info(f"Copia el Chat ID `{_up['chat_id']}` en el campo del contacto correspondiente abajo ↓")

            st.divider()

            _contactos = _get_contactos_notif()
            _todos_usuarios = []
            try:
                _todos_usuarios = listar_usuarios_ejecutivos()
                for _re in ROOTS:
                    _todos_usuarios.insert(0, {'email': _re, 'nombre': 'Root', 'rol': 'root'})
            except: pass
            _contactos_nuevos = dict(_contactos)
            for _idx, _uu in enumerate(_todos_usuarios):
                _ue  = _uu.get('email', '').lower()
                _ur  = _uu.get('rol', 'ejecutivo')
                _un  = _uu.get('nombre', _ue)
                _rol_color = "#7c3aed" if _ur=='root' else ("#8b5cf6" if _ur=='admin' else "#2563eb")
                _rol_txt   = "🔑 Root" if _ur=='root' else ("👑 Admin" if _ur=='admin' else "👤 Ejecutivo")
                _col_nm, _col_em, _col_chat, _col_rol, _col_est = st.columns([1.5, 1.8, 1.5, 1, 0.7])
                with _col_nm:
                    st.markdown(f"<div style='padding:6px 0;font-size:0.88rem;font-weight:600'>{_un}</div>", unsafe_allow_html=True)
                with _col_em:
                    st.markdown(f"<div style='padding:6px 0;font-size:0.78rem;color:#64748b'>{_ue}</div>", unsafe_allow_html=True)
                with _col_chat:
                    _chat_val = _contactos.get(_ue, '')
                    _new_chat = st.text_input("", value=_chat_val, placeholder="@usuario o Chat ID",
                                              key=f"chat_{_idx}_{_ue}", label_visibility="collapsed")
                    _contactos_nuevos[_ue] = _new_chat
                with _col_rol:
                    st.markdown(f"<div style='padding:6px 0;font-size:0.75rem;color:{_rol_color};font-weight:700'>{_rol_txt}</div>", unsafe_allow_html=True)
                with _col_est:
                    _esta = "🟢" if _contactos.get(_ue,'') else "🟡"
                    st.markdown(f"<div style='padding:6px 0;text-align:center'>{_esta}</div>", unsafe_allow_html=True)
            if st.button("💾 Guardar contactos", key="btn_guardar_contactos", type="primary"):
                _set_notif_config('contactos_json', _json_notif.dumps(_contactos_nuevos))
                st.success("✅ Contactos guardados")
                st.rerun()

        # ── 3. Observadores ──
        with st.container(border=True):
            st.markdown("**👁 3 · Observadores externos**")
            st.caption("Sin cuenta en el sistema · Reciben todas las notificaciones")
            _obs_list = _get_observadores_notif()
            _obs_list_edit = list(_obs_list) + [{'nombre': '', 'chat_id': ''}]
            _obs_nuevos = []
            for _oi, _ob in enumerate(_obs_list_edit):
                _oc1, _oc2, _oc3 = st.columns([2, 2, 0.5])
                with _oc1:
                    _on = st.text_input("", value=_ob.get('nombre',''), placeholder="Nombre (ej: Gerente)",
                                        key=f"obs_nm_{_oi}", label_visibility="collapsed")
                with _oc2:
                    _oid = st.text_input("", value=_ob.get('chat_id',''), placeholder="@usuario o Chat ID",
                                         key=f"obs_id_{_oi}", label_visibility="collapsed")
                with _oc3:
                    _del = st.button("✕", key=f"obs_del_{_oi}", help="Eliminar")
                if _on.strip() and not _del:
                    _obs_nuevos.append({'nombre': _on.strip(), 'chat_id': _oid.strip()})
            if st.button("💾 Guardar observadores", key="btn_guardar_obs", type="primary"):
                _set_notif_config('observadores_json', _json_notif.dumps(_obs_nuevos))
                st.success("✅ Observadores guardados")
                st.rerun()

        # ── 4. Grupo ──
        with st.container(border=True):
            st.markdown("**📢 4 · Grupo de Telegram (opcional)**")
            _grupo_id     = _get_notif_config('grupo_chat_id', '')
            _grupo_filtro = _get_notif_config('grupo_filtro', 'todas')
            _gc1, _gc2, _gc3 = st.columns([2, 2, 1])
            with _gc1:
                _g_id_inp = st.text_input("Chat ID del grupo", value=_grupo_id, placeholder="-1001234567890", key="notif_grupo_id")
                st.caption("Agrega el bot al grupo y escribe /start para obtener el ID")
            with _gc2:
                _filtro_opts = {"todas":"Todas las notificaciones","solo_nuevas":"Solo nuevas cotizaciones",
                                "solo_autorizaciones":"Solo autorizaciones","ninguna":"No usar grupo"}
                _filtro_idx  = list(_filtro_opts.keys()).index(_grupo_filtro) if _grupo_filtro in _filtro_opts else 0
                _filtro_sel  = st.selectbox("Qué notificar al grupo", list(_filtro_opts.values()), index=_filtro_idx, key="notif_grupo_filtro")
                _filtro_val  = list(_filtro_opts.keys())[list(_filtro_opts.values()).index(_filtro_sel)]
            with _gc3:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                if st.button("💾 Guardar grupo", key="btn_guardar_grupo", use_container_width=True, type="primary"):
                    _set_notif_config('grupo_chat_id', _g_id_inp)
                    _set_notif_config('grupo_filtro', _filtro_val)
                    st.success("✅ Guardado")

        # ── 5. Mensajes ──
        with st.container(border=True):
            st.markdown("**✏️ 5 · Mensajes personalizables**")

            _msg_defaults = {
                'msg_nueva_cotizacion': "🆕 *Nueva cotización para revisar*\n\n*{ep}* · {ejecutivo}\nCliente: {cliente} · Monto: *{monto}*\nEstado: {estado}",
                'msg_autorizada':       "✅ *¡PRESUPUESTO AUTORIZADO!*\n\n📋 *{ep}* · {cliente}\n💰 Margen aplicado: *{margen}%*\n👤 Autorizado por: *{supervisor}*\n\nYa puedes presentárselo a tu cliente 🎉",
                'msg_margen_removido':  "↩️ La cotización *{ep}* volvió a estado borrador.\nEl supervisor realizó cambios. Revisa el sistema."
            }

            # Guía variables con click para copiar
            st.markdown("""
            <style>
            .var-guide{background:rgba(0,0,0,0.03);border-radius:10px;padding:12px 16px;margin-bottom:14px;}
            .var-guide-title{font-size:0.75rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.08em;margin-bottom:8px;}
            .var-group{margin-bottom:7px;}
            .var-group-label{font-size:0.7rem;color:#94a3b8;margin-bottom:4px;}
            .var-chips{display:flex;flex-wrap:wrap;gap:5px;}
            .var-chip{display:inline-block;background:#f1f5f9;border:1px solid #e2e8f0;border-radius:20px;
                padding:2px 9px;font-size:0.76rem;font-family:monospace;color:#3b82f6;cursor:pointer;
                transition:all 0.15s;user-select:none;}
            .var-chip:hover{background:#dbeafe;border-color:#93c5fd;transform:scale(1.05);}
            .var-chip.copied{background:#dcfce7;border-color:#86efac;color:#16a34a;}
            .fmt-chip{display:inline-block;background:#faf5ff;border:1px solid #e9d5ff;border-radius:20px;
                padding:2px 9px;font-size:0.76rem;font-family:monospace;color:#7c3aed;cursor:pointer;
                transition:all 0.15s;user-select:none;}
            .fmt-chip:hover{background:#ede9fe;border-color:#c4b5fd;transform:scale(1.05);}
            .fmt-chip.copied{background:#dcfce7;border-color:#86efac;color:#16a34a;}
            </style>
            <div class="var-guide">
              <div class="var-guide-title">📋 Variables — click para copiar</div>
              <div class="var-group">
                <div class="var-group-label">🆕 Nueva cotización</div>
                <div class="var-chips">
                  <span class="var-chip" onclick="copyVar(this,'{ep}')">&#123;ep&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{ejecutivo}')">&#123;ejecutivo&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{cliente}')">&#123;cliente&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{monto}')">&#123;monto&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{estado}')">&#123;estado&#125;</span>
                </div>
              </div>
              <div class="var-group">
                <div class="var-group-label">✅ Autorizada</div>
                <div class="var-chips">
                  <span class="var-chip" onclick="copyVar(this,'{ep}')">&#123;ep&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{cliente}')">&#123;cliente&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{margen}')">&#123;margen&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{ejecutivo}')">&#123;ejecutivo&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{supervisor}')">&#123;supervisor&#125;</span>
                </div>
              </div>
              <div class="var-group">
                <div class="var-group-label">↩️ Removido</div>
                <div class="var-chips">
                  <span class="var-chip" onclick="copyVar(this,'{ep}')">&#123;ep&#125;</span>
                  <span class="var-chip" onclick="copyVar(this,'{cliente}')">&#123;cliente&#125;</span>
                </div>
              </div>
              <div class="var-group" style="margin-top:8px;padding-top:8px;border-top:1px solid #e2e8f0;">
                <div class="var-group-label">✨ Formato Telegram</div>
                <div class="var-chips">
                  <span class="fmt-chip" onclick="copyVar(this,'*texto*')">*negrita*</span>
                  <span class="fmt-chip" onclick="copyVar(this,'_texto_')">_cursiva_</span>
                  <span class="fmt-chip" onclick="copyVar(this,'`texto`')">`monospace`</span>
                </div>
              </div>
            </div>
            <script>
            function copyVar(el,txt){
                navigator.clipboard.writeText(txt).catch(function(){
                    var ta=document.createElement('textarea');
                    ta.value=txt;ta.style.position='fixed';ta.style.top='-999px';
                    document.body.appendChild(ta);ta.select();
                    document.execCommand('copy');ta.remove();
                });
                var orig=el.innerHTML;
                el.classList.add('copied');el.innerHTML='✓ copiado';
                setTimeout(function(){el.classList.remove('copied');el.innerHTML=orig;},1200);
            }
            </script>
            """, unsafe_allow_html=True)

            # 3 columnas para los mensajes
            _msg_configs = [
                ('msg_nueva_cotizacion', "🆕 Nueva cotización", "supervisores/admins/obs.", "Al guardar cotización"),
                ('msg_autorizada',       "✅ Cotización autorizada", "ejecutivo + obs.", "Al guardar con margen"),
                ('msg_margen_removido',  "↩️ Margen removido", "ejecutivo", "Al quitar margen"),
            ]
            _msgs_nuevos = {}
            _mcol1, _mcol2, _mcol3 = st.columns(3)
            for (_mk, _mtitulo, _mdest, _mcuando), _mcol in zip(_msg_configs, [_mcol1, _mcol2, _mcol3]):
                _mval = _get_notif_config(_mk, _msg_defaults[_mk])
                with _mcol:
                    st.markdown(f"""
                    <div style='margin-bottom:6px'>
                        <div style='font-size:0.85rem;font-weight:700'>{_mtitulo}</div>
                        <div style='font-size:0.72rem;color:#94a3b8'>→ {_mdest}</div>
                        <div style='font-size:0.7rem;color:#cbd5e1;font-style:italic'>{_mcuando}</div>
                    </div>""", unsafe_allow_html=True)
                    _msgs_nuevos[_mk] = st.text_area("", value=_mval, height=400,
                                                      key=f"msg_{_mk}", label_visibility="collapsed")
            _mb1, _mb2 = st.columns([1, 1])
            with _mb1:
                if st.button("↩️ Restaurar por defecto", key="btn_restaurar_msgs"):
                    for _mk, _mdef in _msg_defaults.items():
                        _set_notif_config(_mk, _mdef)
                    st.success("✅ Mensajes restaurados")
                    st.rerun()
            with _mb2:
                if st.button("💾 Guardar mensajes", key="btn_guardar_msgs", type="primary", use_container_width=True):
                    for _mk, _mv in _msgs_nuevos.items():
                        _set_notif_config(_mk, _mv)
                    st.success("✅ Mensajes guardados")
