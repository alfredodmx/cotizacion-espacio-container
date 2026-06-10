"""
Generadores de PDF de cotizacion: completo (interno) y cliente (sin precios internos).
"""
import io
import random
from datetime import datetime

import streamlit as st
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, Image)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from services.calendario_service import sumar_dias_habiles
from services.excel_service import _get_excel_bytes_activo
from utils.telefono import formatear_telefono
from utils.formato import formato_clp


def _generar_qr_imagen(url: str, size: int = 80):
    """Genera un QR code como BytesIO + tamaño."""
    try:
        import qrcode as _qr
        qr_img = _qr.make(url)
        buf = io.BytesIO()
        qr_img.save(buf, format='PNG')
        buf.seek(0)
        return buf, size
    except Exception:
        return None, size


def _construir_texto_cliente_pdf(datos_cliente: dict, style) -> Paragraph:
    """Construye parrafos del bloque cliente para PDFs, mostrando empresa si es juridica."""
    d = datos_cliente
    tipo = d.get("TipoCliente", "natural")
    lines = []
    if d.get("Nombre"):      lines.append(f"<b>Nombre:</b> {d['Nombre']}")
    if d.get("RUT"):         lines.append(f"<b>RUT:</b> {d['RUT']}")
    if tipo == "juridica":
        if d.get("EmpresaCliente"): lines.append(f"<b>Empresa:</b> {d['EmpresaCliente']}")
        if d.get("RutEmpresa"):     lines.append(f"<b>RUT empresa:</b> {d['RutEmpresa']}")
    if d.get("Correo"):      lines.append(f"<b>Correo:</b> {d['Correo']}")
    if d.get("Teléfono"):    lines.append(f"<b>Teléfono:</b> {d['Teléfono']}")
    if d.get("Dirección"):
        dir_completa = d["Dirección"]
        if d.get("ComunaCliente"): dir_completa += f", {d['ComunaCliente']}"
        if d.get("RegionCliente"): dir_completa += f", {d['RegionCliente']}"
        lines.append(f"<b>Dirección:</b> {dir_completa}")
    if d.get("DireccionProyecto"):
        inst = d["DireccionProyecto"]
        if d.get("ComunaProyecto"): inst += f", {d['ComunaProyecto']}"
        if d.get("RegionProyecto"): inst += f", {d['RegionProyecto']}"
        lines.append(f"<b>Dirección instalación:</b> {inst}")
    if d.get("Observaciones"):
        lines.append(f"<b>Descripción del proyecto:</b> {d['Observaciones']}")
    return Paragraph("<br/>".join(lines), style)


@st.cache_data(ttl=300)
def cargar_visibilidad_impresion() -> dict:
    """Carga hoja Impresion del Excel activo. Retorna dict {item_lower: 'Mostrar'|'Ocultar'}."""
    try:
        import openpyxl as _opx
        src = _get_excel_bytes_activo()
        wb = _opx.load_workbook(src, data_only=True, read_only=True)
        if 'Impresion' not in wb.sheetnames:
            return {}
        ws = wb['Impresion']
        vis = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, cat, pdf = (row[0], row[1], row[2]) if len(row) >= 3 else (None, None, None)
            if item and pdf:
                vis[str(item).strip().lower()] = str(pdf).strip()
        return vis
    except Exception:
        return {}


def generar_pdf_completo(carrito_df, subtotal, iva, total, datos_cliente,
                         fecha_inicio, fecha_termino, dias_validez,
                         datos_asesor, margen=0, numero_cotizacion=None, mostrar_precios=False,
                         fecha_adjudicacion=None, fecha_fidelizacion=None, plazo_obra_dias=45):
    """Genera PDF interno (con o sin precios de compras). Retorna (BytesIO, numero_ep)."""
    buffer = io.BytesIO()

    def _watermark_canvas(canvas_obj, doc_obj):
        canvas_obj.saveState()
        pw, ph = A4
        canvas_obj.setFont('Helvetica-Bold', 28)
        canvas_obj.setFillColorRGB(0.75, 0.05, 0.05)
        canvas_obj.setFillAlpha(0.10)
        for wy in range(-100, int(ph) + 200, 120):
            for wx in range(-100, int(pw) + 200, 220):
                canvas_obj.saveState()
                canvas_obj.translate(wx, wy)
                canvas_obj.rotate(45)
                canvas_obj.drawCentredString(0, 0, 'COMPRAS')
                canvas_obj.restoreState()
        canvas_obj.restoreState()

    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=20, rightMargin=20,
                            topMargin=30, bottomMargin=30, allowSplitting=1)
    elements = []
    styles = getSampleStyleSheet()

    for name, kw in [
        ('SmallFont',         dict(fontSize=8, leading=12, wordWrap='CJK', leftIndent=0, alignment=0)),
        ('HeaderStyle',       dict(fontSize=9, leading=11, alignment=1, textColor=colors.white, fontName='Helvetica-Bold', leftIndent=0)),
        ('TituloPresupuesto', dict(fontSize=16, leading=20, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0)),
        ('TextoNormal',       dict(fontSize=10, leading=14, leftIndent=0, alignment=0, spaceAfter=2)),
        ('TituloSeccion',     dict(fontSize=12, leading=14, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0)),
        ('NotasEstilo',       dict(fontSize=9, leading=13, leftIndent=0, alignment=0, textColor=colors.HexColor('#0d2266'), spaceAfter=2)),
        ('TotalLabel',        dict(fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2)),
        ('TotalValue',        dict(fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2)),
        ('TotalBold',         dict(fontSize=12, leading=16, alignment=2, fontName='Helvetica-Bold', spaceAfter=2, textColor=colors.black)),
        ('EmpresaNombre',     dict(fontSize=10, fontName='Helvetica-Bold', leading=13, textColor=colors.HexColor('#0d2266'), spaceAfter=2)),
        ('QRLabel',           dict(fontSize=7, leading=9, textColor=colors.HexColor('#6b7280'), alignment=1)),
        ('EPTitulo',          dict(fontSize=10, fontName='Helvetica-Bold', leading=13, textColor=colors.HexColor('#0d2266'), spaceAfter=2)),
    ]:
        styles.add(ParagraphStyle(name=name, parent=styles['Normal'], **kw))

    numero_ep = numero_cotizacion or f"EP-{random.randint(1000,9999)}"
    fecha_emision = datetime.now()

    col1 = doc.width * 0.45
    col2 = doc.width * 0.55

    _qr_img, _qr_sz = _generar_qr_imagen("https://www2.sii.cl/stc/noauthz/consulta", size=70)
    try:
        logo = Image("logo.png")
        logo_max = col2 * 0.80
        logo.drawWidth = logo_max
        logo.drawHeight = logo_max * (logo.imageHeight / float(logo.imageWidth))
        logo_cell = logo
    except Exception:
        logo_cell = Paragraph("", styles['EPTitulo'])

    if mostrar_precios and not fecha_adjudicacion:
        styles.add(ParagraphStyle(name='SinAdj', parent=styles['Normal'],
            fontSize=7.5, fontName='Helvetica-Bold', leading=10,
            textColor=colors.HexColor('#6b7280'), alignment=1))
        sin_adj = Table(
            [[Paragraph('Proyecto sin', styles['SinAdj'])],
             [Paragraph('adjudicación', styles['SinAdj'])],
             [Paragraph('Sin fecha de entrega', styles['SinAdj'])]],
            colWidths=[90])
        sin_adj.setStyle(TableStyle([
            ('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#e2e8f0')),
            ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8fafc')),
        ]))
        qr_cell = sin_adj; _qr_sz = 90
    elif mostrar_precios and fecha_adjudicacion:
        styles.add(ParagraphStyle(name='FechaLbl', parent=styles['Normal'],
            fontSize=7, fontName='Helvetica-Bold', leading=9, textColor=colors.HexColor('#6b7280')))
        styles.add(ParagraphStyle(name='FechaVal', parent=styles['Normal'],
            fontSize=8.5, fontName='Helvetica-Bold', leading=11, textColor=colors.HexColor('#0d2266')))
        try:
            fadj = (datetime.fromisoformat(fecha_adjudicacion.replace('Z','')).date()
                    if isinstance(fecha_adjudicacion, str) else fecha_adjudicacion)
            f_ent = sumar_dias_habiles(fadj, plazo_obra_dias)
            dias_r = (f_ent - fadj).days
            meses = ['enero','febrero','marzo','abril','mayo','junio',
                     'julio','agosto','septiembre','octubre','noviembre','diciembre']
            fent_fmt = f"{f_ent.day} {meses[f_ent.month-1]} {f_ent.year}"
            fechas_tbl = Table([
                [Paragraph('Fecha adjudicación',styles['FechaLbl']), Paragraph(fadj.strftime('%d/%m/%Y'),styles['FechaVal'])],
                [Paragraph('Fecha de entrega',styles['FechaLbl']),   Paragraph(fent_fmt,styles['FechaVal'])],
                [Paragraph('Días hábiles',styles['FechaLbl']),        Paragraph(f'{plazo_obra_dias} días hábiles',styles['FechaVal'])],
                [Paragraph('Días reales',styles['FechaLbl']),          Paragraph(f'{dias_r} días',styles['FechaVal'])],
            ], colWidths=[55,65])
            fechas_tbl.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(0,-1),'RIGHT'),('ALIGN',(1,0),(1,-1),'LEFT'),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
                ('LINEBELOW',(0,0),(-1,-2),0.3,colors.HexColor('#e2e8f0')),
                ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#f8faff')),
                ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#1e2447')),
            ]))
            qr_cell = fechas_tbl; _qr_sz = 120
        except Exception:
            qr_cell = Paragraph("", styles['EPTitulo']); _qr_sz = 70
    elif _qr_img:
        from reportlab.platypus import Image as _RLImage
        _qr_img.name = 'qr.png'
        qr_rl = _RLImage(_qr_img, width=_qr_sz, height=_qr_sz)
        qr_cell = Table([[qr_rl],[Paragraph("SII Verifíquenos",styles['QRLabel'])]], colWidths=[_qr_sz+4])
        qr_cell.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
            ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
    else:
        qr_cell = Paragraph("", styles['EPTitulo'])

    qr_col_w  = _qr_sz + 10
    fila1 = Table([[logo_cell, "", qr_cell]], colWidths=[col1, col2-qr_col_w, qr_col_w])
    fila1.setStyle(TableStyle([
        ('ALIGN',(0,0),(0,0),'LEFT'),('ALIGN',(2,0),(2,0),'RIGHT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#e2e8f0')),
    ]))
    elements.append(fila1)

    fila2 = Table([[
        Paragraph(
            f"<font name='Helvetica-Bold' size='14' color='#0d2266'>PRESUPUESTO Nº {numero_ep}</font><br/>"
            f"<font size='9' color='#374151'><b>Fecha Emisión:</b> {fecha_emision.strftime('%d-%m-%Y')}</font><br/>"
            f"<font size='9' color='#374151'><b>Validez:</b> {fecha_inicio.strftime('%d-%m-%Y')} hasta "
            f"{fecha_termino.strftime('%d-%m-%Y')} ({dias_validez} días)</font>", styles['EPTitulo']),
        Paragraph(
            "<b>INVERSIONES CONTAINER HOUSE SPA</b><br/>"
            "RUT: 78.268.851-0<br/>Construcción y Acondicionamiento Containers<br/>"
            "Villasana 2039, Quinta Normal, Santiago", styles['EmpresaNombre'])
    ]], colWidths=[col1, col2])
    fila2.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),('ALIGN',(0,0),(0,-1),'LEFT'),('ALIGN',(1,0),(1,-1),'RIGHT'),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#e2e8f0')),
    ]))
    elements.append(fila2)
    elements.append(Spacer(1, 10))

    asesor_text = "".join(
        f"<b>{k}:</b> {formatear_telefono(v) if k == 'Teléfono Ejecutivo' and v else v}<br/>"
        for k, v in datos_asesor.items() if v
    )
    tabla_ca = Table([
        [Paragraph("<b>DATOS DEL CLIENTE</b>", styles['TituloSeccion']),
         Paragraph("<b>DATOS DEL ASESOR</b>",  styles['TituloSeccion'])],
        [_construir_texto_cliente_pdf(datos_cliente, styles['TextoNormal']),
         Paragraph(asesor_text, styles['TextoNormal'])],
    ], colWidths=[col1, col2])
    tabla_ca.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]))
    elements.append(tabla_ca)
    elements.append(Spacer(1, 20))

    ancho = doc.width
    if mostrar_precios:
        pcts = [15, 50, 8, 13.5, 13.5]
        anchos = [ancho*p/100 for p in pcts]
        data = [[Paragraph(f"<b>{h}</b>", styles['HeaderStyle'])
                 for h in ['Categoría','Item','Cant.','P. Unitario','Subtotal']]]
        for _, row in carrito_df.iterrows():
            if str(row.get("Categoria","")).strip().lower() == "varios": continue
            data.append([
                Paragraph(row["Categoria"], styles['SmallFont']),
                Paragraph(row["Item"],      styles['SmallFont']),
                Paragraph(str(row["Cantidad"]), styles['SmallFont']),
                Paragraph(formato_clp(row["Precio Unitario"]), styles['SmallFont']),
                Paragraph(formato_clp(row["Subtotal"]),        styles['SmallFont']),
            ])
        align_col = ('ALIGN',(2,1),(4,-1),'RIGHT')
    else:
        anchos = [ancho*p/100 for p in [15,70,15]]
        data = [[Paragraph(f"<b>{h}</b>", styles['HeaderStyle'])
                 for h in ['Categoría','Item','Cantidad']]]
        for _, row in carrito_df.iterrows():
            if str(row.get("Categoria","")).strip().lower() == "varios": continue
            data.append([
                Paragraph(row["Categoria"], styles['SmallFont']),
                Paragraph(row["Item"],      styles['SmallFont']),
                Paragraph(str(row["Cantidad"]), styles['SmallFont']),
            ])
        align_col = ('ALIGN',(2,1),(2,-1),'CENTER')

    tbl_prod = Table(data, colWidths=anchos, repeatRows=1, splitByRow=1)
    tbl_prod.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), colors.black),
        ('TEXTCOLOR',(0,0),(-1,0), colors.white),
        ('ALIGN',(0,0),(-1,0),'CENTER'),
        ('FONTSIZE',(0,0),(-1,0),9),
        ('BOTTOMPADDING',(0,0),(-1,0),8),('TOPPADDING',(0,0),(-1,0),8),
        ('BACKGROUND',(0,1),(-1,-1), colors.Color(1,1,1,alpha=0.55)),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        align_col,('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),2),('RIGHTPADDING',(0,0),(-1,-1),2),
        ('TOPPADDING',(0,1),(-1,-1),4),('BOTTOMPADDING',(0,1),(-1,-1),4),
    ]))
    for i in range(1, len(data)):
        if i % 2 == 0:
            tbl_prod.setStyle(TableStyle([('BACKGROUND',(0,i),(-1,i),colors.Color(0.95,0.95,0.95,alpha=0.55))]))
    elements.append(tbl_prod)
    elements.append(Spacer(1, 20))

    bloque_w = (ancho - 20) / 2
    totales_data = [
        [Paragraph("Subtotal:", styles['TotalLabel']), Paragraph(formato_clp(subtotal), styles['TotalValue'])],
        [Paragraph("IVA (19%):", styles['TotalLabel']), Paragraph(formato_clp(iva),      styles['TotalValue'])],
        [Paragraph("TOTAL:",     styles['TotalBold']),  Paragraph(formato_clp(total),    styles['TotalBold'])],
    ]
    totales_tbl = Table(totales_data, colWidths=[100,120])
    totales_tbl.setStyle(TableStyle([
        ('ALIGN',(0,0),(0,-1),'RIGHT'),('ALIGN',(1,0),(1,-1),'RIGHT'),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LINEABOVE',(1,1),(1,1),1,colors.grey),('LINEABOVE',(1,2),(1,2),2,colors.black),
    ]))

    if not mostrar_precios:
        p50  = formato_clp(round(total*0.50))
        p25a = formato_clp(round(total*0.25))
        p25b = formato_clp(round(total*0.25))
        notas = (f"<font color='#0d2266'><b>NOTAS IMPORTANTES:</b></font><br/>"
                 f"1.- Valores incluyen IVA.<br/>"
                 f"2.- Transporte y bases de apoyo <b>no incluidos</b>.<br/>"
                 f"3.- Formas de pago: transferencia - pago contado.<br/>"
                 f"4.- Proceso de pagos:<br/>"
                 f"&nbsp;&nbsp;&nbsp;<b>- 50% inicial</b> correspondiente a <b>{p50}</b><br/>"
                 f"&nbsp;&nbsp;&nbsp;<b>- 25% obra</b> correspondiente a <b>{p25a}</b><br/>"
                 f"&nbsp;&nbsp;&nbsp;<b>- 25% entrega</b> correspondiente a <b>{p25b}</b>.")
        transf = ("&lt;font color='#0d2266'&gt;<b>DATOS PARA TRANSFERENCIA:</b></font><br/>"
                  "Inversiones Container House Spa<br/>RUT: 78.268.851-0<br/>"
                  "Tipo de cuenta: Cuenta Corriente<br/>Banco: Itaú<br/>"
                  "N° de cuenta: 230771767<br/>Correo: jperez@espaciocontainerhouse.cl")
        transf = ("<font color='#0d2266'><b>DATOS PARA TRANSFERENCIA:</b></font><br/>"
                  "Inversiones Container House Spa<br/>RUT: 78.268.851-0<br/>"
                  "Tipo de cuenta: Cuenta Corriente<br/>Banco: Itaú<br/>"
                  "N° de cuenta: 230771767<br/>Correo: jperez@espaciocontainerhouse.cl")
        cordial = Paragraph(
            "<para align='center'>"
            "<font name='Helvetica-Oblique' size='9' color='#718096'>"
            "¡Gracias por confiar en nosotros!<br/>Esperamos que este presupuesto<br/>"
            "supere tus expectativas.<br/><br/>Con cariño,<br/></font>"
            "<font name='Helvetica-BoldOblique' size='9' color='#0d2266'>"
            "La familia<br/>Espacio Container House<br/></font>"
            "<font name='Helvetica-Oblique' size='8' color='#9ca3af'>espaciocontainerhouse.cl</font>"
            "</para>", styles['NotasEstilo'])
        bloques = Table([
            [Paragraph(notas, styles['NotasEstilo']),  totales_tbl],
            [Paragraph(transf, styles['NotasEstilo']), cordial],
        ], colWidths=[bloque_w, bloque_w])
        bloques.setStyle(TableStyle([
            ('ALIGN',(0,0),(0,-1),'LEFT'),('ALIGN',(1,0),(1,-1),'RIGHT'),
            ('VALIGN',(0,0),(0,0),'TOP'),('VALIGN',(0,1),(0,1),'MIDDLE'),
            ('VALIGN',(1,0),(1,0),'MIDDLE'),('VALIGN',(1,1),(1,1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),0),('LEFTPADDING',(1,1),(1,1),40),
            ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
            ('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#e2e8f0')),
        ]))
        elements.append(bloques)
    else:
        elements.append(Spacer(1, 20))
        tbl_tot = Table([['', totales_tbl]], colWidths=[bloque_w, bloque_w])
        tbl_tot.setStyle(TableStyle([('ALIGN',(1,0),(1,0),'RIGHT'),('VALIGN',(0,0),(-1,-1),'TOP')]))
        elements.append(tbl_tot)

    if mostrar_precios:
        doc.build(elements, onFirstPage=_watermark_canvas, onLaterPages=_watermark_canvas)
    else:
        doc.build(elements)
    buffer.seek(0)
    return buffer, numero_ep


def generar_pdf_cliente(carrito_df, subtotal, iva, total, datos_cliente,
                        fecha_inicio, fecha_termino, dias_validez,
                        datos_asesor, margen=0, numero_cotizacion=None, descripciones_ep=None):
    """Genera PDF para el cliente (sin precios, con resumen por categoria). Retorna (BytesIO, ep)."""
    _descs = descripciones_ep or {}
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=20, rightMargin=20,
                            topMargin=30, bottomMargin=30, allowSplitting=1)
    elements = []
    styles = getSampleStyleSheet()

    for name, kw in [
        ('SmallFont',         dict(fontSize=8, leading=12, wordWrap='CJK', leftIndent=0, alignment=0)),
        ('HeaderStyle',       dict(fontSize=9, leading=11, alignment=1, textColor=colors.white, fontName='Helvetica-Bold', leftIndent=0)),
        ('TextoNormal',       dict(fontSize=10, leading=14, leftIndent=0, alignment=0, spaceAfter=2)),
        ('TituloSeccion',     dict(fontSize=12, leading=14, fontName='Helvetica-Bold', spaceAfter=6, leftIndent=0, alignment=0)),
        ('NotasEstilo',       dict(fontSize=9, leading=13, leftIndent=0, alignment=0, textColor=colors.HexColor('#0d2266'), spaceAfter=2)),
        ('TotalLabel',        dict(fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2)),
        ('TotalValue',        dict(fontSize=10, leading=14, alignment=2, fontName='Helvetica', spaceAfter=2)),
        ('TotalBold',         dict(fontSize=12, leading=16, alignment=2, fontName='Helvetica-Bold', spaceAfter=2, textColor=colors.black)),
        ('EmpresaNombre',     dict(fontSize=10, fontName='Helvetica-Bold', leading=13, textColor=colors.HexColor('#0d2266'), spaceAfter=2)),
        ('QRLabel',           dict(fontSize=7, leading=9, textColor=colors.HexColor('#6b7280'), alignment=1)),
        ('EPTitulo',          dict(fontSize=10, fontName='Helvetica-Bold', leading=13, textColor=colors.HexColor('#0d2266'), spaceAfter=2)),
    ]:
        styles.add(ParagraphStyle(name=name, parent=styles['Normal'], **kw))

    numero_ep = numero_cotizacion or f"EP-{random.randint(1000,9999)}"
    fecha_emision = datetime.now()
    col1 = doc.width * 0.45
    col2 = doc.width * 0.55

    qr_img, qr_sz = _generar_qr_imagen("https://www2.sii.cl/stc/noauthz/consulta", size=70)
    try:
        logo = Image("logo.png")
        logo_max = col2 * 0.80
        logo.drawWidth = logo_max
        logo.drawHeight = logo_max * (logo.imageHeight / float(logo.imageWidth))
        logo_cell = logo
    except Exception:
        logo_cell = Paragraph("", styles['EPTitulo'])

    if qr_img:
        from reportlab.platypus import Image as _RLImage
        qr_img.name = 'qr.png'
        qr_rl = _RLImage(qr_img, width=qr_sz, height=qr_sz)
        qr_cell = Table([[qr_rl],[Paragraph("SII Verifíquenos",styles['QRLabel'])]], colWidths=[qr_sz+4])
        qr_cell.setStyle(TableStyle([('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
            ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
    else:
        qr_cell = Paragraph("", styles['EPTitulo'])

    qr_col_w = qr_sz + 10
    fila1 = Table([[logo_cell,"",qr_cell]], colWidths=[col1, col2-qr_col_w, qr_col_w])
    fila1.setStyle(TableStyle([
        ('ALIGN',(0,0),(0,0),'LEFT'),('ALIGN',(2,0),(2,0),'RIGHT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#e2e8f0')),
    ]))
    elements.append(fila1)

    fila2 = Table([[
        Paragraph(
            f"<font name='Helvetica-Bold' size='14' color='#0d2266'>PRESUPUESTO Nº {numero_ep}</font><br/>"
            f"<font size='9' color='#374151'><b>Fecha Emisión:</b> {fecha_emision.strftime('%d-%m-%Y')}</font><br/>"
            f"<font size='9' color='#374151'><b>Validez:</b> {fecha_inicio.strftime('%d-%m-%Y')} hasta "
            f"{fecha_termino.strftime('%d-%m-%Y')} ({dias_validez} días)</font>", styles['EPTitulo']),
        Paragraph(
            "<b>INVERSIONES CONTAINER HOUSE SPA</b><br/>RUT: 78.268.851-0<br/>"
            "Construcción y Acondicionamiento Containers<br/>Villasana 2039, Quinta Normal, Santiago",
            styles['EmpresaNombre'])
    ]], colWidths=[col1, col2])
    fila2.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),('ALIGN',(0,0),(0,-1),'LEFT'),('ALIGN',(1,0),(1,-1),'RIGHT'),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#e2e8f0')),
    ]))
    elements.append(fila2)
    elements.append(Spacer(1, 10))

    asesor_text = "".join(
        f"<b>{k}:</b> {formatear_telefono(v) if k == 'Teléfono Ejecutivo' and v else v}<br/>"
        for k, v in datos_asesor.items() if v
    )
    tabla_ca = Table([
        [Paragraph("<b>DATOS DEL CLIENTE</b>",styles['TituloSeccion']),
         Paragraph("<b>DATOS DEL ASESOR</b>",styles['TituloSeccion'])],
        [_construir_texto_cliente_pdf(datos_cliente, styles['TextoNormal']),
         Paragraph(asesor_text, styles['TextoNormal'])],
    ], colWidths=[col1, col2])
    tabla_ca.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]))
    elements.append(tabla_ca)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("<b>RESUMEN POR CATEGORÍA:</b>", styles['TituloSeccion']))
    elements.append(Spacer(1, 10))

    vis_map = cargar_visibilidad_impresion()
    data_res = []
    for categoria, grupo in carrito_df.groupby('Categoria'):
        desc = (_descs or {}).get(categoria, '').strip()
        if desc:
            data_res.append([Paragraph(categoria, styles['SmallFont']),
                             Paragraph(desc.replace('\n','<br/>'), styles['SmallFont'])])
        else:
            items = [i for i in grupo['Item'].tolist()
                     if vis_map.get(str(i).strip().lower(), 'Mostrar') == 'Mostrar']
            if not items:
                continue
            data_res.append([Paragraph(categoria, styles['SmallFont']),
                             Paragraph("<br/>".join(f"• {i}" for i in items), styles['SmallFont'])])

    ancho_cat  = doc.width * 0.25
    ancho_desc = doc.width * 0.75
    tbl_res = Table(
        [[Paragraph("<b>Categoría</b>",styles['HeaderStyle']),
          Paragraph("<b>Descripción</b>",styles['HeaderStyle'])]] + data_res,
        colWidths=[ancho_cat, ancho_desc], repeatRows=1)
    tbl_res.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.black),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,0),9),('BOTTOMPADDING',(0,0),(-1,0),8),('TOPPADDING',(0,0),(-1,0),8),
        ('BACKGROUND',(0,1),(-1,-1),colors.white),('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
        ('TOPPADDING',(0,1),(-1,-1),5),('BOTTOMPADDING',(0,1),(-1,-1),5),
    ]))
    for i in range(1, len(data_res)+1):
        if i % 2 == 0:
            tbl_res.setStyle(TableStyle([('BACKGROUND',(0,i),(-1,i),colors.Color(0.95,0.95,0.95))]))
    elements.append(tbl_res)
    elements.append(Spacer(1, 20))

    bloque_w = (doc.width - 20) / 2
    p50  = formato_clp(round(total*0.50))
    p25a = formato_clp(round(total*0.25))
    p25b = formato_clp(round(total*0.25))
    notas = (f"<font color='#0d2266'><b>NOTAS IMPORTANTES:</b></font><br/>"
             f"1.- Valores incluyen IVA.<br/>"
             f"2.- Transporte y bases de apoyo <b>no incluidos</b>.<br/>"
             f"3.- Formas de pago: transferencia - pago contado.<br/>"
             f"4.- Proceso de pagos:<br/>"
             f"&nbsp;&nbsp;&nbsp;<b>- 50% inicial</b> correspondiente a <b>{p50}</b><br/>"
             f"&nbsp;&nbsp;&nbsp;<b>- 25% obra</b> correspondiente a <b>{p25a}</b><br/>"
             f"&nbsp;&nbsp;&nbsp;<b>- 25% entrega</b> correspondiente a <b>{p25b}</b>.")
    transf = ("<font color='#0d2266'><b>DATOS PARA TRANSFERENCIA:</b></font><br/>"
              "Inversiones Container House Spa<br/>RUT: 78.268.851-0<br/>"
              "Tipo de cuenta: Cuenta Corriente<br/>Banco: Itaú<br/>"
              "N° de cuenta: 230771767<br/>Correo: jperez@espaciocontainerhouse.cl")
    totales_data = [
        [Paragraph("Subtotal:",styles['TotalLabel']), Paragraph(formato_clp(subtotal),styles['TotalValue'])],
        [Paragraph("IVA (19%):",styles['TotalLabel']), Paragraph(formato_clp(iva),styles['TotalValue'])],
        [Paragraph("TOTAL:",styles['TotalBold']),  Paragraph(formato_clp(total),styles['TotalBold'])],
    ]
    totales_tbl = Table(totales_data, colWidths=[100,120])
    totales_tbl.setStyle(TableStyle([
        ('ALIGN',(0,0),(0,-1),'RIGHT'),('ALIGN',(1,0),(1,-1),'RIGHT'),
        ('LINEABOVE',(1,1),(1,1),1,colors.grey),('LINEABOVE',(1,2),(1,2),2,colors.black),
    ]))
    cordial = Paragraph(
        "<para align='center'><font name='Helvetica-Oblique' size='9' color='#718096'>"
        "¡Gracias por confiar en nosotros!<br/>Esperamos que este presupuesto<br/>"
        "supere tus expectativas.<br/><br/>Con cariño,<br/></font>"
        "<font name='Helvetica-BoldOblique' size='9' color='#0d2266'>La familia<br/>"
        "Espacio Container House<br/></font>"
        "<font name='Helvetica-Oblique' size='8' color='#9ca3af'>espaciocontainerhouse.cl</font>"
        "</para>", styles['NotasEstilo'])
    bloques = Table([
        [Paragraph(notas,  styles['NotasEstilo']), totales_tbl],
        [Paragraph(transf, styles['NotasEstilo']), cordial],
    ], colWidths=[bloque_w, bloque_w])
    bloques.setStyle(TableStyle([
        ('ALIGN',(0,0),(0,-1),'LEFT'),('ALIGN',(1,0),(1,-1),'RIGHT'),
        ('VALIGN',(0,0),(0,0),'TOP'),('VALIGN',(1,0),(1,0),'MIDDLE'),
        ('VALIGN',(0,1),(0,1),'MIDDLE'),('VALIGN',(1,1),(1,1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),0),('LEFTPADDING',(1,1),(1,1),40),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LINEBELOW',(0,0),(-1,0),0.5,colors.HexColor('#e2e8f0')),
    ]))
    elements.append(bloques)
    doc.build(elements)
    buffer.seek(0)
    return buffer, numero_ep
