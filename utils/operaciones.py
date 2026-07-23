"""
Funciones de operaciones: tabla RC, balance Excel/PDF, HTML builder.
Extraídas de app.py para la arquitectura modular.
"""

# ── Iconos SVG inline para el iframe del formulario (reemplazan emoticones) ──
_RC_ICONS = {
    "paperclip": '<path d="m21.44 11.05-9.19 9.19a6 6 0 0 1-8.49-8.49l8.57-8.57A4 4 0 1 1 18 8.84l-8.59 8.57a2 2 0 0 1-2.83-2.83l8.49-8.48"/>',
    "store":     '<path d="m2 7 4.41-4.41A2 2 0 0 1 7.83 2h8.34a2 2 0 0 1 1.42.59L22 7"/><path d="M4 12v8a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2v-8"/><path d="M2 7h20"/><path d="M18 12v.01"/><path d="M6 12v.01"/>',
    "cart":      '<circle cx="8" cy="21" r="1"/><circle cx="19" cy="21" r="1"/><path d="M2.05 2.05h2l2.66 12.42a2 2 0 0 0 2 1.58h9.78a2 2 0 0 0 1.95-1.57l1.65-7.43H5.12"/>',
    "calendar":  '<rect width="18" height="18" x="3" y="4" rx="2" ry="2"/><line x1="16" x2="16" y1="2" y2="6"/><line x1="8" x2="8" y1="2" y2="6"/><line x1="3" x2="21" y1="10" y2="10"/>',
    "clipboard": '<rect width="8" height="4" x="8" y="2" rx="1" ry="1"/><path d="M16 4h2a2 2 0 0 1 2 2v14a2 2 0 0 1-2 2H6a2 2 0 0 1-2-2V6a2 2 0 0 1 2-2h2"/>',
    "note":      '<path d="M11 4H4a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2v-7"/><path d="M18.5 2.5a2.12 2.12 0 0 1 3 3L12 15l-4 1 1-4Z"/>',
    "save":      '<path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"/><polyline points="17 21 17 13 7 13 7 21"/><polyline points="7 3 7 8 15 8"/>',
    "x":         '<path d="M18 6 6 18"/><path d="m6 6 12 12"/>',
    "check":     '<path d="M20 6 9 17l-5-5"/>',
    "file":      '<path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/>',
    "user":      '<path d="M19 21v-2a4 4 0 0 0-4-4H9a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/>',
    "trend-up":  '<polyline points="22 7 13.5 15.5 8.5 10.5 2 17"/><polyline points="16 7 22 7 22 13"/>',
    "trend-down":'<polyline points="22 17 13.5 8.5 8.5 13.5 2 7"/><polyline points="16 17 22 17 22 11"/>',
    "edit":      '<path d="M11 4H4a2 2 0 0 0-2 2v14a2 2 0 0 0 2 2h14a2 2 0 0 0 2-2v-7"/><path d="M18.5 2.5a2.12 2.12 0 0 1 3 3L12 15l-4 1 1-4Z"/>',
    "trash":     '<path d="M3 6h18"/><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"/>',
    "chevron":   '<polyline points="6 9 12 15 18 9"/>',
    "alert":     '<path d="m21.73 18-8-14a2 2 0 0 0-3.48 0l-8 14A2 2 0 0 0 4 21h16a2 2 0 0 0 1.73-3Z"/><line x1="12" x2="12" y1="9" y2="13"/><line x1="12" x2="12.01" y1="17" y2="17"/>',
    "history":   '<circle cx="12" cy="12" r="10"/><polyline points="12 6 12 12 16 14"/>',
    "eye":       '<path d="M2 12s3-7 10-7 10 7 10 7-3 7-10 7-10-7-10-7Z"/><circle cx="12" cy="12" r="3"/>',
    "package":   '<path d="M11 21.73a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73z"/><path d="M3.3 7 12 12l8.7-5"/><path d="M12 22V12"/>',
    "folder":    '<path d="M20 20a2 2 0 0 0 2-2V8a2 2 0 0 0-2-2h-7.9a2 2 0 0 1-1.69-.9L9.6 3.9A2 2 0 0 0 7.93 3H4a2 2 0 0 0-2 2v13a2 2 0 0 0 2 2Z"/>',
}


def _svg_rc(name, color="currentColor", size=14, mr=6, valign=-2, sw=2):
    """SVG inline para el iframe del formulario RC."""
    inner = _RC_ICONS.get(name, "")
    return (
        f'<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" '
        f'stroke="{color}" stroke-width="{sw}" stroke-linecap="round" '
        f'stroke-linejoin="round" style="vertical-align:{valign}px;margin-right:{mr}px;flex-shrink:0;">'
        f'{inner}</svg>'
    )


# CSS de la cuadrícula 2×2 del drawer + del dropdown HTML de proyecto. Va como
# constante de MÓDULO (llaves reales) y se antepone al HTML del iframe, para no
# pelear con el escape de llaves del f-string gigante de build_historial_rc_html.
_RC_GRID_CSS = """<style>
html,body{margin:0;padding:0;height:100%;}
.rcg{display:grid;grid-template-columns:minmax(0,41fr) minmax(0,59fr);
  grid-template-rows:minmax(0,1fr) minmax(0,1fr);gap:11px;height:100vh;
  box-sizing:border-box;padding:2px;font-family:Montserrat,'Segoe UI',sans-serif;}
.rcg-cell{border:1px solid #e6eaf5;border-radius:14px;background:#fff;display:flex;
  flex-direction:column;overflow:hidden;min-height:0;min-width:0;box-shadow:0 1px 3px rgba(15,23,42,0.05);}
.rcg-proj{background:#fbfcff;}
.rcg-head{display:flex;align-items:center;font-family:Montserrat,sans-serif;font-weight:700;
  font-size:0.8rem;letter-spacing:0.05em;text-transform:uppercase;color:#0f172a;
  padding:13px 15px 9px;flex:0 0 auto;}
.rcg-body{flex:1 1 auto;min-height:0;overflow-y:auto;overflow-x:hidden;padding:2px 15px 15px;}
.rcg-body::-webkit-scrollbar{width:8px;}
.rcg-body::-webkit-scrollbar-thumb{background:#cbd5e1;border-radius:8px;}
/* ── Dropdown HTML de proyecto (0 reruns al abrir/buscar/elegir) ── */
.rcpk{font-family:Montserrat,'Segoe UI',sans-serif;}
.rcpk-trg{display:flex;align-items:center;gap:10px;width:100%;box-sizing:border-box;padding:11px 13px;
  border:1.5px solid #e2e8f0;border-radius:12px;background:#fff;cursor:pointer;
  box-shadow:0 1px 3px rgba(15,23,42,0.05);transition:border-color .15s,box-shadow .15s;}
.rcpk-trg:hover{border-color:#5b7cfa;box-shadow:0 6px 16px rgba(91,124,250,0.14);}
.rcpk.open .rcpk-trg{border-color:#5b7cfa;border-bottom-left-radius:0;border-bottom-right-radius:0;}
.rcpk-txt{flex:1;min-width:0;text-align:left;}
.rcpk-t1{font-weight:800;font-size:0.85rem;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.rcpk-t2{font-size:0.7rem;color:#64748b;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:2px;}
.rcpk-chev{flex:0 0 auto;transition:transform .18s;line-height:0;}
.rcpk.open .rcpk-chev{transform:rotate(180deg);}
.rcpk-panel{display:none;border:1.5px solid #5b7cfa;border-top:none;border-radius:0 0 12px 12px;
  background:#fff;box-shadow:0 14px 30px rgba(15,23,42,0.12);overflow:hidden;}
.rcpk.open .rcpk-panel{display:block;}
.rcpk-search{width:100%;box-sizing:border-box;border:none;border-bottom:1px solid #eef2f7;
  padding:10px 13px;font-size:0.82rem;font-family:inherit;color:#0f172a;outline:none;background:#f8fafc;}
.rcpk-list{max-height:230px;overflow-y:auto;}
.rcpk-opt{display:flex;align-items:center;gap:9px;padding:9px 13px;cursor:pointer;border-bottom:1px solid #f4f6fb;}
.rcpk-opt:hover{background:#f5f7ff;}
.rcpk-opt[data-sel="1"]{background:#eef2ff;box-shadow:inset 3px 0 0 #5b7cfa;}
.rcpk-dot{flex:0 0 auto;width:9px;height:9px;border-radius:50%;}
.rcpk-oi{flex:1;min-width:0;}
.rcpk-o1{font-weight:700;font-size:0.8rem;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.rcpk-o2{font-size:0.68rem;color:#64748b;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:2px;}
.rcpk-pct{flex:0 0 auto;font-size:0.62rem;font-weight:800;letter-spacing:.04em;text-transform:uppercase;
  padding:3px 7px;border-radius:6px;background:#f1f5f9;color:#475569;}
.rcpk-empty{display:none;padding:16px;font-size:0.8rem;color:#94a3b8;text-align:center;}
.rcpk-btn{display:flex;align-items:center;justify-content:center;gap:7px;margin-top:9px;padding:11px;
  border-radius:12px;font-family:Montserrat,sans-serif;font-weight:800;font-size:0.76rem;
  letter-spacing:.05em;text-transform:uppercase;cursor:pointer;transition:transform .12s,box-shadow .15s,background .15s;}
.rcpk-load{color:#fff;background:linear-gradient(135deg,#5b7cfa,#4361ee);box-shadow:0 6px 18px rgba(91,124,250,0.30);}
.rcpk-load:hover{transform:translateY(-1px);box-shadow:0 10px 24px rgba(91,124,250,0.38);}
.rcpk-load[data-dis="1"]{opacity:.45;pointer-events:none;box-shadow:none;}
.rcpk-exit{color:#b91c1c;background:#fff;border:1.5px solid #fecaca;}
.rcpk-exit:hover{background:#fef2f2;border-color:#f87171;}
.rcpk-close{color:#475569;background:#fff;border:1.5px solid #e2e8f0;}
.rcpk-close:hover{background:#f8fafc;border-color:#cbd5e1;}
/* ── Mini-resumen del proyecto activo ── */
.rc-res{margin-top:14px;border-top:1px dashed #e2e8f0;padding-top:12px;}
.rc-res-row{display:flex;align-items:center;gap:8px;font-size:0.76rem;color:#475569;margin-bottom:7px;}
.rc-res-row b{color:#0f172a;font-weight:700;}
.rc-res-grid{display:grid;grid-template-columns:1fr 1fr;gap:8px;margin-top:10px;}
.rc-res-kpi{background:#f8fafc;border:1px solid #eef2f7;border-radius:10px;padding:9px 11px;}
.rc-res-kpi .k{font-size:0.6rem;font-weight:700;letter-spacing:.05em;text-transform:uppercase;color:#94a3b8;}
.rc-res-kpi .v{font-family:Montserrat,sans-serif;font-weight:800;font-size:0.92rem;color:#0f172a;margin-top:2px;}
.rc-res-empty{margin-top:14px;padding:16px 12px;border:1.5px dashed #e2e8f0;border-radius:12px;
  text-align:center;font-size:0.76rem;color:#94a3b8;}
</style>"""

# Comportamiento del dropdown + botones Cargar/Salir/Cerrar DENTRO del iframe del
# drawer. El picker vive en el documento LOCAL del iframe; los puentes a Python
# (_rc_pick / _rc_act, text_input ocultos) viven en el documento PADRE, fuera del
# @st.dialog (dentro del fragment el commit no llega a Python). Salir revisa por
# JS si el formulario (otro iframe) tiene datos sin guardar ANTES de avisar.
_RC_PICKER_IFRAME_JS = """
<script>
(function(){
  window.RCGRID=1;                      // desactiva el auto-fit: la grilla es 100vh con scroll interno
  var P=window.parent, PD=P&&P.document, D=document;
  if(P) P._rcClosing=false;             // drawer recién abierto → resetear el guard de cierre
  function root(){ return D.querySelector(".rcpk"); }
  function bridge(sel,v){
    if(!PD) return;
    var inp=PD.querySelector(sel+" input"); if(!inp) return;
    try{
      var setter=Object.getOwnPropertyDescriptor(P.HTMLInputElement.prototype,"value").set;
      inp.focus({preventScroll:true});
      setter.call(inp,v);
      inp.dispatchEvent(new Event("input",{bubbles:true}));
      inp.dispatchEvent(new Event("change",{bubbles:true}));
      inp.dispatchEvent(new KeyboardEvent("keypress",{key:"Enter",keyCode:13,which:13,charCode:13,bubbles:true}));
      inp.dispatchEvent(new KeyboardEvent("keydown",{key:"Enter",keyCode:13,which:13,bubbles:true}));
      inp.dispatchEvent(new KeyboardEvent("keyup",{key:"Enter",keyCode:13,which:13,bubbles:true}));
      inp.dispatchEvent(new FocusEvent("blur",{bubbles:false}));
      inp.dispatchEvent(new FocusEvent("focusout",{bubbles:true}));
      inp.blur();
    }catch(e){}
  }
  function pickEp(){
    var rt=root(); if(!rt) return null;
    if(rt._ep) return rt._ep;
    var s=rt.querySelector('.rcpk-opt[data-sel="1"]');
    return s?s.getAttribute("data-ep"):null;
  }
  function selectOpt(ep){
    var rt=root(); if(!rt) return;
    var opts=rt.querySelectorAll(".rcpk-opt");
    for(var i=0;i<opts.length;i++){
      var on=(opts[i].getAttribute("data-ep")===ep);
      opts[i].setAttribute("data-sel", on?"1":"0");
      if(on){
        var t1=rt.querySelector(".rcpk-t1"), t2=rt.querySelector(".rcpk-t2");
        if(t1) t1.textContent=opts[i].getAttribute("data-t1")||ep;
        if(t2) t2.textContent=opts[i].getAttribute("data-t2")||"";
      }
    }
    rt._ep=ep;
    var lb=rt.querySelector(".rcpk-load"); if(lb) lb.setAttribute("data-dis","0");
  }
  // ¿el formulario (otro iframe del padre) tiene datos sin guardar?
  function formHasData(){
    if(!PD) return false;
    var ifs=PD.querySelectorAll("iframe");
    for(var i=0;i<ifs.length;i++){
      try{ if(ifs[i].contentWindow && typeof ifs[i].contentWindow.rcHasData==="function")
             return !!ifs[i].contentWindow.rcHasData(); }catch(e){}
    }
    return false;
  }
  function confirmExit(onYes){
    if(!PD){ onYes(); return; }
    var old=PD.getElementById("rc-exit-confirm"); if(old) old.remove();
    var ov=PD.createElement("div"); ov.id="rc-exit-confirm";
    ov.style.cssText="position:fixed;inset:0;z-index:2147483600;background:rgba(15,23,42,.5);display:flex;align-items:center;justify-content:center;padding:20px;font-family:Montserrat,'Segoe UI',sans-serif;";
    ov.innerHTML='<div style="background:#fff;border-radius:16px;max-width:440px;width:100%;padding:22px 24px;box-shadow:0 24px 60px rgba(0,0,0,.3);">'
      +'<div style="font-weight:800;font-size:1rem;color:#0f172a;margin-bottom:8px;">Cambios sin guardar</div>'
      +'<div style="font-size:.86rem;color:#475569;line-height:1.5;margin-bottom:18px;">Escribiste precios o cantidades en la tabla que <b>no has guardado</b>. Si sales, se perderan. Deseas salir del proyecto de todas formas?</div>'
      +'<div style="display:flex;gap:10px;"><button id="rc-exit-no" style="flex:1;padding:10px;border:1px solid #e2e8f0;border-radius:10px;background:#f1f5f9;color:#334155;font-weight:700;cursor:pointer;">Cancelar</button>'
      +'<button id="rc-exit-yes" style="flex:1;padding:10px;border:none;border-radius:10px;background:#dc2626;color:#fff;font-weight:700;cursor:pointer;">Si, salir</button></div></div>';
    PD.body.appendChild(ov);
    ov.querySelector("#rc-exit-no").onclick=function(){ ov.remove(); };
    ov.querySelector("#rc-exit-yes").onclick=function(){ ov.remove(); onYes(); };
  }
  // Cierra el drawer con el MISMO efecto a la inversa: desliza el dialog hacia la
  // derecha y, al terminar, dispara el puente (que provoca el rerun que lo quita).
  // Sin esto el rerun elimina el dialog de golpe (sin animación de salida).
  function closeWithAnim(cb){
    if(P&&P._rcClosing) return;   // ya hay un cierre en curso (evita doble disparo)
    if(P) P._rcClosing=true;
    var done=false;
    function fire(){ if(done)return; done=true; try{cb();}catch(e){} }
    var dlg=PD&&PD.querySelector('div[data-testid="stDialog"] div[role="dialog"]');
    if(!dlg){ fire(); return; }
    try{
      if(dlg.getAnimations) dlg.getAnimations().forEach(function(a){ try{a.cancel();}catch(e){} });
      var a=dlg.animate([{transform:"translateX(0)"},{transform:"translateX(100%)"}],
                        {duration:280, easing:"cubic-bezier(.5,0,.75,0)", fill:"forwards"});
      a.onfinish=fire; a.oncancel=fire;
      setTimeout(fire, 340);   // respaldo por si onfinish no dispara
    }catch(e){ fire(); }
  }
  D.addEventListener("click",function(ev){
    var t=ev.target; if(!t||!t.closest) return;
    if(t.closest(".rcpk-trg")){
      var r=root();
      if(r){
        r.classList.toggle("open");
        if(r.classList.contains("open")){
          var sb=r.querySelector(".rcpk-search");
          if(sb){ sb.value=""; sb.dispatchEvent(new Event("input",{bubbles:true})); }
          var sel=r.querySelector('.rcpk-opt[data-sel="1"]'), list=r.querySelector(".rcpk-list");
          if(sel&&list){ var off=sel.offsetTop-list.clientHeight/2+sel.offsetHeight/2; list.scrollTop=Math.max(0,off); }
          if(sb){ try{ sb.focus({preventScroll:true}); }catch(e){} }
        }
      }
      ev.preventDefault(); return;
    }
    var op=t.closest(".rcpk-opt");
    if(op){ selectOpt(op.getAttribute("data-ep")); var r2=root(); if(r2) r2.classList.remove("open"); return; }
    if(t.closest(".rcpk-load")){ var ep=pickEp(); if(ep) bridge(".st-key-_rc_pick",ep+"|"+Date.now()); return; }  /* Cargar NO cierra: el drawer queda abierto y carga la info */
    if(t.closest(".rcpk-exit")){
      var doExit=function(){ closeWithAnim(function(){ bridge(".st-key-_rc_act","exit|"+Date.now()); }); };
      if(formHasData()){ confirmExit(doExit); } else { doExit(); }
      return;
    }
    if(t.closest(".rcpk-close")){ closeWithAnim(function(){ bridge(".st-key-_rc_act","close|"+Date.now()); }); return; }
    if(!t.closest(".rcpk")){ var r3=root(); if(r3) r3.classList.remove("open"); }
  }, true);
  D.addEventListener("input",function(ev){
    var s=ev.target; if(!s||!s.classList||!s.classList.contains("rcpk-search")) return;
    var v=(s.value||"").toLowerCase().trim(), rt=root(); if(!rt) return;
    var opts=rt.querySelectorAll(".rcpk-opt"), n=0;
    for(var i=0;i<opts.length;i++){
      var ok=(!v)||((opts[i].getAttribute("data-s")||"").indexOf(v)>=0);
      opts[i].style.display=ok?"":"none"; if(ok) n++;
    }
    var em=rt.querySelector(".rcpk-empty"); if(em) em.style.display=n?"none":"block";
  }, true);
  // ── Interceptar los cierres NATIVOS de Streamlit (X, clic fuera del drawer,
  // tecla Escape) para animar la salida también en esos casos. Se bindea en el
  // documento PADRE (los botones y el backdrop viven ahí, no en este iframe) en
  // fase de captura + stopImmediatePropagation para ganarle al handler de React.
  // Refs en window.parent → se remueven y re-agregan en cada apertura (el iframe
  // se recrea; sin esto se acumularían listeners).
  function triggerClose(){ closeWithAnim(function(){ bridge(".st-key-_rc_act","close|"+Date.now()); }); }
  function isCloseTarget(t){
    if(!t||!t.closest) return false;
    if(t.closest('button[aria-label="Close"]')) return true;                                        // la X
    if(t.closest('div[data-testid="stDialog"]') && !t.closest('div[role="dialog"]')) return true;   // backdrop (fuera del drawer)
    return false;
  }
  if(PD){
    // El backdrop lo cierra Streamlit en pointerdown/mousedown (ANTES del click),
    // la X en click → interceptar los tres en fase de captura para ganarle a
    // React, animar la salida y recién cerrar. El guard _rcClosing evita que los
    // 3 eventos de un mismo gesto disparen 3 cierres.
    ["pointerdown","mousedown","click"].forEach(function(evt){
      var key="_rcCl_"+evt;
      if(P[key]){ try{ PD.removeEventListener(evt,P[key],true); }catch(e){} }
      P[key]=function(ev){
        if(!PD.querySelector('div[data-testid="stDialog"] div[role="dialog"]')) return;
        if(!isCloseTarget(ev.target)) return;
        ev.preventDefault(); ev.stopImmediatePropagation(); triggerClose();
      };
      PD.addEventListener(evt,P[key],true);
    });
    if(P._rcEscCap){ try{ PD.removeEventListener("keydown",P._rcEscCap,true); }catch(e){} }
    P._rcEscCap=function(ev){
      if(ev.key!=="Escape" && ev.keyCode!==27) return;
      if(!PD.querySelector('div[data-testid="stDialog"] div[role="dialog"]')) return;
      ev.preventDefault(); ev.stopImmediatePropagation(); triggerClose();
    };
    PD.addEventListener("keydown",P._rcEscCap,true);
  }
})();
</script>
"""


def build_historial_rc_html(regs, ep='', supa_url='', supa_key='', picker_html=''):
    """Historial de compras interactivo (iframe): tarjetas con la tabla en HTML
    que entra en modo edición IN-PLACE (las celdas se vuelven inputs, sin crear
    otra tabla). Editar/Eliminar del REGISTRO van SERVER-SIDE vía query param +
    popstate: ?rc_edit=<json> / ?rc_delete=<id>; Python aplica con la service key
    (RLS-safe) y recalcula los totales. La ÚNICA operación con la anon key es
    reemplazar el archivo de factura (subida al bucket de storage 'facturas',
    igual que el formulario de nueva compra); la URL resultante va en el payload
    server-side y Python valida que pertenezca a ese bucket."""
    import json as _json
    regs_json = _json.dumps(regs or [], ensure_ascii=False).replace('<', '\\u003c')
    IC_STORE = _svg_rc('store', color='#475569', size=17, mr=0)
    IC_CAL   = _svg_rc('calendar', color='#94a3b8', size=12, mr=5)
    IC_USER  = _svg_rc('user', color='#94a3b8', size=12, mr=5)
    IC_CART  = _svg_rc('cart', color='#94a3b8', size=12, mr=5)
    IC_FILE  = _svg_rc('file', color='#1d4ed8', size=15, mr=0)
    IC_EDIT  = _svg_rc('edit', color='#1d4ed8', size=14, mr=6)
    IC_TRASH = _svg_rc('trash', color='#dc2626', size=14, mr=6)
    IC_SAVE  = _svg_rc('save', color='#ffffff', size=14, mr=7)
    IC_X     = _svg_rc('x', color='#475569', size=13, mr=6)
    IC_CHEV  = _svg_rc('chevron', color='#94a3b8', size=16, mr=0)
    IC_UP    = _svg_rc('trend-up', color='#dc2626', size=13, mr=5)
    IC_DOWN  = _svg_rc('trend-down', color='#16a34a', size=13, mr=5)
    IC_ALERT = _svg_rc('alert', color='#dc2626', size=14, mr=7)
    IC_CLIP  = _svg_rc('paperclip', color='#1d4ed8', size=14, mr=7)
    IC_PKG   = _svg_rc('package', color='#16a34a', size=16, mr=0)
    IC_INV   = _svg_rc('package', color='#16a34a', size=13, mr=5)

    _hist_html = f"""<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;800&display=swap');
*{{box-sizing:border-box}}
html,body{{margin:0;padding:0;font-family:Montserrat,'Segoe UI',sans-serif;background:transparent}}
.hc-wrap{{display:flex;flex-direction:column;gap:9px;}}
.hc{{border:1px solid #e2e8f0;border-radius:12px;overflow:hidden;background:#fff;box-shadow:0 1px 3px rgba(15,23,42,0.06);transition:box-shadow .18s,border-color .18s;}}
.hc.editing{{border-color:#93c5fd;box-shadow:0 4px 16px rgba(37,99,235,0.16);}}
/* Card expandida (en revisión): "flotando" con borde naranjo 2px + sombra. */
.hc[open]:not(.editing){{border:2px solid #f97316;box-shadow:0 12px 30px rgba(249,115,22,0.26);position:relative;z-index:2;}}
.hc>summary{{list-style:none;cursor:pointer;display:flex;align-items:center;padding:12px 15px;background:#f8fafc;}}
.hc.editing>summary{{background:#eff6ff;}}
.hc>summary::-webkit-details-marker{{display:none}}
.hc[open]>summary{{border-bottom:1px solid #e2e8f0;}}
.hc-lugar{{font-weight:700;font-size:0.84rem;color:#0f172a;margin-left:9px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:45%;}}
.hc-badge{{margin-left:auto;display:inline-flex;align-items:center;font-weight:700;font-size:0.75rem;padding:4px 12px;border-radius:99px;white-space:nowrap;}}
.hc-chev{{margin-left:12px;transition:transform .2s;flex-shrink:0;}}
.hc[open] .hc-chev{{transform:rotate(180deg);}}
.hc-body{{padding:13px 15px;}}
.hc:not([open]) .hc-body{{display:none;}}
.hc-meta{{display:flex;flex-wrap:wrap;gap:16px;font-size:0.72rem;color:#64748b;margin-bottom:11px;}}
.hc-meta span{{display:inline-flex;align-items:center;}}
.hc-tblwrap{{overflow-x:auto;border-radius:9px;border:1px solid #eef2f7;}}
.hc-tbl{{width:100%;border-collapse:collapse;font-size:0.79rem;min-width:520px;}}
.hc-tbl th{{background:#1e2447;color:#fff;text-align:left;padding:7px 11px;font-size:0.62rem;letter-spacing:.05em;text-transform:uppercase;font-weight:700;white-space:nowrap;}}
.hc-tbl th.r,.hc-tbl td.r{{text-align:right;}}
.hc-tbl th.c,.hc-tbl td.c{{text-align:center;}}
.hc-tbl td{{padding:7px 11px;border-bottom:1px solid #eef2f7;color:#475569;white-space:nowrap;}}
.hc-tbl tbody tr:last-child td{{border-bottom:none;}}
.hc-tbl tbody tr:nth-child(even){{background:#f8fafc;}}
.hc-tbl .it{{font-weight:600;color:#0f172a;white-space:normal;}}
.hc-inp{{width:92px;border:1.5px solid #cbd5e1;border-radius:6px;padding:5px 7px;font-size:0.78rem;text-align:right;font-family:inherit;outline:none;background:#fff;}}
.hc-inp:focus{{border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,.18);}}
.hc-rm{{width:17px;height:17px;cursor:pointer;accent-color:#dc2626;}}
tr.rm-on td{{background:#fef2f2 !important;text-decoration:line-through;color:#b91c1c;opacity:.65;}}
.hc-tots{{display:flex;flex-wrap:wrap;gap:18px;margin-top:12px;font-size:0.75rem;color:#64748b;}}
.hc-tots b{{color:#0f172a;font-weight:700;margin-left:4px;}}
.hc-obs{{display:flex;align-items:flex-start;margin-top:10px;font-size:0.74rem;color:#64748b;background:#f8fafc;border-radius:8px;padding:8px 11px;}}
.hc-fac{{display:inline-flex;align-items:center;gap:2px;margin-top:11px;padding:7px 14px;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;border-radius:8px;text-decoration:none;font-size:0.78rem;font-weight:600;}}
.hc-nofac{{display:inline-flex;align-items:center;margin-top:11px;font-size:0.74rem;color:#94a3b8;}}
.hc-editfields{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:9px;margin-bottom:12px;}}
.hc-fld label{{display:block;font-size:0.62rem;font-weight:700;letter-spacing:.04em;text-transform:uppercase;color:#64748b;margin-bottom:4px;}}
.hc-fld input{{width:100%;border:1.5px solid #cbd5e1;border-radius:7px;padding:7px 10px;font-size:0.8rem;font-family:inherit;outline:none;}}
.hc-fld input:focus{{border-color:#2563eb;box-shadow:0 0 0 2px rgba(37,99,235,.18);}}
.hc-hint{{font-size:0.72rem;color:#475569;margin-top:10px;background:#eff6ff;border-left:3px solid #93c5fd;border-radius:0 6px 6px 0;padding:8px 11px;}}
.hc-actions{{display:flex;flex-wrap:wrap;gap:8px;margin-top:12px;}}
.hc-btn{{display:inline-flex;align-items:center;justify-content:center;border:none;border-radius:8px;padding:8px 16px;font-size:0.78rem;font-weight:700;font-family:inherit;cursor:pointer;transition:filter .12s;}}
.hc-btn:hover{{filter:brightness(0.96);}}
.hc-edit{{background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;}}
.hc-del{{background:#fef2f2;color:#dc2626;border:1px solid #fecaca;}}
.hc-save{{background:#2563eb;color:#fff;}}
.hc-cancel{{background:#f1f5f9;color:#475569;border:1px solid #e2e8f0;}}
.hc-confirm{{margin-top:11px;background:#fef2f2;border:1px solid #fecaca;border-left:4px solid #dc2626;border-radius:8px;padding:10px 13px;font-size:0.78rem;color:#991b1b;display:flex;align-items:center;flex-wrap:wrap;gap:10px;}}
.hc-confirm.step2{{background:#fee2e2;border:2px solid #dc2626;box-shadow:0 0 0 3px rgba(220,38,38,0.12);}}
.hc-facedit{{display:flex;flex-wrap:wrap;align-items:center;gap:10px;margin-top:11px;}}
.hc-facbtn{{display:inline-flex;align-items:center;background:#eff6ff;color:#1d4ed8;border:1px dashed #93c5fd;border-radius:8px;padding:7px 14px;font-size:0.78rem;font-weight:600;cursor:pointer;}}
.hc-facbtn:hover{{background:#dbeafe;}}
.hc-tipo{{display:inline-flex;align-items:center;margin-left:10px;font-weight:700;font-size:0.64rem;letter-spacing:.02em;text-transform:uppercase;padding:3px 10px;border-radius:99px;white-space:nowrap;}}
.hc-tipodot{{width:7px;height:7px;border-radius:50%;display:inline-block;margin-right:6px;flex-shrink:0;}}
.hc-inp-c{{width:74px;}}
.hc-filters{{display:flex;flex-direction:column;gap:8px;margin-bottom:12px;}}
.hc-frow{{display:flex;flex-wrap:wrap;align-items:center;gap:7px;}}
.hc-flabel{{font-size:0.68rem;font-weight:700;text-transform:uppercase;letter-spacing:.04em;color:#94a3b8;margin-right:2px;min-width:150px;}}
.hc-fbadge{{display:inline-flex;align-items:center;font-size:0.66rem;font-weight:700;text-transform:uppercase;letter-spacing:.02em;padding:5px 12px;border-radius:99px;border:1.5px solid #e2e8f0;background:#fff;color:#64748b;cursor:pointer;transition:all .12s;white-space:nowrap;}}
.hc-fbadge:hover{{border-color:#cbd5e1;background:#f8fafc;}}
.hc-fbadge.active{{border-color:#1e2447;background:#1e2447;color:#fff;}}
/* ── Cards por proveedor (drill-down) — dentro del mismo iframe para que los
   filtros del historial también las re-agreguen ── */
.pv-title{{display:flex;align-items:center;font-family:Montserrat,sans-serif;font-weight:700;font-size:0.88rem;letter-spacing:0.05em;text-transform:uppercase;color:#0f172a;margin:20px 0 12px;}}
.pv-cards{{display:flex;flex-direction:column;gap:6px;}}
.pv-row{{display:flex;gap:6px;align-items:stretch;}}
.pv-card{{border-radius:9px;padding:10px 13px;min-width:128px;box-sizing:border-box;display:flex;flex-direction:column;background:#fff;box-shadow:0 1px 3px rgba(15,23,42,0.05);cursor:pointer;position:relative;transition:box-shadow .16s,transform .16s,opacity .16s,filter .16s;}}
.pv-card:hover{{box-shadow:0 3px 10px rgba(15,23,42,0.13);}}
.pv-card.sel{{transform:scale(1.045);box-shadow:0 10px 26px rgba(37,99,235,0.30);z-index:3;}}
.pv-cards.has-sel .pv-card:not(.sel){{opacity:.38;filter:saturate(.7);}}
.pv-cards.has-sel .pv-card:not(.sel):hover{{opacity:.7;}}
.pv-name{{font-family:Montserrat,sans-serif;font-size:0.66rem;font-weight:700;text-transform:uppercase;letter-spacing:.03em;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:3px;}}
.pv-total{{font-family:Montserrat,sans-serif;font-size:1rem;font-weight:800;color:#0f172a;}}
.pv-meta{{font-size:0.66rem;color:#64748b;margin-top:2px;display:flex;align-items:center;justify-content:space-between;gap:6px;}}
.pv-mchev{{transition:transform .2s;}}
.pv-card.sel .pv-mchev{{transform:rotate(180deg);}}
.pv-detail{{margin-top:10px;border:1px solid #e2e8f0;border-radius:11px;overflow:hidden;background:#fff;box-shadow:0 2px 10px rgba(15,23,42,0.07);}}
.pv-dhdr{{display:flex;align-items:center;padding:11px 14px;background:#f8fafc;border-bottom:1px solid #e2e8f0;font-family:Montserrat,sans-serif;font-weight:700;font-size:0.8rem;color:#0f172a;}}
.pv-compra{{display:flex;align-items:center;flex-wrap:wrap;gap:8px 14px;padding:11px 14px;border-bottom:1px solid #eef2f7;cursor:pointer;font-size:0.8rem;}}
.pv-compra:last-child{{border-bottom:none;}}
.pv-compra:hover{{background:#f8fafc;}}
.pv-cfecha{{display:inline-flex;align-items:center;color:#475569;white-space:nowrap;}}
.pv-cresp{{display:inline-flex;align-items:center;color:#475569;flex:1;min-width:120px;}}
.pv-cmonto{{font-weight:800;color:#0f172a;font-family:Montserrat,sans-serif;white-space:nowrap;}}
.pv-cchev{{color:#94a3b8;transition:transform .2s;flex-shrink:0;}}
.pv-compra.open .pv-cchev{{transform:rotate(180deg);}}
.pv-compra.open{{background:#eff6ff;}}
.pv-facwrap{{padding:12px 14px;background:#f8fafc;border-bottom:1px solid #eef2f7;}}
.pv-invtbl{{font-family:Montserrat,sans-serif;}}
.pv-invhead,.pv-invrow,.pv-invfoot{{display:grid;grid-template-columns:1fr 60px 110px;gap:8px;align-items:center;padding:9px 14px;}}
.pv-invhead{{font-size:0.64rem;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:#64748b;background:#f0fdf4;border-bottom:1px solid #dcfce7;}}
.pv-invhead span:nth-child(2),.pv-invrow span:nth-child(2),.pv-invfoot span:nth-child(2){{text-align:center;}}
.pv-invhead span:last-child,.pv-invrow span:last-child,.pv-invfoot span:last-child{{text-align:right;}}
.pv-invrow{{border-bottom:1px solid #f1f5f9;font-size:0.82rem;color:#0f172a;}}
.pv-invrow:last-of-type{{border-bottom:none;}}
.pv-invfoot{{background:#f0fdf4;border-top:1px solid #dcfce7;font-weight:800;font-size:0.9rem;color:#166534;}}
.pv-facbox{{display:flex;flex-direction:column;gap:9px;align-items:flex-start;}}
.pv-facimg{{max-width:100%;max-height:640px;border-radius:8px;border:1px solid #e2e8f0;}}
.pv-faclink{{display:inline-flex;align-items:center;gap:2px;padding:7px 14px;background:#eff6ff;color:#1d4ed8;border:1px solid #bfdbfe;border-radius:8px;text-decoration:none;font-size:0.78rem;font-weight:600;}}
.pv-nofac{{display:inline-flex;align-items:center;color:#94a3b8;font-size:0.78rem;}}
.pv-pdfbox{{width:100%;height:600px;overflow:auto;border:1px solid #e2e8f0;border-radius:8px;background:#525659;position:relative;}}
.pv-pdfload{{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:10px;background:#f0f2f5;color:#64748b;font-size:0.85rem;z-index:2;}}
.pv-spin{{width:34px;height:34px;border:4px solid #cbd5e1;border-top-color:#5b7cfa;border-radius:50%;animation:pvspin .8s linear infinite;}}
@keyframes pvspin{{from{{transform:rotate(0)}}to{{transform:rotate(360deg)}}}}
.pv-pdfpages{{padding:10px 0;text-align:center;}}
.pv-pdfpages canvas{{display:block;margin:0 auto 10px;max-width:97%;box-shadow:0 2px 10px rgba(0,0,0,.4);background:#fff;}}
</style>
<div class="rcg">
  <div class="rcg-cell rcg-proj">
    <div class="rcg-head">{_svg_rc('folder', color='#0f172a', size=16, mr=8)}Seleccionar proyecto</div>
    <div class="rcg-body">{picker_html}</div>
  </div>
  <div class="rcg-cell">
    <div class="rcg-head">{_svg_rc('history', color='#0f172a', size=16, mr=8)}Historial de compras</div>
    <div class="rcg-body"><div class="hc-wrap" id="hc-wrap"></div></div>
  </div>
  <div class="rcg-cell">
    <div class="rcg-head">{_svg_rc('file', color='#0f172a', size=16, mr=8)}Informaci&oacute;n de facturas</div>
    <div class="rcg-body"><div class="hc-filters" id="hc-filters"></div></div>
  </div>
  <div class="rcg-cell">
    <div class="rcg-head">{_svg_rc('store', color='#0f172a', size=16, mr=8)}Compras por proveedor</div>
    <div class="rcg-body"><div id="pv-section"></div></div>
  </div>
</div>
<script>
var REGS={regs_json};
var IC={{store:'{IC_STORE}',cal:'{IC_CAL}',user:'{IC_USER}',cart:'{IC_CART}',file:'{IC_FILE}',edit:'{IC_EDIT}',trash:'{IC_TRASH}',save:'{IC_SAVE}',x:'{IC_X}',chev:'{IC_CHEV}',up:'{IC_UP}',down:'{IC_DOWN}',alert:'{IC_ALERT}',clip:'{IC_CLIP}',pkg:'{IC_PKG}',inv:'{IC_INV}'}};
var EP="{ep}";var SUPA_URL="{supa_url}";var SUPA_KEY="{supa_key}";
var editing=-1, confirming=-1, confirmStep=0, _facFile=null, fTipo="", fResp="", fProv="";
var pvSel=-1, pvFacOpen={{}}, PROVS_F=[];
var PVCOLORS=["#3b82f6","#10b981","#f59e0b","#8b5cf6","#ef4444","#06b6d4","#f97316","#84cc16","#ec4899","#6366f1","#14b8a6","#eab308","#dc2626","#7c3aed","#0ea5e9"];
function f(n){{return "$"+Math.round(Math.abs(+n||0)).toLocaleString("de-DE");}}
function esc(s){{return String(s==null?"":s).replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;").replace(/"/g,"&quot;");}}
function nav(param,val){{
  var url=new URL(window.parent.location.href);
  url.searchParams.set(param,val);
  window.parent.history.replaceState({{}},"",url);
  window.parent.dispatchEvent(new PopStateEvent("popstate"));
}}
window.hcEdit=function(i){{_facFile=null;editing=(editing===i?-1:i);confirming=-1;render();}};
window.hcCancel=function(){{_facFile=null;editing=-1;render();}};
window.hcFilter=function(dim,val){{if(dim==="tipo")fTipo=val;else if(dim==="resp")fResp=val;else if(dim==="prov")fProv=val;editing=-1;confirming=-1;_facFile=null;pvSel=-1;pvFacOpen={{}};render();}};
window.pvSelFn=function(i){{pvSel=(pvSel===i?-1:i);pvFacOpen={{}};render();}};
window.pvFacFn=function(i,j){{var k=i+"-"+j;var was=!!pvFacOpen[k];pvFacOpen={{}};if(!was)pvFacOpen[k]=true;render();}};
function pvHexa(h,a){{h=h.replace("#","");return "rgba("+parseInt(h.substr(0,2),16)+","+parseInt(h.substr(2,2),16)+","+parseInt(h.substr(4,2),16)+","+a+")";}}
function pvFacPrev(c){{
  if(!c.factura_url)return '<div class="pv-nofac">'+IC.file+'Sin factura adjunta</div>';
  var link='<a class="pv-faclink" href="'+esc(c.factura_url)+'" target="_blank" rel="noopener noreferrer">'+IC.file+'<span>Abrir factura en pestaña nueva ↗</span></a>';
  var prev=c.is_img?('<img class="pv-facimg" src="'+esc(c.factura_url)+'" loading="lazy"/>'):('<div class="pv-pdfbox" data-pdfurl="'+esc(c.factura_url)+'"><div class="pv-pdfload"><div class="pv-spin"></div><span>Cargando PDF…</span></div><div class="pv-pdfpages"></div></div>');
  return '<div class="pv-facbox">'+prev+link+'</div>';
}}
window.renderPdfs=function(){{
  if(typeof pdfjsLib==="undefined")return;
  pdfjsLib.GlobalWorkerOptions.workerSrc="https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.worker.min.js";
  document.querySelectorAll(".pv-pdfbox[data-pdfurl]").forEach(function(box){{
    if(box.getAttribute("data-done"))return;
    box.setAttribute("data-done","1");
    var pages=box.querySelector(".pv-pdfpages"),load=box.querySelector(".pv-pdfload");
    pdfjsLib.getDocument({{url:box.getAttribute("data-pdfurl"),withCredentials:false}}).promise.then(function(pdf){{
      var seq=Promise.resolve(),first=true;
      for(var i=1;i<=pdf.numPages;i++){{(function(num){{
        seq=seq.then(function(){{return pdf.getPage(num).then(function(page){{
          var vp=page.getViewport({{scale:1.5}});var cv=document.createElement("canvas");var ctx=cv.getContext("2d");
          cv.width=vp.width;cv.height=vp.height;pages.appendChild(cv);
          return page.render({{canvasContext:ctx,viewport:vp}}).promise.then(function(){{if(first){{first=false;if(load)load.style.display="none";fit();}}}});
        }});}});
      }})(i);}}
    }}).catch(function(){{if(load)load.innerHTML='<span style="color:#dc2626;font-size:0.8rem;padding:0 16px;text-align:center;">No se pudo mostrar el PDF. Ábrelo en pestaña nueva.</span>';}});
  }});
}};
function renderProv(matched){{
  var sec=document.getElementById("pv-section");
  if(!sec)return;
  if(!matched.length){{sec.innerHTML="";return;}}
  var agg={{}};
  matched.forEach(function(r){{
    var isInv=(r.inv===true)||((r.lugar||"").toUpperCase()==="INVENTARIO");
    var k=r.lugar||"—";
    if(!agg[k])agg[k]={{total:0,n:0,compras:[],inv:isInv,items:{{}}}};
    agg[k].n++;
    if(isInv){{
      // Card INVENTARIO: total = ahorro (presupuestado de los ítems en stock),
      // y agregamos los productos para el drill-down (en vez de facturas).
      (r.items||[]).forEach(function(it){{
        var isStk=(it.stock===true)||((+it.pr===0)&&!it.sin);
        if(!isStk)return;
        var C=(+it.cant||1);
        var sq=(it.sqty!=null&&+it.sqty>0)?+it.sqty:C; if(sq>C)sq=C;  // unidades en stock
        var ah=(+it.pp||0)*sq;
        agg[k].total+=ah;
        var ik=it.item||"—";
        if(!agg[k].items[ik])agg[k].items[ik]={{item:ik,cat:it.cat||"",cant:0,ahorro:0}};
        agg[k].items[ik].cant+=sq;
        agg[k].items[ik].ahorro+=ah;
      }});
    }} else {{
      agg[k].total+=(+r.tr||0);
    }}
    agg[k].compras.push({{fecha:r.fecha,monto:f(r.tr),responsable:r.usuario,factura_url:r.factura_url,factura_nom:r.factura_nom,is_img:r.is_img}});
  }});
  PROVS_F=Object.keys(agg).map(function(k){{return {{name:k,total:agg[k].total,n:agg[k].n,compras:agg[k].compras,inv:agg[k].inv,items:agg[k].items}};}});
  PROVS_F.sort(function(a,b){{return b.total-a.total;}});
  PROVS_F.forEach(function(p,i){{p.color=p.inv?"#16a34a":PVCOLORS[i%PVCOLORS.length];p.total_fmt=f(p.total);p.nprod=Object.keys(p.items||{{}}).length;}});
  var n=PROVS_F.length, nrows=n<=4?1:(n<=10?2:3), per=Math.ceil(n/nrows);
  var html='<div class="pv-cards'+(pvSel>=0?" has-sel":"")+'">';
  for(var r=0;r<nrows;r++){{
    var start=r*per, end=Math.min(start+per,n); if(start>=end)break;
    var rmax=1; for(var q=start;q<end;q++){{var w=Math.pow(PROVS_F[q].total||1,0.3);if(w>rmax)rmax=w;}}
    html+='<div class="pv-row">';
    for(var q=start;q<end;q++){{
      var c=PROVS_F[q];var grow=Math.max(1,Math.round(Math.pow(c.total||1,0.3)/rmax*1000));var selc=(pvSel===q)?" sel":"";
      var cstyle=c.inv
        ?'border:1.5px solid #86efac;border-left:4px solid #16a34a;background:#f0fdf4;flex:'+grow+' '+grow+' 0;'
        :'border:1.5px solid '+pvHexa(c.color,0.3)+';border-left:4px solid '+c.color+';flex:'+grow+' '+grow+' 0;';
      var meta=c.inv?(c.nprod+(c.nprod===1?" producto":" productos")):(c.n+(c.n===1?" compra":" compras"));
      html+='<div class="pv-card'+selc+'" onclick="pvSelFn('+q+')" style="'+cstyle+'">'
        +'<div class="pv-name" style="color:'+c.color+';">'+(c.inv?IC.inv:"")+esc(c.name)+'</div>'
        +'<div class="pv-total" style="'+(c.inv?"color:#16a34a;":"")+'">'+esc(c.total_fmt)+(c.inv?' <span style="font-size:.58rem;font-weight:700;color:#16a34a;text-transform:uppercase;">ahorro</span>':'')+'</div>'
        +'<div class="pv-meta"><span>'+meta+'</span><span class="pv-mchev">'+IC.chev+'</span></div></div>';
    }}
    html+='</div>';
  }}
  html+='</div>';
  if(pvSel>=0&&pvSel<n){{
    var p=PROVS_F[pvSel];
    if(p.inv){{
      // INVENTARIO seleccionado → TABLA de productos ahorrados (no facturas).
      var prods=Object.keys(p.items).map(function(k){{return p.items[k];}});
      prods.sort(function(a,b){{return b.ahorro-a.ahorro;}});
      html+='<div class="pv-detail" style="border-color:#bbf7d0;"><div class="pv-dhdr" style="background:#f0fdf4;border-bottom-color:#dcfce7;color:#166534;">'+IC.pkg+'&nbsp;Inventario &nbsp;·&nbsp; '+esc(p.total_fmt)+' ahorrado &nbsp;·&nbsp; '+prods.length+(prods.length===1?" producto":" productos")+'</div>';
      html+='<div class="pv-invtbl"><div class="pv-invhead"><span>Producto</span><span>Cant.</span><span>Ahorro</span></div>';
      prods.forEach(function(it){{
        html+='<div class="pv-invrow"><span><b>'+esc(it.item)+'</b><br><small style="color:#64748b;text-transform:uppercase;letter-spacing:.03em;">'+esc(it.cat)+'</small></span><span>'+it.cant+'</span><span style="color:#16a34a;font-weight:800;">'+f(it.ahorro)+'</span></div>';
      }});
      html+='<div class="pv-invfoot"><span>Total ahorrado</span><span></span><span>'+f(p.total)+'</span></div></div></div>';
    }} else {{
      html+='<div class="pv-detail"><div class="pv-dhdr">'+IC.store+'Compras en '+esc(p.name)+' &nbsp;·&nbsp; '+esc(p.total_fmt)+' &nbsp;·&nbsp; '+p.n+(p.n===1?" compra":" compras")+'</div>';
      (p.compras||[]).forEach(function(c,j){{
        var open=!!pvFacOpen[pvSel+"-"+j];
        html+='<div class="pv-compra'+(open?" open":"")+'" onclick="pvFacFn('+pvSel+','+j+')"><span class="pv-cfecha">'+IC.cal+esc(c.fecha)+'</span><span class="pv-cresp">'+IC.user+esc(c.responsable)+'</span><span class="pv-cmonto">'+esc(c.monto)+'</span><span class="pv-cchev">'+IC.chev+'</span></div>';
        if(open)html+='<div class="pv-facwrap">'+pvFacPrev(c)+'</div>';
      }});
      html+='</div>';
    }}
  }}
  sec.innerHTML=html;
  renderPdfs();
}}
window.hcAskDel=function(i){{confirming=i;confirmStep=1;editing=-1;render();}};
window.hcAskDel2=function(){{confirmStep=2;render();}};
window.hcNoDel=function(){{confirming=-1;confirmStep=0;render();}};
window.hcDoDel=function(i){{nav("rc_delete",REGS[i].id);}};
window.hcToggleRm=function(i,j){{
  var tr=document.getElementById("row-"+i+"-"+j);
  var cb=document.getElementById("rm-"+i+"-"+j);
  if(tr)tr.classList.toggle("rm-on",cb.checked);
}};
window.hcFmt=function(el){{var raw=el.value.replace(/[^0-9]/g,"");el.value=raw?("$"+parseInt(raw).toLocaleString("de-DE")):"";}};
window.hcPickFac=function(i,inp){{_facFile=inp.files[0]||null;var l=document.getElementById("facname-"+i);if(l)l.textContent=_facFile?_facFile.name:"Reemplazar factura…";}};
window.hcSave=async function(i){{
  var r=REGS[i];var items=[];
  r.items.forEach(function(it,j){{
    var p=document.getElementById("p-"+i+"-"+j);
    var c=document.getElementById("c-"+i+"-"+j);
    var rm=document.getElementById("rm-"+i+"-"+j);
    var praw=p?parseInt((p.value+"").replace(/[^0-9]/g,"")):0;
    var cval=c?(parseInt(c.value)||0):(Math.round(it.cant)||0);
    items.push({{i:j,c:cval,p:praw||0,rm:rm?rm.checked:false}});
  }});
  var g=function(id){{var e=document.getElementById(id);return e?e.value:"";}};
  var payload={{id:r.id,lugar:g("lugar-"+i),obs:g("obs-"+i),fent:g("fent-"+i),items:items}};
  var btn=document.getElementById("savebtn-"+i);
  if(_facFile){{
    if(btn){{btn.disabled=true;btn.style.opacity="0.6";btn.textContent="Subiendo factura...";}}
    try{{
      var ext=(_facFile.name.split(".").pop()||"pdf");
      var path="cotizacion-"+encodeURIComponent(EP)+"/"+Date.now()+"."+ext;
      var up=await fetch(SUPA_URL+"/storage/v1/object/facturas/"+path,{{method:"POST",headers:{{"Authorization":"Bearer "+SUPA_KEY,"apikey":SUPA_KEY,"Content-Type":_facFile.type||"application/octet-stream","x-upsert":"true"}},body:_facFile}});
      if(!up.ok) throw new Error("HTTP "+up.status);
      payload.factura_url=SUPA_URL+"/storage/v1/object/public/facturas/"+path;
      payload.factura_nom=_facFile.name;
    }}catch(e){{
      if(btn){{btn.disabled=false;btn.style.opacity="1";btn.innerHTML=IC.save+"Guardar cambios";}}
      alert("No se pudo subir la factura: "+e.message);return;
    }}
  }}
  nav("rc_edit",JSON.stringify(payload));
}};
function badge(r){{
  var ah=(+r.balance||0)>=0;
  var col=ah?"#16a34a":"#dc2626";var bg=ah?"#f0fdf4":"#fef2f2";
  var lbl=ah?"Ahorro":"Sobrecosto";var ic=ah?IC.down:IC.up;
  return '<span class="hc-badge" style="background:'+bg+';color:'+col+';">'+ic+lbl+' '+f(r.balance)+'</span>';
}}
function viewRows(r){{
  if(!r.items.length)return '<tr><td colspan="5" style="text-align:center;color:#94a3b8;">Sin ítems.</td></tr>';
  var pill='font-size:.6rem;font-weight:700;background:#dcfce7;color:#166534;padding:1px 6px;border-radius:20px;text-transform:uppercase;letter-spacing:.04em;vertical-align:middle;';
  return r.items.map(function(it){{
    var C=(+it.cant||1),SQ=(it.sqty!=null?+it.sqty:0);
    var prCell;
    if(it.stock===true&&SQ>=C){{
      prCell='<td class="r" style="font-weight:700;color:#16a34a;">$0 <span style="'+pill+'">En stock</span></td>';
    }} else if(it.stock===true&&SQ>0){{
      prCell='<td class="r" style="font-weight:700;">'+f(it.pr)+' <span style="'+pill+'">'+SQ+' en stock</span></td>';
    }} else {{
      prCell='<td class="r" style="font-weight:700;">'+f(it.pr)+'</td>';
    }}
    return '<tr><td>'+esc(it.cat)+'</td><td class="it">'+esc(it.item)+'</td>'
      +'<td class="r">'+esc(it.cant)+'</td><td class="r">'+f(it.pp)+'</td>'
      +prCell+'</tr>';
  }}).join("");
}}
function editRows(i,r){{
  return r.items.map(function(it,j){{
    var cantCell=it.sin
      ?'<td class="r"><input class="hc-inp hc-inp-c" id="c-'+i+'-'+j+'" type="number" min="0" step="1" value="'+(Math.round(it.cant)||0)+'" title="Adicional sin registro — cantidad editable"/></td>'
      :'<td class="r">'+esc(it.cant)+'</td>';
    return '<tr id="row-'+i+'-'+j+'"><td>'+esc(it.cat)+'</td><td class="it">'+esc(it.item)+'</td>'
      +cantCell
      +'<td class="r">'+f(it.pp)+'</td>'
      +'<td class="r"><input class="hc-inp" id="p-'+i+'-'+j+'" type="text" inputmode="numeric" value="'+f(it.pr)+'" oninput="hcFmt(this)"/></td>'
      +'<td class="c"><input class="hc-rm" id="rm-'+i+'-'+j+'" type="checkbox" onchange="hcToggleRm('+i+','+j+')" title="Quitar este ítem"/></td></tr>';
  }}).join("");
}}
function valOf(r,dim){{return dim==="tipo"?(r.tipo_lbl||""):dim==="resp"?(r.usuario||""):dim==="prov"?(r.lugar||""):"";}}
function filterRow(label,dim,order,tcolors,activeVal){{
  var seen={{}};
  REGS.forEach(function(r){{var v=valOf(r,dim);if(v){{seen[v]=(seen[v]||0)+1;}}}});
  var keys=Object.keys(seen);
  if(order){{keys.sort(function(a,b){{var ia=order.indexOf(a),ib=order.indexOf(b);ia=ia<0?99:ia;ib=ib<0?99:ib;return ia-ib||(a.toLowerCase()<b.toLowerCase()?-1:1);}});}}
  else{{keys.sort(function(a,b){{return a.toLowerCase()<b.toLowerCase()?-1:1;}});}}
  if(keys.length<2)return "";
  var h='<div class="hc-frow"><span class="hc-flabel">'+label+'</span>';
  h+='<span class="hc-fbadge'+(activeVal===""?" active":"")+'" data-dim="'+dim+'" data-val="">Todas ('+REGS.length+')</span>';
  keys.forEach(function(k){{
    var act=(activeVal===k);
    var col=tcolors?tcolors[k]:null;
    var sty=(act&&col)?('background:'+col.bg+';color:'+col.fg+';border-color:'+col.fg+';'):'';
    var dot=col?('<span class="hc-tipodot" style="background:'+col.fg+';"></span>'):'';
    h+='<span class="hc-fbadge'+(act&&!col?" active":"")+'" style="'+sty+'" data-dim="'+dim+'" data-val="'+esc(k)+'">'+dot+esc(k)+' ('+seen[k]+')</span>';
  }});
  h+='</div>';return h;
}}
function renderFilters(){{
  var el=document.getElementById("hc-filters");if(!el)return;
  var tcolors={{}};
  REGS.forEach(function(r){{if(r.tipo_lbl&&!tcolors[r.tipo_lbl])tcolors[r.tipo_lbl]={{bg:r.tipo_bg,fg:r.tipo_fg}};}});
  var html="";
  html+=filterRow("Filtrar por tipo:","tipo",["Normal","Adicional con registro","Adicional sin registro","Mixto"],tcolors,fTipo);
  html+=filterRow("Filtrar por responsable:","resp",null,null,fResp);
  html+=filterRow("Filtrar por proveedor:","prov",null,null,fProv);
  el.innerHTML=html;
}}
function fit(){{
  if(window.RCGRID)return;   // modo grilla: alto fijo 100vh con scroll interno
  try{{
    // Medir SOLO document.body.scrollHeight = alto real del CONTENIDO. NO usar
    // documentElement.scrollHeight ni getBoundingClientRect: devuelven el alto del
    // VIEWPORT del iframe (el <html> llena el iframe) → nunca encogen y dejan
    // huecos. El <body> no llena el viewport, así que da el alto real → auto-fit.
    var h=Math.max(document.body.scrollHeight, 60)+2;
    var fe=window.frameElement;
    if(fe){{fe.style.height=h+"px";fe.setAttribute("height",h);}}
  }}catch(e){{}}
}}
function render(){{
  renderFilters();
  var w=document.getElementById("hc-wrap");w.innerHTML="";
  var matched=[];
  REGS.forEach(function(r,i){{
    if(fTipo&&(r.tipo_lbl||"")!==fTipo)return;
    if(fResp&&(r.usuario||"")!==fResp)return;
    if(fProv&&(r.lugar||"")!==fProv)return;
    matched.push(r);
    var ed=(editing===i);
    var d=document.createElement("details");d.className="hc"+(ed?" editing":"");d.open=ed||(confirming===i);
    d.addEventListener("toggle",fit);
    var meta='<div class="hc-meta"><span>'+IC.cal+esc(r.fecha)+'</span><span>'+IC.user+esc(r.usuario)+'</span><span>'+IC.cart+esc(r.tipo)+'</span></div>';
    var fac=r.factura_url
      ?'<a class="hc-fac" href="'+esc(r.factura_url)+'" target="_blank" rel="noopener noreferrer">'+IC.file+'<span style="margin-left:2px;">Ver factura: '+esc(r.factura_nom)+'</span></a>'
      :'<div class="hc-nofac">'+IC.alert+'<span style="margin-left:4px;">Sin factura adjunta</span></div>';
    var obs=(r.obs&&!ed)?'<div class="hc-obs">'+esc(r.obs)+'</div>':'';
    var tipoB=r.tipo_lbl?('<span class="hc-tipo" style="background:'+(r.tipo_bg||"#e2e8f0")+';color:'+(r.tipo_fg||"#334155")+';"><span class="hc-tipodot" style="background:'+(r.tipo_fg||"#334155")+';"></span>'+esc(r.tipo_lbl)+'</span>'):'';
    var head='<summary>'+IC.store+'<span class="hc-lugar">'+esc(r.lugar||"Compra")+'</span>'+tipoB+badge(r)+IC.chev+'</summary>';
    var body;
    if(ed){{
      body='<div class="hc-body">'
        +'<div class="hc-editfields">'
        +'<div class="hc-fld"><label>Lugar de compra</label><input id="lugar-'+i+'" value="'+esc(r.lugar)+'"/></div>'
        +'<div class="hc-fld"><label>Fecha de entrega</label><input id="fent-'+i+'" value="'+esc(r.fent)+'"/></div>'
        +'<div class="hc-fld"><label>Observaciones</label><input id="obs-'+i+'" value="'+esc(r.obs)+'"/></div>'
        +'</div>'
        +'<div class="hc-tblwrap"><table class="hc-tbl"><thead><tr><th>Categoría</th><th>Ítem</th><th class="r">Cantidad</th><th class="r">Presupuestado</th><th class="r">Precio real</th><th class="c">Quitar</th></tr></thead><tbody>'+editRows(i,r)+'</tbody></table></div>'
        +'<div class="hc-hint">Corrige el <b>precio real</b> mal digitado, o marca <b>Quitar</b> para eliminar un ítem. El balance se recalcula al guardar.</div>'
        +'<div class="hc-facedit">'
        +(r.factura_url?'<a class="hc-fac" style="margin-top:0;" href="'+esc(r.factura_url)+'" target="_blank" rel="noopener noreferrer">'+IC.file+'<span style="margin-left:2px;">Factura actual: '+esc(r.factura_nom)+'</span></a>':'<span class="hc-nofac" style="margin-top:0;">'+IC.alert+'<span style="margin-left:4px;">Sin factura adjunta</span></span>')
        +'<label class="hc-facbtn">'+IC.clip+'<span id="facname-'+i+'">'+(r.factura_url?"Reemplazar factura…":"Adjuntar factura…")+'</span><input type="file" accept=".pdf,image/*" style="display:none" onchange="hcPickFac('+i+',this)"/></label>'
        +'</div>'
        +'<div class="hc-actions"><button id="savebtn-'+i+'" class="hc-btn hc-save" onclick="hcSave('+i+')">'+IC.save+'Guardar cambios</button><button class="hc-btn hc-cancel" onclick="hcCancel()">'+IC.x+'Cancelar</button></div>'
        +'</div>';
    }} else {{
      var conf='';
      if(confirming===i){{
        if(confirmStep===1){{
          conf='<div class="hc-confirm">'+IC.alert+'<span><b>¿Eliminar esta compra por completo?</b> Sus ítems volverán a quedar pendientes.</span><span style="margin-left:auto;display:flex;gap:8px;"><button class="hc-btn hc-del" onclick="hcAskDel2()">Sí, eliminar</button><button class="hc-btn hc-cancel" onclick="hcNoDel()">Cancelar</button></span></div>';
        }}else{{
          conf='<div class="hc-confirm step2">'+IC.alert+'<span><b>Confirmación final:</b> esta acción NO se puede deshacer. ¿Eliminar la compra definitivamente?</span><span style="margin-left:auto;display:flex;gap:8px;"><button class="hc-btn hc-del" onclick="hcDoDel('+i+')">'+IC.trash+'Sí, eliminar definitivamente</button><button class="hc-btn hc-cancel" onclick="hcNoDel()">Cancelar</button></span></div>';
        }}
      }}
      body='<div class="hc-body">'+meta
        +'<div class="hc-tblwrap"><table class="hc-tbl"><thead><tr><th>Categoría</th><th>Ítem</th><th class="r">Cantidad</th><th class="r">Presupuestado</th><th class="r">Precio real</th></tr></thead><tbody>'+viewRows(r)+'</tbody></table></div>'
        +'<div class="hc-tots"><span>Presupuestado <b>'+f(r.tp)+'</b></span><span>Real <b>'+f(r.tr)+'</b></span><span>Balance <b style="color:'+((+r.balance||0)>=0?"#16a34a":"#dc2626")+';">'+((+r.balance||0)>=0?"Ahorro":"Sobrecosto")+' '+f(r.balance)+'</b></span></div>'
        +obs+fac
        +'<div class="hc-actions"><button class="hc-btn hc-edit" onclick="hcEdit('+i+')">'+IC.edit+'Editar</button><button class="hc-btn hc-del" style="margin-left:auto;" onclick="hcAskDel('+i+')">'+IC.trash+'Eliminar</button></div>'
        +conf
        +'</div>';
    }}
    d.innerHTML=head+body;
    w.appendChild(d);
  }});
  renderProv(matched);
  fit();
}}
render();
// Delegación de clicks en los badges de filtro (una sola vez; el contenedor
// persiste aunque su innerHTML se re-renderice).
(function(){{
  var fEl=document.getElementById("hc-filters");
  if(fEl)fEl.addEventListener("click",function(e){{
    var b=e.target&&e.target.closest?e.target.closest(".hc-fbadge"):null;
    if(!b)return;
    hcFilter(b.getAttribute("data-dim")||"", b.getAttribute("data-val")||"");
  }});
}})();
window.addEventListener("load",fit);
try{{new ResizeObserver(function(){{fit();}}).observe(document.body);}}catch(e){{}}
setTimeout(fit,60);setTimeout(fit,350);
</script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.min.js" onload="window.renderPdfs&&window.renderPdfs()" onerror="document.querySelectorAll('.pv-pdfload').forEach(function(l){{l.innerHTML='<span style=\\'color:#dc2626;font-size:0.8rem;\\'>No se pudo cargar el visor PDF. Ábrelo en pestaña nueva.</span>';}})"></script>"""
    # CSS de la grilla 2×2 + dropdown ANTES (llaves reales); JS del picker DESPUÉS.
    return _RC_GRID_CSS + _hist_html + _RC_PICKER_IFRAME_JS



# ── CÁLCULO DE TOTALES ────────────────────────────────────────────────────────

def calcular_totales_rc(productos_presupuesto, registros, incluir_varios=False):
    """
    Calcula totales igual al JS calc() de la tabla RC.
    tP: precio presupuesto * cantidad para ítems normales
    tR: precio real * cantidad + adicional * precio real para TODOS los ítems
    tA: precio real * cantidad para adicionales con registro
    tS: precio real * cantidad para adicionales sin registro
    """
    import json as _jct
    _todos = list(productos_presupuesto or [])
    _pn = {str(p.get('Item', '')) for p in _todos}
    _pu_map = {str(p.get('Item', '')): round(float(p.get('Precio Unitario', 0) or 0))
               for p in _todos}

    if incluir_varios:
        prods = _todos
    else:
        prods = [p for p in _todos
                 if str(p.get('Categoria', '')).strip().lower() != 'varios']

    _comprados = {}
    for reg in (registros or []):
        items_r = reg.get('items') or []
        if isinstance(items_r, str):
            try:
                items_r = _jct.loads(items_r)
            except Exception:
                items_r = []
        for it in items_r:
            nombre = str(it.get('item', ''))
            pr = float(it.get('precio_real', 0) or 0)
            if pr > 0 and nombre:
                _comprados[nombre] = {
                    'real': pr,
                    'cant': float(it.get('cantidad', 1) or 1),
                    'adic': int(it.get('adicional', 0) or 0),
                    'es_adicional': it.get('es_adicional', False),
                    'sin_registro': it.get('sin_registro', False),
                }

    tP = 0; tR = 0; tA = 0; tS = 0

    for nombre, data in _comprados.items():
        re = data['real']
        c = data['cant']
        ad = data['adic']
        isSinReg = data['sin_registro']
        isAdic = (nombre not in _pn) and not isSinReg

        if isSinReg:
            tS += re * c
        elif isAdic:
            tA += re * c
        else:
            pu = _pu_map.get(nombre, 0)
            tR += re * c + ad * re
            if nombre in _pn:
                tP += pu * c
            continue
        tR += re * c + ad * re

    for p in prods:
        nombre = str(p.get('Item', ''))
        if nombre not in _comprados:
            pu = round(float(p.get('Precio Unitario', 0) or 0))
            c = round(float(p.get('Cantidad', 1) or 1))
            tP += pu * c

    return {'tP': tP, 'tR': tR, 'tA': tA, 'tS': tS}


# ── EXCEL BALANCE ─────────────────────────────────────────────────────────────

def generar_excel_balance(cotizacion_numero, registros, productos_presupuesto, incluir_varios=False):
    """Genera Excel con precios reales consolidados por ítem para nutrir BD."""
    import io, json
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance de Precios"

    azul_oscuro = "1E2447"
    verde = "16A34A"
    rojo = "DC2626"
    gris_claro = "F8FAFC"
    borde_gris = "E2E8F0"

    hdr_font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor=azul_oscuro)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color=borde_gris)
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:E1')
    ws['A1'] = f'Balance de Precios — {cotizacion_numero}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color=azul_oscuro)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Categoría', 'Ítem', 'Precio Unitario (Presupuestado)', 'Precio Real (Compra)', 'Diferencia']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = borde
    ws.row_dimensions[2].height = 32

    if incluir_varios:
        prods_valid = list(productos_presupuesto or [])
    else:
        prods_valid = [p for p in (productos_presupuesto or [])
                       if str(p.get('Categoria', '')).strip().lower() != 'varios']

    precios_reales = {}
    for reg in registros:
        items_r = reg.get('items') or []
        if isinstance(items_r, str):
            try:
                items_r = json.loads(items_r)
            except Exception:
                items_r = []
        for it in items_r:
            nombre = str(it.get('item', ''))
            real = float(it.get('precio_real', 0) or 0)
            if real > 0:
                precios_reales[nombre] = {
                    'real': real,
                    'categoria': it.get('categoria', ''),
                    'presup': float(it.get('precio_presupuestado', 0) or 0)
                }

    row = 3
    alt = False
    for prod in prods_valid:
        item_nombre = str(prod.get('Item', ''))
        if item_nombre not in precios_reales:
            continue
        datos = precios_reales[item_nombre]
        cat = datos['categoria'] or str(prod.get('Categoria', ''))
        pp = datos['presup'] or round(float(prod.get('Precio Unitario', 0) or 0))
        pr = datos['real']
        dif = pp - pr

        bg = PatternFill("solid", fgColor="FFFFFF" if not alt else gris_claro)
        alt = not alt

        vals = [cat, item_nombre, pp, pr, dif]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = bg
            cell.border = borde
            cell.font = Font(name='Calibri', size=9)
            if col in (3, 4, 5):
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            if col == 5:
                cell.font = Font(name='Calibri', size=9, bold=True,
                                 color=verde if dif >= 0 else rojo)
        ws.row_dimensions[row].height = 18
        row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    if row > 3:
        ws.cell(row=row, column=1, value='').border = borde
        ws.cell(row=row, column=2, value='TOTAL ÍTEMS COMPRADOS').font = Font(bold=True, name='Calibri', size=9)
        ws.cell(row=row, column=2).border = borde
        ws.cell(row=row, column=3, value=f'{row - 3} de {len(prods_valid)} ítems').font = Font(name='Calibri', size=9, color="64748B")
        ws.cell(row=row, column=3).border = borde
        ws.cell(row=row, column=4).border = borde
        ws.cell(row=row, column=5).border = borde
        ws.row_dimensions[row].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── PDF BALANCE ───────────────────────────────────────────────────────────────

def generar_pdf_balance(cotizacion_numero, datos_cliente, datos_asesor, registros,
                        productos_presupuesto, incluir_varios=False):
    """Genera PDF de balance de compras consolidando todos los registros."""
    import io, json
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from datetime import datetime, timezone, timedelta

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm,
                            topMargin=2 * cm, bottomMargin=2 * cm)
    elements = []
    styles = getSampleStyleSheet()
    tz_cl = timezone(timedelta(hours=-3))

    def _sty(name, **kw):
        try:
            styles.add(ParagraphStyle(name=name, parent=styles['Normal'], **kw))
        except Exception:
            pass
        return styles[name]

    _sty('BTitle', fontSize=18, fontName='Helvetica-Bold', spaceAfter=4, textColor=colors.HexColor('#1e2447'))
    _sty('BSubtitle', fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=2)
    _sty('BSection', fontSize=11, fontName='Helvetica-Bold', spaceAfter=4, textColor=colors.HexColor('#1e2447'), spaceBefore=10)
    _sty('BLabel', fontSize=9, textColor=colors.HexColor('#64748b'))
    _sty('BValue', fontSize=9, fontName='Helvetica-Bold')
    _sty('BSmall', fontSize=8, textColor=colors.HexColor('#64748b'))
    _sty('BRegHeader', fontSize=9, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e2447'), spaceBefore=8, spaceAfter=2)

    now_str = datetime.now(tz_cl).strftime('%d/%m/%Y %H:%M')

    _logo_cell = ""
    try:
        from reportlab.platypus import Image as _RLImage
        _logo = _RLImage("logo.png")
        _logo_w = 4 * cm
        _logo_aspect = _logo.imageHeight / float(_logo.imageWidth)
        _logo.drawWidth = _logo_w
        _logo.drawHeight = _logo_w * _logo_aspect
        _logo_cell = _logo
    except Exception:
        pass

    header_data = [[
        _logo_cell,
        Paragraph("<b>BALANCE DE COMPRAS" + (" (CON VARIOS)" if incluir_varios else "") + "</b>", styles['BTitle']),
        Paragraph(f"Generado: {now_str}", styles['BSmall'])
    ]]
    header_tbl = Table(header_data, colWidths=[4.5 * cm, 9 * cm, 4 * cm])
    header_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(header_tbl)
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(HRFlowable(width='100%', thickness=2, color=colors.HexColor('#1e2447'), spaceBefore=4, spaceAfter=8))

    _nombre = datos_cliente.get('Nombre', '')
    _rut = datos_cliente.get('RUT', '')
    _asesor = datos_asesor.get('Nombre Ejecutivo', '')
    info_data = [
        [Paragraph('<b>N° Presupuesto</b>', styles['BLabel']), Paragraph(cotizacion_numero, styles['BValue']),
         Paragraph('<b>Cliente</b>', styles['BLabel']), Paragraph(_nombre, styles['BValue'])],
        [Paragraph('<b>RUT</b>', styles['BLabel']), Paragraph(_rut, styles['BValue']),
         Paragraph('<b>Ejecutivo</b>', styles['BLabel']), Paragraph(_asesor, styles['BValue'])],
    ]
    info_tbl = Table(info_data, colWidths=[3 * cm, 6 * cm, 3 * cm, 5.5 * cm])
    info_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 0.3 * cm))

    if incluir_varios:
        prods_valid = list(productos_presupuesto or [])
    else:
        prods_valid = [p for p in (productos_presupuesto or [])
                       if str(p.get('Categoria', '')).strip().lower() != 'varios']
    total_items = len(prods_valid)
    items_en_registros = set()
    for reg in registros:
        items_r = reg.get('items') or []
        if isinstance(items_r, str):
            try:
                items_r = json.loads(items_r)
            except Exception:
                items_r = []
        for it in items_r:
            if float(it.get('precio_real', 0) or 0) > 0:
                items_en_registros.add(str(it.get('item', '')))
    comprados = sum(1 for p in prods_valid if str(p.get('Item', '')) in items_en_registros)
    pct = round(comprados / total_items * 1000) / 10 if total_items > 0 else 0
    pct_col = colors.HexColor('#3b82f6') if pct >= 100 else (
              colors.HexColor('#16a34a') if pct >= 66.6 else (
              colors.HexColor('#eab308') if pct >= 33.3 else colors.HexColor('#dc2626')))
    pct_lbl = 'Compra finalizada' if pct >= 100 else f'{pct}% comprado'

    prog_data = [[
        Paragraph('<b>Progreso de compra</b>', styles['BLabel']),
        Paragraph(f'<b>{pct_lbl}</b>', ParagraphStyle('_pc', parent=styles['Normal'],
            fontSize=11, fontName='Helvetica-Bold', textColor=pct_col)),
        Paragraph(f'{comprados} de {total_items} ítems', styles['BSmall']),
    ]]
    prog_tbl = Table(prog_data, colWidths=[3.5 * cm, 6 * cm, 3 * cm])
    prog_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (0, -1), 8),
    ]))
    elements.append(prog_tbl)
    elements.append(Spacer(1, 0.4 * cm))

    elements.append(Paragraph('Detalle de Registros de Compra', styles['BSection']))
    elements.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#e2e8f0'), spaceAfter=6))

    _tipo_labels = {'online': 'Compra Online', 'presencial': 'Compra Presencial'}
    _subtipo_labels = {'retiro': 'Retiro', 'despacho': 'Despacho',
                       'completo': 'Retiro Completo', 'parcial': 'Retiro Parcial'}
    col_azul = colors.HexColor('#1e2447')
    col_gris = colors.HexColor('#64748b')
    col_verde = colors.HexColor('#16a34a')
    col_rojo = colors.HexColor('#dc2626')

    tbl_header = ['Categoría', 'Ítem', 'Cant.', 'Presup.', 'Real', 'Adic.', 'Diferencia']
    col_ws = [2.5 * cm, 5.5 * cm, 1.2 * cm, 2.2 * cm, 2.2 * cm, 1.2 * cm, 2.7 * cm]

    for idx_r, reg in enumerate(registros):
        try:
            fecha_reg = datetime.fromisoformat(reg['fecha_registro'].replace('Z', '+00:00')).astimezone(tz_cl).strftime('%d/%m/%Y %H:%M')
        except Exception:
            fecha_reg = '—'
        lugar = reg.get('lugar_compra', '') or '—'
        tipo = _tipo_labels.get(reg.get('tipo_compra', ''), reg.get('tipo_compra', '') or '—')
        subtipo = _subtipo_labels.get(reg.get('subtipo_compra', ''), reg.get('subtipo_compra', '') or '')
        tipo_full = f"{tipo} — {subtipo}" if subtipo else tipo
        fecha_ent = reg.get('fecha_entrega_compra', '') or ''
        falto = reg.get('falto_retirar', '') or ''
        obs = reg.get('observaciones', '') or 'Sin observaciones'
        factura = reg.get('factura_nombre', '') or '—'

        reg_info = f"Registro #{idx_r + 1} — {fecha_reg} | {lugar} | {tipo_full}"
        if fecha_ent:
            reg_info += f" | Para: {fecha_ent}"
        elements.append(Paragraph(reg_info, styles['BRegHeader']))

        if falto:
            elements.append(Paragraph(f"Faltó retirar: {falto}", styles['BSmall']))
        elements.append(Paragraph(f"Observación: {obs}", styles['BSmall']))
        _factura_url = reg.get('factura_url', '') or ''
        if _factura_url:
            elements.append(Paragraph(
                f'Factura: <link href="{_factura_url}"><u><font color="#3b82f6">{factura}</font></u></link>',
                styles['BSmall']
            ))
        else:
            elements.append(Paragraph(f"Factura: {factura}", styles['BSmall']))
        elements.append(Spacer(1, 0.2 * cm))

        items_r = reg.get('items') or []
        if isinstance(items_r, str):
            try:
                items_r = json.loads(items_r)
            except Exception:
                items_r = []

        if items_r:
            rows = [tbl_header]
            row_types = ['header']
            sub_p = 0; sub_r = 0; sub_a = 0; sub_s = 0
            _pn_pdf = {str(p.get('Item', '')) for p in (productos_presupuesto or [])}
            _pp_map = {str(p.get('Item', '')): round(float(p.get('Precio Unitario', 0) or 0)) for p in (productos_presupuesto or [])}
            for it in items_r:
                pp = float(it.get('precio_presupuestado', 0) or 0)
                pr = float(it.get('precio_real', 0) or 0)
                cant = float(it.get('cantidad', 1) or 1)
                adic = int(it.get('adicional', 0) or 0)
                # Stock parcial: solo se pagaron (cant - stock_cantidad) unidades;
                # el resto es inventario ($0, ahorro puro).
                _es_stk = bool(it.get('stock', False))
                _stk_c = float(it.get('stock_cantidad', cant) or 0) if _es_stk else 0.0
                _stk_c = max(0.0, min(cant, _stk_c))
                _purch = (cant - _stk_c) if _es_stk else cant
                dif = pp * cant - pr * _purch - (adic * pr)
                _is_sin = it.get('sin_registro', False)
                _is_con = (it.get('es_adicional', False) or str(it.get('item', '')) not in _pn_pdf) and not _is_sin
                pp_real = _pp_map.get(str(it.get('item', '')), pp) if not _is_con and not _is_sin else pp
                if _is_sin:
                    sub_s += pr * cant
                elif _is_con:
                    sub_a += pr * cant
                else:
                    sub_p += pp_real * cant; sub_r += pr * _purch + adic * pr
                dif_str = f"${abs(dif):,.0f} {'▼' if dif >= 0 else '▲'}".replace(',', '.')
                rows.append([it.get('categoria', ''), it.get('item', ''), str(int(cant)),
                    f"${pp_real:,.0f}".replace(',', '.'), f"${pr:,.0f}".replace(',', '.'),
                    str(adic), dif_str])
                row_types.append('sin' if _is_sin else ('con' if _is_con else 'normal'))
            bal_r = sub_p - sub_r
            rows.append(['', 'SUBTOTAL PRESUPUESTO', '', f"${sub_p:,.0f}".replace(',', '.'),
                         f"${sub_r:,.0f}".replace(',', '.'), '',
                         f"${abs(bal_r):,.0f} {'▼' if bal_r >= 0 else '▲'}".replace(',', '.')])
            row_types.append('subtotal')
            if sub_a > 0:
                rows.append(['', 'ADICIONALES CON REGISTRO', '', '—', f"${sub_a:,.0f}".replace(',', '.'), '', ''])
                row_types.append('subtotal_con')
            if sub_s > 0:
                rows.append(['', 'ADICIONALES SIN REGISTRO', '', '—', f"${sub_s:,.0f}".replace(',', '.'), '', ''])
                row_types.append('subtotal_sin')
            tbl = Table(rows, colWidths=col_ws, repeatRows=1)
            tbl_style = [
                ('BACKGROUND', (0, 0), (-1, 0), col_azul),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (1, -1), 'LEFT'),
                ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]
            for ri, rtype in enumerate(row_types):
                if rtype == 'con':
                    tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#fff7ed')))
                    tbl_style.append(('TEXTCOLOR', (0, ri), (1, ri), colors.HexColor('#c2410c')))
                elif rtype == 'sin':
                    tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#fdf2f8')))
                    tbl_style.append(('TEXTCOLOR', (0, ri), (1, ri), colors.HexColor('#9d174d')))
                elif rtype == 'subtotal':
                    tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#f1f5f9')))
                    tbl_style.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                elif rtype == 'subtotal_con':
                    tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#fff3e0')))
                    tbl_style.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                    tbl_style.append(('TEXTCOLOR', (0, ri), (-1, ri), colors.HexColor('#c2410c')))
                elif rtype == 'subtotal_sin':
                    tbl_style.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#fdf2f8')))
                    tbl_style.append(('FONTNAME', (0, ri), (-1, ri), 'Helvetica-Bold'))
                    tbl_style.append(('TEXTCOLOR', (0, ri), (-1, ri), colors.HexColor('#9d174d')))
                if ri > 0 and row_types[ri] not in ('subtotal', 'subtotal_con', 'subtotal_sin') and len(rows[ri]) > 6 and rows[ri][6]:
                    is_ahorro = '▼' in rows[ri][6]
                    tbl_style.append(('TEXTCOLOR', (6, ri), (6, ri), col_verde if is_ahorro else col_rojo))
            tbl.setStyle(TableStyle(tbl_style))
            elements.append(tbl)

        elements.append(Spacer(1, 0.4 * cm))
        if idx_r < len(registros) - 1:
            elements.append(HRFlowable(width='100%', thickness=0.3, color=colors.HexColor('#e2e8f0'), spaceAfter=4))

    elements.append(Paragraph('Resumen Final Consolidado', styles['BSection']))
    elements.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#e2e8f0'), spaceAfter=8))

    _tots = calcular_totales_rc(productos_presupuesto, registros, incluir_varios=incluir_varios)
    total_p = _tots['tP']; total_r = _tots['tR']
    total_adic_con = _tots['tA']; total_adic_sin = _tots['tS']

    iva_p = total_p * 0.19; iva_r = total_r * 0.19
    bal = total_p - total_r; iva_bal = iva_p - iva_r
    bal_col = col_verde if bal >= 0 else col_rojo
    bal_lbl = 'AHORRO' if bal >= 0 else 'SOBRECOSTO'

    def _fmt(v): return f"${abs(v):,.0f}".replace(',', '.')

    iva_adic_con = total_adic_con * 0.19; iva_adic_sin = total_adic_sin * 0.19

    resumen_rows = [
        ['', 'PRESUPUESTADO', 'REAL', 'BALANCE', 'ADIC. C/REG.', 'ADIC. S/REG.'],
        ['Subtotal neto', _fmt(total_p), _fmt(total_r), _fmt(bal), _fmt(total_adic_con), _fmt(total_adic_sin)],
        ['IVA (19%)', _fmt(iva_p), _fmt(iva_r), _fmt(iva_bal), _fmt(iva_adic_con), _fmt(iva_adic_sin)],
        ['Total con IVA', _fmt(total_p + iva_p), _fmt(total_r + iva_r), _fmt(bal + iva_bal),
         _fmt(total_adic_con + iva_adic_con), _fmt(total_adic_sin + iva_adic_sin)],
    ]
    res_tbl = Table(resumen_rows, colWidths=[3 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 2.5 * cm, 2.5 * cm])
    res_style = [
        ('BACKGROUND', (0, 0), (-1, 0), col_azul),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TEXTCOLOR', (3, 1), (3, -1), bal_col),
        ('TEXTCOLOR', (4, 1), (4, -1), colors.HexColor('#c2410c')),
        ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#ea580c')),
        ('BACKGROUND', (4, 1), (4, -1), colors.HexColor('#fff7ed')),
        ('TEXTCOLOR', (4, 0), (4, 0), colors.white),
        ('TEXTCOLOR', (5, 1), (5, -1), colors.HexColor('#9d174d')),
        ('BACKGROUND', (5, 0), (5, 0), colors.HexColor('#db2777')),
        ('BACKGROUND', (5, 1), (5, -1), colors.HexColor('#fdf2f8')),
        ('TEXTCOLOR', (5, 0), (5, 0), colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    res_tbl.setStyle(TableStyle(res_style))
    elements.append(res_tbl)
    elements.append(Spacer(1, 0.3 * cm))

    badge_data = [[Paragraph(
        f"<b>{bal_lbl}: {_fmt(bal + iva_bal)} (con IVA)</b>",
        ParagraphStyle('_badge', parent=styles['Normal'], fontSize=12,
            fontName='Helvetica-Bold', textColor=bal_col, alignment=1)
    )]]
    badge_tbl = Table(badge_data, colWidths=[16 * cm])
    badge_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0fdf4') if bal >= 0 else colors.HexColor('#fef2f2')),
        ('BOX', (0, 0), (-1, -1), 1, bal_col),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(badge_tbl)

    doc.build(elements)
    buf.seek(0)
    return buf.read()


# Menú contextual (click derecho en la tabla o click en el hint) para agregar un
# producto CON o SIN registro reutilizando el #add-section (que ahora es un popup
# oculto por defecto → interfaz más limpia). Va como <script> aparte (string
# normal, llaves reales) que se concatena al html del iframe, así no hay que
# escapar llaves dentro del f-string gigante de build_rc_html.
_RC_ADD_MENU_JS = """
<script>
(function(){
  function openAddPopup(mode){
    var sec=document.getElementById("add-section");
    var bd=document.getElementById("rc-add-backdrop");
    if(!sec||!bd) return;
    sec.style.cssText="display:block;position:fixed;left:50%;top:50%;transform:translate(-50%,-50%);width:min(800px,95vw);max-height:88vh;overflow:auto;background:#fff;border:1px solid #e2e8f0;border-radius:14px;box-shadow:0 26px 64px rgba(15,23,42,.34);z-index:99999;padding:18px 22px;box-sizing:border-box";
    bd.style.display="block";
    if(window.switchAddTab) window.switchAddTab(mode);
    setTimeout(function(){var f=document.getElementById(mode==="reg"?"add-cat":"sin-cat"); if(f){try{f.focus();}catch(e){}}},40);
  }
  window.openAddPopup=openAddPopup;
  function openSavePopup(){
    var sec=document.getElementById("save-section");
    var bd=document.getElementById("rc-add-backdrop");
    if(!sec||!bd) return;
    var add=document.getElementById("add-section"); if(add) add.style.display="none";
    sec.style.cssText="display:block;position:fixed;left:50%;top:50%;transform:translate(-50%,-50%);width:min(920px,95vw);max-height:88vh;overflow:auto;background:#f8fafc;border:1px solid #e2e8f0;border-radius:14px;box-shadow:0 26px 64px rgba(15,23,42,.34);z-index:99999;padding:18px 22px;box-sizing:border-box";
    bd.style.display="block";
    if(window.checkSaveBtn) window.checkSaveBtn();
  }
  window.openSavePopup=openSavePopup;
  window.closeAddPopup=function(){
    var add=document.getElementById("add-section"); if(add) add.style.display="none";
    var save=document.getElementById("save-section"); if(save) save.style.display="none";
    var bd=document.getElementById("rc-add-backdrop"); if(bd) bd.style.display="none";
  };
  function rcHasData(){
    var rows=document.querySelectorAll("tr[data-idx]");
    for(var i=0;i<rows.length;i++){
      var r=rows[i];
      if(r.dataset.comprado==="1") continue;
      var inp=r.querySelector(".rc-real");
      var re=inp?(parseFloat(inp.dataset.val)||0):0;
      var c=+r.dataset.cant||1;
      var isStock=r.getAttribute("data-stock")==="1";
      var stockSaved=r.getAttribute("data-stock-saved")==="1";
      var sq=isStock?(parseInt(r.getAttribute("data-stock-qty"))||c):0;
      var buyAttr=r.getAttribute("data-buy");
      var compr=(buyAttr!=null)?(parseInt(buyAttr)||0):(isStock?(c-sq):c);
      var newStock=isStock&&!stockSaved&&sq>0;
      if(newStock||(re>0&&compr>0)) return true;
    }
    return false;
  }
  window.rcHasData=rcHasData;
  // ¿Hay ALGO escrito sin guardar en la tabla? (para habilitar "Deshacer": precio
  // real, adicional o un "en stock" recién marcado, en filas editables no guardadas.)
  function rcHasAnyInput(){
    var rows=document.querySelectorAll("tr[data-idx]");
    for(var i=0;i<rows.length;i++){
      var r=rows[i]; if(r.dataset.comprado==="1") continue;
      var real=r.querySelector(".rc-real");
      if(real && !real.readOnly && (parseFloat(real.dataset.val)||0)>0) return true;
      var adic=r.querySelector(".rc-adic");
      if(adic && !adic.readOnly && (parseInt(adic.value)||0)>0) return true;
      var stk=r.querySelector(".rc-stock");
      if(stk && stk.checked && r.getAttribute("data-stock-saved")!=="1") return true;
    }
    return false;
  }
  // "Deshacer": limpia lo escrito SIN guardar (no toca filas ya guardadas).
  function rcUndo(){
    var rows=document.querySelectorAll("tr[data-idx]");
    for(var i=0;i<rows.length;i++){
      var r=rows[i]; if(r.dataset.comprado==="1") continue;
      // "en stock" recién marcado (no guardado) → desmarcar (resetea real + attrs)
      var stk=r.querySelector(".rc-stock");
      if(stk && stk.checked && r.getAttribute("data-stock-saved")!=="1"){
        stk.checked=false; if(window.toggleStock) window.toggleStock(stk);
      }
      var real=r.querySelector(".rc-real");
      if(real && !real.readOnly){ real.value=""; real.dataset.val="0"; }
      var adic=r.querySelector(".rc-adic");
      if(adic && !adic.readOnly){ adic.value=""; }
    }
    if(window.calc) window.calc();
    if(window.checkSaveBtn) window.checkSaveBtn();
  }
  window.rcUndo=rcUndo;
  var SVG_SAVE='<svg viewBox="0 0 24 24" width="15" height="15" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M19 21H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11l5 5v11a2 2 0 0 1-2 2z"/><polyline points="17 21 17 13 7 13 7 21"/><polyline points="7 3 7 8 15 8"/></svg>';
  var SVG_UNDO='<svg viewBox="0 0 24 24" width="15" height="15" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M3 7v6h6"/><path d="M21 17a9 9 0 0 0-9-9 9 9 0 0 0-6 2.3L3 13"/></svg>';
  function closeMenu(){ var m=document.getElementById("rc-add-ctxmenu"); if(m) m.remove(); }
  function menuItem(m,label,color,dot,disabled,cb,svg){
    var b=document.createElement("button");
    b.style.cssText="display:flex;align-items:center;gap:9px;width:100%;background:transparent;border:none;border-radius:8px;padding:9px 11px;cursor:"+(disabled?"not-allowed":"pointer")+";text-align:left;font-family:Montserrat,'Segoe UI',sans-serif;font-size:12px;font-weight:700;color:"+(disabled?"#cbd5e1":color);
    var ic = svg
      ? '<span style="width:16px;height:16px;display:inline-flex;align-items:center;justify-content:center;flex:0 0 auto;color:'+(disabled?"#cbd5e1":color)+'">'+svg+'</span>'
      : '<span style="width:9px;height:9px;border-radius:50%;background:'+(disabled?"#e2e8f0":dot)+';flex:0 0 auto"></span>';
    b.innerHTML=ic+'<span>'+label+'</span>';
    if(!disabled){
      b.addEventListener("mouseenter",function(){b.style.background="#f8fafc";});
      b.addEventListener("mouseleave",function(){b.style.background="transparent";});
      b.addEventListener("click",function(ev){ev.stopPropagation();closeMenu();cb();});
    }
    m.appendChild(b);
  }
  function buildMenu(x,y){
    closeMenu();
    var m=document.createElement("div"); m.id="rc-add-ctxmenu";
    m.style.cssText="position:fixed;z-index:100000;background:#fff;border:1px solid #e5e9f2;border-radius:12px;box-shadow:0 14px 40px rgba(15,23,42,.24);padding:6px;min-width:256px";
    menuItem(m,"Agregar producto CON registro","#c2410c","#f97316",false,function(){openAddPopup("reg");});
    menuItem(m,"Agregar producto SIN registro","#be185d","#ec4899",false,function(){openAddPopup("sin");});
    var sep=document.createElement("div"); sep.style.cssText="height:1px;background:#eef1f6;margin:5px 4px"; m.appendChild(sep);
    menuItem(m,"Deseo guardar","#15803d","#22c55e",!rcHasData(),function(){openSavePopup();},SVG_SAVE);
    menuItem(m,"Deshacer","#475569","#94a3b8",!rcHasAnyInput(),function(){rcUndo();},SVG_UNDO);
    document.body.appendChild(m);
    var r=m.getBoundingClientRect(), px=x, py=y;
    if(px+r.width>window.innerWidth) px=window.innerWidth-r.width-8;
    if(py+r.height>window.innerHeight) py=window.innerHeight-r.height-8;
    m.style.left=px+"px"; m.style.top=py+"px";
  }
  window.rcAddMenuAt=function(ev){ if(ev){try{ev.preventDefault();ev.stopPropagation();}catch(e){}} buildMenu(ev?ev.clientX:80, ev?ev.clientY:80); };
  var wrap=document.getElementById("tbl-wrap");
  if(wrap){ wrap.addEventListener("contextmenu",function(ev){ ev.preventDefault(); buildMenu(ev.clientX, ev.clientY); }); }
  document.addEventListener("click",function(ev){ var t=ev.target; if(!t.closest||!t.closest("#rc-add-ctxmenu")) closeMenu(); });
  document.addEventListener("keydown",function(ev){ if(ev.key==="Escape"){ closeMenu(); window.closeAddPopup(); } });
})();
</script>
"""

# Herramientas de la tabla: descargar CSV (desde el doc padre, el iframe está
# sandboxed) + pantalla completa (frameElement). Mismos iconos que PRESUPUESTO.
# String normal (llaves reales) concatenado al html; sin backslashes (String.fromCharCode).
_RC_TABLE_TOOLS_JS = """
<script>
(function(){
  var P=window.parent, IFR=null;
  try{ IFR=window.frameElement; }catch(e){}
  if(!IFR){ try{ var ifs=P.document.querySelectorAll("iframe"); for(var i=0;i<ifs.length;i++){ if(ifs[i].contentWindow===window){ IFR=ifs[i]; break; } } }catch(e){} }
  function clean(s){ s=String(s==null?"":s); var sp=String.fromCharCode(32); s=s.split(String.fromCharCode(10)).join(sp); s=s.split(String.fromCharCode(13)).join(sp); s=s.split(String.fromCharCode(9)).join(sp); while(s.indexOf(sp+sp)>=0) s=s.split(sp+sp).join(sp); return s.trim(); }
  function cell(el){ var inp=el.querySelector("input"); if(inp){ if(inp.type==="checkbox") return inp.checked?"Si":""; return clean(inp.value||""); } var c=el.cloneNode(true); var hs=c.querySelectorAll("script,style,button,svg"); for(var i=0;i<hs.length;i++)hs[i].remove(); return clean(c.innerText||c.textContent||""); }
  function csvVal(v){ v=String(v==null?"":v); var q=String.fromCharCode(34),nl=String.fromCharCode(10); if(v.indexOf(q)>=0||v.indexOf(",")>=0||v.indexOf(";")>=0||v.indexOf(nl)>=0){ v=q+v.split(q).join(q+q)+q; } return v; }
  function dlCSV(){
    var t=document.querySelector("#tbl-wrap table"); if(!t)return;
    var rows=[]; var hs=t.querySelectorAll("thead th"); var hd=[]; for(var i=0;i<hs.length;i++)hd.push(cell(hs[i])); rows.push(hd);
    var trs=t.querySelectorAll("tbody tr"); for(var j=0;j<trs.length;j++){ if(trs[j].style.display==="none")continue; var tds=trs[j].querySelectorAll("td"); if(!tds.length)continue; var r=[]; for(var k=0;k<tds.length;k++)r.push(cell(tds[k])); rows.push(r); }
    var lines=[]; for(var mm=0;mm<rows.length;mm++){ var cols=[]; for(var nn=0;nn<rows[mm].length;nn++)cols.push(csvVal(rows[mm][nn])); lines.push(cols.join(",")); }
    var csv=String.fromCharCode(65279)+lines.join(String.fromCharCode(10));
    var dt=new Date(),pp=function(x){return (x<10?"0":"")+x;},fn="registro_compras_"+dt.getFullYear()+pp(dt.getMonth()+1)+pp(dt.getDate())+".csv";
    try{
      var blob=new P.Blob([csv],{type:"text/csv;charset=utf-8;"}); var url=P.URL.createObjectURL(blob);
      var a=P.document.createElement("a"); a.href=url; a.download=fn; P.document.body.appendChild(a); a.click();
      setTimeout(function(){ a.remove(); P.URL.revokeObjectURL(url); },1500);
    }catch(e){}
  }
  var cb=document.getElementById("_rc_csvbtn"); if(cb) cb.onclick=dlCSV;
  var EXP='<svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M8 3H5a2 2 0 0 0-2 2v3"/><path d="M21 8V5a2 2 0 0 0-2-2h-3"/><path d="M3 16v3a2 2 0 0 0 2 2h3"/><path d="M16 21h3a2 2 0 0 0 2-2v-3"/></svg>';
  var SHR='<svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M4 14h6v6"/><path d="M20 10h-6V4"/><path d="M14 10l7-7"/><path d="M3 21l7-7"/></svg>';
  var fb=document.getElementById("_rc_fsbtn");
  var PROPS=[["position","fixed"],["top","0"],["left","0"],["width","100vw"],["height","100vh"],["z-index","2147483000"],["border","none"],["border-radius","0"],["margin","0"],["background","#fff"]];
  function isFS(){ return P._rcFsActive===true; }
  function applyFS(){ if(!IFR)return; for(var i=0;i<PROPS.length;i++) IFR.style.setProperty(PROPS[i][0],PROPS[i][1],"important"); try{document.documentElement.classList.add("rc-fs");}catch(e){} P._rcFsActive=true; if(fb){fb.innerHTML=SHR;fb.title="Salir de pantalla completa";} }
  function removeFS(){ if(IFR){for(var i=0;i<PROPS.length;i++) IFR.style.removeProperty(PROPS[i][0]);} try{document.documentElement.classList.remove("rc-fs");}catch(e){} P._rcFsActive=false; if(fb){fb.innerHTML=EXP;fb.title="Pantalla completa";} }
  function toggleFS(){ if(isFS())removeFS(); else applyFS(); }
  if(fb){ fb.onclick=toggleFS; fb.innerHTML=isFS()?SHR:EXP; }
  if(isFS()) applyFS();
  document.addEventListener("keydown",function(e){ if(e.key==="Escape"&&isFS()) removeFS(); });
  try{ if(P._rcFsEsc) P.document.removeEventListener("keydown",P._rcFsEsc,true); P._rcFsEsc=function(e){ if(e.key==="Escape"&&isFS()) removeFS(); }; P.document.addEventListener("keydown",P._rcFsEsc,true); }catch(e){}
})();
</script>
"""


# Toggle "Modo Admin" DENTRO del formulario: al pulsar el switch escribe su nuevo
# estado ("1|ts"/"0|ts") en el puente oculto _rc_admin_tg del documento padre →
# Python re-filtra (incluye/excluye Varios) y reconstruye el iframe. Refleja el
# estado al instante (aunque igual se recargue tras el rerun).
_RC_ADMIN_TOGGLE_JS = """
<script>
(function(){
  var P=window.parent, PD=P&&P.document; if(!PD) return;
  var btn=document.getElementById("_rc_admtg"); if(!btn) return;
  btn.addEventListener("click",function(){
    var newOn = btn.getAttribute("data-on")==="1" ? "0" : "1", on = newOn==="1";
    btn.setAttribute("data-on",newOn); btn.setAttribute("aria-checked", on?"true":"false");
    btn.style.background = on ? "#5b7cfa" : "#cbd5e1";
    var kn=btn.querySelector("span"); if(kn) kn.style.left = on ? "19px" : "3px";
    var inp=PD.querySelector(".st-key-_rc_admin_tg input"); if(!inp) return;
    try{
      var setter=Object.getOwnPropertyDescriptor(P.HTMLInputElement.prototype,"value").set;
      inp.focus({preventScroll:true});
      setter.call(inp,newOn+"|"+Date.now());
      inp.dispatchEvent(new Event("input",{bubbles:true}));
      inp.dispatchEvent(new Event("change",{bubbles:true}));
      inp.dispatchEvent(new KeyboardEvent("keypress",{key:"Enter",keyCode:13,which:13,charCode:13,bubbles:true}));
      inp.dispatchEvent(new KeyboardEvent("keydown",{key:"Enter",keyCode:13,which:13,bubbles:true}));
      inp.dispatchEvent(new KeyboardEvent("keyup",{key:"Enter",keyCode:13,which:13,bubbles:true}));
      inp.dispatchEvent(new FocusEvent("blur",{bubbles:false}));
      inp.dispatchEvent(new FocusEvent("focusout",{bubbles:true}));
      inp.blur();
    }catch(e){}
  });
})();
</script>
"""

# Toggle "Resultados detallados" DENTRO del formulario: muestra/oculta el resumen
# (#tots) 100% client-side (SIN rerun ni recarga del iframe → no pierde lo escrito
# en la tabla). Por defecto OCULTO (interfaz limpia). El estado se recuerda en
# window.parent (sobrevive recargas del iframe; se resetea sólo en recarga total).
_RC_DETAILS_TOGGLE_JS = """
<script>
(function(){
  var P=window.parent;
  var btn=document.getElementById("_rc_dettg"); if(!btn) return;
  var tots=document.getElementById("tots");
  function apply(on){
    btn.setAttribute("data-on", on?"1":"0"); btn.setAttribute("aria-checked", on?"true":"false");
    btn.style.background = on?"#5b7cfa":"#cbd5e1";
    var kn=btn.querySelector("span"); if(kn) kn.style.left = on?"19px":"3px";
    if(tots) tots.style.display = on ? "grid" : "none";
  }
  var on=false; try{ on=!!(P&&P._rcDetails); }catch(e){}
  apply(on);
  btn.addEventListener("click",function(){ on=!on; try{ if(P) P._rcDetails=on; }catch(e){} apply(on); });
})();
</script>
"""


# ── HTML BUILDER REGISTRO DE COMPRAS ─────────────────────────────────────────

def build_rc_html(rc_prods, rc_cat_json, rc_prev, items_comprados=None, es_admin=False,
                  supa_url='', supa_key='', ep='', usuario='', items_ya_comprados_json='[]',
                  total_items_presupuesto=0, cats_cards_html='', proveedores=None,
                  cat_colors=None, admin_on=False):
    rows = ""
    items_comprados = items_comprados or {}
    cat_colors = cat_colors or {}
    # Fila de toggles DENTRO del formulario, entre el buscador y la tabla:
    #  · "Modo admin (incluye Varios)" — solo admin; escribe su estado en el puente
    #    _rc_admin_tg → Python re-filtra (ver _RC_ADMIN_TOGGLE_JS).
    #  · "Resultados detallados" — TODOS; muestra/oculta el resumen (#tots). Es 100%
    #    client-side (sin rerun): por defecto OCULTO para despejar la interfaz; el
    #    estado se recuerda en window.parent (ver _RC_DETAILS_TOGGLE_JS).
    def _rc_switch(_id, _on, _label):
        _c = "#5b7cfa" if _on else "#cbd5e1"
        _x = "19px" if _on else "3px"
        return (
            '<div style="display:flex;align-items:center;gap:10px">'
            f'<button id="{_id}" type="button" role="switch" aria-checked="{str(bool(_on)).lower()}" '
            f'data-on="{"1" if _on else "0"}" style="position:relative;width:38px;height:22px;border-radius:99px;'
            f'border:none;cursor:pointer;flex:0 0 auto;transition:background .18s;background:{_c}">'
            f'<span style="position:absolute;top:3px;left:{_x};width:16px;height:16px;border-radius:50%;'
            'background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.25);transition:left .18s"></span></button>'
            f'<span style="font-family:Montserrat,sans-serif;font-weight:700;font-size:0.7rem;letter-spacing:0.06em;'
            f'text-transform:uppercase;color:#475569">{_label}</span></div>'
        )
    _toggles = ""
    if es_admin:
        _toggles += _rc_switch("_rc_admtg", admin_on, "Modo admin (incluye Varios)")
    _toggles += _rc_switch("_rc_dettg", False, "Resultados detallados")
    admin_toggle_html = (
        '<div style="display:flex;align-items:center;gap:26px;flex-wrap:wrap;padding:9px 12px;'
        'border-bottom:1px solid #eef2f7;background:#f8fafc;flex-shrink:0">' + _toggles + '</div>'
    )

    def _rc_badge_style(_hex):
        """Badge de categoría: fondo tenue + texto en el color (legible)."""
        _h = str(_hex or '#64748b').lstrip('#')
        try:
            _r, _g, _b = int(_h[0:2], 16), int(_h[2:4], 16), int(_h[4:6], 16)
        except Exception:
            _r, _g, _b = 100, 116, 139
        # texto un poco más oscuro que el color base para contraste sobre fondo claro
        _dr, _dg, _db = int(_r * 0.72), int(_g * 0.72), int(_b * 0.72)
        return (f'background:rgba({_r},{_g},{_b},0.13);color:rgb({_dr},{_dg},{_db});'
                f'border:1px solid rgba({_r},{_g},{_b},0.22);')

    # Datalist de proveedores ya usados (autocompletar "¿Dónde compraste?").
    # Son los lugar_compra distintos del historial; si se escribe uno nuevo, al
    # guardarse queda automáticamente en la lista la próxima vez.
    def _esc_prov(_s):
        return (str(_s).replace('&', '&amp;').replace('<', '&lt;')
                .replace('>', '&gt;').replace('"', '&quot;'))
    # INVENTARIO siempre disponible como "proveedor" (stock propio, ahorro puro).
    _prov_names = [str(_p).strip() for _p in (proveedores or [])
                   if str(_p).strip() and str(_p).strip().upper() != 'INVENTARIO']
    _prov_opts = '<option value="INVENTARIO"></option>' + ''.join(
        f'<option value="{_esc_prov(_p)}"></option>' for _p in _prov_names
    )
    _prov_datalist = f'<datalist id="rc-proveedores">{_prov_opts}</datalist>'
    for ri, prod in enumerate(rc_prods):
        cat = str(prod.get('Categoria', ''))
        item = str(prod.get('Item', ''))
        cant = round(float(prod.get('Cantidad', 1) or 1))
        pu = round(float(prod.get('Precio Unitario', 0) or 0))
        _es_adicional = bool(prod.get('_adicional', False))
        _es_sin_reg = bool(prod.get('_sin_registro', False))
        _ic = items_comprados.get(item, {})
        # Cobertura acumulada del ítem (unidades ya en stock + ya compradas).
        _stock_units = int(_ic.get('stock_units', 0) or 0) if _ic else 0
        _bought_units = int(_ic.get('bought_units', 0) or 0) if _ic else 0
        _covered = _stock_units + _bought_units
        _has_stock = bool(_ic and _ic.get('stock')) and _stock_units > 0
        _fully = bool(_ic) and _covered >= cant            # ítem completo
        _partial = bool(_ic) and 0 < _covered < cant       # falta comprar unidades
        _falta = max(0, cant - _covered)
        _real_price = float(_ic.get('real', 0) or 0) if _ic else 0
        _stock_qty_saved = min(cant, _stock_units) if _has_stock else 0
        # Comprado/bloqueado SOLO si el ítem está COMPLETO (o es adicional). Un
        # PARCIAL queda editable para poder comprar lo que falta.
        _ya_comprado = _fully or _es_adicional
        _pure_full_stock = _fully and _has_stock and _bought_units == 0
        # Un ítem COMPLETO queda en solo lectura SIEMPRE (aunque sea admin): ya se
        # cubrió toda la cantidad presupuestada (por compra, inventario o mezcla);
        # para corregir se edita/elimina el registro en el historial.
        _readonly = _ya_comprado

        if _es_sin_reg:
            bg = '#fdf2f8'
        elif _es_adicional:
            bg = '#fff3e0'
        elif _has_stock:
            bg = '#dcfce7'   # verde más intenso: fila con inventario
        elif _ya_comprado:
            bg = '#f0fdf4'
        elif ri % 2 == 0:
            bg = '#ffffff'
        else:
            bg = '#f8fafc'
        _row_extra = 'border-left:3px solid #16a34a;' if _has_stock else ''

        pu_fmt = '$' + f'{pu:,}'.replace(',', '.')
        pv = rc_prev.get(str(ri), {})
        vreal = _real_price if (_ya_comprado or _partial) else (pv.get('real', 0) or 0)
        vadic = int(_ic.get('adicional', 0)) if _ya_comprado else (pv.get('adic', 0) or 0)
        vreal_fmt = ('$' + f'{int(vreal):,}'.replace(',', '.')) if vreal else ('$0' if _pure_full_stock else '')

        _dc_attr = 'data-comprado="1"' if _ya_comprado else ""
        _da_attr = 'data-adicional="1"' if _es_adicional else ""
        _ds_attr = 'data-sin-registro="1"' if _es_sin_reg else ""
        # Fila CONTABLE para el progreso principal: ítem del presupuesto (no
        # adicional, no sin-registro, no varios). El % principal se calcula solo
        # sobre estas → coincide con el % del dropdown. Los adicionales van aparte.
        _es_varios = cat.strip().lower() == 'varios'
        _dpresup_attr = 'data-presup="1"' if (not _es_adicional and not _es_sin_reg and not _es_varios) else ""
        _dstock_attr = 'data-stock="1"' if _has_stock else ""
        _dstockqty_attr = f'data-stock-qty="{_stock_qty_saved}"' if _has_stock else ""
        # Stock ya guardado en BD (no se re-guarda): full o parcial recargados.
        _dsaved_attr = 'data-stock-saved="1"' if (_has_stock and (_fully or _partial)) else ""
        _dbuy_attr = f'data-buy="{_falta}"' if _partial else ""
        _ph_attr = 'placeholder="compra lo que falta"' if _partial else ''
        # Nota "faltan N por comprar" bajo la cantidad (se actualiza en vivo por JS
        # al marcar stock / cambiar la cantidad; sembrada aquí para el parcial recargado).
        _falta_inner = f'faltan {_falta} por comprar' if _partial else ''
        _falta_note = (f'<div class="rc-falta" style="font-size:9px;color:#ea580c;'
                       f'font-weight:600;margin-top:2px;white-space:nowrap;">{_falta_inner}</div>')
        _inp_border = "#86efac" if (_has_stock or _ya_comprado) else "#cbd5e1"
        _cat_badge = f'<span class="rc-cat" style="{_rc_badge_style(cat_colors.get(cat))}">{cat}</span>'
        # Celda "En stock": checkbox (+ stepper de cantidad) para ítems sin cobertura;
        # o badge verde con la cantidad en stock para ítems ya inventariados.
        if not _es_adicional and not _es_sin_reg and _covered == 0:
            _qty_input = (
                '<div class="rc-stockqty-wrap" style="display:none;margin-top:4px;white-space:nowrap;">'
                '<span class="rc-step">'
                '<button type="button" onclick="window.stockStep(this,-1);event.stopPropagation()" title="menos">&#8722;</button>'
                f'<span class="rc-stepval">{cant}</span>'
                '<button type="button" onclick="window.stockStep(this,1);event.stopPropagation()" title="m&#225;s">+</button>'
                f'</span><span class="rc-steptot">/ {cant}</span></div>'
            ) if cant > 1 else ''
            _stock_ck = (
                '<td style="text-align:center;white-space:nowrap"><input type="checkbox" class="rc-stock" '
                f'data-idx="{ri}" onchange="window.toggleStock(this)" '
                'title="Ya lo tengo en stock — precio real $0 (ahorro puro)"/>'
                f'{_qty_input}</td>'
            )
        elif _has_stock:
            _sq_lbl = ('' if _stock_units >= cant
                       else f'<span class="rc-steptot" style="color:#166534;">{_stock_units}/{cant}</span>')
            _stock_ck = (
                '<td style="text-align:center;white-space:nowrap" title="En stock — ahorro puro">'
                + _svg_rc("check", color="#16a34a", size=15, mr=0, valign=-3)
                + _sq_lbl + '</td>'
            )
        else:
            _stock_ck = '<td></td>'
        rows += f"""<tr style="background:{bg};{_row_extra}" data-idx="{ri}" data-pu="{pu}" data-cant="{cant}" {_dc_attr} {_da_attr} {_ds_attr} {_dstock_attr} {_dstockqty_attr} {_dsaved_attr} {_dbuy_attr} {_dpresup_attr}>
<td>{_cat_badge}</td>
<td class="rc-item">{item}</td>
<td class="r rc-cant">{cant}{_falta_note}</td>
<td class="r rc-pu">{pu_fmt}</td>
<td style="padding:4px 8px"><input type="text" inputmode="numeric" value="{vreal_fmt}" class="rc-real" data-idx="{ri}" data-val="{vreal}" {"readonly" if _readonly else ""} {_ph_attr} style="width:100%;{("border-color:%s" % _inp_border) if _inp_border!="#cbd5e1" else ""};{"background:#f0fdf4;color:#15803d;cursor:default" if _readonly else ""}"/></td>
<td style="padding:4px 8px"><input type="number" min="0" step="1" value="{vadic}" class="rc-adic" data-idx="{ri}" {"readonly" if _readonly else ""} style="width:100%;border-color:{"#86efac" if (_has_stock or _ya_comprado) else "#fca5a5"};background:{"#f0fdf4" if _readonly else "#fff7f7"};{"pointer-events:none" if _readonly else ""}"/></td>
<td class="r rc-dif rc-difv" data-idx="{ri}" style="color:#16a34a">-</td>
{_stock_ck}
</tr>"""

    html = f"""<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700&display=swap');
html,body{{margin:0;padding:0;font-family:Montserrat,'Segoe UI',sans-serif;font-size:13px;height:100%;overflow:hidden;color:#334155;-webkit-font-smoothing:antialiased}}
body{{display:flex;flex-direction:column}}
table{{width:100%;border-collapse:collapse}}
th{{background:#1e2447;color:#cbd5e1;padding:9px 12px;font-size:10px;font-weight:600;letter-spacing:.07em;text-transform:uppercase;white-space:nowrap;position:sticky;top:0;z-index:1;text-align:left}}
th.r{{text-align:right}}
tbody td{{padding:5px 12px;font-size:12.5px;vertical-align:middle;border-bottom:1px solid #f1f5f9}}
td.r{{text-align:right}}
.rc-cat{{display:inline-block;padding:2px 10px;border-radius:20px;font-size:10.5px;font-weight:600;letter-spacing:.01em;white-space:nowrap;line-height:1.55}}
.rc-item{{font-weight:600;color:#1e293b}}
.rc-cant,.rc-pu{{font-weight:500;color:#475569;font-variant-numeric:tabular-nums}}
.rc-difv{{font-weight:600;font-variant-numeric:tabular-nums;white-space:nowrap}}
/* Inputs limpios, sin flechas nativas */
input.rc-real,input.rc-adic{{border:1px solid #cbd5e1;border-radius:7px;padding:6px 9px;font-size:12.5px;text-align:right;box-sizing:border-box;font-family:inherit;color:#0f172a;-webkit-appearance:none;-moz-appearance:textfield;appearance:textfield;transition:border-color .12s,box-shadow .12s}}
input.rc-real::-webkit-inner-spin-button,input.rc-real::-webkit-outer-spin-button,input.rc-adic::-webkit-inner-spin-button,input.rc-adic::-webkit-outer-spin-button{{-webkit-appearance:none;margin:0}}
input.rc-real:focus,input.rc-adic:focus{{outline:none;border-color:#5b7cfa;box-shadow:0 0 0 3px rgba(91,124,250,.14)}}
input.rc-adic::placeholder{{color:#cbd5e1}}
/* Checkbox de stock */
.rc-stock{{width:17px;height:17px;cursor:pointer;accent-color:#16a34a;vertical-align:middle}}
/* Stepper de cantidad en stock (reemplaza las flechas nativas) */
.rc-step{{display:inline-flex;align-items:center;border:1px solid #86efac;border-radius:7px;overflow:hidden;height:23px;vertical-align:middle;background:#fff}}
.rc-step button{{border:none;background:#f0fdf4;color:#16a34a;width:20px;height:100%;font-size:14px;font-weight:700;cursor:pointer;line-height:1;padding:0;display:flex;align-items:center;justify-content:center}}
.rc-step button:hover{{background:#dcfce7}}
.rc-step button:active{{background:#bbf7d0}}
.rc-stepval{{min-width:20px;text-align:center;font-size:11.5px;font-weight:700;color:#166534;font-variant-numeric:tabular-nums}}
.rc-steptot{{font-size:10px;color:#94a3b8;font-weight:600;margin-left:4px;vertical-align:middle}}
/* Tarjetas de totales (panel inferior) — livianas y elegantes */
#tots{{background:#f8fafc;border-top:1px solid #eef2f7}}
.tot-card{{background:#fff;border-radius:12px;padding:12px 14px;box-shadow:0 1px 3px rgba(15,23,42,.05);border:1px solid #f1f5f9;display:flex;flex-direction:column;min-width:0}}
.tot-h{{font-size:10px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;margin-bottom:7px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.tot-sub{{font-size:8.5px;font-weight:600;margin-top:-5px;margin-bottom:7px;letter-spacing:.03em;text-transform:uppercase;opacity:.9}}
.tot-l{{font-size:10px;color:#94a3b8;margin-top:6px;font-weight:500}}
.tot-v{{font-size:13.5px;font-weight:700;color:#334155;font-variant-numeric:tabular-nums}}
.tot-v2{{font-size:12px;font-weight:600;color:#64748b;font-variant-numeric:tabular-nums}}
.tot-big{{font-size:16.5px;font-weight:800;color:#0f172a;font-variant-numeric:tabular-nums;margin-top:1px}}
/* Panel de guardar — claro (antes azul oscuro) */
.rc-field{{margin-bottom:9px}}
.rc-lbl{{font-size:11px;font-weight:600;color:#64748b;margin-bottom:4px;display:flex;align-items:center}}
.rc-inp{{width:100%;border:1px solid #cbd5e1;border-radius:8px;padding:8px 11px;font-size:13px;background:#fff;color:#0f172a;box-sizing:border-box;outline:none;font-family:inherit;transition:border-color .12s,box-shadow .12s}}
.rc-inp::placeholder{{color:#94a3b8}}
.rc-inp:focus{{border-color:#5b7cfa;box-shadow:0 0 0 3px rgba(91,124,250,.14)}}
.rc-sel{{width:100%;border:1px solid #cbd5e1;border-radius:8px;padding:8px 11px;font-size:13px;background:#fff;color:#0f172a;box-sizing:border-box;outline:none;cursor:pointer;font-family:inherit}}
.rc-sel:focus{{border-color:#5b7cfa;box-shadow:0 0 0 3px rgba(91,124,250,.14)}}
.rc-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}
.rc-hidden{{display:none}}
/* Fullscreen de la tabla: se ocultan cards y resumen → solo buscador + tabla. */
html.rc-fs #rc-cards{{display:none!important}}
html.rc-fs #tots{{display:none!important}}
html.rc-fs body{{margin:0!important;padding:0!important}}
</style>
<div id="rc-cards">{cats_cards_html}</div>
<div style="border:1px solid #e2e8f0;border-radius:8px;display:flex;flex-direction:column;flex:1;overflow:hidden;min-height:0;background:#fff">
  <div style="display:flex;align-items:center;gap:8px;padding:8px 10px;border-bottom:1px solid #eef2f7;flex-shrink:0">
    <input id="rc-search" type="text" placeholder="Buscar item..." oninput="window.filterRows(this.value)" style="flex:1;min-width:0;border:1px solid #cbd5e1;border-radius:7px;padding:7px 11px;font-size:13px;box-sizing:border-box"/>
    <button id="_rc_csvbtn" type="button" title="Descargar tabla como CSV" style="width:34px;height:32px;border:1px solid #e2e8f0;border-radius:7px;background:#fff;color:#475569;cursor:pointer;display:inline-flex;align-items:center;justify-content:center;flex:0 0 auto"><svg viewBox="0 0 24 24" width="16" height="16" fill="none" stroke="currentColor" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" x2="12" y1="15" y2="3"/></svg></button>
    <button id="_rc_fsbtn" type="button" title="Pantalla completa" style="width:34px;height:32px;border:1px solid #e2e8f0;border-radius:7px;background:#fff;color:#475569;cursor:pointer;display:inline-flex;align-items:center;justify-content:center;flex:0 0 auto"></button>
  </div>
  {admin_toggle_html}
  <div id="tbl-wrap" style="overflow:auto;flex:1;min-height:0"><table>
    <thead><tr>
      <th>Categor&#237;a</th><th>&#205;tem</th><th class="r">Cant.</th>
      <th class="r">Presup. unit.</th><th class="r">Real unit.</th>
      <th class="r">Adicional</th><th class="r">Diferencia</th><th style="text-align:center">En stock</th>
    </tr></thead>
    <tbody>{rows}</tbody>
  </table></div>
  <div id="tots" style="display:none;grid-template-columns:1fr 1fr 1fr 1fr 1fr 1fr 1fr;gap:10px;padding:14px 16px;flex-shrink:0">
    <div class="tot-card" style="border-left:3px solid #cbd5e1">
      <div class="tot-h" style="color:#64748b">Presupuestado</div>
      <div class="tot-l">Subtotal neto</div><div class="tot-v" id="tp-n">$0</div>
      <div class="tot-l">IVA (19%)</div><div class="tot-v2" id="tp-i">$0</div>
      <div class="tot-l">Total con IVA</div><div class="tot-big" id="tp-t">$0</div>
    </div>
    <div class="tot-card" style="border-left:3px solid #94a3b8">
      <div class="tot-h" style="color:#64748b">Real</div>
      <div class="tot-l">Subtotal neto</div><div class="tot-v" id="tr-n">$0</div>
      <div class="tot-l">IVA (19%)</div><div class="tot-v2" id="tr-i">$0</div>
      <div class="tot-l">Total con IVA</div><div class="tot-big" id="tr-t">$0</div>
    </div>
    <div class="tot-card" id="b-card" style="border-left:3px solid #16a34a">
      <div class="tot-h" id="b-hdr" style="color:#16a34a">Balance</div>
      <div class="tot-l" id="b-lbl1">Neto</div><div class="tot-v" id="b-n">$0</div>
      <div class="tot-l" id="b-lbl2">IVA</div><div class="tot-v2" id="b-i">$0</div>
      <div class="tot-l" id="b-icon">Ahorro</div><div class="tot-big" id="b-t">$0</div>
    </div>
    <div class="tot-card" style="border-left:3px solid #16a34a;background:#f0fdf4">
      <div class="tot-h" style="color:#16a34a">Inventario</div>
      <div class="tot-sub" style="color:#22c55e">Ahorro puro</div>
      <div class="tot-l" style="color:#4ade80">Productos</div><div class="tot-v" style="color:#16a34a" id="ti-c">0</div>
      <div class="tot-l" style="color:#4ade80">Total ahorrado</div><div class="tot-big" style="color:#16a34a;font-size:15.5px" id="ti-t">$0</div>
      <button onclick="window.verInventario()" id="ti-ver" style="margin-top:9px;background:#16a34a;color:#fff;border:none;border-radius:7px;padding:5px 12px;font-size:11px;font-weight:600;cursor:pointer;display:inline-flex;align-items:center;gap:5px;box-shadow:0 2px 6px rgba(22,163,74,.22);align-self:flex-start">{_svg_rc('eye', color='#fff', size=12, mr=0)}Ver</button>
    </div>
    <div class="tot-card" style="border-left:3px solid #f97316;background:#fff7ed">
      <div class="tot-h" style="color:#f97316">Adicionales</div>
      <div class="tot-sub" style="color:#fb923c">Con registro</div>
      <div class="tot-l" style="color:#fdba74">Subtotal neto</div><div class="tot-v" style="color:#ea580c" id="ta-n">$0</div>
      <div class="tot-l" style="color:#fdba74">IVA (19%)</div><div class="tot-v2" style="color:#f97316" id="ta-i">$0</div>
      <div class="tot-l" style="color:#fdba74">Total con IVA</div><div class="tot-big" style="color:#ea580c;font-size:15px" id="ta-t">$0</div>
    </div>
    <div class="tot-card" style="border-left:3px solid #ec4899;background:#fdf2f8">
      <div class="tot-h" style="color:#ec4899">Adicionales</div>
      <div class="tot-sub" style="color:#f472b6">Sin registro</div>
      <div class="tot-l" style="color:#f9a8d4">Subtotal neto</div><div class="tot-v" style="color:#db2777" id="ts-n">$0</div>
      <div class="tot-l" style="color:#f9a8d4">IVA (19%)</div><div class="tot-v2" style="color:#ec4899" id="ts-i">$0</div>
      <div class="tot-l" style="color:#f9a8d4">Total con IVA</div><div class="tot-big" style="color:#db2777;font-size:15px" id="ts-t">$0</div>
    </div>
    <div class="tot-card" style="align-items:center;justify-content:center;text-align:center;border-left:3px solid #e2e8f0">
      <div class="tot-h" style="color:#64748b;margin-bottom:10px">Progreso</div>
      <div id="prog-pct" style="font-size:38px;font-weight:800;line-height:1;color:#dc2626;font-variant-numeric:tabular-nums">0%</div>
      <div id="prog-lbl" style="font-size:11px;font-weight:600;color:#dc2626;margin-top:6px">Sin compras</div>
      <div style="width:100%;background:#f1f5f9;border-radius:99px;height:7px;margin-top:11px;overflow:hidden">
        <div id="prog-bar" style="height:100%;width:0%;border-radius:99px;background:#dc2626;transition:width .4s ease,background .4s ease"></div>
      </div>
      <div id="prog-adic" style="margin-top:7px;display:flex;flex-direction:column;align-items:center"></div>
    </div>
  </div>
  <div id="rc-add-backdrop" style="display:none;position:fixed;inset:0;background:rgba(15,23,42,.42);z-index:99998" onclick="window.closeAddPopup()"></div>
  <div id="add-section" style="display:none">
    <div style="display:flex;align-items:center;margin-bottom:12px">
      <div style="font-family:Montserrat,'Segoe UI',sans-serif;font-weight:700;font-size:.8rem;letter-spacing:.04em;text-transform:uppercase;color:#0f172a">Agregar producto</div>
      <button onclick="window.closeAddPopup()" style="margin-left:auto;background:#f1f5f9;border:none;border-radius:8px;width:30px;height:30px;cursor:pointer;color:#64748b;font-size:15px;line-height:1">&#10005;</button>
    </div>
    <div style="display:flex;gap:8px;margin-bottom:8px;">
      <button onclick="window.switchAddTab('reg')" id="tab-reg" style="font-size:11px;font-weight:700;padding:4px 12px;border-radius:6px;border:1.5px solid #f97316;background:#fff7ed;color:#f97316;cursor:pointer"><span style="color:#f97316;font-size:24px;line-height:1;vertical-align:middle;">&#9679;</span> Con registro</button>
      <button onclick="window.switchAddTab('sin')" id="tab-sin" style="font-size:11px;font-weight:700;padding:4px 12px;border-radius:6px;border:1.5px solid #e2e8f0;background:#fff;color:#94a3b8;cursor:pointer"><span style="color:#ec4899;font-size:24px;line-height:1;vertical-align:middle;">&#9679;</span> Sin registro</button>
    </div>
    <div id="add-con-reg" style="display:grid;grid-template-columns:1.5fr 3fr 0.8fr 1.2fr auto;gap:6px;align-items:end">
      <div><div style="font-size:10px;color:#64748b;margin-bottom:3px">Categor&#237;a</div>
        <select id="add-cat" style="width:100%;border:1px solid #cbd5e1;border-radius:6px;padding:5px;font-size:12px"><option value="">Seleccionar...</option></select></div>
      <div><div style="font-size:10px;color:#64748b;margin-bottom:3px">&#205;tem</div>
        <select id="add-item" style="width:100%;border:1px solid #cbd5e1;border-radius:6px;padding:5px;font-size:12px"><option value="">Seleccionar categor&#237;a primero</option></select></div>
      <div><div style="font-size:10px;color:#64748b;margin-bottom:3px">Cant.</div>
        <input id="add-cant" type="number" min="1" value="1" style="width:100%;border:1px solid #cbd5e1;border-radius:6px;padding:5px;font-size:12px;text-align:right"/></div>
      <div><div style="font-size:10px;color:#64748b;margin-bottom:3px">Presup. unit.</div>
        <div id="add-precio" style="border:1px solid #e2e8f0;border-radius:6px;padding:5px 8px;font-size:12px;font-weight:600;background:#f8fafc;text-align:right">$0</div></div>
      <div style="padding-bottom:1px">
        <button onclick="window.addRow()" style="background:#f97316;color:#fff;border:none;border-radius:6px;padding:6px 16px;font-size:12px;font-weight:700;cursor:pointer">+ Agregar</button></div>
    </div>
    <div id="add-sin-reg" style="display:none;grid-template-columns:1.5fr 3fr 0.8fr 1.2fr auto;gap:6px;align-items:end">
      <div style="padding-right:2px"><div style="font-size:10px;color:#ec4899;margin-bottom:3px">Categor&#237;a *</div>
        <input id="sin-cat" type="text" placeholder="Ej: Herramientas" style="width:100%;border:1px solid #fbcfe8;border-radius:6px;padding:5px;font-size:12px;background:#fdf2f8;box-sizing:border-box"/></div>
      <div style="padding-right:2px"><div style="font-size:10px;color:#ec4899;margin-bottom:3px">Nombre del &#237;tem *</div>
        <input id="sin-item" type="text" placeholder="Ej: Taladro percutor" style="width:100%;border:1px solid #fbcfe8;border-radius:6px;padding:5px;font-size:12px;background:#fdf2f8;box-sizing:border-box"/></div>
      <div style="padding-right:2px"><div style="font-size:10px;color:#ec4899;margin-bottom:3px">Cant.</div>
        <input id="sin-cant" type="number" min="1" value="1" style="width:100%;border:1px solid #fbcfe8;border-radius:6px;padding:5px;font-size:12px;text-align:right;background:#fdf2f8;box-sizing:border-box"/></div>
      <div style="padding-right:2px"><div style="font-size:10px;color:#ec4899;margin-bottom:3px">Precio real *</div>
        <input id="sin-precio" type="text" inputmode="numeric" placeholder="$0" style="width:100%;border:1px solid #fbcfe8;border-radius:6px;padding:5px;font-size:12px;text-align:right;background:#fdf2f8;box-sizing:border-box"/></div>
      <div style="padding-bottom:1px">
        <button onclick="window.addRowSinReg()" style="background:#ec4899;color:#fff;border:none;border-radius:6px;padding:6px 16px;font-size:12px;font-weight:700;cursor:pointer">+ Agregar</button></div>
    </div>
  </div>
  <div id="save-section" style="display:none">
    <div style="font-size:11px;font-weight:700;color:#334155;letter-spacing:.05em;text-transform:uppercase;margin-bottom:11px;display:flex;align-items:center">{_svg_rc('paperclip', color='#64748b', size=14, mr=8)}Adjuntar Factura y Guardar<button onclick="window.closeAddPopup()" style="margin-left:auto;background:#eef2f7;border:none;border-radius:8px;width:28px;height:28px;cursor:pointer;color:#64748b;font-size:14px;line-height:1">&#10005;</button></div>
    <div class="rc-grid">
      <div class="rc-field">
        <div class="rc-lbl">{_svg_rc('store', color='#94a3b8', size=13)}&#191;D&#243;nde compraste? *</div>
        <input id="lugar-compra" type="text" class="rc-inp" list="rc-proveedores" autocomplete="off" placeholder="Escribe un proveedor o elige INVENTARIO..." oninput="window.onLugarChange()"/>{_prov_datalist}
      </div>
      <div class="rc-field">
        <div class="rc-lbl">{_svg_rc('cart', color='#94a3b8', size=13)}Tipo de compra *</div>
        <select id="tipo-compra" class="rc-sel" onchange="window.onTipoChange()">
          <option value="">Seleccionar...</option>
          <option value="online">Compra Online</option>
          <option value="presencial">Compra Presencial</option>
          <option value="stock">Inventario / Stock propio</option>
        </select>
      </div>
    </div>
    <div id="subtipo-wrap" class="rc-field rc-hidden">
      <div class="rc-lbl" id="subtipo-lbl">Modalidad *</div>
      <select id="subtipo-compra" class="rc-sel" onchange="window.onSubtipoChange()">
      </select>
    </div>
    <div id="fecha-wrap" class="rc-field rc-hidden">
      <div class="rc-lbl" id="fecha-lbl">{_svg_rc('calendar', color='#94a3b8', size=13)}&#191;Para cu&#225;ndo? *</div>
      <input id="fecha-compra" type="date" class="rc-inp" oninput="window.checkSaveBtn()" onchange="window.checkSaveBtn()"/>
    </div>
    <div id="falt&#243;-wrap" class="rc-field rc-hidden">
      <div class="rc-lbl">{_svg_rc('clipboard', color='#94a3b8', size=13)}&#191;Qu&#233; falt&#243; por retirar? *</div>
      <textarea id="falto-texto" class="rc-inp" rows="2" placeholder="Describe los &#237;tems que faltaron..." oninput="window.checkSaveBtn()" style="resize:vertical"></textarea>
    </div>
    <div class="rc-field">
      <div class="rc-lbl">{_svg_rc('note', color='#94a3b8', size=13)}Observaciones adicionales (opcional)</div>
      <textarea id="obs-compra" class="rc-inp" rows="2" placeholder="Notas, motivos u observaciones de esta compra..." style="resize:vertical"></textarea>
    </div>
    <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;margin-top:2px">
      <span id="factura-wrap" style="display:inline-flex;align-items:center;gap:10px;">
        <label id="factura-label" style="background:#fff;color:#475569;border:1.5px dashed #cbd5e1;border-radius:8px;padding:7px 14px;font-size:12px;font-weight:500;cursor:pointer;white-space:nowrap;display:inline-flex;align-items:center;transition:border-color .12s">{_svg_rc('paperclip', color='#64748b', size=13)}Seleccionar factura PDF
          <input id="factura-input" type="file" accept=".pdf" style="display:none"/>
        </label>
        <button id="factura-clear" onclick="window.clearFactura()" style="display:none;background:#fef2f2;color:#dc2626;border:1px solid #fecaca;border-radius:8px;padding:6px 11px;font-size:12px;font-weight:500;cursor:pointer;white-space:nowrap;align-items:center">{_svg_rc('x', color='#dc2626', size=12, mr=4)}Quitar</button>
      </span>
      <span id="inv-note" style="display:none;align-items:center;gap:6px;background:#f0fdf4;color:#166534;border:1px solid #bbf7d0;border-radius:8px;padding:7px 12px;font-size:12px;font-weight:600;white-space:nowrap">{_svg_rc('check', color='#16a34a', size=13)}Inventario — no requiere factura (ahorro puro)</span>
      <div id="save-status" style="font-size:12px;color:#64748b;flex:1"></div>
      <button id="save-btn" onclick="window.guardarRegistro()" disabled style="background:#16a34a;color:#fff;border:none;border-radius:9px;padding:10px 26px;font-size:13px;font-weight:600;cursor:pointer;opacity:0.5;white-space:nowrap;display:inline-flex;align-items:center;justify-content:center;box-shadow:0 3px 10px rgba(22,163,74,.22)">{_svg_rc('save', color='#fff', size=14, mr=7)}Guardar compra</button>
    </div>
  </div>
    </div>
  </div>
</div>
<div id="inv-modal" style="display:none;position:fixed;inset:0;background:rgba(15,23,42,0.55);z-index:9999;align-items:center;justify-content:center;padding:20px;box-sizing:border-box" onclick="if(event.target===this)window.cerrarInventario()">
  <div style="background:#fff;border-radius:14px;max-width:560px;width:100%;max-height:82vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,0.35);overflow:hidden">
    <div style="display:flex;align-items:center;justify-content:space-between;gap:10px;padding:14px 18px;background:#f0fdf4;border-bottom:1px solid #bbf7d0">
      <div style="display:flex;align-items:center;font-family:Montserrat,'Segoe UI',sans-serif;font-weight:700;font-size:.9rem;letter-spacing:.03em;text-transform:uppercase;color:#166534">{_svg_rc('package', color='#16a34a', size=16, mr=8)}Productos en inventario · Ahorro puro</div>
      <button onclick="window.cerrarInventario()" style="background:none;border:none;color:#64748b;font-size:22px;line-height:1;cursor:pointer;padding:0 4px">&times;</button>
    </div>
    <div id="inv-modal-body" style="overflow:auto;padding:0;font-family:Montserrat,'Segoe UI',sans-serif"></div>
    <div id="inv-modal-foot" style="padding:12px 18px;background:#f8fafc;border-top:1px solid #e2e8f0;display:flex;justify-content:space-between;align-items:center;font-family:Montserrat,'Segoe UI',sans-serif;font-weight:800;font-size:1rem"></div>
  </div>
</div>
<script>(function(){{
var SUPA_URL="{supa_url}";
var SUPA_KEY="{supa_key}";
var EP_NUM="{ep}";
var USUARIO="{usuario}";
var ITEMS_YA_COMPRADOS={items_ya_comprados_json};
var TOTAL_ITEMS={total_items_presupuesto};
var _facturaFile=null;
var _facturaUrl="";
var _facturaNom="";
var ICO_CAL='{_svg_rc('calendar', color='#94a3b8', size=13)}';
var ICO_CLIP='{_svg_rc('paperclip', color='#64748b', size=13)}';
var CAT={rc_cat_json};
var addCat=document.getElementById("add-cat");
Object.keys(CAT).sort().forEach(function(c){{var o=document.createElement("option");o.value=c;o.textContent=c;addCat.appendChild(o);}});
addCat.addEventListener("change",function(){{
  var sel=document.getElementById("add-item");
  sel.innerHTML='<option value="">Seleccionar item...</option>';
  document.getElementById("add-precio").textContent="$0";
  var items=CAT[this.value]||[];
  items.forEach(function(it){{var o=document.createElement("option");o.value=JSON.stringify(it);o.textContent=it.item;sel.appendChild(o);}});
}});
document.getElementById("add-item").addEventListener("change",function(){{
  try{{var it=JSON.parse(this.value);document.getElementById("add-precio").textContent=f(it.precio);}}catch(e){{}}
}});
var _addIdx=10000;
window.addRow=function(){{
  var catEl=document.getElementById("add-cat");
  var itemEl=document.getElementById("add-item");
  var cantEl=document.getElementById("add-cant");
  if(!catEl.value||!itemEl.value)return;
  var it=JSON.parse(itemEl.value);
  var cant=parseInt(cantEl.value)||1;
  var pu=it.precio;
  var tr=document.createElement("tr");
  tr.style.cssText="background:#fff3e0;border-bottom:1px solid #eef0f6;border-left:3px solid #f97316";
  tr.dataset.idx=String(_addIdx);tr.dataset.pu=String(pu);tr.dataset.cant=String(cant);
  tr.dataset.adicional="1";
  tr.innerHTML="<td style='padding:5px 8px;font-size:.75rem;color:#64748b'>"+catEl.value+"</td>"
    +"<td style='padding:5px 8px;font-size:.82rem'>"+it.item+"</td>"
    +"<td style='padding:5px 8px;text-align:right'>"+cant+"</td>"
    +"<td style='padding:5px 8px;text-align:right;font-weight:600'>"+f(pu)+"</td>"
    +'<td style="padding:3px 4px"><input type="text" inputmode="numeric" value="" class="rc-real" data-idx="'+_addIdx+'" data-val="0" style="width:100%;border:1px solid #cbd5e1;border-radius:6px;padding:5px;font-size:13px;text-align:right;box-sizing:border-box"/></td>'
    +'<td style="padding:3px 4px"><input type="number" min="0" step="1" value="0" class="rc-adic" data-idx="'+_addIdx+'" style="width:100%;border:1px solid #fca5a5;border-radius:6px;padding:5px;font-size:13px;text-align:right;background:#fff5f5;box-sizing:border-box"/></td>'
    +'<td class="rc-dif" style="padding:5px 8px;text-align:right;font-weight:700;color:#16a34a;white-space:nowrap">-</td>'
    +'<td style="padding:3px 6px;text-align:center"><button onclick="window.removeRow(this)" style="background:none;border:none;color:#ef4444;font-size:14px;cursor:pointer;padding:2px 4px;line-height:1;" title="Eliminar">&#10005;</button></td>';
  document.querySelector("tbody").appendChild(tr);
  attachListeners(tr.querySelector(".rc-real"), tr.querySelector(".rc-adic"));
  _addIdx++;cantEl.value="1";itemEl.selectedIndex=0;
  document.getElementById("add-precio").textContent="$0";calc();
}};
function attachListeners(inp, adic){{
  inp.addEventListener("input",function(){{
    var raw=this.value.replace(/[^0-9]/g,"");
    this.dataset.val=raw||"0";
    if(!raw){{this.value="";calc();return;}}
    this.value="$"+parseInt(raw).toLocaleString("de-DE");calc();checkSaveBtn();
  }});
  inp.addEventListener("focus",function(){{
    var r=this.dataset.val||"0";
    this.value=r==="0"?"":r;
  }});
  inp.addEventListener("blur",function(){{
    var n=parseInt(this.dataset.val)||0;
    this.dataset.val=String(n);
    this.value=n>0?"$"+n.toLocaleString("de-DE"):"";calc();checkSaveBtn();
  }});
  adic.addEventListener("input",function(){{calc();checkSaveBtn();}});
}}
function f(n){{return "$"+Math.round(Math.abs(n)).toLocaleString("de-DE");}}
function calc(){{
  var tP=0,tR=0,tA=0,tS=0,tI=0,nI=0,vals=[];
  var pTot=0,pComp=0,nAc=0,nSc=0;   // progreso: presupuesto vs adicionales
  document.querySelectorAll("tr[data-idx]").forEach(function(r){{
    var idx=parseInt(r.dataset.idx)||0;
    var pu=+r.dataset.pu||0,c=+r.dataset.cant||1;
    var re=parseFloat(r.querySelector(".rc-real").dataset.val)||0;
    var ad=+r.querySelector(".rc-adic").value||0;
    var isSinReg=r.getAttribute("data-sin-registro")==="1";
    var isAdic=r.dataset.adicional==="1"&&!isSinReg;
    var isStock=r.getAttribute("data-stock")==="1";
    // Stock: sq = unidades en stock (ahorro). Unidades a comprar: data-buy si
    // viene (parcial recargado = lo que falta), si no, c - sq (fila fresca).
    var sq=isStock?(parseInt(r.getAttribute("data-stock-qty"))||c):0;
    if(sq>c)sq=c; if(sq<0)sq=0;
    var buyAttr=r.getAttribute("data-buy");
    var comprUnits=(buyAttr!=null)?(parseInt(buyAttr)||0):(isStock?(c-sq):c);
    if(comprUnits<0)comprUnits=0;
    // Diferencia (ahorro). Stock: ahorro puro de lo inventariado + diferencia de
    // precio de lo comprado (solo si hay precio; lo que falta comprar no es ahorro).
    var d=isStock?(pu*sq+(re>0?(pu-re)*comprUnits:0)):((pu-re)*c-(ad*re));
    var td=r.querySelector(".rc-dif");
    td.textContent=f(d)+(d>=0?" ▼":" ▲");
    td.style.color=d>=0?"#16a34a":"#dc2626";
    // Nota "faltan N por comprar" (unidades en stock que aún no se han comprado).
    var faltan=(isStock&&re<=0)?comprUnits:0;
    var fEl=r.querySelector(".rc-falta");
    if(fEl)fEl.textContent=faltan>0?("faltan "+faltan+" por comprar"):"";
    if(isSinReg){{tS+=re*c;}}
    else if(isAdic){{tA+=re*c;}}
    else{{tP+=pu*c;}}
    tR+=re*comprUnits+ad*re;
    // Inventario (stock, $0): ahorro puro = presupuestado de las unidades en stock.
    if(isStock&&idx<10000){{tI+=pu*sq;nI++;}}
    // Progreso PRINCIPAL: SOLO ítems del presupuesto (data-presup) → coincide con
    // el % del dropdown. Los adicionales (con/sin registro) se cuentan aparte.
    if(r.getAttribute("data-presup")==="1"){{pTot++;if(re>0||isStock)pComp++;}}
    else if(isSinReg&&re>0){{nSc++;}}
    else if(isAdic&&re>0){{nAc++;}}
    // Unidades cubiertas (ya en stock/compradas + lo que se compra ahora). La
    // línea está COMPLETA solo si se cubre toda la cantidad (parcial ≠ completo).
    var coveredUnits=(c-comprUnits)+(re>0?comprUnits:0);
    vals.push({{idx:+r.dataset.idx,real:re,adic:ad,dif:d,stock:isStock,complete:coveredUnits>=c}});
  }});
  var _tiT=document.getElementById("ti-t");if(_tiT)_tiT.textContent=f(tI);
  var _tiC=document.getElementById("ti-c");if(_tiC)_tiC.textContent=nI;
  var iP=tP*.19,iR=tR*.19,b=tP-tR,ib=iP-iR;
  var col=b>=0?"#16a34a":"#dc2626";
  var iA=tA*.19,iS=tS*.19;
  var ids=["tp-n","tp-i","tp-t","tr-n","tr-i","tr-t","b-n","b-i","b-t","ta-n","ta-i","ta-t","ts-n","ts-i","ts-t"];
  var v=[tP,iP,tP+iP,tR,iR,tR+iR,b,ib,b+ib,tA,iA,tA+iA,tS,iS,tS+iS];
  ids.forEach(function(id,i){{var el=document.getElementById(id);if(el)el.textContent=f(v[i]);}});
  ["b-hdr","b-n","b-i","b-lbl1","b-lbl2","b-icon","b-t"].forEach(function(id){{var el=document.getElementById(id);if(el)el.style.color=col;}});
  var _bc=document.getElementById("b-card");if(_bc)_bc.style.borderLeftColor=col;
  var bi=document.getElementById("b-icon");
  if(bi)bi.textContent=b>=0?"Ahorro":"Sobrecosto";
  // % principal = SOLO presupuesto (mismo criterio que el dropdown).
  var pct=pTot>0?Math.round(pComp/pTot*1000)/10:0;
  var pctCol=pct>=100?"#3b82f6":pct>=66.6?"#16a34a":pct>=33.3?"#eab308":"#dc2626";
  var pctLbl=pct>=100?"Compra finalizada":pct>0?(pct.toFixed(1)+"% comprado"):"Sin compras";
  var pp=document.getElementById("prog-pct");
  var pl=document.getElementById("prog-lbl");
  var pb=document.getElementById("prog-bar");
  if(pp){{pp.textContent=pct>=100?"100%":(pct.toFixed(1)+"%");pp.style.color=pctCol;}}
  if(pl){{pl.textContent=pctLbl;pl.style.color=pctCol;}}
  if(pb){{pb.style.width=Math.min(pct,100)+"%";pb.style.background=pctCol;}}
  // Adicionales (NO están en el presupuesto) → se muestran aparte como % extra
  // sobre el total presupuestado: naranjo con registro, rosado sin registro.
  var pa=document.getElementById("prog-adic");
  if(pa){{
    var _ah="";
    if(nAc>0)_ah+='<div style="font-size:9.5px;font-weight:700;color:#ea580c;white-space:nowrap;margin-top:2px;">+'+(pTot>0?(Math.round(nAc/pTot*1000)/10):0).toFixed(1)+'% con registro</div>';
    if(nSc>0)_ah+='<div style="font-size:9.5px;font-weight:700;color:#db2777;white-space:nowrap;margin-top:1px;">+'+(pTot>0?(Math.round(nSc/pTot*1000)/10):0).toFixed(1)+'% sin registro</div>';
    pa.innerHTML=_ah;
  }}
  window.parent.postMessage({{type:"rc_vals",vals:vals,tP:tP,tR:tR}},"*");
}}
var _rcCatFiltro='';
window.applyFilters=function(catOverride){{
  if(catOverride!==undefined) _rcCatFiltro=catOverride;
  var q=document.getElementById('rc-search')?document.getElementById('rc-search').value.toLowerCase():'';
  document.querySelectorAll("tbody tr").forEach(function(r){{
    var txt=r.cells[1]?r.cells[1].textContent.toLowerCase():"";
    var cat=r.cells[0]?r.cells[0].textContent.trim():"";
    var matchTxt=!q||txt.indexOf(q)>-1;
    var matchCat=!_rcCatFiltro||cat===_rcCatFiltro;
    r.style.display=(matchTxt&&matchCat)?"":"none";
  }});
}}
window.filterRows=function(q){{applyFilters();}};
function _rcRgba(h,a){{h=h.replace('#','');return 'rgba('+parseInt(h.substr(0,2),16)+','+parseInt(h.substr(2,2),16)+','+parseInt(h.substr(4,2),16)+','+a+')';}}
window.rcFilterCat=function(el){{
  var cat=el.getAttribute('data-cat')||'';
  _rcCatFiltro=(_rcCatFiltro===cat)?'':cat;
  document.querySelectorAll('.rc-cat-card').forEach(function(c){{
    var cc=c.getAttribute('data-color')||'#6366f1';
    var nm=c.getAttribute('data-name')||'';
    var active=(_rcCatFiltro!==''&&c.getAttribute('data-cat')===_rcCatFiltro);
    c.style.background=active?_rcRgba(cc,0.15):'#fff';
    c.style.border=active?('2px solid '+cc):('1.5px solid '+_rcRgba(cc,0.3));
    c.style.borderLeft='4px solid '+cc;
    var cn=c.querySelector('.rc-cname');
    if(cn)cn.textContent=nm+(active?' \\u2713':'');
  }});
  applyFilters(_rcCatFiltro);
}};
document.querySelectorAll(".rc-real").forEach(function(inp){{
  attachListeners(inp, inp.closest("tr").querySelector(".rc-adic"));
}});
window.addEventListener("load",function(){{calc();}});
var sinPrecioEl=document.getElementById("sin-precio");
if(sinPrecioEl){{
  sinPrecioEl.addEventListener("input",function(){{
    var raw=this.value.replace(/[^0-9]/g,"");
    this.value=raw?"$"+parseInt(raw).toLocaleString("de-DE"):"";
  }});
  sinPrecioEl.addEventListener("blur",function(){{
    var raw=this.value.replace(/[^0-9]/g,"");
    this.value=raw?"$"+parseInt(raw).toLocaleString("de-DE"):"";
  }});
}}
window.switchAddTab=function(tab){{
  var reg=document.getElementById("add-con-reg");
  var sin=document.getElementById("add-sin-reg");
  var tbReg=document.getElementById("tab-reg");
  var tbSin=document.getElementById("tab-sin");
  if(tab==="reg"){{
    reg.style.display="grid";sin.style.display="none";
    tbReg.style.cssText="font-size:11px;font-weight:700;padding:4px 12px;border-radius:6px;border:1.5px solid #f97316;background:#fff7ed;color:#f97316;cursor:pointer";
    tbSin.style.cssText="font-size:11px;font-weight:700;padding:4px 12px;border-radius:6px;border:1.5px solid #e2e8f0;background:#fff;color:#94a3b8;cursor:pointer";
  }}else{{
    reg.style.display="none";sin.style.display="grid";
    tbSin.style.cssText="font-size:11px;font-weight:700;padding:4px 12px;border-radius:6px;border:1.5px solid #ec4899;background:#fdf2f8;color:#ec4899;cursor:pointer";
    tbReg.style.cssText="font-size:11px;font-weight:700;padding:4px 12px;border-radius:6px;border:1.5px solid #e2e8f0;background:#fff;color:#94a3b8;cursor:pointer";
  }}
}};
window.addRowSinReg=function(){{
  var catEl=document.getElementById("sin-cat");
  var itemEl=document.getElementById("sin-item");
  var cantEl=document.getElementById("sin-cant");
  var precioEl=document.getElementById("sin-precio");
  if(!catEl.value.trim()||!itemEl.value.trim()){{alert("Ingresa categoría y nombre del ítem");return;}}
  var rawPrecio=precioEl.value.replace(/[^0-9]/g,"");
  var precio=parseInt(rawPrecio)||0;
  var cant=parseInt(cantEl.value)||1;
  var tr=document.createElement("tr");
  tr.style.cssText="background:#fdf2f8;border-bottom:1px solid #eef0f6;border-left:3px solid #ec4899";
  tr.dataset.idx=String(_addIdx);tr.dataset.pu="0";tr.dataset.cant=String(cant);
  tr.dataset.adicional="1";tr.setAttribute("data-sin-registro","1");
  tr.innerHTML="<td style='padding:5px 8px;font-size:.75rem;color:#ec4899'>"+catEl.value.trim()+"</td>"
    +"<td style='padding:5px 8px;font-size:.82rem'>"+itemEl.value.trim()+"</td>"
    +"<td style='padding:5px 8px;text-align:right'>"+cant+"</td>"
    +"<td style='padding:5px 8px;text-align:right;font-weight:600'>—</td>"
    +'<td style="padding:3px 4px"><input type="text" inputmode="numeric" value="'+("$"+precio.toLocaleString("de-DE"))+'" class="rc-real" data-idx="'+_addIdx+'" data-val="'+precio+'" style="width:100%;border:1.5px solid #fbcfe8;border-radius:6px;padding:5px;font-size:13px;text-align:right;background:#fdf2f8;box-sizing:border-box"/></td>'
    +'<td style="padding:3px 4px"><input type="number" min="0" step="1" value="0" class="rc-adic" data-idx="'+_addIdx+'" style="width:100%;border:1.5px solid #fbcfe8;border-radius:6px;padding:5px;font-size:13px;text-align:right;background:#fdf2f8;box-sizing:border-box"/></td>'
    +'<td class="rc-dif" style="padding:5px 8px;text-align:right;font-weight:700;color:#ec4899;white-space:nowrap">—</td>'
    +'<td style="padding:3px 6px;text-align:center"><button onclick="window.removeRow(this)" style="background:none;border:none;color:#ef4444;font-size:14px;cursor:pointer;padding:2px 4px;line-height:1;" title="Eliminar">&#10005;</button></td>';
  document.querySelector("tbody").appendChild(tr);
  attachListeners(tr.querySelector(".rc-real"),tr.querySelector(".rc-adic"));
  _addIdx++;
  catEl.value="";itemEl.value="";cantEl.value="1";precioEl.value="";
  calc();checkSaveBtn();
}};
// Escribir/elegir INVENTARIO en "¿Dónde compraste?" cambia el tipo a Stock.
window.onLugarChange=function(){{
  var lu=document.getElementById("lugar-compra");
  var tipoSel=document.getElementById("tipo-compra");
  if((lu.value||"").trim().toUpperCase()==="INVENTARIO"){{
    if(tipoSel.value!=="stock"){{tipoSel.value="stock";window.onTipoChange();return;}}
  }} else if(tipoSel.value==="stock"){{
    // Cambió el lugar a un proveedor real → salir del modo inventario.
    tipoSel.value="";window.onTipoChange();return;
  }}
  window.checkSaveBtn();
}};
window.onTipoChange=function(){{
  var tipo=document.getElementById("tipo-compra").value;
  var sw=document.getElementById("subtipo-wrap");
  var ss=document.getElementById("subtipo-compra");
  var facW=document.getElementById("factura-wrap");
  var invNote=document.getElementById("inv-note");
  // Modo INVENTARIO / Stock propio: sin factura, sin fecha, sin subtipo/falta.
  if(tipo==="stock"){{
    sw.className="rc-field rc-hidden";
    var lu=document.getElementById("lugar-compra");
    if((lu.value||"").trim().toUpperCase()!=="INVENTARIO")lu.value="INVENTARIO";
    document.getElementById("fecha-wrap").className="rc-field rc-hidden";
    document.getElementById("faltó-wrap").className="rc-field rc-hidden";
    if(facW)facW.style.display="none";
    if(invNote)invNote.style.display="inline-flex";
    window.checkSaveBtn();
    return;
  }}
  if(facW)facW.style.display="";
  if(invNote)invNote.style.display="none";
  sw.className="rc-field"+(tipo?"" : " rc-hidden");
  ss.innerHTML="";
  if(tipo==="online"){{
    [["retiro","Retiro"],["despacho","Despacho"]].forEach(function(o){{
      var opt=document.createElement("option");opt.value=o[0];opt.textContent=o[1];ss.appendChild(opt);
    }});
  }}else if(tipo==="presencial"){{
    [["completo","Retiro Completo"],["parcial","Retiro Parcial"],["despacho","Despacho"]].forEach(function(o){{
      var opt=document.createElement("option");opt.value=o[0];opt.textContent=o[1];ss.appendChild(opt);
    }});
  }}
  window.onSubtipoChange();
}};
window.onSubtipoChange=function(){{
  var tipo=document.getElementById("tipo-compra").value;
  var sub=document.getElementById("subtipo-compra").value;
  var fw=document.getElementById("fecha-wrap");
  var fl=document.getElementById("fecha-lbl");
  var pw=document.getElementById("faltó-wrap");
  var needFecha=(tipo==="online")||(tipo==="presencial"&&sub!=="completo");
  fw.className="rc-field"+(needFecha?"":" rc-hidden");
  if(tipo==="presencial"&&sub==="parcial"){{
    fl.innerHTML=ICO_CAL+"¿Para cuándo llega lo que faltó? *";
    pw.className="rc-field";
  }}else{{
    fl.innerHTML=ICO_CAL+"¿Para cuándo? *";
    pw.className="rc-field rc-hidden";
  }}
  window.checkSaveBtn();
}};
window.removeRow=function(btn){{
  var tr=btn.parentElement;
  while(tr&&tr.tagName!=="TR")tr=tr.parentElement;
  if(tr)tr.remove();
  calc();checkSaveBtn();
}};
function checkSaveBtn(){{
  var hasVals=false,hasReal=false;
  document.querySelectorAll("tr[data-idx]").forEach(function(r){{
    if(r.dataset.comprado==="1") return; // ya completo → no se re-guarda
    var re=parseFloat(r.querySelector(".rc-real").dataset.val)||0;
    var c=+r.dataset.cant||1;
    var isStock=r.getAttribute("data-stock")==="1";
    var stockSaved=r.getAttribute("data-stock-saved")==="1"; // stock ya en BD
    var sq=isStock?(parseInt(r.getAttribute("data-stock-qty"))||c):0;
    var buyAttr=r.getAttribute("data-buy");
    var compr=(buyAttr!=null)?(parseInt(buyAttr)||0):(isStock?(c-sq):c);
    // Hay algo NUEVO que guardar: stock fresco (aún no en BD) o una compra real.
    var newStock=isStock&&!stockSaved&&sq>0;
    if(newStock||(re>0&&compr>0)) hasVals=true;
    // Compra real (exige factura): unidades compradas con precio > 0.
    if(re>0&&compr>0) hasReal=true;
  }});
  var lugar=document.getElementById("lugar-compra");
  var hasLugar=lugar&&lugar.value.trim().length>0;
  var hasFactura=_facturaFile!==null;
  var tipo=document.getElementById("tipo-compra");
  var hasTipo=tipo&&tipo.value.length>0;
  var fw=document.getElementById("fecha-wrap");
  var fechaOk=!fw||fw.className.indexOf("rc-hidden")>-1||(document.getElementById("fecha-compra")&&document.getElementById("fecha-compra").value.length>0);
  var pw=document.getElementById("faltó-wrap");
  var faltóOk=!pw||pw.className.indexOf("rc-hidden")>-1||(document.getElementById("falto-texto")&&document.getElementById("falto-texto").value.trim().length>0);
  // La factura solo se exige si hay compra real (hasReal). En modo INVENTARIO
  // (tipo=stock) hasReal es false → no pide factura; fecha/falta van ocultas
  // (fechaOk/faltóOk = true). Un ítem parcial ya salió del modo stock, así que
  // exige factura como cualquier compra.
  var ok=hasVals&&hasLugar&&hasTipo&&fechaOk&&faltóOk&&(hasFactura||!hasReal);
  var btn=document.getElementById("save-btn");
  if(btn){{btn.disabled=!ok;btn.style.opacity=ok?"1":"0.5";}}
}}
window.checkSaveBtn=checkSaveBtn;
// "En stock": el usuario ya tiene el producto (no lo compra). Fija precio real $0,
// lo deja readonly/verde y marca la fila data-stock="1" → cuenta como comprado
// (progreso) y es ahorro puro. Desmarcar revierte a input normal editable.
// Aplica el estado visual de la fila según data-stock / data-stock-qty:
//  · full stock (compradas=0)  → real $0 bloqueado, verde.
//  · stock parcial (compradas>0)→ real editable (precio de lo comprado).
//  · sin stock                  → input normal editable.
function _applyStockRow(tr){{
  var inp=tr.querySelector(".rc-real");
  var adic=tr.querySelector(".rc-adic");
  var isStock=tr.getAttribute("data-stock")==="1";
  var c=+tr.dataset.cant||1;
  var sq=isStock?(parseInt(tr.getAttribute("data-stock-qty"))||c):0;
  var compr=isStock?(c-sq):c;
  if(isStock&&compr<=0){{
    if(inp){{inp.dataset.val="0";inp.value="$0";inp.readOnly=true;inp.style.background="#f0fdf4";inp.style.color="#15803d";inp.style.borderColor="#86efac";inp.style.cursor="default";}}
    if(adic){{adic.value="0";adic.readOnly=true;adic.style.background="#f0fdf4";adic.style.pointerEvents="none";}}
  }} else if(isStock&&compr>0){{
    if(inp){{inp.readOnly=false;inp.style.background="";inp.style.color="";inp.style.borderColor="#86efac";inp.style.cursor="";}}
    if(adic){{adic.readOnly=false;adic.style.background="#fff5f5";adic.style.pointerEvents="";}}
  }} else {{
    if(inp){{inp.readOnly=false;inp.style.background="";inp.style.color="";inp.style.borderColor="#cbd5e1";inp.style.cursor="";}}
    if(adic){{adic.readOnly=false;adic.style.background="#fff5f5";adic.style.pointerEvents="";}}
  }}
}}
window.toggleStock=function(cb){{
  var tr=cb.closest("tr[data-idx]");
  if(!tr) return;
  var inp=tr.querySelector(".rc-real");
  var c=+tr.dataset.cant||1;
  var wrap=tr.querySelector(".rc-stockqty-wrap");
  var sv=tr.querySelector(".rc-stepval");
  if(cb.checked){{
    tr.setAttribute("data-stock","1");
    tr.setAttribute("data-stock-qty",String(c)); // por defecto: todo en stock
    if(sv)sv.textContent=String(c);
    if(wrap)wrap.style.display=(c>1?"inline-block":"none");
  }} else {{
    tr.removeAttribute("data-stock");
    tr.removeAttribute("data-stock-qty");
    if(wrap)wrap.style.display="none";
    if(inp){{inp.dataset.val="0";inp.value="";}}
  }}
  _applyStockRow(tr);
  calc();checkSaveBtn();
}};
// Stepper de cantidad en stock (−/+). V = unidades en stock (de C). Si V<C quedan
// (C−V) por comprar; el precio real se habilita (comprar el resto es OPCIONAL, no
// se sale del modo inventario). Reemplaza las flechas nativas feas del input.
window.stockStep=function(btn,delta){{
  var tr=btn.closest("tr[data-idx]");
  if(!tr) return;
  var c=+tr.dataset.cant||1;
  var cur=parseInt(tr.getAttribute("data-stock-qty"))||c;
  var v=cur+delta; if(v<1)v=1; if(v>c)v=c;
  tr.setAttribute("data-stock-qty",String(v));
  var sv=tr.querySelector(".rc-stepval"); if(sv)sv.textContent=String(v);
  var inp=tr.querySelector(".rc-real");
  if(v<c && inp && inp.dataset.val==="0"){{inp.value="";}}
  _applyStockRow(tr);
  calc();checkSaveBtn();
}};
function _escH(s){{var d=document.createElement("div");d.textContent=s==null?"":String(s);return d.innerHTML;}}
// "Ver" inventario: lista los productos marcados en stock (item, categoría,
// cantidad, ahorro = presupuestado) en un modal, ordenados por mayor ahorro.
window.verInventario=function(){{
  var items=[];
  document.querySelectorAll("tr[data-idx]").forEach(function(r){{
    if(r.getAttribute("data-stock")!=="1")return;
    if((parseInt(r.dataset.idx)||0)>=10000)return;
    var pu=+r.dataset.pu||0,c=+r.dataset.cant||1;
    var sq=parseInt(r.getAttribute("data-stock-qty"))||c; if(sq>c)sq=c; if(sq<1)sq=1;
    items.push({{item:r.cells[1]?r.cells[1].textContent.trim():"",cat:r.cells[0]?r.cells[0].textContent.trim():"",cant:sq,ahorro:pu*sq}});
  }});
  items.sort(function(a,b){{return b.ahorro-a.ahorro;}});
  var body=document.getElementById("inv-modal-body");
  var foot=document.getElementById("inv-modal-foot");
  if(!items.length){{
    body.innerHTML='<div style="padding:26px 18px;text-align:center;color:#94a3b8;font-size:.9rem">A&#250;n no marcaste productos en stock. Usa la casilla <b>En stock</b> de cada &#237;tem que ya tengas.</div>';
    foot.innerHTML="";
  }} else {{
    var rows="",tot=0;
    items.forEach(function(it){{
      tot+=it.ahorro;
      rows+='<div style="display:flex;align-items:center;justify-content:space-between;gap:10px;padding:10px 18px;border-bottom:1px solid #f1f5f9">'
        +'<div style="min-width:0"><div style="font-weight:700;font-size:.9rem;color:#0f172a;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">'+_escH(it.item)+'</div>'
        +'<div style="font-size:.72rem;color:#64748b;text-transform:uppercase;letter-spacing:.04em;margin-top:2px">'+_escH(it.cat)+' &middot; '+it.cant+' u.</div></div>'
        +'<div style="text-align:right;white-space:nowrap"><div style="font-weight:800;font-size:.92rem;color:#16a34a">'+f(it.ahorro)+'</div>'
        +'<div style="font-size:.62rem;color:#16a34a;text-transform:uppercase;letter-spacing:.05em">Ahorro</div></div></div>';
    }});
    body.innerHTML=rows;
    foot.innerHTML='<span style="color:#166534;font-size:.82rem;text-transform:uppercase;letter-spacing:.04em">'+items.length+' producto'+(items.length!==1?"s":"")+'</span><span style="color:#16a34a;font-size:1.05rem">'+f(tot)+'</span>';
  }}
  document.getElementById("inv-modal").style.display="flex";
}};
window.cerrarInventario=function(){{var m=document.getElementById("inv-modal");if(m)m.style.display="none";}};
document.getElementById("factura-input") && document.getElementById("factura-input").addEventListener("change",function(){{
  _facturaFile=this.files[0]||null;
  var lbl=document.getElementById("factura-label");
  var clr=document.getElementById("factura-clear");
  if(_facturaFile){{
    if(lbl)lbl.innerHTML=ICO_CLIP+_facturaFile.name+'<input id="factura-input" type="file" accept=".pdf" style="display:none"/>';
    if(clr)clr.style.display="inline-flex";
  }}else{{
    if(lbl)lbl.innerHTML=ICO_CLIP+'Seleccionar factura PDF<input id="factura-input" type="file" accept=".pdf" style="display:none"/>';
    if(clr)clr.style.display="none";
  }}
  checkSaveBtn();
}});
window.clearFactura=function(){{
  _facturaFile=null;
  var lbl=document.getElementById("factura-label");
  var clr=document.getElementById("factura-clear");
  if(lbl)lbl.innerHTML=ICO_CLIP+'Seleccionar factura PDF<input id="factura-input" type="file" accept=".pdf" style="display:none"/>';
  if(clr)clr.style.display="none";
  var ni=document.getElementById("factura-input");
  if(ni)ni.addEventListener("change",function(){{
    _facturaFile=this.files[0]||null;
    var l2=document.getElementById("factura-label");
    var c2=document.getElementById("factura-clear");
    if(_facturaFile){{
      if(l2)l2.innerHTML=ICO_CLIP+_facturaFile.name+'<input id="factura-input" type="file" accept=".pdf" style="display:none"/>';
      if(c2)c2.style.display="inline-block";
    }}
    checkSaveBtn();
  }});
  checkSaveBtn();
}};
window.guardarRegistro=async function(){{
  var btn=document.getElementById("save-btn");
  var status=document.getElementById("save-status");
  btn.disabled=true;btn.textContent="Verificando compras previas...";
  // Ya comprados: Python ya lo pasa server-side (service key) en ITEMS_YA_COMPRADOS
  // → evitamos el fetch desde el navegador (que RLS bloquearía).
  var itemsYaComprados=(ITEMS_YA_COMPRADOS||[]).slice();
  var _tipoEl=document.getElementById("tipo-compra");
  var _stockMode=_tipoEl&&_tipoEl.value==="stock";
  var items=[];
  document.querySelectorAll("tr[data-idx]").forEach(function(r){{
    var idx=parseInt(r.dataset.idx)||0;
    var inp=r.querySelector(".rc-real");
    var re=parseFloat(inp.dataset.val)||0;
    var isStock=r.getAttribute("data-stock")==="1";
    var stockSaved=r.getAttribute("data-stock-saved")==="1"; // stock ya en BD
    var c=+r.dataset.cant||1;
    var sq=isStock?(parseInt(r.getAttribute("data-stock-qty"))||c):0;
    if(sq>c)sq=c; if(sq<0)sq=0;
    var buyAttr=r.getAttribute("data-buy");
    var comprUnits=(buyAttr!=null)?(parseInt(buyAttr)||0):(isStock?(c-sq):c);
    if(comprUnits<0)comprUnits=0;
    // Ítem normal ya COMPLETO → no re-guardar.
    if(idx<10000 && r.dataset.comprado==="1") return;
    // Modo INVENTARIO: solo ítems en stock.
    if(_stockMode && !isStock) return;
    // Fila normal sin precio y sin stock → nada que guardar.
    if(!isStock && re<=0) return;
    var pu=+r.dataset.pu||0;
    var ad=+r.querySelector(".rc-adic").value||0;
    var cat=r.cells[0]?r.cells[0].textContent.trim():"";
    var nom=r.cells[1]?r.cells[1].textContent.trim():"";
    var esAdic=idx>=10000||r.dataset.adicional==="1";
    var esSin=r.getAttribute("data-sin-registro")==="1";
    // Cada registro guarda unidades DISJUNTAS (evita doble conteo al completar):
    // ── entrada de STOCK (unidades en inventario, $0) — solo stock NUEVO ──
    if(isStock && !stockSaved && sq>0){{
      items.push({{categoria:cat,item:nom,cantidad:sq,precio_presupuestado:pu,precio_real:0,adicional:0,diferencia:pu*sq,es_adicional:false,sin_registro:false,stock:true,stock_cantidad:sq}});
    }}
    // ── entrada de COMPRA (unidades compradas ahora, con precio real) ──
    if(re>0 && comprUnits>0){{
      items.push({{categoria:cat,item:nom,cantidad:comprUnits,precio_presupuestado:pu,precio_real:re,adicional:ad,diferencia:(pu-re)*comprUnits-(ad*re),es_adicional:esAdic,sin_registro:esSin,stock:false}});
    }}
  }});
  if(items.length===0){{status.textContent="Marca ítems en stock o ingresa un precio real para comprar";status.style.color="#dc2626";btn.disabled=false;btn.textContent="Guardar compra";return;}}
  // ¿Hay al menos una compra real (>$0)? Solo entonces se exige/sube factura.
  // Un registro PURO de stock (todo $0, ahorro puro) no lleva factura.
  var hasReal=items.some(function(it){{return it.precio_real>0;}});
  if(hasReal && !_facturaFile){{status.textContent="Debes subir una factura primero";status.style.color="#dc2626";btn.disabled=false;btn.textContent="Guardar compra";return;}}
  status.textContent="";
  try{{
    if(_facturaFile){{
      btn.disabled=true;btn.textContent="Subiendo factura...";
      var ext=_facturaFile.name.split(".").pop();
      var path="cotizacion-"+EP_NUM+"/"+Date.now()+"."+ext;
      var uploadResp=await fetch(SUPA_URL+"/storage/v1/object/facturas/"+path,{{
        method:"POST",
        headers:{{"Authorization":"Bearer "+SUPA_KEY,"apikey":SUPA_KEY,"Content-Type":_facturaFile.type,"x-upsert":"true"}},
        body:_facturaFile
      }});
      if(!uploadResp.ok) throw new Error("Error subiendo factura: "+uploadResp.status);
      _facturaUrl=SUPA_URL+"/storage/v1/object/public/facturas/"+path;
      _facturaNom=_facturaFile.name;
    }}
    var tP=0,tR=0;
    items.forEach(function(it){{
      var _sq=it.stock?(it.stock_cantidad||it.cantidad):0;
      var _compr=it.cantidad-_sq;   // unidades compradas (stock: C-sq; normal: C)
      tP+=it.precio_presupuestado*it.cantidad;
      tR+=(it.precio_real*_compr)+(it.adicional*it.precio_real);
    }});
    btn.textContent="Guardando registro...";
    var lugarEl=document.getElementById("lugar-compra");
    var lugarVal=lugarEl?lugarEl.value.trim():"";
    var tipoVal=document.getElementById("tipo-compra")?document.getElementById("tipo-compra").value:"";
    var subtipoVal=document.getElementById("subtipo-compra")?document.getElementById("subtipo-compra").value:"";
    var fechaVal=document.getElementById("fecha-compra")?document.getElementById("fecha-compra").value:"";
    var faltóVal=document.getElementById("falto-texto")?document.getElementById("falto-texto").value.trim():"";
    var obsVal=document.getElementById("obs-compra")?document.getElementById("obs-compra").value.trim():"";
    // Guardado DIRECTO en Supabase con la anon key (RLS permite el INSERT, igual
    // que la subida de la factura). Es 100% fiable (fetch REST), sin depender de
    // que Streamlit lea un query param ni de reruns sintéticos (que fallaban de
    // forma intermitente). Tras insertar, se refresca la vista clickeando el botón
    // nativo oculto (que limpia la cache y re-renderiza el formulario/historial).
    btn.textContent="Guardando registro...";
    var _regData={{
      cotizacion_numero:EP_NUM,
      usuario_registro:USUARIO,
      lugar_compra:lugarVal,
      tipo_compra:tipoVal,
      subtipo_compra:subtipoVal,
      fecha_entrega_compra:fechaVal,
      falto_retirar:faltóVal,
      observaciones:obsVal,
      factura_url:_facturaUrl,
      factura_nombre:_facturaNom,
      items:items,
      total_presupuestado:tP,
      total_real:tR,
      balance:(tP-tR)
    }};
    var insResp=await fetch(SUPA_URL+"/rest/v1/registro_compras",{{
      method:"POST",
      headers:{{"apikey":SUPA_KEY,"Authorization":"Bearer "+SUPA_KEY,"Content-Type":"application/json","Prefer":"return=minimal"}},
      body:JSON.stringify(_regData)
    }});
    if(!insResp.ok){{
      var _et=""; try{{_et=await insResp.text();}}catch(e){{}}
      throw new Error("No se pudo guardar ("+insResp.status+") "+_et.slice(0,120));
    }}
    btn.textContent="Guardado";btn.style.background="#16a34a";
    status.textContent="Guardado. Actualizando...";
    status.style.color="#16a34a";
    setTimeout(function(){{
      try{{ var b=window.parent.document.querySelector('.st-key-_rc_apply button'); if(b) b.click(); }}catch(e){{}}
    }},250);
  }}catch(e){{
    btn.disabled=false;btn.textContent="Guardar compra";
    status.textContent="Error: "+e.message;status.style.color="#dc2626";
  }}
}};
calc();
}})();</script>"""
    html = html + _RC_ADD_MENU_JS + _RC_TABLE_TOOLS_JS + _RC_ADMIN_TOGGLE_JS + _RC_DETAILS_TOGGLE_JS
    return html
