"""
Gestión del Excel activo desde Supabase Storage y exportaciones.
Extraídas de app.py para la arquitectura modular.
"""
import io
import pandas as pd
import streamlit as st
from config.supabase import supabase_admin as _supa_admin


# ── EXCEL ACTIVO DESDE SUPABASE ───────────────────────────────────────────────

@st.cache_data(ttl=60, show_spinner=False)
def get_excel_bytes_activo():
    """Descarga el Excel activo desde Supabase Storage. Cache 60s."""
    try:
        _resp = _supa_admin.table('excel_versiones').select('archivo_url')\
            .eq('activa', True).limit(1).execute()
        if _resp.data:
            _url = _resp.data[0]['archivo_url']
            import requests as _rq
            _r = _rq.get(_url, timeout=15)
            _r.raise_for_status()
            return io.BytesIO(_r.content)
    except Exception:
        pass
    return "cotizador.xlsx"


@st.cache_data(ttl=300, show_spinner=False)
def _excel_src():
    """Retorna la fuente del Excel (BytesIO desde Supabase o path local)."""
    if 'excel_bytes_cache' not in st.session_state:
        st.session_state.excel_bytes_cache = get_excel_bytes_activo()
    return st.session_state.excel_bytes_cache


@st.cache_data(ttl=300, show_spinner=False)
def leer_hoja_excel(nombre_hoja):
    """Lee y cachea una hoja del Excel — evita re-parsear en cada render."""
    try:
        return pd.read_excel(_excel_src(), sheet_name=nombre_hoja)
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=300, show_spinner=False)
def leer_bd_total():
    """Lee y cachea la hoja BD Total."""
    try:
        return pd.read_excel(_excel_src(), sheet_name="BD Total")[["Item", "P. Unitario real"]]
    except Exception:
        return pd.DataFrame(columns=["Item", "P. Unitario real"])


def cargar_visibilidad_impresion():
    """Carga hoja Impresion del Excel activo. Retorna dict {item_lower: 'Mostrar'|'Ocultar'}."""
    try:
        import openpyxl as _opx
        _src = get_excel_bytes_activo()
        _wb = _opx.load_workbook(_src, data_only=True, read_only=True)
        if 'Impresion' not in _wb.sheetnames:
            return {}
        _ws = _wb['Impresion']
        _vis = {}
        for _row in _ws.iter_rows(min_row=2, values_only=True):
            _item, _cat, _pdf = (_row[0], _row[1], _row[2]) if len(_row) >= 3 else (None, None, None)
            if _item and _pdf:
                _vis[str(_item).strip().lower()] = str(_pdf).strip()
        return _vis
    except Exception:
        return {}


# ── EXPORTACIONES ─────────────────────────────────────────────────────────────

def exportar_csv_completo():
    """Exporta todas las cotizaciones de Supabase a CSV (bytes UTF-8-sig)."""
    try:
        response = _supa_admin.table('cotizaciones').select(
            'numero', 'fecha_creacion', 'fecha_modificacion',
            'cliente_nombre', 'cliente_rut', 'cliente_email', 'cliente_telefono',
            'cliente_direccion', 'cliente_comuna', 'cliente_region',
            'cliente_tipo', 'cliente_empresa', 'cliente_rut_empresa',
            'proyecto_direccion', 'proyecto_comuna', 'proyecto_region',
            'asesor_nombre', 'asesor_email', 'asesor_telefono',
            'config_margen',
            'total_subtotal_sin_margen', 'total_subtotal_con_margen',
            'total_iva', 'total_total', 'total_margen_valor',
            'total_comision_vendedor', 'total_comision_supervisor', 'total_utilidad_real',
            'estado', 'plano_nombre', 'plano_url',
            'contrato_generado', 'contrato_fecha'
        ).order('fecha_creacion', desc=True).execute()
        if not response.data:
            return None
        df = pd.DataFrame(response.data)
        df.columns = [
            'N° Presupuesto', 'Fecha Creación', 'Fecha Modificación',
            'Cliente', 'RUT', 'Email Cliente', 'Teléfono Cliente',
            'Dirección Cliente', 'Comuna Cliente', 'Región Cliente',
            'Tipo Cliente', 'Empresa', 'RUT Empresa',
            'Dirección Proyecto', 'Comuna Proyecto', 'Región Proyecto',
            'Asesor', 'Email Asesor', 'Teléfono Asesor',
            'Margen %',
            'Subtotal sin Margen', 'Subtotal con Margen',
            'IVA', 'Total con IVA', 'Valor Margen',
            'Comisión Vendedor', 'Comisión Supervisor', 'Utilidad Real',
            'Estado', 'Nombre Plano', 'URL Plano',
            'Contrato Generado', 'Fecha Contrato'
        ]
        return df.to_csv(index=False).encode('utf-8-sig')
    except Exception as e:
        st.error(f"Error al exportar: {e}")
        return None
