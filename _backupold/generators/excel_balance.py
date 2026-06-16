"""
Generador de Excel de balance de compras (precios reales vs presupuestados).
"""
import io
import json


def generar_excel_balance(cotizacion_numero, registros, productos_presupuesto, incluir_varios=False):
    """Genera Excel con precios reales consolidados por item para nutrir BD."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precios Reales"

    azul_oscuro = "1e2447"
    azul_claro  = "dbeafe"
    verde_claro = "f0fdf4"
    rojo_claro  = "fef2f2"
    gris_claro  = "f8fafc"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", fgColor=azul_oscuro)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    normal_font  = Font(name="Arial", size=9)
    bold_font    = Font(name="Arial", bold=True, size=9)
    thin_border  = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )

    headers   = ["Categoría", "Ítem", "Precio Unitario Presup.", "Precio Real Unitario", "Diferencia"]
    col_widths = [20, 45, 22, 22, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30

    items_consolidados = {}
    adicionales_con = []
    adicionales_sin = []
    if incluir_varios:
        prods_valid = list(productos_presupuesto or [])
    else:
        prods_valid = [p for p in (productos_presupuesto or [])
                       if str(p.get('Categoria', '')).strip().lower() != 'varios']
    pn_xls = {str(p.get('Item', '')) for p in prods_valid}

    for reg in registros:
        items_r = reg.get('items') or []
        if isinstance(items_r, str):
            try:
                items_r = json.loads(items_r)
            except Exception:
                items_r = []
        for it in items_r:
            nombre = str(it.get('item', ''))
            real = float(it.get('precio_real', 0) or 0)
            if real > 0 and nombre:
                if it.get('sin_registro'):
                    if not any(a['nombre'] == nombre for a in adicionales_sin):
                        adicionales_sin.append({'nombre': nombre, 'cat': it.get('categoria', ''), 'real': real})
                elif nombre not in pn_xls or it.get('es_adicional'):
                    if not any(a['nombre'] == nombre for a in adicionales_con):
                        adicionales_con.append({
                            'nombre': nombre, 'cat': it.get('categoria', ''), 'real': real,
                            'pp': float(it.get('precio_presupuestado', 0) or 0)
                        })
                else:
                    items_consolidados[nombre] = {
                        'categoria': it.get('categoria', ''),
                        'precio_presupuesto': float(it.get('precio_presupuestado', 0) or 0),
                        'precio_real': real,
                    }

    naranja = "fff3e0"; rosa = "fdf2f8"
    naranja_txt = "c2410c"; rosa_txt = "9d174d"

    row = 2
    for p in prods_valid:
        nombre = str(p.get('Item', ''))
        cat = str(p.get('Categoria', ''))
        pp = round(float(p.get('Precio Unitario', 0) or 0))
        pr = items_consolidados.get(nombre, {}).get('precio_real', 0)
        dif = pp - pr if pr > 0 else None
        bg = verde_claro if (dif is not None and dif >= 0 and pr > 0) else (rojo_claro if (dif is not None and dif < 0) else gris_claro)
        fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=row, column=1, value=cat).font = normal_font
        ws.cell(row=row, column=2, value=nombre).font = normal_font
        ws.cell(row=row, column=3, value=pp).font = normal_font
        ws.cell(row=row, column=4, value=pr if pr > 0 else "").font = bold_font if pr > 0 else normal_font
        ws.cell(row=row, column=5, value=dif if dif is not None else "").font = bold_font
        for col in range(1, 6):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.fill = fill
            c.alignment = Alignment(horizontal="left" if col <= 2 else "right", vertical="center")
        row += 1

    total_pp = sum(round(float(p.get('Precio Unitario', 0) or 0)) for p in prods_valid)
    total_pr = sum(v['precio_real'] for v in items_consolidados.values())
    for col, val in [(1, 'TOTAL PRESUPUESTO'), (3, total_pp), (4, total_pr), (5, total_pp - total_pr)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=azul_oscuro)
        c.border = thin_border
        c.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
    row += 1

    if adicionales_con:
        ws.cell(row=row, column=1, value="ADICIONALES CON REGISTRO").font = Font(name="Arial", bold=True, size=9, color=naranja_txt)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=naranja)
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=naranja)
            ws.cell(row=row, column=col).border = thin_border
        row += 1
        for a in adicionales_con:
            ws.cell(row=row, column=1, value=a['cat']).font = Font(name="Arial", size=9, color=naranja_txt)
            ws.cell(row=row, column=2, value=a['nombre']).font = Font(name="Arial", size=9, color=naranja_txt)
            ws.cell(row=row, column=3, value="—").font = normal_font
            ws.cell(row=row, column=4, value=a['real']).font = Font(name="Arial", bold=True, size=9, color=naranja_txt)
            ws.cell(row=row, column=5, value="—").font = normal_font
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=naranja)
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left" if col <= 2 else "right", vertical="center")
            row += 1

    if adicionales_sin:
        ws.cell(row=row, column=1, value="ADICIONALES SIN REGISTRO").font = Font(name="Arial", bold=True, size=9, color=rosa_txt)
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=rosa)
            ws.cell(row=row, column=col).border = thin_border
        row += 1
        for a in adicionales_sin:
            ws.cell(row=row, column=1, value=a['cat']).font = Font(name="Arial", size=9, color=rosa_txt)
            ws.cell(row=row, column=2, value=a['nombre']).font = Font(name="Arial", size=9, color=rosa_txt)
            ws.cell(row=row, column=3, value="—").font = normal_font
            ws.cell(row=row, column=4, value=a['real']).font = Font(name="Arial", bold=True, size=9, color=rosa_txt)
            ws.cell(row=row, column=5, value="—").font = normal_font
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=rosa)
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="left" if col <= 2 else "right", vertical="center")
            row += 1

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
